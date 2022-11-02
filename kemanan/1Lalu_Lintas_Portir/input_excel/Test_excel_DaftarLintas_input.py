
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
import pytest
import time
import platform
import os

class TestDaftarLaluLintas_Input():
    @pytest.fixture()
    def test_setup(self):
        global driver
        global wb
        swin = Service(r'C:/Users/user/Documents/TRCH/chromedriver.exe')
        smac = Service('/Users/will/Documents/chromedriver')

        if platform.system() == 'Darwin':
            driver = webdriver.Chrome(service=smac)
            # url = "http://kumbang.torche.id:32400/"
            url = "http://192.168.2.11:32400/"

            driver.get(url)
            # seting windows nya jadi max
            wb = load_workbook(filename=r"/Users/will/Documents/work/Automationpython/Filexel/Keamanan.xlsx")
            driver.maximize_window()
            driver.implicitly_wait(5)
            yield
            print(' ===== done pak will berhasil uyeee =====')
            time.sleep(5)
            driver.close()
            driver.quit()

        elif platform.system() == 'Windows':
            driver = webdriver.Chrome(service=swin)
            # url = "http://kumbang.torche.id:32400/"
            url = "http://192.168.2.11:32400/"

            driver.get(url)
            # seting windows nya jadi max   
            wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Filexel\Keamanan.xlsx")
            driver.maximize_window()
            driver.implicitly_wait(5)
            yield
            print(' ===== done pak will berhasil uyeee =====')
            time.sleep(5)
            driver.close()
            driver.quit()

    def test_DaftarLaluLintas_Input(self,test_setup):
        driver.implicitly_wait(10)
        sheetrange = wb ['DaftarLaluLintas_Input']
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div/div[2]/div/div/div[1]/canvas")))

        driver.find_element(By.XPATH, "//div/span").click()

        driver.find_element(By.ID, "username").click()
        driver.find_element(By.ID, "username").send_keys("test-user")
        driver.find_element(By.ID, "password").send_keys("password")
        driver.find_element(By.ID, "kc-login").click()
        print('login done')
        nav1 = driver.find_element(By.XPATH, '//*[@id="app"]/div/nav/ul/li[2]/div')
        actions = ActionChains(driver)
        actions.move_to_element(nav1).perform()

        element2 = driver.find_element(By.XPATH, "//div[4]/div/ul/li/div")
        time.sleep(1)
        actions2 = ActionChains(driver)
        actions2.move_to_element(element2).perform()
        time.sleep(1)

        driver.find_element(By.LINK_TEXT, 'Daftar Lalu Lintas').click()



        i = 2


        while i <= len(sheetrange['A']):

            Drpdownsearch                         = sheetrange['A'+str(i)].value
            Nama                                  = sheetrange['B'+str(i)].value
            JenisKeluar                           = sheetrange['C'+str(i)].value
            TanggalKeluar                         = sheetrange['D'+str(i)].value
            TanggalHarusKembali                   = sheetrange['E'+str(i)].value
            deskripsi                             = sheetrange['F'+str(i)].value



            try:
                time.sleep(1)
                WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="createButton"]')))
                driver.find_element(By.XPATH, '//*[@id="createButton"]').click()
                WebDriverWait(driver,10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".h-5 > path")))
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div/div[1]/label/div/div/div/input').click()
                if Drpdownsearch == 'Nama' :
                    driver.find_element(By.XPATH, "//li[contains(.,\'Nama\')]").click()

                elif Drpdownsearch == 'NoRegistrasi' :
                    driver.find_element(By.XPATH, "//li[contains(.,'No. Registrasi')]").click()

                WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div/div[1]/div/div/div/input')))
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div/div[1]/div/div/div/input').send_keys(Nama)

                WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div/div[1]/div/div/button')))
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div/div[1]/div/div/button').click()

                time.sleep(2)
                WebDriverWait(driver,10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".h-5 > path")))
                WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".h-5 > path")))
                driver.find_element(By.CSS_SELECTOR, ".h-5 > path").click()

                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/div[3]/div/div/div[2]/div/div/div/div/div[2]/button').click()

                WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="keluarKeamanan"]')))
                driver.find_element(By.XPATH, '//*[@id="keluarKeamanan"]').send_keys(TanggalKeluar)
                driver.find_element(By.XPATH, '//*[@id="keluarKeamanan"]').send_keys(Keys.ENTER)
                
                driver.find_element(By.XPATH, '//*[@id="tanggalKembali"]').send_keys(TanggalHarusKembali)
                driver.find_element(By.XPATH, '//*[@id="tanggalKembali"]').send_keys(Keys.ENTER)
                
                WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="jenisKeluar"]')))
                driver.find_element(By.XPATH, '//*[@id="jenisKeluar"]').click()
                time.sleep(1)
                if JenisKeluar == 'Cuti Bersyarat' :
                    driver.find_element(By.XPATH, "//li[contains(.,\'Cuti Bersyarat\')]").click()
                elif JenisKeluar == 'Bebas dari Tuntutan' :
                    driver.find_element(By.XPATH, "//li[contains(.,\'Bebas dari Tuntutan\')]").click()
                elif JenisKeluar == 'Cuti Menjelang Bebas':
                    driver.find_element(By.XPATH, "//li[contains(.,\'Cuti Menjelang Bebas\')]").click()

                driver.find_element(By.XPATH, '//*[@id="deskripsi"]').send_keys(deskripsi)
                
                time.sleep(3)
                WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="buttonSubmit"]')))
                driver.find_element(By.XPATH, '//*[@id="buttonSubmit"]').click()
                WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="createButton"]')))

            except TimeoutException:
                print("MASIH ADA ERROR, CEK LAGI PAK WIL")
                pass
            i = i + 1
        print("DONE PAK WILDAN, SEBATS DULU")
