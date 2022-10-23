from codecs import namereplace_errors
from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
import os
import pyautogui
import pytest
import time
import platform
import subprocess
from pathlib import Path

@pytest.fixture()
def test_setup():
    global driver
    global wb
    swin = Service(r'C:/Users/user/Documents/TRCH/chromedriver.exe')
    smac = Service('/Users/will/Downloads/chromedriver')

    if platform.system() == 'Darwin':
        driver = webdriver.Chrome(service=smac)
        # url = "http://kumbang.torche.id:32400/"
        url = "http://192.168.2.11:32400/"
        
        driver.get(url)
        # seting windows nya jadi max
        wb = load_workbook(filename=r"/Users/will/Documents/work/Automationpython/Filexel/Keamanan.xlsx")   
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
    elif platform.system() == 'Windows':
        driver = webdriver.Chrome(service=swin)
        # url = "http://kumbang.torche.id:32400/"
        url = "http://192.168.2.11:32400/"
        
        driver.get(url)
        # seting windows nya jadi max   
        wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Filexel\Keamanan.xlsx")
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()

def test_DaftarLaluLintas(test_setup):
    driver.implicitly_wait(10)
    sheetrange = wb ['DaftarLaluLintas']
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div/div[2]/div/div/div[1]/canvas")))
        
    driver.find_element(By.XPATH, "//div/span").click()
        
    driver.find_element(By.ID, "username").click()
    driver.find_element(By.ID, "username").send_keys("wildan")
    driver.find_element(By.ID, "password").send_keys("wildan")
    driver.find_element(By.ID, "kc-login").click()
        
    nav1 = driver.find_element(By.XPATH, '//*[@id="app"]/div/nav/ul/li[2]/div')
    actions = ActionChains(driver)
    actions.move_to_element(nav1).perform()

    element2 = driver.find_element(By.XPATH, "//div[4]/div/ul/li/div")
    time.sleep(1)
    actions2 = ActionChains(driver)
    actions2.move_to_element(element2).perform()
    time.sleep(1)

    driver.find_element(By.LINK_TEXT, 'Portir').click()
    
    i = 2

    while i == len(sheetrange['A']):
        nama                        = sheetrange['A'+str(i)].value



        try:
            time.sleep(1)


        except TimeoutException:
            print("MASIH ADA ERROR, CEK LAGI PAK WIL")
            pass
        time.sleep(5)
        i = i + 1
    print("DONE PAK WILDAN, SEBATS DULU")



     