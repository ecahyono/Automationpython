from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
import pytest
from openpyxl import load_workbook
import time
from selenium.webdriver.common.keys import Keys 
import pyautogui

@pytest.fixture()
def test_setup():
    global driver
    s = Service('/Users/will/Downloads/chromedriver')
    #s = Service()
    driver = webdriver.Chrome(service=s)
   # url = "http://kumbang.torche.id:32400/"
    url = "http://192.168.2.11:32400/"
    driver.get(url)
    #driver.maximize_window()
    driver.implicitly_wait(10)
    yield
    #driver.close()
    #driver.quit()
def test_BarangBawaan(test_setup):
    wb = load_workbook(filename=r"/Users/will/Documents/work/Automationpython/Filexel/Registrasi.xlsx")
    #wb = load_workbook(filename=r"C:\Users\wilda\Documents\Automationpython\Filexel/Registrasi.xlsx")
    
    sheetrange = wb['Catatan']
    WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, '//div/span')))
        
    driver.find_element(By.XPATH, "//div/span").click()
    # ini masuk ke form input username
    driver.find_element(By.ID, "username").click()
    # masukin input username
    driver.find_element(By.ID, "username").send_keys("wildan")
    # masukin input password
    driver.find_element(By.ID, "password").send_keys("wildan")
    # click button login
    driver.find_element(By.ID, "kc-login").click()
    time.sleep(1)
    element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[1]/div")
    time.sleep(1)
    actions = ActionChains(driver)
    time.sleep(1)
    actions.move_to_element(element).perform()

    driver.find_element(By.LINK_TEXT, "Catatan").click()
    #WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/div[1]/h1")))
       
    i = 2

    # nge baca mulai dari tabel A
    while i <= len(sheetrange['A']):
        # deklarasi bahwa NIP itu ada di A 
        NoInduk = sheetrange['A'+str(i)].value
        JenisCatatan = sheetrange['B'+str(i)].value
        TanggalKejadian = sheetrange['C'+str(i)].value
        Deskripsi = sheetrange['D'+str(i)].value
        
        
        #======================== Halaman Index ============================
        time.sleep(2)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/button").click()
        #WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div[1]/div/div/div/h1")))
        time.sleep(2)
        
        try:      
            #======================== Halaman Cari ===========================
            driver.find_element(By.CSS_SELECTOR, ".el-input__inner").click()
            time.sleep(1)
            driver.find_element(By.CSS_SELECTOR, ".el-input__inner").send_keys(NoInduk)
            time.sleep(3)
            #WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="el-popper-container-2384"]/div[3]/div/div')))
            
            driver.find_element(By.CSS_SELECTOR, ".el-select-dropdown__item:nth-child(1) tr:nth-child(2) > .el-descriptions__cell:nth-child(1)").click()
            time.sleep(1)
            #======================== Button Cari ============================
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div[1]/div/form/button[1]").click()
            #======================== masuk ke halaman input ============================
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[1]/div/div/div/div/input').click()
            if JenisCatatan == 'Catatan Kesehatan' :
                dropdown = driver.find_element(By.XPATH, "//li[contains(.,'Catatan Kesehatan')]")
                dropdown.find_element(By.XPATH, "//li[contains(.,'Catatan Kesehatan')]").click()
            elif JenisCatatan == 'Catatan Perkembangan Pembinaan' :
                dropdown = driver.find_element(By.XPATH, "//span[contains(.,\'Catatan Perkembangan Pembinaan\')]")
                dropdown.find_element(By.XPATH, "//span[contains(.,\'Catatan Perkembangan Pembinaan\')]").click()
            elif JenisCatatan == 'Catatan Prestasi' :
                dropdown = driver.find_element(By.XPATH, "//span[contains(.,\'Catatan Prestasi\')]")
                dropdown.find_element(By.XPATH, "//span[contains(.,\'Catatan Prestasi\')]").click()
            elif JenisCatatan == 'Catatan Penghargaan' :
                dropdown = driver.find_element(By.XPATH, "//span[contains(.,\'Catatan Penghargaan\')]")
                dropdown.find_element(By.XPATH, "//span[contains(.,\'Catatan Penghargaan\')]").click()
            elif JenisCatatan == 'Catatan Barang Titipan' :
                dropdown = driver.find_element(By.XPATH, "//span[contains(.,\'Catatan Barang Titipan\')]")
                dropdown.find_element(By.XPATH, "//span[contains(.,\'Catatan Barang Titipan\')]").click()
            elif JenisCatatan == 'Catatan Lain-Lain' :
                dropdown = driver.find_element(By.XPATH, "//span[contains(.,\'Catatan Lain-Lain\')]")
                dropdown.find_element(By.XPATH, "//span[contains(.,\'Catatan Lain-Lain\')]").click()

            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[2]/div/div[1]/input').send_keys(TanggalKejadian)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[2]/div/div[1]/input').send_keys(Keys.ENTER)

            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[3]/div/div/textarea').send_keys(Deskripsi)
            
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[4]/button[2]').click()
           
        except TimeoutException:
            print("MASIH ADA ERROR, CEK LAGI PAK WIL")
            pass
        time.sleep(5)
        i = i + 1
    print("DONE PAK WILDAN, SEBATS DULU")
