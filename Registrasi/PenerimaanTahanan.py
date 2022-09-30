from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
import pytest
from openpyxl import load_workbook
import time

@pytest.fixture()
def test_setup():
    global driver
    #s = Service('/Users/will/Downloads/chromedriver')
    s = Service()
    driver = webdriver.Chrome(service=s)
    url = "http://kumbang.torche.id:32400/"
    driver.get(url)
    driver.maximize_window()
    driver.implicitly_wait(10)
    yield
    #driver.close()
    #driver.quit()
def test_negara(test_setup):
    
    #wb = load_workbook(filename=r"/Users/will/Documents/Automationpython/Filexel/Registrasi.xlsx")
    wb = load_workbook(filename=r"C:\Users\wilda\Documents\Automationpython\Filexel/Registrasi.xlsx")
    
    sheetrange = wb['reg']
    WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, '//div/span')))
        
    driver.find_element(By.XPATH, "//div/span").click()
    # ini masuk ke form input username
    driver.find_element(By.ID, "username").click()
    # masukin input username
    driver.find_element(By.ID, "username").send_keys("test-user")
    # masukin input password
    driver.find_element(By.ID, "password").send_keys("password")
    # click button login
    driver.find_element(By.ID, "kc-login").click()
    time.sleep(1)
    #======================== Menu Registrasi ============================
    element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[1]/div")
    time.sleep(1)
    actions = ActionChains(driver)
    actions.move_to_element(element).perform()
    time.sleep(1)
    #======================== menu Penerimaan dan Penolakan ============================
    element2 = driver.find_element(By.XPATH, "//div/ul/li/div")
    time.sleep(2)
    actions2 = ActionChains(driver)
    actions2.move_to_element(element2).perform()
    driver.find_element(By.LINK_TEXT, "Penerimaan").click()

    #======================== Halaman Create ============================

    WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/div[1]/h1")))
    
    i = 2

    # nge baca mulai dari tabel A
    while i <= len(sheetrange['A']):
        # deklarasi bahwa NIP itu ada di A 
        NoInduk = sheetrange['A'+str(i)].value
        noregis = sheetrange['B'+str(i)].value
        TglSuratPenahanan = sheetrange['C'+str(i)].value
        NomorSuratPenahanan = sheetrange['D'+str(i)].value
        NamaPetugasInstansi = sheetrange['E'+str(i)].value
        Kejaksaan = sheetrange['F'+str(i)].value
        AsalInstansi = sheetrange['G'+str(i)].value
        Keterangan = sheetrange['H'+str(i)].value
        Penyidik = sheetrange['I'+str(i)].value
        menahan10 = sheetrange['J'+str(i)].value
        menahan3 = sheetrange['K'+str(i)].value
        menahan1 = sheetrange['L'+str(i)].value
        LokasiDokumen = sheetrange['M'+str(i)].value

        #======================== Halaman Index ============================
        time.sleep(2)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/div[2]/button[2]").click()
        time.sleep(2)
        
        try:      
            #======================== Halaman Cari ============================
            WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div[1]/div/form/button[2]")))
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[1]/div/div/input").click()
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[1]/div/div/input").send_keys("22/05/2021 10:10:10")
        except TimeoutException:
            print("MASIH ADA ERROR, CEK LAGI PAK WIL")
            pass
        time.sleep(5)
        i = i + 1
    print("DONE PAK WILDAN, SEBATS DULU")