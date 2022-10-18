from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
import os
import pyautogui
import pytest
import time
import platform
from pathlib import Path

@pytest.fixture()
def test_setup():
    global driver
    global wb
    swin = Service(r'C:/Users/user/Documents/TRCH/chromedriver.exe')
    smac = Service('/Users/will/Downloads/chromedriver')
    if platform.system() == 'Darwin':
        driver = webdriver.Chrome(service=smac)
        # url = 'http://kumbang.torche.id:32400/'
        url = 'http://192.168.2.11:32400/'
        
        driver.get(url)
        # seting windows nya jadi max
        wb = load_workbook(filename='/Users/will/Documents/work/Automationpython/Filexel/Registrasi.xlsx')   
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
    elif platform.system() == 'Windows':
        driver = webdriver.Chrome(service=swin)
        # url = 'http://kumbang.torche.id:32400/'
        url = 'http://192.168.2.11:32400/'
        
        driver.get(url)
        # seting windows nya jadi max   
        wb = load_workbook(filename=r'C:\Users\user\Documents\TRCH\Automationpython\Filexel\Registrasi.xlsx')
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
def test_Web(test_setup):
    # read excel
    sheetrange = wb ['Identitas']
    # Menuju login
    driver.find_element(By.XPATH, '//div/span').click()
    driver.find_element(By.ID, 'username').click()
    driver.find_element(By.ID, 'username').send_keys('waru')
    driver.find_element(By.ID, 'password').send_keys('waru')
    # click button login
    driver.find_element(By.ID, 'kc-login').click()
    time.sleep(5) 
    #Registrasi
    element = driver.find_element(By.XPATH, '//*[@id="app"]/div/nav/ul/li[1]/div')                                   
    time.sleep(1)
    actions = ActionChains(driver)
    time.sleep(1)
    actions.move_to_element(element).perform()
    time.sleep(1)
    #Identitas
    driver.find_element(By.LINK_TEXT, 'Daftar Identitas').click()
    time.sleep(1)

    i = 3
    while i <= len(sheetrange['A']):
        #deklarasi per colom pada sheet
        #------------------------------------------------------
        #Tab Biodata-------------------------------------------
        #------------------------------------------------------
        Residivis               = sheetrange['A'+str(i)].value
        Rke                     = sheetrange['B'+str(i)].value
        Nama_Lengkap            = sheetrange['C'+str(i)].value
        Nama_Alias1             = sheetrange['D'+str(i)].value
        Nama_Alias2             = sheetrange['E'+str(i)].value
        Nama_Alias3             = sheetrange['F'+str(i)].value
        Nama_Kecil1             = sheetrange['G'+str(i)].value
        Nama_Kecil2             = sheetrange['H'+str(i)].value
        Nama_Kecil3             = sheetrange['I'+str(i)].value
        chcktab1                = sheetrange['J'+str(i)].value
        chcktab2                = sheetrange['K'+str(i)].value
        Kewarganegaraan         = sheetrange['L'+str(i)].value
        nik                     = sheetrange['M'+str(i)].value
        Tempat_Asal             = sheetrange['N'+str(i)].value
        Tempat_lahir            = sheetrange['O'+str(i)].value
        Tanggal_lahir           = sheetrange['P'+str(i)].value
        Jenis_kelamin           = sheetrange['Q'+str(i)].value
        Negara                  = sheetrange['R'+str(i)].value
        Agama                   = sheetrange['S'+str(i)].value
        Agama_lain              = sheetrange['T'+str(i)].value
        suku                    = sheetrange['U'+str(i)].value
        Status_perkawinan       = sheetrange['v'+str(i)].value
        Provinsi                = sheetrange['W'+str(i)].value
        Kota                    = sheetrange['X'+str(i)].value
        Alamat_rumah            = sheetrange['Y'+str(i)].value
        Telepon                 = sheetrange['Z'+str(i)].value
        Kode_pos                = sheetrange['AA'+str(i)].value
        Alamat_lain             = sheetrange['AB'+str(i)].value
        #-------------------------------------------------------
        #Tab Pekerjaan------------------------------------------
        #-------------------------------------------------------
        Jenis_Pekerjaan         = sheetrange['AC'+str(i)].value
        namaipemerintah         = sheetrange['AD'+str(i)].value
        noindpegawai            = sheetrange['AE'+str(i)].value
        Bekerjadi               = sheetrange['AF'+str(i)].value
        Keterangan_pekerjaan    = sheetrange['AG'+str(i)].value
        Tingkat_penghasilan     = sheetrange['AH'+str(i)].value
        Tingkat_pendidikan      = sheetrange['AI'+str(i)].value
        Keahlian1               = sheetrange['AJ'+str(i)].value
        Level_keahlian1         = sheetrange['AK'+str(i)].value
        Keahlian2               = sheetrange['AL'+str(i)].value
        Levelkeahlian2          = sheetrange['AM'+str(i)].value
        Minat                   = sheetrange['AN'+str(i)].value
        latin                   = sheetrange['AO'+str(i)].value
        quran                   = sheetrange['AP'+str(i)].value
        Jenis_Pekerjaan_Lain    = sheetrange['CG'+str(i)].value
        #--------------------------------------------------------
        #Tab Keluarga--------------------------------------------
        #--------------------------------------------------------
        Nama_ayah               = sheetrange['AQ'+str(i)].value
        Alamat_ayah             = sheetrange['AR'+str(i)].value
        Nama_ibu                = sheetrange['AS'+str(i)].value
        Alamat_ibu              = sheetrange['AT'+str(i)].value
        Anak_ke                 = sheetrange['AU'+str(i)].value
        Dari                    = sheetrange['AV'+str(i)].value
        Nama_saudara1           = sheetrange['AW'+str(i)].value
        Nama_saudara2           = sheetrange['AX'+str(i)].value
        Nama_saudara3           = sheetrange['AY'+str(i)].value
        Nama_saudara4           = sheetrange['AZ'+str(i)].value
        jml_istrsuam            = sheetrange['BA'+str(i)].value
        Nm_istrsuam             = sheetrange['BB'+str(i)].value
        alm_istrsuam            = sheetrange['BC'+str(i)].value
        jumlah_anak             = sheetrange['BD'+str(i)].value
        Nama_anak1              = sheetrange['BE'+str(i)].value
        Nama_anak2              = sheetrange['BF'+str(i)].value
        Nama_anak3              = sheetrange['CH'+str(i)].value
        Telepon_keluarga        = sheetrange['BG'+str(i)].value
        #--------------------------------------------------------
        #Tab Data Fisik------------------------------------------
        # -------------------------------------------------------
        Tinggi_badan            = sheetrange['BH'+str(i)].value
        Berat_badan             = sheetrange['BI'+str(i)].value
        Bentuk_rambut           = sheetrange['BJ'+str(i)].value
        Warna_rambut            = sheetrange['BK'+str(i)].value
        Bentuk_bibir            = sheetrange['BL'+str(i)].value
        Berkacamata             = sheetrange['BM'+str(i)].value
        Bentuk_mata             = sheetrange['BN'+str(i)].value
        Warna_mata              = sheetrange['BO'+str(i)].value
        Hidung                  = sheetrange['BP'+str(i)].value
        Raut_muka               = sheetrange['BQ'+str(i)].value
        Telinga                 = sheetrange['BR'+str(i)].value
        Mulut                   = sheetrange['BS'+str(i)].value
        Lengan                  = sheetrange['BT'+str(i)].value
        Tangan                  = sheetrange['BU'+str(i)].value
        Kaki                    = sheetrange['BV'+str(i)].value
        Warna_kulit             = sheetrange['BW'+str(i)].value
        Cacat_tubuh             = sheetrange['BX'+str(i)].value
        Catatancirikhusus1      = sheetrange['BY'+str(i)].value
        Catatancirikhusus2      = sheetrange['BZ'+str(i)].value
        Catatancirikhusus3      = sheetrange['CA'+str(i)].value
        #---------------------------------------------------------
        #Tab Sidik Jari-------------------------------------------
        # --------------------------------------------------------
        Nopaspor                = sheetrange['CB'+str(i)].value
        Rumus                   = sheetrange['CC'+str(i)].value
        Nodaktolskopi           = sheetrange['CD'+str(i)].value
        Pengambilansidikjari    = sheetrange['CE'+str(i)].value
        Tanggalpengambilan      = sheetrange['CF'+str(i)].value
        time.sleep(5) 
        #button +Tambah
        driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/div[1]/button').click()                                 
        try:
            WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/div/div/h1')))
            #======================================================================
            driver.find_element(By.ID, 'tab-1').click()
            #========================Input Tab Biodata ============================
            driver.find_element(By.ID, 'btn_residivis').click()
            time.sleep(2)
            driver.find_element(By.ID, 'btn_residivis').send_keys(Residivis)
            driver.find_element(By.ID, 'btn_residivis').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'btn_residivis').send_keys(Keys.ENTER)
            

            if Residivis == 'Ya':
                driver.find_element(By.XPATH, '//*[@id="btn_residivis_counter"]/div/input').click()
                pyautogui.hotkey('backspace')
                driver.find_element(By.XPATH, '//*[@id="btn_residivis_counter"]/div/input').send_keys(Rke)
            elif Residivis == 'Tidak':
                print ('Residivis yang ke', Rke)
            #--------------------------------------------------------------                                      
            driver.find_element(By.ID, 'btn_nama_lengkap').send_keys(Nama_Lengkap)
            #----------------------------------------------------------------
            # driver.find_element(By.ID, 'btn_nama_alias1').send_keys(Nama_Alias1)
            # # #--------------------------------------------------------------
            # driver.find_element(By.ID, 'btn_nama_alias2').send_keys(Nama_Alias2)
            # # #--------------------------------------------------------------
            # driver.find_element(By.ID, 'btn_nama_alias3').send_keys(Nama_Alias3)
            # # #--------------------------------------------------------------
            # driver.find_element(By.ID, 'btn_nama_kecil1').send_keys(Nama_Kecil1)
            # # #---------------------------------------------------------------
            # driver.find_element(By.ID, 'btn_nama_kecil2').send_keys(Nama_Kecil2)
            # # #----------------------------------------------------------------
            # driver.find_element(By.ID, 'btn_nama_kecil3').send_keys(Nama_Kecil3)
            # # #--------------------------------------------------------------
            if chcktab1 == 'tidak' :
                driver.find_element(By.ID, 'btn_is_wbp_beresiko_tinggi').click()
                print ('uncheckh')
            elif chcktab1 == 'ya' :
                print ('wbp_beresiko_tinggi Masih default')
            if chcktab2 == 'tidak' :
                driver.find_element(By.ID, 'btn_is_pengaruh_terhadap_masyarakat').click()
                print ('uncheckh')
            elif chcktab2 == 'ya' :
                print ('pengaruh_terhadap_masyarakat Masih default')
            # --------------------------------------------------------------
            driver.find_element(By.ID, 'btn_id_jenis_warganegara').click()
            time.sleep(2) 
            if Kewarganegaraan == 'WNI':
                pyautogui.typewrite(Kewarganegaraan)
                pyautogui.keyDown('down')
                pyautogui.press('enter')
            elif Kewarganegaraan == 'WNA':
                pyautogui.typewrite(Kewarganegaraan)
                pyautogui.keyDown('down')
                pyautogui.press('enter')
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'btn_nik').send_keys(nik) 
            #-------------------------------------------------------------- 
            if Kewarganegaraan == 'WNI':
                driver.find_element(By.ID, 'btn_id_tempat_asal').click()   
                time.sleep(2)
                driver.find_element(By.ID, 'btn_id_tempat_asal').send_keys(Tempat_Asal)                                
                driver.find_element(By.ID, 'btn_id_tempat_asal').send_keys(Keys.DOWN)                                
                driver.find_element(By.ID, 'btn_id_tempat_asal').send_keys(Keys.ENTER)
            elif Kewarganegaraan == 'WNA':
                # driver.find_element(By.ID, 'btn_id_tempat_asal_lain').click()   
                driver.find_element(By.ID, 'btn_id_tempat_asal_lain').send_keys(Tempat_Asal)
                                        
            # --------------------------------------------------------------
            time.sleep(2) 
            if Kewarganegaraan == 'WNI':
                driver.find_element(By.ID, 'btn_id_tempat_lahir').click()          
                time.sleep(2)   
                driver.find_element(By.ID, 'btn_id_tempat_lahir').send_keys(Tempat_lahir)          
                driver.find_element(By.ID, 'btn_id_tempat_lahir').send_keys(Keys.DOWN)        
                driver.find_element(By.ID, 'btn_id_tempat_lahir').send_keys(Keys.ENTER)          
            elif Kewarganegaraan == 'WNA':
                # driver.find_element(By.ID, '').click()          
                driver.find_element(By.ID, 'btn_id_tempat_lahir_lain').send_keys(Tempat_lahir)
            #------untuk tanggal Data format exel di sesuaikan-----------------------------
            driver.find_element(By.XPATH, '//div[5]/div/div/div/div/div/input').click()
            time.sleep(2) 
            driver.find_element(By.XPATH, '//div[5]/div/div/div/div/div/input').send_keys(Tanggal_lahir)
            driver.find_element(By.XPATH, '//div[5]/div/div/div/div/div/input').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'btn_id_jenis_kelamin').click()
            time.sleep(2) 
            driver.find_element(By.ID, 'btn_id_jenis_kelamin').send_keys(Jenis_kelamin)
            driver.find_element(By.ID, 'btn_id_jenis_kelamin').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'btn_id_jenis_kelamin').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'btn_id_negara_asing').click()
            time.sleep(2)
            driver.find_element(By.ID, 'btn_id_negara_asing').send_keys(Negara)
            driver.find_element(By.ID, 'btn_id_negara_asing').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'btn_id_negara_asing').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'btn_id_jenis_agama').click()
            time.sleep(2)
            driver.find_element(By.ID, 'btn_id_jenis_agama').send_keys(Agama)
            driver.find_element(By.ID, 'btn_id_jenis_agama').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'btn_id_jenis_agama').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            if Agama == 'Lain-lain':
                driver.find_element(By.ID, 'btn_id_jenis_agama_lain').click()
                driver.find_element(By.ID, 'btn_id_jenis_agama_lain').send_keys(Agama_lain)
            elif Agama == Agama:
                print ('agama masih default')
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'btn_id_jenis_suku').click()
            time.sleep(2)
            driver.find_element(By.ID, 'btn_id_jenis_suku').send_keys(suku) 
            driver.find_element(By.ID, 'btn_id_jenis_suku').send_keys(Keys.DOWN) 
            driver.find_element(By.ID, 'btn_id_jenis_suku').send_keys(Keys.ENTER) 
            #------------------------------------------------------------------------------
            driver.find_element(By.ID, 'btn_id_jenis_status_perkawinan').click()
            time.sleep(2) 
            driver.find_element(By.ID, 'btn_id_jenis_status_perkawinan').send_keys(Status_perkawinan)
            driver.find_element(By.ID, 'btn_id_jenis_status_perkawinan').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'btn_id_jenis_status_perkawinan').send_keys(Keys.ENTER)
            #------------------------------------------------------------------------------
            if Kewarganegaraan == 'WNI':
                driver.find_element(By.ID, 'btn_id_propinsi').click()
                time.sleep(2)
                driver.find_element(By.ID, 'btn_id_propinsi').send_keys(Provinsi)
                driver.find_element(By.ID, 'btn_id_propinsi').send_keys(Keys.DOWN)
                driver.find_element(By.ID, 'btn_id_propinsi').send_keys(Keys.ENTER)
            elif Kewarganegaraan == 'WNA':
                # driver.find_element(By.ID, 'btn_id_propinsi_lain').click()
                driver.find_element(By.ID, 'btn_id_propinsi_lain').send_keys(Provinsi)
            #------------------------------------------------------------------------------
            if Kewarganegaraan == 'WNI':
                driver.find_element(By.ID, 'btn_id_kota').click()
                time.sleep(2) 
                driver.find_element(By.ID, 'btn_id_kota').send_keys(Kota)
                driver.find_element(By.ID, 'btn_id_kota').send_keys(Keys.DOWN)
                driver.find_element(By.ID, 'btn_id_kota').send_keys(Keys.ENTER)
            elif Kewarganegaraan == 'WNA':
                # driver.find_element(By.ID, 'btn_id_kota_lain').click()
                driver.find_element(By.ID, 'btn_id_kota_lain').send_keys(Kota)
            #------------------------------------------------------------------------------
            driver.find_element(By.ID, 'btn_alamat').click()       
            pyautogui.typewrite(Alamat_rumah)
            # # ------------------------------------------------------------------------------
            # driver.find_element(By.ID, 'btn_telepon').click()
            driver.find_element(By.ID, 'btn_telepon').send_keys(Telepon)
            # #------------------------------------------------------------------------------
            # driver.find_element(By.ID, 'btn_kodepos').click()
            driver.find_element(By.ID, 'btn_kodepos').send_keys(Kode_pos)
            # # ------------------------------------------------------------------------------
            # driver.find_element(By.ID, 'btn_alamat_alternatif').click()
            driver.find_element(By.ID, 'btn_alamat_alternatif').send_keys(Alamat_lain)
            # ======================================================================
            driver.find_element(By.ID, 'tab-2').click()
            # ========================Input Tab Pekerjaan===========================
            driver.find_element(By.ID, 'id_jenis_pekerjaan').click()
            time.sleep(2) 
            driver.find_element(By.ID, 'id_jenis_pekerjaan').send_keys(Jenis_Pekerjaan)
            driver.find_element(By.ID, 'id_jenis_pekerjaan').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_jenis_pekerjaan').send_keys(Keys.ENTER)
            # --------------------------------------------------------------
            if Jenis_Pekerjaan == 'pegawai negeri sipil':
                # driver.find_element(By.ID, 'nama_instansi_pns').click()
                driver.find_element(By.ID, 'nama_instansi_pns').send_keys(namaipemerintah)
                # driver.find_element(By.ID, 'nip').click()
                driver.find_element(By.ID, 'nip').send_keys(noindpegawai)        
            elif Jenis_Pekerjaan == 'lain-lain':
                driver.find_element(By.ID, 'id_jenis_pekerjaan_lain').click()
                driver.find_element(By.ID, 'id_jenis_pekerjaan_lain').send_keys(Jenis_Pekerjaan_Lain)
            # -----------------------------------------------------------
            # driver.find_element(By.ID, 'alamat_pekerjaan').click()
            driver.find_element(By.ID, 'alamat_pekerjaan').send_keys(Bekerjadi)
            #--------------------------------------------------------------
            # driver.find_element(By.ID, 'keterangan_pekerjaan').click()
            driver.find_element(By.ID, 'keterangan_pekerjaan').send_keys(Keterangan_pekerjaan)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_tingkat_penghasilan').click()        
            time.sleep(2) 
            driver.find_element(By.ID, 'id_tingkat_penghasilan').send_keys(Tingkat_penghasilan)
            driver.find_element(By.ID, 'id_tingkat_penghasilan').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_tingkat_penghasilan').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_jenis_pendidikan').click()        
            time.sleep(2) 
            driver.find_element(By.ID, 'id_jenis_pendidikan').send_keys(Tingkat_pendidikan)
            driver.find_element(By.ID, 'id_jenis_pendidikan').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_jenis_pendidikan').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_jenis_keahlian_1').click()        
            time.sleep(2)
            driver.find_element(By.ID, 'id_jenis_keahlian_1').send_keys(Keahlian1)
            driver.find_element(By.ID, 'id_jenis_keahlian_1').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_jenis_keahlian_1').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_jenis_level_1').click()        
            time.sleep(2)
            driver.find_element(By.ID, 'id_jenis_level_1').send_keys(Level_keahlian1)
            driver.find_element(By.ID, 'id_jenis_level_1').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_jenis_level_1').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_jenis_keahlian_2').click()
            time.sleep(2)
            driver.find_element(By.ID, 'id_jenis_keahlian_2').send_keys(Keahlian2)
            driver.find_element(By.ID, 'id_jenis_keahlian_2').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_jenis_keahlian_2').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_jenis_level_2').click()
            time.sleep(2)        
            driver.find_element(By.ID, 'id_jenis_level_2').send_keys(Levelkeahlian2)
            driver.find_element(By.ID, 'id_jenis_level_2').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_jenis_level_2').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            if quran == 'ya' :
                driver.find_element(By.XPATH, '//*[@id="is_baca_quran"]/span[1]/span').click()
                print (quran)
            elif quran == 'Tidak':
                print ('tidak di check')
            # #----------------------------------------------------------------------
            if latin == 'ya' :
                driver.find_element(By.XPATH, '//*[@id="is_baca_latin"]').click()
                print (latin)
            elif latin == 'Tidak':
                print ('tidak di check')
            #----------------------------------------------------------------------
            # driver.find_element(By.ID, 'minat').click()
            driver.find_element(By.ID, 'minat').send_keys(Minat)
            #======================================================================
            driver.find_element(By.ID, 'tab-3').click()
            #========================Input Tab Keluarga============================ 
            #------------------------------------------------------------------------------
            # driver.find_element(By.ID, 'nm_ayah').click()        
            driver.find_element(By.ID, 'nm_ayah').send_keys(Nama_ayah)
            #------------------------------------------------------------------------------
            # driver.find_element(By.ID, 'tmp_tgl_ayah').click()      
            driver.find_element(By.ID, 'tmp_tgl_ayah').send_keys(Alamat_ayah)
            #------------------------------------------------------------------------------
            # driver.find_element(By.ID, 'nm_ibu').click()
            driver.find_element(By.ID, 'nm_ibu').send_keys(Nama_ibu)
            #------------------------------------------------------------------------------
            # driver.find_element(By.ID, 'tmp_tgl_ibu').click()        
            driver.find_element(By.ID, 'tmp_tgl_ibu').send_keys(Alamat_ibu)
            #------------------------------------------------------------------------------
            driver.find_element(By.XPATH, '//*[@id="anakke"]/div/input').click()
            pyautogui.hotkey('backspace')        
            driver.find_element(By.XPATH, '//*[@id="anakke"]/div/input').send_keys(Anak_ke)
            pyautogui.press('enter')
            driver.find_element(By.XPATH, '//*[@id="jml_saudara"]/div/input').click()
            pyautogui.hotkey('backspace')        
            driver.find_element(By.XPATH, '//*[@id="jml_saudara"]/div/input').send_keys(Dari)
            pyautogui.press('enter')
            if Dari == 2:
                # driver.find_element(By.ID, 'nm_saudara_1').click()    
                driver.find_element(By.ID, 'nm_saudara_1').send_keys(Nama_saudara1)
            elif Dari == 3:
                # driver.find_element(By.ID, 'nm_saudara_1').click()    
                driver.find_element(By.ID, 'nm_saudara_1').send_keys(Nama_saudara1)
                # driver.find_element(By.ID, 'nm_saudara_2').click()    
                driver.find_element(By.ID, 'nm_saudara_2').send_keys(Nama_saudara2)
            elif Dari == 4:
                # driver.find_element(By.ID, 'nm_saudara_1').click()    
                driver.find_element(By.ID, 'nm_saudara_1').send_keys(Nama_saudara1)
                # driver.find_element(By.ID, 'nm_saudara_2').click()    
                driver.find_element(By.ID, 'nm_saudara_2').send_keys(Nama_saudara2)
                # driver.find_element(By.ID, 'nm_saudara_3').click()    
                driver.find_element(By.ID, 'nm_saudara_3').send_keys(Nama_saudara3)
            elif Dari == 5:
                # driver.find_element(By.ID, 'nm_saudara_1').click()    
                driver.find_element(By.ID, 'nm_saudara_1').send_keys(Nama_saudara1)
                # driver.find_element(By.ID, 'nm_saudara_2').click()    
                driver.find_element(By.ID, 'nm_saudara_2').send_keys(Nama_saudara2)
                # driver.find_element(By.ID, 'nm_saudara_3').click()    
                driver.find_element(By.ID, 'nm_saudara_3').send_keys(Nama_saudara3)
                # driver.find_element(By.ID, 'nm_saudara_4').click()    
                driver.find_element(By.ID, 'nm_saudara_4').send_keys(Nama_saudara4)
            elif Dari == 1 :
                print ('anak satu satunya')
            if Status_perkawinan == 'Belum Kawin':
                print('default')
            elif Status_perkawinan == 'Duda':
                driver.find_element(By.ID, 'jml_istri_suami').click()
                pyautogui.hotkey('backspace')
                driver.find_element(By.XPATH, '//*[@id="jml_istri_suami"]/div/input' ).send_keys(jml_istrsuam) 
                pyautogui.press('enter')
                driver.find_element(By.ID, 'nm_istri_suami_1').click()
                pyautogui.typewrite(Nm_istrsuam)
                driver.find_element(By.ID, 'tmp_tgl_istri_suami').click()
                pyautogui.typewrite(alm_istrsuam)
                driver.find_element(By.XPATH, '//*[@id="jml_anak"]/div/input').click()
                pyautogui.hotkey('backspace')
                jumlah = driver.find_element(By.XPATH, '//*[@id="jml_anak"]/div/input' ).send_keys(jumlah_anak) 
                pyautogui.press('enter')
                if jumlah_anak == 1:
                    # driver.find_element(By.ID, 'nm_anak_1').click()    
                    driver.find_element(By.ID, 'nm_anak_1').send_keys(Nama_anak1)
                elif jumlah_anak == 2:
                    # driver.find_element(By.ID, 'nm_anak_1').click()    
                    driver.find_element(By.ID, 'nm_anak_1').send_keys(Nama_anak1)        
                    # driver.find_element(By.ID, 'nm_anak_2').click()    
                    driver.find_element(By.ID, 'nm_anak_2').send_keys(Nama_anak2)
                elif jumlah_anak == 3:
                    # driver.find_element(By.ID, 'nm_anak_1').click()    
                    driver.find_element(By.ID, 'nm_anak_1').send_keys(Nama_anak1)        
                    # driver.find_element(By.ID, 'nm_anak_2').click()    
                    driver.find_element(By.ID, 'nm_anak_2').send_keys(Nama_anak2)
                    # driver.find_element(By.ID, 'nm_anak_3').click()    
                    driver.find_element(By.ID, 'nm_anak_3').send_keys(Nama_anak3)
                elif  jumlah_anak == 0: 
                    print('masa ga punya anak')
                #--------------------------------------------------------------
                # driver.find_element(By.XPATH, '//*[@id="telephone_keluarga"]').send_keys(Telepon_keluarga)
            elif Status_perkawinan == 'Janda':
                driver.find_element(By.ID, 'jml_istri_suami').click()
                pyautogui.hotkey('backspace')
                driver.find_element(By.XPATH, '//*[@id="jml_istri_suami"]/div/input' ).send_keys(jml_istrsuam) 
                pyautogui.press('enter')
                driver.find_element(By.ID, 'nm_istri_suami_1').click()
                pyautogui.typewrite(Nm_istrsuam)
                driver.find_element(By.ID, 'tmp_tgl_istri_suami').click()
                pyautogui.typewrite(alm_istrsuam)
                driver.find_element(By.XPATH, '//*[@id="jml_anak"]/div/input').click()
                pyautogui.hotkey('backspace')
                jumlah = driver.find_element(By.XPATH, '//*[@id="jml_anak"]/div/input' ).send_keys(jumlah_anak) 
                pyautogui.press('enter')
                if jumlah_anak == 1:
                    # driver.find_element(By.ID, 'nm_anak_1').click()    
                    driver.find_element(By.ID, 'nm_anak_1').send_keys(Nama_anak1)
                elif jumlah_anak == 2:
                    # driver.find_element(By.ID, 'nm_anak_1').click()    
                    driver.find_element(By.ID, 'nm_anak_1').send_keys(Nama_anak1)        
                    # driver.find_element(By.ID, 'nm_anak_2').click()    
                    driver.find_element(By.ID, 'nm_anak_2').send_keys(Nama_anak2)
                elif jumlah_anak == 3:
                    # driver.find_element(By.ID, 'nm_anak_1').click()    
                    driver.find_element(By.ID, 'nm_anak_1').send_keys(Nama_anak1)        
                    # driver.find_element(By.ID, 'nm_anak_2').click()    
                    driver.find_element(By.ID, 'nm_anak_2').send_keys(Nama_anak2)
                    # driver.find_element(By.ID, 'nm_anak_3').click()    
                    driver.find_element(By.ID, 'nm_anak_3').send_keys(Nama_anak3)
                elif  jumlah_anak == 0: 
                    print('masa ga punya anak')
                #--------------------------------------------------------------
                driver.find_element(By.XPATH, '//*[@id="telephone_keluarga"]').send_keys(Telepon_keluarga)
            elif Status_perkawinan == 'Kawin':
                driver.find_element(By.ID, 'jml_istri_suami').click()
                pyautogui.hotkey('backspace')
                driver.find_element(By.XPATH, '//*[@id="jml_istri_suami"]/div/input' ).send_keys(jml_istrsuam) 
                pyautogui.press('enter')
                driver.find_element(By.ID, 'nm_istri_suami_1').click()
                pyautogui.typewrite(Nm_istrsuam)
                driver.find_element(By.ID, 'tmp_tgl_istri_suami').click()
                pyautogui.typewrite(alm_istrsuam)
                driver.find_element(By.XPATH, '//*[@id="jml_anak"]/div/input').click()
                pyautogui.hotkey('backspace')
                jumlah = driver.find_element(By.XPATH, '//*[@id="jml_anak"]/div/input' ).send_keys(jumlah_anak) 
                pyautogui.press('enter')
                if jumlah_anak == 1:
                    # driver.find_element(By.ID, 'nm_anak_1').click()    
                    driver.find_element(By.ID, 'nm_anak_1').send_keys(Nama_anak1)
                elif jumlah_anak == 2:
                    # driver.find_element(By.ID, 'nm_anak_1').click()    
                    driver.find_element(By.ID, 'nm_anak_1').send_keys(Nama_anak1)        
                    # driver.find_element(By.ID, 'nm_anak_2').click()    
                    driver.find_element(By.ID, 'nm_anak_2').send_keys(Nama_anak2)
                elif jumlah_anak == 3:
                    # driver.find_element(By.ID, 'nm_anak_1').click()    
                    driver.find_element(By.ID, 'nm_anak_1').send_keys(Nama_anak1)        
                    # driver.find_element(By.ID, 'nm_anak_2').click()    
                    driver.find_element(By.ID, 'nm_anak_2').send_keys(Nama_anak2)
                    # driver.find_element(By.ID, 'nm_anak_3').click()    
                    driver.find_element(By.ID, 'nm_anak_3').send_keys(Nama_anak3)
                elif  jumlah_anak == 0: 
                    print('masa ga punya anak')
                #--------------------------------------------------------------
                driver.find_element(By.XPATH, '//*[@id="telephone_keluarga"]').send_keys(Telepon_keluarga)
            #======================================================================
            driver.find_element(By.ID, 'tab-4').click()
            #========================Input Tab Data Fisik========================== 
            driver.find_element(By.XPATH, '//*[@id="tinggi"]/div/input').click()
            pyautogui.hotkey('backspace')
            driver.find_element(By.XPATH, '//*[@id="tinggi"]/div/input' ).send_keys(Tinggi_badan) 
            #--------------------------------------------------------------
            driver.find_element(By.XPATH, '//*[@id="berat"]/div/input').click()
            pyautogui.hotkey('backspace')
            driver.find_element(By.XPATH, '//*[@id="berat"]/div/input').send_keys(Berat_badan)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_bentukrambut').click()
            time.sleep(2) 
            driver.find_element(By.ID, 'id_bentukrambut').send_keys(Bentuk_rambut)
            driver.find_element(By.ID, 'id_bentukrambut').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_bentukrambut').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_jenis_rambut').click()
            time.sleep(2)
            driver.find_element(By.ID, 'id_jenis_rambut').send_keys(Warna_rambut)
            driver.find_element(By.ID, 'id_jenis_rambut').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_jenis_rambut').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_bentukbibir').click()
            time.sleep(2) 
            driver.find_element(By.ID, 'id_bentukbibir').send_keys(Bentuk_bibir)
            driver.find_element(By.ID, 'id_bentukbibir').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_bentukbibir').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_kacamata').click()
            time.sleep(2) 
            driver.find_element(By.ID, 'id_kacamata').send_keys(Berkacamata)
            driver.find_element(By.ID, 'id_kacamata').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_kacamata').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_bentuk_mata').click()
            time.sleep(2) 
            driver.find_element(By.ID, 'id_bentuk_mata').send_keys(Bentuk_mata)
            driver.find_element(By.ID, 'id_bentuk_mata').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_bentuk_mata').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_warna_mata').click()
            time.sleep(2) 
            driver.find_element(By.ID, 'id_warna_mata').send_keys(Warna_mata)
            driver.find_element(By.ID, 'id_warna_mata').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_warna_mata').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_jenis_hidung').click()
            time.sleep(2)
            driver.find_element(By.ID, 'id_jenis_hidung').send_keys(Hidung)
            driver.find_element(By.ID, 'id_jenis_hidung').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_jenis_hidung').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_jenis_muka').click()
            time.sleep(2) 
            driver.find_element(By.ID, 'id_jenis_muka').send_keys(Raut_muka)
            driver.find_element(By.ID, 'id_jenis_muka').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_jenis_muka').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_telinga').click()
            time.sleep(2) 
            driver.find_element(By.ID, 'id_telinga').send_keys(Telinga)
            driver.find_element(By.ID, 'id_telinga').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_telinga').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_jenis_mulut').click()
            time.sleep(2)
            driver.find_element(By.ID, 'id_jenis_mulut').send_keys(Mulut)
            driver.find_element(By.ID, 'id_jenis_mulut').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_jenis_mulut').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_lengan').click()
            time.sleep(2) 
            driver.find_element(By.ID, 'id_lengan').send_keys(Lengan)
            driver.find_element(By.ID, 'id_lengan').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_lengan').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_jenis_tangan').click()
            time.sleep(2) 
            driver.find_element(By.ID, 'id_jenis_tangan').send_keys(Tangan)
            driver.find_element(By.ID, 'id_jenis_tangan').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_jenis_tangan').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_jenis_kaki').click()
            time.sleep(2) 
            driver.find_element(By.ID, 'id_jenis_kaki').send_keys(Kaki)
            driver.find_element(By.ID, 'id_jenis_kaki').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_jenis_kaki').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            driver.find_element(By.ID, 'id_warnakulit').click()
            time.sleep(2) 
            driver.find_element(By.ID, 'id_warnakulit').send_keys(Warna_kulit)
            driver.find_element(By.ID, 'id_warnakulit').send_keys(Keys.DOWN)
            driver.find_element(By.ID, 'id_warnakulit').send_keys(Keys.ENTER)
            #--------------------------------------------------------------
            # driver.find_element(By.ID, 'cacat').click()
            driver.find_element(By.ID, 'cacat').send_keys(Cacat_tubuh)
            # #--------------------------------------------------------------
            # time.sleep(2)
            # driver.find_element(By.XPATH, '//*[@id="upload_foto_ciri_1"]').click()
            # time.sleep(2) 
            # pyautogui.write(r'C:\Users\user\Documents\TRCH\Automationpython\Filefoto\lim.jpg')
            # pyautogui.press('enter')
            # time.sleep(2) 
            # driver.find_element(By.XPATH, '//*[@id="upload_foto_ciri_2"]').click()
            # time.sleep(2) 
            # pyautogui.write(r'C:\Users\user\Documents\TRCH\Automationpython\Filefoto\lim.jpg')
            # pyautogui.press('enter')
            # time.sleep(2) 
            # driver.find_element(By.XPATH, '//*[@id="upload_foto_ciri_3"]').click()
            # time.sleep(2) 
            # pyautogui.write(r'C:\Users\user\Documents\TRCH\Automationpython\Filefoto\lim.jpg')
            # pyautogui.press('enter')
            # time.sleep(2)
            # #--------------------------------------------------------------
            # driver.find_element(By.ID, 'ciri').click()
            driver.find_element(By.ID, 'ciri').send_keys(Catatancirikhusus1)  
            # #--------------------------------------------------------------
            # driver.find_element(By.ID, 'ciri2').click()
            driver.find_element(By.ID, 'ciri2').send_keys(Catatancirikhusus2) 
            # #--------------------------------------------------------------
            # driver.find_element(By.ID, 'ciri3').click()
            driver.find_element(By.ID, 'ciri3').send_keys(Catatancirikhusus3)
            # # ======================================================================
            # driver.find_element(By.ID, 'tab-5').click()
            # # ========================Input Tab Sidik Jari==========================
            # #--------------------------------------------------------------
            # driver.find_element(By.ID, 'no_paspor').click()
            # driver.find_element(By.ID, 'no_paspor').send_keys(Nopaspor)
            # #--------------------------------------------------------------
            # driver.find_element(By.ID, 'rumus_daktil').click()
            # driver.find_element(By.ID, 'rumus_daktil').send_keys(Rumus)
            # #--------------------------------------------------------------
            # driver.find_element(By.ID, 'nomor_daktil').click()
            # driver.find_element(By.ID, 'nomor_daktil').send_keys(Nopaspor)
            # #--------------------------------------------------------------
            # driver.find_element(By.ID, 'pengambil_sj').click()
            # driver.find_element(By.ID, 'pengambil_sj').send_keys(Pengambilansidikjari)
            # #--------------------------------------------------------------
            # driver.find_element(By.XPATH, '//*[@id="pane-5"]/div/form/div/div[2]/div[2]/div/div/input').click()
            # driver.find_element(By.XPATH, '//*[@id="pane-5"]/div/form/div/div[2]/div[2]/div/div/input').send_keys(Tanggalpengambilan)
            #======================================================================
            driver.find_element(By.ID, 'tab-6').click()
            #========================Input Tab Foto========================== 
            time.sleep(2)
            driver.find_element(By.XPATH,   '//*[@id="pane-6"]/form/div/div[1]/div/div/div/div/div[1]/button').click()
            time.sleep(2) 
            pyautogui.write(r'C:\Users\user\Documents\TRCH\Automationpython\Filefoto\lim.jpg')
            pyautogui.press('enter')
            time.sleep(2) 
            driver.find_element(By.XPATH,   '//*[@id="pane-6"]/form/div/div[2]/div/div/div/div/div[1]/button').click()
            time.sleep(2) 
            pyautogui.write(r'C:\Users\user\Documents\TRCH\Automationpython\Filefoto\lim.jpg')
            pyautogui.press('enter')
            time.sleep(2) 
            driver.find_element(By.XPATH,   '//*[@id="pane-6"]/form/div/div[3]/div/div/div/div/div[1]/button').click()
            time.sleep(2) 
            pyautogui.write(r'C:\Users\user\Documents\TRCH\Automationpython\Filefoto\lim.jpg')
            pyautogui.press('enter')
            time.sleep(2)
            #======================================================================
            driver.find_element(By.ID, 'tab-7').click()
            #========================Input Tab Identitas lama========================== 
            #Submit
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[3]/div/div/button[2]').click()
            WebDriverWait(driver,15).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/div/div/h1')))
            time.sleep(4)
        except TimeoutException:
            print('ERROR')
            pass
        time.sleep(2)   
        i = i + 1
    print ('Success Created')
