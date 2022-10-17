from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
import os
import pyautogui
import pytest
import time
import platform
import subprocess

from pathlib import Path

@pytest.fixture()
def test_setup():
    global driver
    global wb
    swin = Service(r'C:/Users/user/Documents/TRCH/chromedriver.exe')
    smac = Service('/Users/will/Downloads/chromedriver')

    if platform.system() == 'Darwin':
        driver = webdriver.Chrome(service=smac)
        # url = "http://kumbang.torche.id:32400/"
        url = "http://192.168.2.11:32400/"
        
        driver.get(url)
        # seting windows nya jadi max
        wb = load_workbook(filename=r"/Users/will/Documents/work/Automationpython/Filexel/Registrasi.xlsx")   
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
    elif platform.system() == 'Windows':
        driver = webdriver.Chrome(service=swin)
        # url = "http://kumbang.torche.id:32400/"
        url = "http://192.168.2.11:32400/"
        
        driver.get(url)
        # seting windows nya jadi max   
        wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Filexel\Registrasi.xlsx")
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
        
def test_Registrasi(test_setup):
    
    driver.implicitly_wait(5)
    sheetrange = wb['reg']
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div/div[2]/div/div/div[1]/canvas")))
        
    driver.find_element(By.XPATH, "//div/span").click()
    # ini masuk ke form input username
    driver.find_element(By.ID, "username").click()
    # masukin input username
    driver.find_element(By.ID, "username").send_keys("wildan")
    # masukin input password
    driver.find_element(By.ID, "password").send_keys("wildan")
    # click button login
    driver.find_element(By.ID, "kc-login").click()
    time.sleep(1)
    element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[1]/div")
    time.sleep(1)
    actions = ActionChains(driver)
    actions.move_to_element(element).perform()
    time.sleep(1)
    element2 = driver.find_element(By.XPATH, "//div/ul/li[2]/div")
    time.sleep(2)
    actions2 = ActionChains(driver)
    actions2.move_to_element(element2).perform()
    time.sleep(2)
    driver.find_element(By.LINK_TEXT, "Registrasi Tahanan/ Narapidana").click()
    WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/div[1]/h1")))
       
    i = 2

    # nge baca mulai dari tabel A
    while i <= len(sheetrange['A']):
        # deklarasi bahwa NIP itu ada di A 
        NoInduk = sheetrange['A'+str(i)].value
        NoRegistrasi = sheetrange['B'+str(i)].value
        TglSuratPenahanan = sheetrange['C'+str(i)].value
        NomorSuratPenahanan = sheetrange['D'+str(i)].value
        NamaPetugasInstansi = sheetrange['E'+str(i)].value
        Kejaksaan = sheetrange['F'+str(i)].value
        AsalInstansi = sheetrange['G'+str(i)].value
        Keterangan = sheetrange['H'+str(i)].value
        Penyidik = sheetrange['I'+str(i)].value
        Menahan10hari = sheetrange['J'+str(i)].value
        Menahan3hari = sheetrange['K'+str(i)].value
        Menahan1hari = sheetrange['L'+str(i)].value
        LokasiDokumen = sheetrange['M'+str(i)].value
        AsalTahanan = sheetrange['N'+str(i)].value
        Kepolisian = sheetrange['O'+str(i)].value
        JenisPutusan = sheetrange['P'+str(i)].value
        NamaLengkap= sheetrange['Q'+str(i)].value
        TanggalLahir = sheetrange['R'+str(i)].value
        PutusSidang = sheetrange['S'+str(i)].value
        TerhitungSejakPertamaDitangkap= sheetrange['T'+str(i)].value
        EksekusiJaksa= sheetrange['U'+str(i)].value
        TanggalBA8 = sheetrange['V'+str(i)].value
        TglMenjalaniPutusanAkhir = sheetrange['W'+str(i)].value
        TglPertamaKaliDitahan = sheetrange['X'+str(i)].value
        TglPertamaKaliDitahan1 = sheetrange['Y'+str(i)].value
        TglPertamaDitahanGolonganBI = sheetrange['Z'+str(i)].value
        TglEkspirasiAwal = sheetrange['AA'+str(i)].value
        TglEkspirasiAkhir = sheetrange['AB'+str(i)].value
        TanggalmenjalaniPidanaPencabutanPB = sheetrange['AC'+str(i)].value
        #======================== PERKARA ============================
        TidakAdaTanggalKejadian = sheetrange['AD'+str(i)].value # switch
        TempatRisalahKejadian = sheetrange['AE'+str(i)].value
        RisalahKejadianPerkara = sheetrange['AF'+str(i)].value
        TanggalKejadian = sheetrange['AG'+str(i)].value
        JamKejadian = sheetrange['AH'+str(i)].value 
        KejahatanUtama = sheetrange['AI'+str(i)].value # switch
        UraianKejahatan = sheetrange['AJ'+str(i)].value
        Undangundang = sheetrange['AK'+str(i)].value
        PasalUtama = sheetrange['AL'+str(i)].value
        PasalTambahan = sheetrange['AM'+str(i)].value
        JenisKejahatan = sheetrange['AN'+str(i)].value
        TempatPenangkapan = sheetrange['AO'+str(i)].value
        #======================== Putusan Pengadilan Negeri ============================
        KetersediaanVonis = sheetrange['AP'+str(i)].value # switch
        TanggalPutusan = sheetrange['AQ'+str(i)].value
        NomorPutusan = sheetrange['AR'+str(i)].value
        Pasal = sheetrange['AS'+str(i)].value
        NamaHakimKetua = sheetrange['AT'+str(i)].value
        NamaHakimAnggota1 = sheetrange['AU'+str(i)].value
        NamaHakimAnggota2 = sheetrange['AV'+str(i)].value
        NamaPaniteraPengganti = sheetrange['AW'+str(i)].value
        NamaJaksa = sheetrange['AX'+str(i)].value
        PengadilanNegeri = sheetrange['AY'+str(i)].value
        TanggalPutusanDijalankan = sheetrange['AZ'+str(i)].value
        PeranDalamKejahatan = sheetrange['BA'+str(i)].value
        JenisHukuman = sheetrange['BB'+str(i)].value
        PidanaTahun = sheetrange['BC'+str(i)].value
        PidanaBulan = sheetrange['BD'+str(i)].value
        PidanaHari = sheetrange['BE'+str(i)].value
        KategoriRemisi = sheetrange['BF'+str(i)].value
        Denda = sheetrange['BG'+str(i)].value
        PidanaKurunganTahun1= sheetrange['BH'+str(i)].value
        PidanaKurunganBulan1= sheetrange['BI'+str(i)].value
        PidanaKurunganHari1= sheetrange['BJ'+str(i)].value
        UangPengganti= sheetrange['BK'+str(i)].value
        PidanaKurunganTahun2= sheetrange['BL'+str(i)].value
        PidanaKurunganBulan2= sheetrange['BM'+str(i)].value
        PidanaKurunganHari2= sheetrange['BN'+str(i)].value
        Restitusi= sheetrange['BO'+str(i)].value
        PidanaKurunganTahun3= sheetrange['BP'+str(i)].value
        PidanaKurunganBulan3= sheetrange['BQ'+str(i)].value
        PidanaKurunganHari3= sheetrange['BR'+str(i)].value
        TanggalEkspirasiPerkiraan= sheetrange['BS'+str(i)].value
        
        #======================== Halaman Index ============================
        time.sleep(2)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/div[2]/button[3]/i").click()
        #WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div[1]/div/div/div/h1")))
        time.sleep(2)
        
        try:      
            #======================== Halaman Cari ============================

            WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div[1]/div/form/button[2]")))
        
            
            driver.find_element(By.CSS_SELECTOR, ".el-form-item:nth-child(1) .el-input__inner").click()
            time.sleep(2)
            driver.find_element(By.XPATH, "//li[contains(.,\'B I\')]").click()
            driver.find_element(By.XPATH, "(//input[@type=\'text\'])[2]").click()
            driver.find_element(By.XPATH, "(//input[@type=\'text\'])[2]").send_keys(NoInduk)
            time.sleep(3)
            WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".el-select-dropdown__item:nth-child(1) tr:nth-child(2) > .el-descriptions__cell:nth-child(1)")))
            driver.find_element(By.CSS_SELECTOR, ".el-select-dropdown__item:nth-child(1) tr:nth-child(2) > .el-descriptions__cell:nth-child(1)").click()
            time.sleep(1)
            #======================== Button Cari ============================
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div[1]/div/form/button[1]").click()
            #======================== masuk ke halaman input ============================
            
            WebDriverWait(driver,10).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.el-dialog__close > svg')))
            WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.el-dialog__close > svg')))
            
            driver.find_element(By.CSS_SELECTOR, ".el-dialog__close > svg").click()
            WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[10]/div/button[1]")))
            
            
       
            
            driver.find_element(By.CSS_SELECTOR, ".el-col-lg-12 > .is-required:nth-child(1) .el-input__inner").send_keys(NoRegistrasi)
           
            
            
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[1]/div[1]/div[2]/div/div/input").send_keys(TglSuratPenahanan)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[1]/div[1]/div[2]/div/div/input").send_keys(Keys.ENTER)
           
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[1]/div[1]/div[3]/div/div/input").send_keys(NomorSuratPenahanan)


            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[1]/div[1]/div[4]/div/div/input").send_keys(NamaPetugasInstansi)
            
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[1]/div[1]/div[5]/div/div/div/div/input").click()
            if Kejaksaan == 'Kejaksaan Negeri Bandung' :
                dropdown = driver.find_element(By.XPATH, "//li[contains(.,'Kejaksaan Negeri Bandung')]")
                dropdown.find_element(By.XPATH, "//li[contains(.,'Kejaksaan Negeri Bandung')]").click()
            elif Kejaksaan == 'Kejaksaan Negeri Jakarta Pusat' :
                dropdown = driver.find_element(By.XPATH, "//li[contains(.,'Kejaksaan Negeri Jakarta Pusat')]")
                dropdown.find_element(By.XPATH, "//li[contains(.,'Kejaksaan Negeri Jakarta Pusat')]").click()

            
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[1]/div[1]/div[6]/div/div/input").send_keys(AsalInstansi)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[1]/div[1]/div[7]/div/div/textarea").send_keys(Keterangan)

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[1]/div[1]/div[8]/div/div/div/div/input").click()

            if Penyidik == 'POLDA JABAR' :
                dropdown = driver.find_element(By.XPATH, "//li[contains(.,'POLDA JABAR')]")
                dropdown.find_element(By.XPATH, "//li[contains(.,'POLDA JABAR')]").click()
            elif Penyidik == 'POLRES BANDUNG' :
                dropdown = driver.find_element(By.XPATH, "//li[contains(.,'POLRES BANDUNG')]")
                dropdown.find_element(By.XPATH, "//li[contains(.,'POLRES BANDUNG')]").click()
            

            
            
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[1]/div[2]/div[2]/div/div/input").send_keys(Menahan3hari)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[1]/div[2]/div[2]/div/div/input").send_keys(Keys.ENTER)

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[1]/div[2]/div[3]/div/div/input").send_keys(Menahan1hari)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[1]/div[2]/div[3]/div/div/input").send_keys(Keys.ENTER)

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[1]/div[2]/div[1]/div/div/input").send_keys(Menahan10hari)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[1]/div[2]/div[1]/div/div/input").send_keys(Keys.ENTER)

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[1]/div[2]/div[4]/div/div[1]/textarea").send_keys(LokasiDokumen)

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[1]/div[2]/div[6]/div/div[1]/input").send_keys(AsalTahanan)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[1]/div[2]/div[7]/div/div/input").send_keys(Kepolisian)


            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[2]/div[2]/div[1]/div[1]/div/div/div/div/input").click()
            
            if JenisPutusan == 'Pilih Jenis Putusan' :
                dropdown = driver.find_element(By.XPATH, "//li[contains(.,'Pilih Jenis Putusan')]")
                dropdown.find_element(By.XPATH, "//li[contains(.,'Pilih Jenis Putusan')]").click()
            elif JenisPutusan == 'Putusan Pengadilan Negeri' :
                dropdown = driver.find_element(By.XPATH, "//li[contains(.,'Putusan Pengadilan Negeri')]")
                dropdown.find_element(By.XPATH, "//li[contains(.,'Putusan Pengadilan Negeri')]").click()
            elif JenisPutusan == 'Putusan Pengadilan Negeri' :
                dropdown = driver.find_element(By.XPATH, "//li[contains(.,'Putusan Pengadilan Tinggi')]")
                dropdown.find_element(By.XPATH, "//li[contains(.,'Putusan Pengadilan Tinggi')]").click()
            elif JenisPutusan == 'Putusan Pengadilan Tinggi' :
                dropdown = driver.find_element(By.XPATH, "//li[contains(.,'Putusan MA')]")
                dropdown.find_element(By.XPATH, "//li[contains(.,'Putusan MA')]").click()
            elif JenisPutusan == 'Peninjauan Kembali' :
                dropdown = driver.find_element(By.XPATH, "//li[contains(.,'Peninjauan Kembali')]")
                dropdown.find_element(By.XPATH, "//li[contains(.,'Peninjauan Kembali')]").click()

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[2]/div[2]/div[1]/div[2]/div/div[1]/div/input").send_keys(NamaLengkap)
            
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[2]/div[2]/div[1]/div[2]/div/div[2]/div/input").send_keys(TanggalLahir)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[2]/div[2]/div[1]/div[2]/div/div[2]/div/input").send_keys(Keys.ENTER)

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[2]/div[2]/div[1]/div[5]/div/div/div/div/input").click()
            time.sleep(1)
            if PutusSidang == 'Sudah' :
                dropdown = driver.find_element(By.XPATH, "//li[contains(.,'Sudah')]")
                dropdown.find_element(By.XPATH, "//li[contains(.,'Sudah')]").click()
            elif PutusSidang == 'Belum' :
                dropdown = driver.find_element(By.XPATH, "//li[contains(.,'Belum')]")
                dropdown.find_element(By.XPATH, "//li[contains(.,'Belum')]").click()

            

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[2]/div[2]/div[2]/div[3]/div/div/input").send_keys(TerhitungSejakPertamaDitangkap)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[2]/div[2]/div[2]/div[3]/div/div/input").send_keys(Keys.ENTER)

            

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[2]/div[2]/div[1]/div[7]/div/div/div/div/input").click()
            
            if EksekusiJaksa == 'Sudah' :
                driver.find_element(By.CSS_SELECTOR, ".el-popper:nth-child(22) .el-select-dropdown__item:nth-child(2)").click()
        
            elif EksekusiJaksa == 'Belum' :
                driver.find_element(By.CSS_SELECTOR, ".el-popper:nth-child(22) .el-select-dropdown__item:nth-child(1)").click()
        
            

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[2]/div[2]/div[2]/div[4]/div/div/input").send_keys(TanggalBA8)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[2]/div[2]/div[2]/div[4]/div/div/input").send_keys(Keys.ENTER)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[2]/div[2]/div[2]/div[5]/div/div/input").send_keys(TglMenjalaniPutusanAkhir)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[2]/div[2]/div[2]/div[5]/div/div/input").send_keys(Keys.ENTER)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[3]/div[3]/div[1]/div/div/div[1]/input").send_keys(TglPertamaKaliDitahan)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[3]/div[3]/div[1]/div/div/div[1]/input").send_keys(Keys.ENTER)
            time.sleep(1)
            WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tab-perkara"]')))
            


               
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[3]/div[4]/div[1]/div/div/div[1]/input").send_keys(TglPertamaKaliDitahan1)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[3]/div[4]/div[1]/div/div/div[1]/input").send_keys(Keys.ENTER)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[3]/div[5]/div[1]/div/div/div[1]/input").send_keys(TglPertamaDitahanGolonganBI)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[3]/div[5]/div[1]/div/div/div[1]/input").send_keys(Keys.ENTER)

            #driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[3]/div[6]/div[1]/div/div/div[1]/input").send_keys(TglEkspirasiAwal)
            #time.sleep(1)
            #driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[3]/div[6]/div[1]/div/div/div[1]/input").send_keys(Keys.ENTER)

            #driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[3]/div[6]/div[2]/div/div/div[1]/input").send_keys(TglEkspirasiAkhir)
            #time.sleep(1)
            #driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[3]/div[6]/div[2]/div/div/div[1]/input").send_keys(Keys.ENTER)

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[2]/div[22]/div/div/input").send_keys(TanggalmenjalaniPidanaPencabutanPB)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[2]/div[22]/div/div/input").send_keys(Keys.ENTER)
         
            #======================== Tab Perkara ============================

            
            driver.find_element(By.XPATH, "//*[@id=\"tab-perkara\"]").click()
        
            
            
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[4]/div[2]/div[1]/div[1]/div/div/input").send_keys(TempatRisalahKejadian)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[4]/div[2]/div[1]/div[2]/div/div/textarea").send_keys(RisalahKejadianPerkara)

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[4]/div[2]/div[2]/div/div[1]/div/div[1]/input").send_keys(TanggalKejadian)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[4]/div[2]/div[2]/div/div[1]/div/div[1]/input").send_keys(Keys.ENTER)

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[4]/div[2]/div[2]/div/div[2]/div/div/input").send_keys(JamKejadian)
            
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[4]/div[2]/div[2]/div/div[2]/div/div/input").send_keys(Keys.ENTER)

            #======================== Hukuman ============================
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[4]/div[4]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[2]/div/div[1]/input").send_keys(UraianKejahatan)
            driver.find_element(By.XPATH, " //*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[4]/div[4]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[3]/div/div[1]/input").send_keys(Undangundang)
            driver.find_element(By.XPATH, " //*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[4]/div[4]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[4]/div/div[1]/input").send_keys(PasalUtama)
            driver.find_element(By.XPATH, " //*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[4]/div[4]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[5]/div/div/input").send_keys(PasalTambahan)
            time.sleep(2)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[4]/div[4]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[6]/div/div[1]/div/div/input").send_keys(JenisKejahatan)
            
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[4]/div[4]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[6]/div/div[1]/div/div/input").send_keys(Keys.DOWN)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[4]/div[4]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[6]/div/div[1]/div/div/input").send_keys(Keys.ENTER)

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[4]/div[4]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[7]/div/div[1]/input").send_keys(TempatPenangkapan)
            #======================== tab-putusan_pengadilan_negeri ============================
            
            
            driver.find_element(By.XPATH,"//*[@id=\"tab-putusan_pengadilan_negeri\"]").click()
                    
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[2]/div/div/input").send_keys(TanggalPutusan)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[2]/div/div/input").send_keys(Keys.ENTER)

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[3]/div/div/input").send_keys(NomorPutusan)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[4]/div/div/input").send_keys(Pasal)
            
            

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[5]/div/div/input").send_keys(NamaHakimKetua)

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[6]/div/div/input").send_keys(NamaHakimAnggota1)

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[7]/div/div/input").send_keys(NamaHakimAnggota2)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[8]/div/div/input").send_keys(NamaPaniteraPengganti)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[9]/div/div/input").send_keys(NamaJaksa)
            
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[10]/div/div/div/div/input").send_keys(PengadilanNegeri)
            if PengadilanNegeri == 'Pengadilan Negeri Bandung' :
                driver.find_element(By.XPATH, "//li[contains(.,'Pengadilan Negeri Bandung')]").click()
        
            elif PengadilanNegeri == 'Pengadilan Negeri Jakarta Utara' :
                driver.find_element(By.CSS_SELECTOR, ".el-popper:nth-child(22) .el-select-dropdown__item:nth-child(1)").click()




            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[10]/div/div/div/div/input").send_keys(Keys.DOWN)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[10]/div/div/div/div/input").send_keys(Keys.ENTER)
            time.sleep(2)
    


            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[11]/div/div/input").send_keys(TanggalPutusanDijalankan)
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[11]/div/div/input").send_keys(Keys.ENTER)

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[12]/div/div/input").send_keys(PeranDalamKejahatan)
            

            
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[14]/div/div/div/div/div/div/input").click()
            time.sleep(1)
            pyautogui.typewrite(JenisHukuman)
            pyautogui.keyDown('down')
            time.sleep(1)
            #pyautogui.press('enter')
            #driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[14]/div/div/div/div/div/div/input").send_keys(Keys.DOWN)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[14]/div/div/div/div/div/div/input").send_keys(Keys.ENTER)
            time.sleep(2)

         

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[15]/div/div/div/div/input").send_keys(KategoriRemisi)
            time.sleep(2)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[15]/div/div/div/div/input").send_keys(Keys.DOWN)
            
            time.sleep(2)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[15]/div/div/div/div/input").send_keys(Keys.ENTER)

            
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[14]/div/div[2]/div/div/div[1]/div/div/div/input").send_keys(Keys.DELETE)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[14]/div/div[2]/div/div/div[1]/div/div/div/input").send_keys(PidanaTahun)
            

            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div[3]/div/div/form/div[5]/div[18]/div/div[1]/input").click()
            

            WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//div/button[2]/span')))
            driver.find_element(By.XPATH, "//div/button[2]/span").click()

            
            

    
        except TimeoutException:
            print("MASIH ADA ERROR, CEK LAGI PAK WIL")
            pass
        time.sleep(5)
        i = i + 1
    print("DONE PAK WILDAN, SEBATS DULU")
