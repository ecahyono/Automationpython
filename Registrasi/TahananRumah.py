




from select import kevent
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
import pytest
from openpyxl import load_workbook
import time
from selenium.webdriver.common.keys import Keys 
import pyautogui

@pytest.fixture()
def test_setup():
    global driver
    s = Service('/Users/will/Downloads/chromedriver')
    #s = Service()
    driver = webdriver.Chrome(service=s)
   # url = "http://kumbang.torche.id:32400/"
    url = "http://192.168.2.11:32400/"
    driver.get(url)
    #driver.maximize_window()
    driver.implicitly_wait(10)
    yield
    #driver.close()
    #driver.quit()
def test_Tahanan_Rumah(test_setup):
    wb = load_workbook(filename=r"/Users/will/Documents/work/Automationpython/Filexel/Registrasi.xlsx")
    #wb = load_workbook(filename=r"C:\Users\wilda\Documents\Automationpython\Filexel/Registrasi.xlsx")
    
    sheetrange = wb['TahananRumah']
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div/div[2]/div/div/div[1]/canvas")))
        
    driver.find_element(By.XPATH, "//div/span").click()
    # ini masuk ke form input username
    driver.find_element(By.ID, "username").click()
    # masukin input username
    driver.find_element(By.ID, "username").send_keys("wildan")
    # masukin input password
    driver.find_element(By.ID, "password").send_keys("wildan")
    # click button login
    driver.find_element(By.ID, "kc-login").click()
    time.sleep(1)
    element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[1]/div")
    time.sleep(1)
    actions = ActionChains(driver)
    actions.move_to_element(element).perform()
    time.sleep(1)
    element2 = driver.find_element(By.XPATH, "//div/ul/li[2]/div")
    time.sleep(2)
    actions2 = ActionChains(driver)
    actions2.move_to_element(element2).perform()
    

    element3 = driver.find_element(By.XPATH, "//div[2]/div/ul/li[2]/div")
    time.sleep(2)
    actions3 = ActionChains(driver)
    actions3.move_to_element(element3).perform()

    time.sleep(1)
    driver.find_element(By.LINK_TEXT, "Tahanan Rumah/Kota & Pembantaran").click()
    time.sleep(3)

    i = 2

    # nge baca mulai dari tabel A
    while i <= len(sheetrange['A']):
        # deklarasi bahwa NIP itu ada di A 
        NoInduk                                     = sheetrange['A'+str(i)].value
        StatusTahananLuar                           = sheetrange['B'+str(i)].value
        NomorSurat                                  = sheetrange['C'+str(i)].value
        TanggalSurat                                = sheetrange['D'+str(i)].value
        Instansi                                    = sheetrange['E'+str(i)].value
        TanggalMulai                                = sheetrange['F'+str(i)].value
        UraianTambahan                              = sheetrange['G'+str(i)].value
        Nama                                        = sheetrange['H'+str(i)].value
        Telepon                                     = sheetrange['I'+str(i)].value
        Alamat                                      = sheetrange['J'+str(i)].value
        NamaDokumen                                 = sheetrange['J'+str(i)].value
        DirFile                                     = sheetrange['J'+str(i)].value
        Keterangan                                  = sheetrange['J'+str(i)].value
        
        #======================== Halaman Index ============================
    
            
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/button')))
            
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/button").click()
        
        
        try:      
            #======================== Halaman Cari ===========================
            WebDriverWait(driver,20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/button')))
            driver.find_element(By.CSS_SELECTOR, ".el-input__inner").click()
            time.sleep(1)
            driver.find_element(By.CSS_SELECTOR, ".el-input__inner").send_keys(NoInduk)
            time.sleep(3)
            
            driver.find_element(By.CSS_SELECTOR, ".el-select-dropdown__item:nth-child(1) tr:nth-child(2) > .el-descriptions__cell:nth-child(1)").click()
            time.sleep(1)
            #======================== Button Cari ============================
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/button').click()
            #======================== Button Cari ============================
            WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[1]/div[1]/div[3]/div/div[1]/input')))

            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[1]/div[1]/div[3]/div/div[1]/input').send_keys(NomorSurat)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[1]/div[1]/div[4]/div/div/input').send_keys(TanggalSurat)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[1]/div[1]/div[4]/div/div/input').send_keys(Keys.ENTER)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[1]/div[2]/div[1]/div/div/input').send_keys(Instansi)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[1]/div[2]/div[2]/div/div/input').send_keys(TanggalMulai)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[1]/div[2]/div[2]/div/div/input').send_keys(Keys.ENTER)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[1]/div[2]/div[3]/div/div/textarea').send_keys(UraianTambahan)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[2]/div[1]/div[1]/div/div/input').send_keys(Nama)
            

            driver.find_element(By.XPATH, "//input[@type='text']").click()
            WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="el-popper-container-1722"]/div[3]/div/div/div[1]')))
            if StatusTahananLuar == 'Pembantaran' :
                driver.find_element(By.XPATH, "//li[contains(.,'Pembantaran')]").click()
            elif StatusTahananLuar == 'Penangguhan' :
                driver.find_element(By.XPATH, "//li[contains(.,'Penangguhan')]").click()
            elif StatusTahananLuar == 'Tahanan Kota' :
                driver.find_element(By.XPATH, "//li[contains(.,'Tahanan Kota')]").click()
            elif StatusTahananLuar == 'Tahanan Rumah' :
                driver.find_element(By.XPATH, "//li[contains(.,'Tahanan Rumah')]").click()

         
    


        except TimeoutException:
            print("MASIH ADA ERROR, CEK LAGI PAK WIL")
            pass
        time.sleep(5)
        i = i + 1
    print("DONE PAK WILDAN, SEBATS DULU")
