from json import load
from lib2to3.pgen2 import driver
import string
from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys 

from openpyxl import load_workbook
import time

#target halaman excel ada dimana , wb = variablenya
# wb = load_workbook(filename=r"/Users/will/Documents/Automationpython/Filexel/registrasi.xlsx")
wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Backup\contohkasus.xlsx")

# jadi ini bisa read sheet yang dibawah itu yang di excel
sheetrange = wb['dropdown']

# ini web driver disimpen dimana, kalo disimpen di path kosongin aja
# driver = webdriver.Chrome(r"/Users/will/Downloads/chromedriver")
driver = webdriver.Chrome(r"C:\Users\user\Documents\TRCH\chromedriver.exe")

# link nya ini dimana
driver.get("http://kumbang.torche.id:32400/")
# seting windows nya jadi max   
driver.maximize_window()
# script gakan di eksekusi kalo web ga muncul. kalo lebih dari 10 detik ga muncul error
driver.implicitly_wait(10)
# ini letak xpath icon login
driver.find_element(By.XPATH, "//div/span").click()
# ini masuk ke form input username
driver.find_element(By.ID, "username").click()
# masukin input username
driver.find_element(By.ID, "username").send_keys("rono")
# masukin input password
driver.find_element(By.ID, "password").send_keys("rene")
# click button login
driver.find_element(By.ID, "kc-login").click()
time.sleep(2)
# variable element nyari dimana letak menu lain lain
element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[1]/div")
time.sleep(2)
actions = ActionChains(driver)
actions.move_to_element(element).perform()
time.sleep(2)
#WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"el-popper-container-7971\"]/div[3]/div/ul/li[2]")))
time.sleep(2)

#xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx anak ke 2
element2 = driver.find_element(By.XPATH, "//div/ul/li[2]/div")
#xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx anak ke 2

time.sleep(2)
actions2 = ActionChains(driver)
actions2.move_to_element(element2).perform()
driver.find_element(By.LINK_TEXT, "Registrasi Tahanan/ Narapidana").click()

i = 2

# nge baca mulai dari tabel A
while i <= len(sheetrange['A']):
    # deklarasi bahwa NIP itu ada di A 
    jenisreg = sheetrange['A'+str(i)].value
    NoInduk = sheetrange['B'+str(i)].value
    
    #========================Input No Induk============================
    WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/div[1]/h1")))
    time.sleep(2)
    driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/div[2]/button[3]/i").click()
    time.sleep(3)
    WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div[1]/div/div/div/h1")))
    time.sleep(2)
    driver.find_element(By.CSS_SELECTOR, ".el-form-item:nth-child(1) .el-input__inner").click()
    time.sleep(4)

    try:
        #cropdown  u/ pilih jenis reg
        WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div[1]/div")))
        time.sleep(4)
        driver.find_element(By.XPATH, "//input[@type='text']").click()
        if jenisreg == 'A I':
            driver.find_element(By.XPATH, "//div[11]/div/div/div[1]/ul/li[1]").click()
        elif jenisreg == 'B II A':
            driver.find_element(By.XPATH, "//div[11]/div/div/div[1]/ul/li[10]").click()
        elif jenisreg == 'B I':
            driver.find_element(By.XPATH, "//div[11]/div/div/div[1]/ul/li[9]").click()
        print('dipilih')

        driver.find_element(By.XPATH, "(//input[@type='text'])[2]").click()
        driver.find_element(By.XPATH, "(//input[@type='text'])[2]").send_keys(NoInduk)
        time.sleep(4)
        driver.find_element(By.CSS_SELECTOR, ".el-select-dropdownitem:nth-child(1) tr:nth-child(2) > .el-descriptionscell:nth-child(1)").click()
        driver.find_element(By.XPATH, "//button[contains(.,'Cari')]").click()                                
        
    except TimeoutException:
        print("ga muncul")
        pass
    time.sleep(5)
    i = i + 1
print("done")