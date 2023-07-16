from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime
import pytest
from selenium import webdriver

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("data"))
    file_path = environ.get("FakerRemisi")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
import random

from Settings.setupPembinaan import *
from Settings.Login.LoginRemisi import *
from Settings.Page.remisi import *
import random
import logging

sheetrangeIndex = wb['listWbpSumedang']

random1 = random.randint(1,170)
random2= random.randint(1,170)
random3 = random.randint(1,170)

NamaInput1                                 = sheetrangeIndex['A'+str(random1)].value
NamaInput2                                 = sheetrangeIndex['B'+str(random2)].value
NamaInput3                                 = sheetrangeIndex['C'+str(random3)].value
time.sleep(3)

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'RemisiTambah'

fake = Faker('id_ID')
peraturanPP                         = ["Peraturan Menteri No. 7/2022","Peraturan Menteri No. 3/2018","Keputusan Presiden No. 120/1955"]
peraturanPPJenis                    = ["1","2","3"]
jenisPermen                         = ["1","2","3"]
JenisRemSyarat                      = ["Remisi Dasawarsa","Remisi Umum","Remisi Kepramukaan","Non suscipit veritatis.","Natus ut quasi."]
KategoriRemisi                      = ["Kategori1","Kategori2","Kategori3",'Kategori4','Kategori0']
JenisRegistrasi                     = ["A I","A II","A III","A II Terorisme","A I Terorisme","A IV","Tahanan Militer","Tahanan Militer","B I","10","B II A","B II B","B III","C","Hukuman Mati","Anak Negara","Anak Sipil"]
Operator                            = ["=","<","<=",'>','>=']
DropHari                            = ["dropHari","dropMinggu","dropBulan","dropMP","dropMT"]
SudahPutusan                        = ["SPtidak","SPiya","SPyaHukuman"]
BerkelakuanBaik                     = ["=Baik","<Baik","<=Baik",'>Baik','>=Baik']
KelakuanBaikWaktu                   = ["HariBaik","Minggu Baik","Bulan Baik","MPBaik","MTBaik"]
SyaratUsiA                          = ["UsiaTidakAda","Usia<18","Usia>70"]
AsalSK                              = ["Menkumham","ditjenPas","Kaupt"]
MelunasiDenda                       = ["DendaYa","DendaTidak","dendaSebaiknyaAda"]
LamaRemisi                          = ["1Lam","1/2Lam","1/3Lam"]
JenisRegistrasiWBP                  = ['B III','B II B','B II A','B II B','A V','B I','Tahanan Militer','A IV','A I Terorisme','A II Terorisme','A III','A II','A I','Hukuman Mati','Anak Negara','Anak Sipil']

for i in range(1):
    peraturanPPFaker                = random.choice(peraturanPP)
    peraturanPPJenisFaker           = random.choice(peraturanPPJenis)
    JenisRemisiFaker                = fake.text(max_nb_chars=25)
    KeteranganFaker                 = fake.text(max_nb_chars=50)
    jenisPermenFaker                = random.choice(jenisPermen)
    JenisRemSyaratFaker             = random.choice(JenisRemSyarat)
    KategoriRemisiFaker             = random.choice(KategoriRemisi)
    JenisRegistrasiFaker            = random.choice(JenisRegistrasi)
    OperatorFaker                   = random.choice(Operator)
    MenjalaniPidanaFaker            = random.randint(1,30)
    KelakuanBaikBilanganFaker       = random.randint(1,30)
    DropHariFaker                   = random.choice(DropHari)
    SudahPutusanFaker               = random.choice(SudahPutusan)
    BerkelaunBaikFaker              = random.choice(BerkelakuanBaik)
    KelakuanBaikWaktu               = random.choice(KelakuanBaikWaktu)
    SyaratUsiAFaker                 = random.choice(SyaratUsiA)
    AsalSKFaker                     = random.choice(AsalSK)
    MelunasiDendaFaker              = random.choice(MelunasiDenda)
    LamaRemisiFaker                 = random.choice(LamaRemisi)




    worksheet.append([
        peraturanPPFaker,
        peraturanPPJenisFaker,
        JenisRemisiFaker,
        KeteranganFaker,
        jenisPermenFaker,
        JenisRemSyaratFaker,
        KategoriRemisiFaker,
        JenisRegistrasiFaker,
        OperatorFaker,
        MenjalaniPidanaFaker,
        KelakuanBaikBilanganFaker,
        DropHariFaker,
        SudahPutusanFaker,
        BerkelaunBaikFaker,
        KelakuanBaikWaktu,
        SyaratUsiAFaker,
        AsalSKFaker,
        MelunasiDendaFaker,
        LamaRemisiFaker,
        JenisRegistrasiFaker


        ])
workbook.save(file_path)

workbook = load_workbook(filename=file_path)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=1, values_only=True):
    peraturanPPExcel                            = row[0]
    peraturanPPJenisExcel                       = row[1]
    JenisRemisiExcel                            = row[2]
    KeteranganExcel                             = row[3]
    jenisPermenExcel                            = row[4]
    JenisRemSyaratExcel                         = row[5]
    KategoriRemisiExcel                         = row[6]
    JenisRegistrasiExcel                        = row[7]
    OperatorExcel                               = row[8]
    MenjalaniPidanaExcel                        = row[9]
    KelakuanBaikBilanganExcel                   = row[10]
    DropHariExcel                               = row[11]
    SudahPutusanExcel                           = row[12]
    BerkelaunBaikExcel                          = row[13]
    KelakuanBaikWaktu                           = row[14]
    SyaratUsiAExcel                             = row[15]
    AsalSKExcel                                 = row[16]
    MelunasiDendaExcel                          = row[17]
    LamaRemisiExcel                             = row[18]
    JenisRegistrasiWBPExcel                     = row[19]



global driver, pathData
driver = initDriver()
pathData = loadDataPath()      
