from distutils.archive_util import make_archive
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("RegSelenium"))

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("RegSelenium"))


from Settings.setupPembinaan import *
from Settings.loginPembinaan import *

import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Registrasi.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)


sheetrange = wb['RegSelenium']

@mark.webtest()
def test_1_setupOS():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')

@mark.webtest()
def test_2_login():
    (driver)
    Log.info('Login')
    sumedang(driver)


@mark.webtest()
def test_Input_Registrasi():
    driver.implicitly_wait(60)
    driver.implicitly_wait(30)
    nav1 = driver.find_element(By.XPATH, pathData['AksesMenu']['Registrasi']['MainText'])
    ActionChains(driver).move_to_element(nav1).perform()
    element2 = driver.find_element(By.XPATH, pathData['AksesMenu']['Registrasi']['child']['ManajemenRegistrasi']['MainText'])
    time.sleep(1)
    ActionChains(driver).move_to_element(element2).perform()
    time.sleep(1)
    driver.find_element(By.LINK_TEXT, 'Registrasi Tahanan/ Narapidana').click()
    Log.info('Akses Menu Registrasi')
    attach(data=driver.get_screenshot_as_png())

    i = 2
        # nge baca mulai dari tabel A
    while i <= len(sheetrange['A']):
        # deklarasi bahwa NIP itu ada di A 
        NamaWBP                             = sheetrange['A'+str(i)].value
        NoRegistrasi                        = sheetrange['B'+str(i)].value
        NomorSuratPenahanan                 = sheetrange['C'+str(i)].value
        Tgl                                 = sheetrange['D'+str(i)].value
        Pidana                              = sheetrange['E'+str(i)].value
        namapetugas                         = sheetrange['F'+str(i)].value
        AsalInstansi                        = sheetrange['G'+str(i)].value
        LokasiDokumen                       = sheetrange['H'+str(i)].value
        asaltahanan                         = sheetrange['I'+str(i)].value
        kepolisian                          = sheetrange['J'+str(i)].value
        kejahatanutama                      = sheetrange['K'+str(i)].value
        uud                                 = sheetrange['L'+str(i)].value
        pasalutama                          = sheetrange['M'+str(i)].value
        tempat                              = sheetrange['N'+str(i)].value
        nomorputusan                        = sheetrange['O'+str(i)].value
        namahakimketua                      = sheetrange['P'+str(i)].value
        HakimAnggota1                       = sheetrange['Q'+str(i)].value
        HakimAnggota2                       = sheetrange['R'+str(i)].value
        panitera                            = sheetrange['S'+str(i)].value
        NamaJaksa                           = sheetrange['T'+str(i)].value
        tersangka                           = sheetrange['U'+str(i)].value
       
        pasal                               = sheetrange['V'+str(i)].value
      
        driver.implicitly_wait(30)


        try:

    
            input('enter')
            
            # WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.el-dialog__headerbtn')))
            # driver.find_element(By.CSS_SELECTOR, ".el-dialog__headerbtn").click()
            # Log.info('close Pop Up')

            WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, "tgl_srt_thn")))

            driver.find_element(By.ID, "jatuh_vonis").click()
            driver.find_element(By.XPATH, "//li[contains(.,'Sudah')]").click()
            driver.find_element(By.ID, "eksekusi_jaksa").click()
            driver.find_element(By.ID, "EKS1").click()

            driver.find_element(By.ID, "tab-putusan_pengadilan_negeri").click()
            driver.find_element(By.ID, "peranan_kejahatan-0").send_keys('pelaku utama')
            
            driver.find_element(By.ID, "tab-putusan_pengadilan_tinggi").click()
            driver.find_element(By.ID, "peranan_kejahatan-1").clear()
            driver.find_element(By.ID, "peranan_kejahatan-1").send_keys('pelaku utama')

            





            driver.find_element(By.ID, "tab-file_dokumen").click()

            WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, "chooseFile-0")))
            driver.find_element(By.ID, "chooseFile-0").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-1").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-2").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-3").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-4").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-5").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-6").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-7").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-8").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-15").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-16").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-17").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-20").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-21").click()
            upload(driver)


            driver.find_element(By.ID, "submitButton").click()
            Log.info('Klik Button Submit')

            
        except TimeoutException:
            print("MASIH ADA ERROR, CEK LAGI PAK WIL")
            pass
            quit(driver)
        time.sleep(5)
        i = i + 1

  
        

@mark.webtest()
def test_exit():
    quit(driver)




