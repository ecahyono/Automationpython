from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
import os, sys
from os import environ, path
import pyautogui
from pytest import mark
import pytest
import time
import platform
from pathlib import Path
import logging
from faker import Faker
from datetime import datetime
import openpyxl
from faker.providers import date_time
from datetime import datetime, timedelta
import random


fake = Faker('id_ID')
from dotenv import load_dotenv
load_dotenv()

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("RegSelenium"))

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("RegSelenium"))

from Settings.SetupAll import *
from Settings.login import *
from Settings.Page.accessmenu import *



import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Registrasi.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
#Log.addHandler(fh)

sheetrange = wb['RegSelenium']
baris = int(input('Mulai dari baris ke : '))
ValueJenisReg = int(input("\n 00. Urutan Sheet \n 1. A I \n 2. A II \n 3. A III \n 4. A IV \n 5. A V \n 6. B I \n 7. B II A \n 8. B II B \n 9. B III \n 10. C \n 11. Hukuman Mati \n 12. Random Narapidana \n Pilih Jenis Registrasi :  "))
		



# if ValueJenisReg == 00:
# 	jenis_registrasi 			= jenis_registrasi
if ValueJenisReg == 1:
	jenis_registrasi 			= random.choice(['A I'])
elif ValueJenisReg == 2:
	jenis_registrasi 			= random.choice(['A II'])
elif ValueJenisReg == 3:
	jenis_registrasi 			= random.choice(['A III'])
elif ValueJenisReg == 4:
	jenis_registrasi 			= random.choice(['A IV'])
elif ValueJenisReg == 5:
	jenis_registrasi 			= random.choice(['A V'])
elif ValueJenisReg == 6:
	jenis_registrasi 			= random.choice(['B I'])
elif ValueJenisReg == 7:
	jenis_registrasi 			= random.choice(['B II A'])
elif ValueJenisReg == 8:
	jenis_registrasi 			= random.choice(['B II B'])
elif ValueJenisReg == 9:
	jenis_registrasi 			= random.choice(['B III'])
elif ValueJenisReg == 10:
	jenis_registrasi 			= random.choice(['C'])
elif ValueJenisReg == 11:
	jenis_registrasi 			= random.choice(['Hukuman Mati'])
elif ValueJenisReg == 12:
	jenis_registrasi 			= random.choice(['B I','B II A','B II B'])

print(jenis_registrasi)
@mark.webtest()
def test_1setupOS():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
	
@mark.webtest()
def test_2loggin():
	# login(driver)
	sorong(driver)

@mark.webtest
def test_3aksesmenu():
	driver.implicitly_wait(20)
	registrasitahanannarapidana(driver)

@mark.webtest
def test_4Registrasi():
	
	i = baris
	while i <= len(sheetrange['A']):
		NamaWBP                             		= sheetrange['A'+str(i)].value
		# jenis_registrasi 							= sheetrange['B'+str(i)].value
		
			
		print("nomor urut :")
		print(i)
		nums 						= '01234567891011121415161716'

		

		NoRegistrasiBI					= "B.I-/"					+ fake.isbn10() + ".PASS" + ".PASS" + random.choice(nums) +".WL." + fake.date_between(start_date='today', end_date='today').strftime('%d.%m.%Y') + "-" + random.choice(nums)
		NoRegistrasiBIIA				= "B.II.A-/"				+ fake.isbn10() + ".PASS" + ".PASS" + random.choice(nums) +".WL." + fake.date_between(start_date='today', end_date='today').strftime('%d.%m.%Y') + "-" + random.choice(nums)
		NoRegistrasiBIIB				= "B.II.B-/"				+ fake.isbn10() + ".PASS" + ".PASS" + random.choice(nums) +".WL." + fake.date_between(start_date='today', end_date='today').strftime('%d.%m.%Y') + "-" + random.choice(nums)
		NoRegistrasiBIII				= "B.III.-/"				+ fake.isbn10() + ".PASS" + ".PASS" + random.choice(nums) +".WL." + fake.date_between(start_date='today', end_date='today').strftime('%d.%m.%Y') + "-" + random.choice(nums)
		NoRegistrasiBI					= "B.I-/"					+ fake.isbn10() + ".PASS" + ".PASS" + random.choice(nums) +".WL." + fake.date_between(start_date='today', end_date='today').strftime('%d.%m.%Y') + "-" + random.choice(nums)
		NoRegistrasiAI					= "A.I-/"					+ fake.isbn10() + ".PASS" + ".PASS" + random.choice(nums) +".WL." + fake.date_between(start_date='today', end_date='today').strftime('%d.%m.%Y') + "-" + random.choice(nums)
		NoRegistrasiC					= "C.-/"					+ fake.isbn10() + ".PASS" + ".PASS" + random.choice(nums) +".WL." + fake.date_between(start_date='today', end_date='today').strftime('%d.%m.%Y') + "-" + random.choice(nums)
		NoRegistrasiHukumanMati         = "Hukuman Mati-/"          + fake.isbn10() + ".PASS" + ".PASS" + random.choice(nums) +".WL." + fake.date_between(start_date='-1years', end_date='-1years').strftime('%d.%m.%Y') + "-" + random.choice(nums)
		NoRegistrasiAII					= "A.II-/"					+ fake.isbn10() + ".PASS" + ".PASS" + random.choice(nums) +".WL." + fake.date_between(start_date='today', end_date='today').strftime('%d.%m.%Y') + "-" + random.choice(nums)
		NoRegistrasiAIII				= "A.III-/"					+ fake.isbn10() + ".PASS" + ".PASS" + random.choice(nums) +".WL." + fake.date_between(start_date='today', end_date='today').strftime('%d.%m.%Y') + "-" + random.choice(nums)
		NoRegistrasiAIITeroris			= "A.II Terorisme-/"		+ fake.isbn10() + ".PASS" + ".PASS" + random.choice(nums) +".WL." + fake.date_between(start_date='today', end_date='today').strftime('%d.%m.%Y') + "-" + random.choice(nums)
		NoRegistrasiAIV					= "A.IV-/"					+ fake.isbn10() + ".PASS" + ".PASS" + random.choice(nums) +".WL." + fake.date_between(start_date='today', end_date='today').strftime('%d.%m.%Y') + "-" + random.choice(nums)
		NoRegistrasiAV					= "A.V -/"					+ fake.isbn10() + ".PASS" + ".PASS" + random.choice(nums) +".WL." + fake.date_between(start_date='today', end_date='today').strftime('%d.%m.%Y') + "-" + random.choice(nums)
		
		tgl_BA8							= fake.date_between(start_date='-600d', end_date='-400d').strftime('%d.%m.%Y') #masih belum tau tanggal ba8 itu di dapat dari mana, rulesnya seperti apa
		tgl_BA8GolonganA				= fake.date_between(start_date='-10d', end_date='today').strftime('%d.%m.%Y') #masih belum tau tanggal ba8 itu di dapat dari mana, rulesnya seperti apa
		tgl_Surat_Penahanan				= tgl_BA8
		tgl_Surat_PenahananGolonganA	= tgl_BA8GolonganA

		NomorSuratPenahananAI			= "SP.PAS/A.I-" 			+ fake.date_between(start_date='-1years', end_date='-1years').strftime('%d.%m.%Y')
		NomorSuratPenahananBI			= "SP.PAS/B.I-" 			+ fake.date_between(start_date='-1years', end_date='-1years').strftime('%d.%m.%Y')
		NomorSuratPenahananAIITeroris	= "SP.PAS/A.II Terorisme" 	+ fake.date_between(start_date='-1years', end_date='-1years').strftime('%d.%m.%Y')

		penyidikAI						= random.choice(['penyidikOption-0','penyidikOption-1','penyidikOption-2','penyidikOption-3','penyidikOption-4','penyidikOption-5'])
		JenisKejahatanAI				= random.choice(['jenisKejahatan-0-0','jenisKejahatan-0-1','jenisKejahatan-0-2','jenisKejahatan-0-3','jenisKejahatan-0-4','jenisKejahatan-0-5'])
		namapetugas						= fake.name()
		Kejaksaan						= random.choice(['#kejaksaanOption-0 > span','#kejaksaanOption-1 > span','#kejaksaanOption-2 > span','#kejaksaanOption-3 > span','#kejaksaanOption-4 > span','#kejaksaanOption-5 > span'])
		instansiPenyidik				= random.choice(['id_instansi_penyidik-0','id_instansi_penyidik-1','id_instansi_penyidik-2','id_instansi_penyidik-3','id_instansi_penyidik-4','id_instansi_penyidik-5'])
		LokasiDokumen					= fake.address()
		AsalTahanan 					= random.choice(['APH ','RUTAN ','LAPAS ']) + fake.administrative_unit()
		tgl_Penangkapan					= tgl_Surat_Penahanan #apakah harus ada validasi mengenai tanggal penangkapan terharap surat penahanan
		tgl_PenangkapanGolonganA		= tgl_Surat_PenahananGolonganA #apakah harus ada validasi mengenai tanggal penangkapan terharap surat penahanan
		tgl_putusan_akhir				= tgl_BA8 #tanggal putusan akhir tidak tau rulesnya seperti apa
		# Jenis_putusan					= random.choice(['PPN','PPT','PMA','PK'])
		Jenis_putusan					= random.choice(['PPN'])
		tgl_menjalani_cabutpb			= tgl_BA8 

		#TAB PERKARA
		KejahatanUtama					= random.choice(['testing '])
		terminologi 					= random.choice(['kejahatan.0.id_terminologi-0','kejahatan.0.id_terminologi-1','kejahatan.0.id_terminologi-2','kejahatan.0.id_terminologi-3','kejahatan.0.id_terminologi-4','kejahatan.0.id_terminologi-5'])
		tempatKejahatan					= fake.administrative_unit()
		Tanggal_Putusan					= tgl_BA8

		#TAB PUTUSAN
		NomorPutusan					= "PTS"+ fake.isbn10() + ".PASS" + ".PASS" + random.choice(nums) +".WL." + tgl_BA8 + "-" + random.choice(nums)
		namahakimketua					= fake.name()
		HakimAnggota1					= fake.name()
		HakimAnggota2					= fake.name()
		panitera						= fake.name()
		NamaJaksa						= fake.name()

		# jenis_hukuman					=random.choice(['id_jenis_hukuman-0-0','id_jenis_hukuman-0-1','id_jenis_hukuman-0-2'])
		jenis_hukuman					= random.choice(['id_jenis_hukuman-0-1','id_jenis_hukuman-0-2'])
		PidanaTahun						= fake.random_int(min=2, max=3)
		# jenisRemisi					= random.choice(['jenis_remisi-0-0','jenis_remisi-0-1','jenis_remisi-0-2','jenis_remisi-0-3','jenis_remisi-0-4'])
		# jenisRemisi						= random.choice(['jenis_remisi-0-0','jenis_remisi-0-1','jenis_remisi-0-2'])
		jenisRemisi						= random.choice(['jenis_remisi-0-0'])
		wait=WebDriverWait
		find=driver.find_element
		input('Tekan Enter Untuk Melanjutkan')

		wait(driver,1050).until(EC.element_to_be_clickable((By.ID, 'cari')))
		Log.info('Click Button Cari')
		input('Tekan Enter Untuk Melanjutkan')
		wait(driver,120).until(EC.element_to_be_clickable((By.XPATH, '//span[contains(.,\'Registrasi Baru\')]')))
		input('Tekan Enter Untuk Melanjutkan')
		find(By.XPATH, "//span[contains(.,\'Registrasi Baru\')]").click()x
		#Log.info('Click Button Registrasi')
		hold(driver)
						 
		try:
			input('Tekan Enter Untuk Melanjutkan')
			wait(driver,120).until(EC.element_to_be_clickable((By.ID, 'findButton')))
			find(By.ID, "jenisRegistrasi").send_keys(jenis_registrasi)



			if jenis_registrasi == 'A I' or jenis_registrasi == 'A II' or jenis_registrasi == 'A III' or jenis_registrasi == 'B I' or jenis_registrasi == 'B II A' or jenis_registrasi == 'B II B' or jenis_registrasi == 'B III' or jenis_registrasi == 'A IV' or jenis_registrasi == 'A V' or jenis_registrasi == 'Hukuman Mati':
				wait(driver,20).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'"+jenis_registrasi+"')]"))).click()
				time.sleep(0.5)
			elif jenis_registrasi == 'C':
				wait(driver,20).until(EC.element_to_be_clickable((By.ID, "jenisRegistrasiOption-12"))).click()
				time.sleep(0.5)
		
			#Log.info("Click Button Jenis Registrasi '"+ jenis_registrasi +"' ")
			hold(driver)

			find(By.ID, "noRegistrasi").click()
			# input('Tekan Enter Untuk Melanjutkan')
			#Log.info("Input No Registrasi")		
			hold(driver)
			time.sleep(0.5)
			find(By.ID, "noRegistrasi").send_keys(NamaWBP)	
			hold(driver)

			wait(driver,20).until(EC.element_to_be_clickable((By.XPATH, "//li[@id='noRegistrasiOption-0']/div/div/table/tbody/tr[2]/td"))).click()
				#Log.info("Click Button No Registrasi")
			hold(driver)

			find(By.ID, "findButton").click()
			#Log.info("Click Button cari")
			hold(driver)

			wait(driver,20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.el-dialog__headerbtn')))
			find(By.CSS_SELECTOR, ".el-dialog__headerbtn").click()
			#Log.info('close Pop Up')
			hold(driver)

# ============================================================================================================================================================================
#REGISTRASI UTAMA A
			if jenis_registrasi == 'A I' or jenis_registrasi == 'A II' or jenis_registrasi == 'A III' or jenis_registrasi == 'C' or jenis_registrasi == 'A IV' or jenis_registrasi == 'A V':

				if jenis_registrasi == 'A I':
					find(By.ID, "nmr_reg_gol").send_keys(NoRegistrasiAI)
				elif jenis_registrasi == 'A II':
					find(By.ID, "nmr_reg_gol").send_keys(NoRegistrasiAII)
				elif jenis_registrasi == 'A III':
					find(By.ID, "nmr_reg_gol").send_keys(NoRegistrasiAIII)
				elif jenis_registrasi == 'A IV':
					find(By.ID, "nmr_reg_gol").send_keys(NoRegistrasiAIV)
				elif jenis_registrasi == 'A V':
					find(By.ID, "nmr_reg_gol").send_keys(NoRegistrasiAV)
				elif jenis_registrasi == 'C':
					find(By.ID, "nmr_reg_gol").send_keys(NoRegistrasiC)


				find(By.ID, "tgl_srt_thn").click()
				find(By.ID, "tgl_srt_thn").send_keys(tgl_Surat_PenahananGolonganA)
				find(By.ID, "tgl_srt_thn").send_keys(Keys.ENTER)
				#Log.info('input tanggal surat ')
				hold(driver)

				find(By.ID, "nmr_srt_thn").send_keys(NomorSuratPenahananAI)
				#Log.info('input nomor surat penahanan')
				hold(driver)

				find(By.ID, "nm_pjbt_thn").send_keys(namapetugas)
				#Log.info('input nama petugas')
				hold(driver)

				find(By.ID, "kejaksaan").click()
				wait(driver,20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, Kejaksaan))).click()

				find(By.ID, "instansi_thn").send_keys("Kejaksaan Agung Republik Indonesia")
				#Log.info('Input Instansi')
				hold(driver)

				find(By.ID, "keterangan").send_keys(fake.text())

				find(By.ID, "id_instansi_penyidik").click()
				if jenis_registrasi == 'A IV':
					wait(driver,20).until(EC.element_to_be_clickable((By.ID, penyidikAI))).click()
				else:
					wait(driver,20).until(EC.element_to_be_clickable((By.ID, instansiPenyidik))).click()
				#Log.info('input penyidik')
				hold(driver)

				find(By.ID, "lokasi_dokumen").send_keys(LokasiDokumen)
				#Log.info('input lokasi dokumen')
				hold(driver)

				find(By.ID, "asal_tahanan").send_keys(AsalTahanan)
				#Log.info('input asal tahanan')
				hold(driver)

				find(By.ID, "kepolisian").send_keys(AsalTahanan)
			
				#Log.info('input kepolisian')
				hold(driver)
				find(By.ID, "tgl_pertama_ditahan").click()
				find(By.ID, "tgl_pertama_ditahan").send_keys(tgl_PenangkapanGolonganA)
				find(By.ID, "tgl_pertama_ditahan").send_keys(Keys.ENTER)
				#Log.info('input tanggal pertama ditahan')
				hold(driver)

				time.sleep(1)	
				find(By.ID, "tgl_msk_lapas").click()
				find(By.ID, "tgl_msk_lapas").send_keys(tgl_PenangkapanGolonganA)
				find(By.ID, "tgl_msk_lapas").send_keys(Keys.ENTER)
				#Log.info('input tanggal masuk rutan')
				hold(driver)

				find(By.ID, "tgl_akhir_ditahan").click()
				find(By.ID, "tgl_akhir_ditahan").send_keys(tgl_PenangkapanGolonganA)
				find(By.ID, "tgl_akhir_ditahan").send_keys(Keys.ENTER)
				#Log.info('Input tanggal akhir ditahan')
				hold(driver)

				find(By.ID, "tgl_awal_tahan_golongan").click()
				find(By.ID, "tgl_awal_tahan_golongan").send_keys(tgl_PenangkapanGolonganA)
				find(By.ID, "tgl_awal_tahan_golongan").send_keys(Keys.ENTER)
				#Log.info('input tanggal pertama ditahan AI')
				hold(driver)
			
				
#TAB PERKARA SEMUA GOLONGAN A

				# time.sleep(3)
				# print('pilih TAB perkara Golongan')
				# find(By.ID, "tab-perkara").click()
				# time.sleep(1)
				# wait(driver,20).until(EC.element_to_be_clickable((By.ID, "jam_kejadian"))).click()
				# find(By.ID, "tgl_kejadian").send_keys(tgl_BA8)
				# #Log.info('input tanggal kejadian')
				# hold(driver)
				# find(By.ID, "tempat_kejadian").send_keys(tempatKejahatan)
				# #Log.info('input tempat kejadian')
				# hold(driver)
				# find(By.ID, "risalah_kejadian_perkara").send_keys(fake.text())
				# #Log.info('input risalah kejadian perkara')
				# hold(driver)
				
				# if jenis_registrasi == 'A III':
				# 	find(By.CSS_SELECTOR, ".is-required .el-switch__action").click()
				# 	hold(driver)

				# 	find(By.ID, "kejahatan.0.deskripsi").click()
				# 	find(By.ID, "kejahatan.0.deskripsi").send_keys(KejahatanUtama)
				# 	find(By.ID, "kejahatan.0.uu_kejahatan").send_keys(KejahatanUtama)
				# 	#Log.info('input kejahatan utama')
				# 	hold(driver)

				# 	find(By.ID, "kejahatan.0.pasal_utama").click()
				# 	find(By.ID, "kejahatan.0.pasal_utama").send_keys("pasal " + KejahatanUtama)
				# 	#Log.info('input pasal utama')
				# 	hold(driver)

				# 	find(By.ID, "kejahatan.0.id_terminologi").click()
				# 	wait(driver,20).until(EC.element_to_be_clickable((By.ID, terminologi))).click()
		
				# 	#Log.info('input Kejahatan')
				# 	hold(driver)

				# 	find(By.ID, "kejahatan.0.wilayah").click()
				# 	find(By.ID, "kejahatan.0.wilayah").send_keys(tempatKejahatan)
				# 	#Log.info('input tempat kejadian')
				# 	hold(driver)
				# else:
				
				# 	find(By.CSS_SELECTOR, "#uraianKejahatan0").send_keys(KejahatanUtama)
				# 	#Log.info('uraian kejahatan utama')
				# 	hold(driver)

				# 	find(By.CSS_SELECTOR, "#undangUndang0").send_keys(KejahatanUtama)
				# 	#Log.info('undang undang')
				# 	hold(driver)

				# 	find(By.CSS_SELECTOR, "#pasalUtama0").send_keys(KejahatanUtama)

				# 	find(By.CSS_SELECTOR, "#jenisKejahatan0").click()
				# 	wait(driver,20).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,'Terhadap Kepala Negara')]"))).click()
				
				# 	find(By.CSS_SELECTOR, "#tempatPenangkapan0").send_keys(tempatKejahatan)

# ============================================================================================================================================================================
#REGISTRASI UTAMA B
			elif jenis_registrasi == 'B I' or jenis_registrasi == 'B II A' or jenis_registrasi == 'B II B' or jenis_registrasi == 'B III' or jenis_registrasi == 'Hukuman Mati':

				find(By.ID, "nmr_reg_gol").click()
				if jenis_registrasi == 'B I':
					find(By.ID, "nmr_reg_gol").send_keys(NoRegistrasiBI)
				elif jenis_registrasi == 'B II A':
					find(By.ID, "nmr_reg_gol").send_keys(NoRegistrasiBIIA)
				elif jenis_registrasi == 'B II B':
					find(By.ID, "nmr_reg_gol").send_keys(NoRegistrasiBIIB)
				elif jenis_registrasi == 'B III':
					find(By.ID, "nmr_reg_gol").send_keys(NoRegistrasiBIII)
				elif jenis_registrasi == 'Hukuman Mati':
					find(By.ID, "nmr_reg_gol").send_keys(NoRegistrasiHukumanMati)

				#Log.info('input no registrasi ')
				hold(driver)
				find(By.ID, "tgl_srt_thn").click()
				find(By.ID, "tgl_srt_thn").send_keys(tgl_Surat_Penahanan)
				find(By.ID, "tgl_srt_thn").send_keys(Keys.ENTER)
				#Log.info('input tanggal surat ')
				hold(driver)

				find(By.ID, "nmr_srt_thn").click()
				find(By.CSS_SELECTOR, ".el-col > .is-required:nth-child(3)").click()
				find(By.ID, "nmr_srt_thn").send_keys(NomorSuratPenahananBI)
				find(By.CSS_SELECTOR, ".el-row:nth-child(4) > .el-col:nth-child(1)").click()
				find(By.CSS_SELECTOR, ".is-success:nth-child(3) > .el-form-item__label").click()
				#Log.info('input nomor surat penahanan')
				hold(driver)

				find(By.ID, "nm_pjbt_thn").click()
				find(By.ID, "nm_pjbt_thn").send_keys(namapetugas)
				#Log.info('input nama petugas')
				hold(driver)

				find(By.ID, "kejaksaan").click()
				wait(driver,20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, Kejaksaan))).click()
				#Log.info('Pilih Kejaksaan')
				hold(driver)

				find(By.ID, "instansi_thn").send_keys("Kejaksaan Agung Republik Indonesia")
				#Log.info('Input Instansi')
				hold(driver)

				find(By.ID, "id_instansi_penyidik").click()
				wait(driver,20).until(EC.element_to_be_clickable((By.ID, instansiPenyidik))).click()
				#Log.info('pilih instansi penyidik')
				hold(driver)

				find(By.ID, "lokasi_dokumen").send_keys(LokasiDokumen)
				#Log.info('input lokasi dokumen')
				hold(driver)

				find(By.ID, "asal_tahanan").send_keys(AsalTahanan)
				#Log.info('input asal tahanan')
				hold(driver)

				find(By.ID, "kepolisian").send_keys(AsalTahanan)
				#Log.info('input kepolisian')
				hold(driver)

				# find(By.ID, "tgl_penangkapan").click()
				# find(By.ID, "tgl_penangkapan").send_keys(tgl_Penangkapan)
				# find(By.ID, "tgl_penangkapan").send_keys(Keys.ENTER)
				# Log.info('input tanggal penangkapan')
				# input('Tekan Enter Untuk Melanjutkan')
				hold(driver)

				find(By.ID, "tgl_ba8").click()
				find(By.ID, "tgl_ba8").send_keys(tgl_BA8)
				find(By.ID, "tgl_ba8").send_keys(Keys.ENTER)
				#Log.info('input tanggal BA 8')
				hold(driver)

				find(By.ID, "tgl_menjalani_putusan_akhir").click()
				find(By.ID, "tgl_menjalani_putusan_akhir").send_keys(tgl_putusan_akhir)
				find(By.ID, "tgl_menjalani_putusan_akhir").send_keys(Keys.ENTER)
				#Log.info('input tanggal menjalani putusan akhir')
				hold(driver)

				find(By.CSS_SELECTOR, ".el-tooltip__trigger:nth-child(5) > .el-form-item__label").click()
				find(By.ID, "jenis_putusan").click()
				find(By.ID, Jenis_putusan).click()
				#Log.info('input jenis putusan')
				hold(driver)

				# find(By.ID, "tgl_menjalani_cabutpb").click()
				# find(By.ID, "tgl_menjalani_cabutpb").send_keys(tgl_menjalani_cabutpb)
				# find(By.ID, "tgl_menjalani_cabutpb").send_keys(Keys.ENTER)
				#Log.info('tanggal menjalani pencaputan PB')
				hold(driver)


				find(By.ID, "jatuh_vonis").click()
				wait(driver,20).until(EC.element_to_be_clickable((By.ID, "EKS3"))).click()
		
				#Log.info('jatuh vonis')

				find(By.CSS_SELECTOR, ".el-form-item:nth-child(6) .el-switch__core").click()

				find(By.ID, "eksekusi_jaksa").click()
				find(By.ID, "EKS1").click()
				#Log.info('input eksekusi jaksa')
				hold(driver)

			
				find(By.CSS_SELECTOR, ".el-form-item:nth-child(22) > .el-form-item__content").click()

				find(By.CSS_SELECTOR, ".el-row:nth-child(2) > .el-col-xs-24:nth-child(2)").click()
				hold(driver)

# ============================================================================================================================================================================
# TAB PERKARA
			print('pilih TAB perkara' )
			find(By.ID, "tab-perkara").click()
			Log.info('pilih TAB perkara')
			hold(driver)

			# find(By.CSS_SELECTOR, "div:nth-child(4) > .el-form-item .el-switch__action").click()
		
			time.sleep(1)
			wait(driver,20).until(EC.element_to_be_clickable((By.ID, "jam_kejadian"))).click()

			find(By.ID, "tgl_kejadian").send_keys(tgl_BA8)
			#Log.info('input tanggal kejadian')
			hold(driver)

			find(By.ID, "tempat_kejadian").send_keys(tempatKejahatan)
			#Log.info('input tempat kejadian')
			hold(driver)

			find(By.ID, "risalah_kejadian_perkara").send_keys(fake.text())
			#Log.info('input risalah kejadian perkara')
			hold(driver)
			
			# input('Tekan Enter Untuk Melanjutkan')
			# find(By.CSS_SELECTOR, ".is-required .el-switch__action").click()

			hold(driver)

			find(By.ID, "kejahatan.0.deskripsi").click()
			find(By.ID, "kejahatan.0.deskripsi").send_keys(KejahatanUtama)
			find(By.ID, "kejahatan.0.uu_kejahatan").send_keys(KejahatanUtama)
			#Log.info('input kejahatan utama')
			hold(driver)

			find(By.ID, "kejahatan.0.pasal_utama").click()
			find(By.ID, "kejahatan.0.pasal_utama").send_keys("pasal " + KejahatanUtama)
			#Log.info('input pasal utama')
			hold(driver)

			find(By.ID, "kejahatan.0.id_terminologi").click()
			wait(driver,20).until(EC.element_to_be_clickable((By.ID, terminologi))).click()

			#Log.info('input Kejahatan')
			hold(driver)

			find(By.ID, "kejahatan.0.wilayah").click()
			find(By.ID, "kejahatan.0.wilayah").send_keys(tempatKejahatan)
			#Log.info('input tempat kejadian')
			hold(driver)
			

			if jenis_registrasi == 'B I' or jenis_registrasi == 'B II A' or jenis_registrasi == 'B II B' or jenis_registrasi == 'B III':
# ============================================================================================================================================================================
# TAB PUTUSAN
				if Jenis_putusan == 'PPN':
					find(By.ID, "tab-putusan_pengadilan_negeri").click()
					#Log.info('pilih jenis putusan Pengadilan Negeri')
					hold(driver)

					find(By.XPATH, "//form/div[5]/div/div/div/span").click()
					hold(driver)

					find(By.ID, "tgl_putusan-0").click()
					find(By.ID, "tgl_putusan-0").send_keys(Tanggal_Putusan)
					#Log.info('Input tanggal putusan akhir')
					hold(driver)

					find(By.CSS_SELECTOR, "div:nth-child(5) > .el-form-item:nth-child(1) > .el-form-item__content:nth-child(2)").click()

					find(By.ID, "nmr_putusan-0").send_keys(NomorPutusan)
					#Log.info('input nomor putusan ')
					hold(driver)

					find(By.CSS_SELECTOR, ".px-2").click()

					find(By.ID, "pasal-0").send_keys(KejahatanUtama)
					#Log.info('input pasal')
					hold(driver)

					find(By.ID, "hakim_ketua-0").send_keys(namahakimketua)
					#Log.info('Input nama hakim ketua')
					hold(driver)


					find(By.ID, "hakim_anggota1-0").send_keys(HakimAnggota1)
					#Log.info('input hakim anggota 1')
					hold(driver)

					find(By.ID, "hakim_anggota2-0").send_keys(HakimAnggota2)
					#Log.info('input hakim anggota 2')
					hold(driver)

					find(By.CSS_SELECTOR, ".el-form > div:nth-child(5)").click()
					find(By.ID, "panitera-0").click()
					find(By.ID, "panitera-0").send_keys(panitera)
					#Log.info('input panitera')
					hold(driver)

					find(By.ID, "jaksa-0").click()
					find(By.ID, "jaksa-0").send_keys(NamaJaksa)
					#Log.info('input nama jaksa')
					hold(driver)

					find(By.XPATH, "//input[@id='instansi-0']").click()
					find(By.XPATH, "//li[contains(.,\'Pengadilan Negeri Jakarta Pusat\')]").click()
					#Log.info('input pengadilan')
					hold(driver)

					find(By.ID, "tgl_dijalankan_ptsn-0").click()
					find(By.ID, "tgl_dijalankan_ptsn-0").send_keys(tgl_BA8)
					find(By.ID, "tgl_dijalankan_ptsn-0").send_keys(Keys.ENTER)
					#Log.info('input tanggal dijalankan putusan')
					hold(driver)

					find(By.CSS_SELECTOR, "div:nth-child(5) > .el-form-item:nth-child(11) > .el-form-item__content").click()
					find(By.ID, "peranan_kejahatan-0").click()
					find(By.ID, "peranan_kejahatan-0").send_keys("pelaku utama")
					#Log.info('Input Tersangka')
					hold(driver)

					find(By.ID, "status_hukum_basan_baran-0").click()
					find(By.ID, "status_hukum_basan_baran-0-0").click()
					#Log.info('status hukum basan baran')
					hold(driver)

					find(By.ID, "id_jenis_hukuman-0").click()


					hold(driver)

					if jenis_hukuman == 'id_jenis_hukuman-0-0':
						find(By.ID, "id_jenis_hukuman-0-0").click()
						#Log.info('pilih jenis hukuman  mati')
						hold(driver)

					else :
						find(By.ID, "id_jenis_hukuman-0-1").click()
						#Log.info('pilih jenis hukuman pidana ')
						hold(driver)
						time.sleep(1)
						wait(driver,20).until(EC.element_to_be_clickable((By.ID, "thn_kurung-0")))
						find(By.ID, "thn_kurung-0").click()
						if jenis_registrasi == 'B II B':
							find(By.ID, "bln_kurung-0").send_keys(fake.random_int(min=2, max=3))
						elif jenis_registrasi == 'B II A':
							find(By.ID, "thn_kurung-0").send_keys(fake.random_int(min=1, max=1))
						elif jenis_registrasi == 'B I':
							find(By.ID, "thn_kurung-0").send_keys(fake.random_int(min=2, max=3))
						else:
							find(By.ID, "thn_kurung-0").send_keys(fake.random_int(min=2, max=3))

						
					
						#Log.info("input pidana")

						find(By.CSS_SELECTOR, ".px-2").click()
						find(By.ID, "jenis_remisi-0").click()
						find(By.ID, jenisRemisi).click()
						find(By.CSS_SELECTOR, ".el-form > div:nth-child(5)").click()
						#Log.info('input jenis remisi')

					# Log.info('input Denda')
					# find(By.ID, "denda-0").send_keys("1000000")
					# find(By.ID, "bln_sub_denda-0").send_keys(fake.random_int(min=1, max=2))

					# find(By.ID, "up-0").send_keys("1000000")
					# find(By.ID, "bln_sub_denda-0").send_keys(fake.random_int(min=1, max=2))

					# find(By.ID, "restitusi-0").send_keys("1000000")
					# find(By.ID, "bln-0").send_keys(fake.random_int(min=1, max=2))

				elif Jenis_putusan =='PPT':
					find(By.ID, "tab-putusan_pengadilan_tinggi").click()
					#Log.info('pilih jenis putusan')


				elif Jenis_putusan =='PMA':
					find(By.ID, "tab-putusan_mahkamah_agung").click()
	

				find(By.CSS_SELECTOR, ".is-always-shadow > .el-card__body > .el-form").click()
				find(By.ID, "tab-registrasi").click()
				
				find(By.ID, "tgl_msk_lapas").click()
				find(By.ID, "tgl_msk_lapas").send_keys(tgl_BA8)
				find(By.ID, "tgl_msk_lapas").send_keys(Keys.ENTER)

				find(By.ID, "tgl_akhir_ditahan").click()
				find(By.ID, "tgl_akhir_ditahan").send_keys(tgl_BA8)
				find(By.ID, "tgl_akhir_ditahan").send_keys(Keys.ENTER)
				#Log.info('input tanggal pertama ditahan')
				hold(driver)
				
				find(By.ID, "tgl_awal_tahan_golongan").click()
				find(By.ID, "tgl_awal_tahan_golongan").send_keys(tgl_BA8)
				find(By.ID, "tgl_awal_tahan_golongan").send_keys(Keys.ENTER)
				#Log.info('Input tanggal awal ditahan golongan')
				hold(driver)

				find(By.ID, "tgl_pertama_ditahan").click()
				find(By.ID, "tgl_pertama_ditahan").send_keys(tgl_BA8)
				find(By.ID, "tgl_pertama_ditahan").send_keys(Keys.ENTER)
				#Log.info('input tanggal pertama ditahan')
				hold(driver)



				InputDocRegis(driver)
			else:
				pass
			
			# input('Tekan Enter Untuk Melanjutkan')
			find(By.ID, 'submitButton').click() 
			wait(driver,50).until(EC.element_to_be_clickable((By.ID, 'cari')))
			# print(i + 'nomor urutan')

		

		except TimeoutException:
			print('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!LOADING TERLALU LAMA!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
			pass
		i = i + 1
		
print ('Success Created')
