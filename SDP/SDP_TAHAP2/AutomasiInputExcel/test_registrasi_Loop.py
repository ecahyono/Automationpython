from distutils.archive_util import make_archive
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("RegSelenium"))

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("RegSelenium"))


from Settings.setupkeamanan import initDriver, loadDataPath, quit, sleep
from Settings.loginkeamanan import *
from Settings.setupPembinaan import *

import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Registrasi.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)


sheetrange = wb['RegSelenium']

@mark.webtest()
def test_1_setupOS():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')

@mark.webtest()
def test_2_login():
    (driver)
    Log.info('Login')
    loginSumedang(driver)


@mark.webtest()
def test_Input_Registrasi():
    driver.implicitly_wait(60)
    driver.implicitly_wait(30)
    nav1 = driver.find_element(By.XPATH, pathData['AksesMenu']['Registrasi']['MainText'])
    ActionChains(driver).move_to_element(nav1).perform()
    element2 = driver.find_element(By.XPATH, pathData['AksesMenu']['Registrasi']['child']['ManajemenRegistrasi']['MainText'])
    time.sleep(1)
    ActionChains(driver).move_to_element(element2).perform()
    time.sleep(1)
    driver.find_element(By.LINK_TEXT, 'Registrasi Tahanan/ Narapidana').click()
    Log.info('Akses Menu Registrasi')
    attach(data=driver.get_screenshot_as_png())

    i = 2
        # nge baca mulai dari tabel A
    while i <= len(sheetrange['A']):
        # deklarasi bahwa NIP itu ada di A 
        NamaWBP                             = sheetrange['A'+str(i)].value
        NoRegistrasi                        = sheetrange['B'+str(i)].value
        NomorSuratPenahanan                 = sheetrange['C'+str(i)].value
        Tgl                                 = sheetrange['D'+str(i)].value
        Pidana                              = sheetrange['E'+str(i)].value
        namapetugas                         = sheetrange['F'+str(i)].value
        AsalInstansi                        = sheetrange['G'+str(i)].value
        LokasiDokumen                       = sheetrange['H'+str(i)].value
        asaltahanan                         = sheetrange['I'+str(i)].value
        kepolisian                          = sheetrange['J'+str(i)].value
        kejahatanutama                      = sheetrange['K'+str(i)].value
        uud                                 = sheetrange['L'+str(i)].value
        pasalutama                          = sheetrange['M'+str(i)].value
        tempat                              = sheetrange['N'+str(i)].value
        nomorputusan                        = sheetrange['O'+str(i)].value
        namahakimketua                      = sheetrange['P'+str(i)].value
        HakimAnggota1                       = sheetrange['Q'+str(i)].value
        HakimAnggota2                       = sheetrange['R'+str(i)].value
        panitera                            = sheetrange['S'+str(i)].value
        NamaJaksa                           = sheetrange['T'+str(i)].value
        tersangka                           = sheetrange['U'+str(i)].value
       
        pasal                               = sheetrange['V'+str(i)].value
      
        driver.implicitly_wait(30)


        try:
     
            driver.implicitly_wait(60)
            WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//span[contains(.,\'Registrasi Baru\')]')))
            driver.find_element(By.XPATH, "//span[contains(.,\'Registrasi Baru\')]").click()
            Log.info('Click Button Registrasi')
    
            WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'findButton')))
            driver.find_element(By.ID, "jenisRegistrasi").click()
            WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//li[contains(.,\'B I\')]')))
            time.sleep(0.5)
            driver.find_element(By.XPATH, "//li[contains(.,\'B I\')]").click()
            Log.info('Click Button Jenis Registrasi B1')

            driver.find_element(By.ID, "noRegistrasi").click()
            driver.find_element(By.ID, "noRegistrasi").send_keys(NamaWBP)
            Log.info('Input nomor registrasi wbp')
            
            WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[@id='noRegistrasiOption-0']/div/div/table/tbody/tr[2]/td")))
            driver.find_element(By.XPATH, "//li[@id='noRegistrasiOption-0']/div/div/table/tbody/tr[2]/td").click()
            Log.info('Click Pilih Data Registrasi')

            driver.find_element(By.ID, "findButton").click()
            Log.info('Click Button search')

            
            WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.el-dialog__headerbtn')))
            driver.find_element(By.CSS_SELECTOR, ".el-dialog__headerbtn").click()
            Log.info('close Pop Up')


            driver.find_element(By.ID, "nmr_reg_gol").click()
            driver.find_element(By.ID, "nmr_reg_gol").send_keys(NoRegistrasi)
            Log.info('input no registrasi ')

            driver.find_element(By.ID, "tgl_srt_thn").click()
            driver.find_element(By.ID, "tgl_srt_thn").send_keys(Tgl)
            driver.find_element(By.ID, "tgl_srt_thn").send_keys(Keys.ENTER)
            Log.info('input tanggal surat ')
  
            driver.find_element(By.ID, "nmr_srt_thn").click()
            driver.find_element(By.CSS_SELECTOR, ".el-col > .is-required:nth-child(3)").click()
            driver.find_element(By.ID, "nmr_srt_thn").send_keys(NomorSuratPenahanan)
            driver.find_element(By.CSS_SELECTOR, ".el-row:nth-child(4) > .el-col:nth-child(1)").click()
            driver.find_element(By.CSS_SELECTOR, ".is-success:nth-child(3) > .el-form-item__label").click()
            Log.info('input nomor surat penahanan')

            driver.find_element(By.ID, "nm_pjbt_thn").click()
            driver.find_element(By.ID, "nm_pjbt_thn").send_keys(namapetugas)
            Log.info('input nama petugas')

            driver.find_element(By.ID, "kejaksaan").click()
            driver.find_element(By.CSS_SELECTOR, "#kejaksaanOption-0 > span").click()
            Log.info('Pilih Kejaksaan')

            driver.find_element(By.ID, "instansi_thn").click()
            driver.find_element(By.ID, "instansi_thn").send_keys(AsalInstansi)
            Log.info('pilih asal instansi')

            driver.find_element(By.ID, "id_instansi_penyidik").click()
            driver.find_element(By.ID, "id_instansi_penyidik-0").click()
            Log.info('pilih instansi penyidik')

            driver.find_element(By.ID, "lokasi_dokumen").click()
            driver.find_element(By.ID, "lokasi_dokumen").send_keys(LokasiDokumen)
            Log.info('input lokasi dokumen')

            driver.find_element(By.ID, "asal_tahanan").click()
            driver.find_element(By.ID, "asal_tahanan").send_keys(asaltahanan)
            Log.info('input asal tahanan')

            driver.find_element(By.ID, "kepolisian").click()
            driver.find_element(By.ID, "kepolisian").send_keys(kepolisian)
            Log.info('Input asal kepolisian')

            driver.find_element(By.ID, "tgl_penangkapan").click()
            driver.find_element(By.ID, "tgl_penangkapan").send_keys(Tgl)
            driver.find_element(By.ID, "tgl_penangkapan").send_keys(Keys.ENTER)
            Log.info('input tanggal penangkapan')

            driver.find_element(By.ID, "tgl_ba8").click()
            driver.find_element(By.ID, "tgl_ba8").send_keys(Tgl)
            driver.find_element(By.ID, "tgl_ba8").send_keys(Keys.ENTER)
            Log.info('input tanggal BA 8')

            driver.find_element(By.ID, "tgl_menjalani_putusan_akhir").click()
            driver.find_element(By.ID, "tgl_menjalani_putusan_akhir").send_keys(Tgl)
            driver.find_element(By.ID, "tgl_menjalani_putusan_akhir").send_keys(Keys.ENTER)
            Log.info('input tanggal menjalani putusan akhir')


            driver.find_element(By.CSS_SELECTOR, ".el-tooltip__trigger:nth-child(5) > .el-form-item__label").click()
            driver.find_element(By.ID, "jenis_putusan").click()
            driver.find_element(By.ID, "PPN").click()
            Log.info('input jenis putusan')

            driver.find_element(By.ID, "tgl_menjalani_cabutpb").click()
            driver.find_element(By.ID, "tgl_menjalani_cabutpb").send_keys(Tgl)
            driver.find_element(By.ID, "tgl_menjalani_cabutpb").send_keys(Keys.ENTER)
            Log.info('tanggal menjalani pencaputan PB')

            driver.find_element(By.CSS_SELECTOR, ".el-form-item:nth-child(22) > .el-form-item__content").click()

            driver.find_element(By.ID, "tgl_ba8").click()
            driver.find_element(By.ID, "tgl_ba8").send_keys(Tgl)
            driver.find_element(By.ID, "tgl_ba8").send_keys(Keys.ENTER)

            driver.find_element(By.CSS_SELECTOR, ".el-row:nth-child(2) > .el-col-xs-24:nth-child(2)").click()
            
            driver.find_element(By.ID, "tab-perkara").click()
            driver.find_element(By.CSS_SELECTOR, "div:nth-child(4) > .el-form-item .el-switch__action").click()

            driver.find_element(By.CSS_SELECTOR, ".is-required .el-switch__action").click()

            driver.find_element(By.ID, "kejahatan.0.deskripsi").click()
            driver.find_element(By.ID, "kejahatan.0.deskripsi").send_keys(kejahatanutama)
            driver.find_element(By.ID, "kejahatan.0.uu_kejahatan").send_keys(uud)
            Log.info('input kejahatan utama')

            driver.find_element(By.ID, "kejahatan.0.pasal_utama").click()
            driver.find_element(By.ID, "kejahatan.0.pasal_utama").send_keys(pasalutama)
            driver.find_element(By.ID, "kejahatan.0.id_terminologi").click()
            Log.info('input pasal utama')

            driver.find_element(By.ID, "kejahatan.0.id_terminologi").send_keys("kesusilaan")
            driver.find_element(By.ID, "kejahatan.0.id_terminologi-7").click()
            Log.info('input Kejahatan')

            driver.find_element(By.ID, "kejahatan.0.wilayah").click()
            driver.find_element(By.ID, "kejahatan.0.wilayah").send_keys(tempat)
            Log.info('input tempat kejadian')

            driver.find_element(By.ID, "tab-putusan_pengadilan_negeri").click()
            Log.info('pilih jenis putusan')

            driver.find_element(By.ID, "tgl_putusan-0").click()
            driver.find_element(By.ID, "tgl_putusan-0").send_keys(Tgl)
            Log.info('Input tanggal putusan akhir')

            driver.find_element(By.CSS_SELECTOR, "div:nth-child(5) > .el-form-item:nth-child(1) > .el-form-item__content:nth-child(2)").click()

            time.sleep(2)
            driver.find_element(By.ID, "nmr_putusan-0").click()
            driver.find_element(By.ID, "nmr_putusan-0").send_keys(nomorputusan)
            Log.info('input nomor putusan ')

            driver.find_element(By.CSS_SELECTOR, ".px-2").click()

            driver.find_element(By.ID, "pasal-0").send_keys(pasal)
            Log.info('input pasal')

            time.sleep(2)
            driver.find_element(By.ID, "hakim_ketua-0").click()
            driver.find_element(By.ID, "hakim_ketua-0").send_keys(namahakimketua)
            Log.info('Input nama hakim ketua')

            time.sleep(2)
            driver.find_element(By.CSS_SELECTOR, ".px-2").click()
            driver.find_element(By.ID, "hakim_anggota1-0").send_keys(HakimAnggota1)
            Log.info('input hakim anggota 1')

            time.sleep(2)
            driver.find_element(By.ID, "hakim_anggota2-0").click()
            driver.find_element(By.ID, "hakim_anggota2-0").send_keys(HakimAnggota2)
            Log.info('input hakim anggota 2')

            driver.find_element(By.CSS_SELECTOR, ".el-form > div:nth-child(5)").click()
            driver.find_element(By.ID, "panitera-0").click()
            driver.find_element(By.ID, "panitera-0").send_keys(panitera)
            Log.info('input panitera')

            driver.find_element(By.ID, "jaksa-0").click()
            driver.find_element(By.ID, "jaksa-0").send_keys(NamaJaksa)
            Log.info('input nama jaksa')

            driver.find_element(By.XPATH, "//input[@id='instansi-0']").click()
            driver.find_element(By.XPATH, "//li[contains(.,\'Pengadilan Negeri Jakarta Pusat\')]").click()
            Log.info('input pengadilan')

            driver.find_element(By.ID, "tgl_dijalankan_ptsn-0").click()
            driver.find_element(By.ID, "tgl_dijalankan_ptsn-0").send_keys(Tgl)
            driver.find_element(By.ID, "tgl_dijalankan_ptsn-0").send_keys(Keys.ENTER)
            Log.info('input tanggal dijalankan putusan')

            driver.find_element(By.CSS_SELECTOR, "div:nth-child(5) > .el-form-item:nth-child(11) > .el-form-item__content").click()
            driver.find_element(By.ID, "peranan_kejahatan-0").click()
            driver.find_element(By.ID, "peranan_kejahatan-0").send_keys(tersangka)
            Log.info('Input Tersangka')

            driver.find_element(By.ID, "status_hukum_basan_baran-0").click()
            driver.find_element(By.ID, "status_hukum_basan_baran-0-0").click()
            Log.info('status hukum basan baran')

            driver.find_element(By.ID, "id_jenis_hukuman-0").click()
            driver.find_element(By.ID, "id_jenis_hukuman-0-1").click()
            Log.info('jenis hukuman')

            driver.find_element(By.ID, "thn_kurung-0").click()
            driver.find_element(By.ID, "thn_kurung-0").send_keys(Pidana)
            Log.info("input pidana")
            
            driver.find_element(By.CSS_SELECTOR, ".px-2").click()
            driver.find_element(By.ID, "jenis_remisi-0").click()
            driver.find_element(By.ID, "jenis_remisi-0-0").click()
            driver.find_element(By.CSS_SELECTOR, ".el-form > div:nth-child(5)").click()
            Log.info('input jenis remisi')

            driver.find_element(By.ID, "tgl_ekspirasi_perkiraan-0").click()
            driver.find_element(By.ID, "tgl_ekspirasi_perkiraan-0").send_keys(Tgl)
            driver.find_element(By.ID, "tgl_ekspirasi_perkiraan-0").send_keys(Keys.ENTER)
            Log.info('Input tanggal ekspirasi perkiraan')

            driver.find_element(By.CSS_SELECTOR, ".is-always-shadow > .el-card__body > .el-form").click()
            driver.find_element(By.ID, "tab-registrasi").click()
            driver.find_element(By.ID, "tgl_pertama_ditahan").click()
            driver.find_element(By.ID, "tgl_pertama_ditahan").send_keys(Tgl)
            driver.find_element(By.ID, "tgl_pertama_ditahan").send_keys(Keys.ENTER)
            time.sleep(3)
            Log.info('Input tanggal pertama ditahan')


            driver.find_element(By.ID, "tgl_akhir_ditahan").click()
            driver.find_element(By.ID, "tgl_akhir_ditahan").send_keys(Tgl)
            driver.find_element(By.ID, "tgl_akhir_ditahan").send_keys(Keys.ENTER)
            Log.info('Input tanggal akhir ditahan')

            driver.find_element(By.CSS_SELECTOR, ".el-row:nth-child(4) .el-form-item__content").click()

            driver.find_element(By.ID, "tgl_awal_tahan_golongan").click()
            driver.find_element(By.ID, "tgl_awal_tahan_golongan").send_keys(Tgl)
            driver.find_element(By.ID, "tgl_awal_tahan_golongan").send_keys(Keys.ENTER)
            Log.info('Input tanggal awal ditahan golongan')

            driver.find_element(By.CSS_SELECTOR, ".el-row:nth-child(4) .el-form-item__content").click()
            driver.find_element(By.CSS_SELECTOR, ".el-row:nth-child(3) > .el-col:nth-child(2) svg").click()

            driver.find_element(By.ID, "tgl_msk_lapas").click()
            driver.find_element(By.ID, "tgl_msk_lapas").send_keys(Tgl)
            driver.find_element(By.ID, "tgl_msk_lapas").send_keys(Keys.ENTER)
            time.sleep(4)
            Log.info('Input tanggal masuk lapas')

            driver.find_element(By.ID, "tab-file_dokumen").click()



            WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, "chooseFile-0")))
            driver.find_element(By.ID, "chooseFile-0").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-1").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-3").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-5").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-6").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-15").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-8").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-16").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-17").click()
            upload(driver)

            driver.find_element(By.ID, "chooseFile-20").click()
            upload(driver)

            

            

            driver.find_element(By.ID, "submitButton").click()
            Log.info('Klik Button Submit')

            
        except TimeoutException:
            print("MASIH ADA ERROR, CEK LAGI PAK WIL")
            pass
            quit(driver)
        time.sleep(5)
        i = i + 1

  
        

@mark.webtest()
def test_exit():
    quit(driver)




