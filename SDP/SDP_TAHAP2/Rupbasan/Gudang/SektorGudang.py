# import imp
from json import load
import pyautogui
import string
from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
import pyautogui

from openpyxl import load_workbook
import time

#target halaman excel ada dimana , wb = variablenya
# wb = load_workbook(filename="C:\chromedriver\Data.xlsx")
wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Filexel\Rupbasan.xlsx")

# jadi ini bisa read sheet yang dibawah itu yang di excel
sheetrange = wb['Sektor Gudang']

# ini web driver disimpen dimana, kalo disimpen di path kosongin aja
driver = webdriver.Chrome(r'C:\Users\user\Documents\TRCH\chromedriver.exe')

# link nya ini dimana
# driver.get("http://192.168.2.11:32400/")
driver.get("http://kumbang.torche.id:32400/")
# seting windows nya jadi max   
driver.maximize_window()
# script gakan di eksekusi kalo web ga muncul. kalo lebih dari 10 detik ga muncul error
driver.implicitly_wait(6)
# ini letak xpath icon login
driver.find_element(By.XPATH, "//div/span").click()
# ini masuk ke form input username
driver.find_element(By.ID, "username").click()
# masukin input username
driver.find_element(By.ID, "username").send_keys("oprupbasanbdg")
# masukin input password
driver.find_element(By.ID, "password").send_keys("password")
# click button login
driver.find_element(By.ID, "kc-login").click()
time.sleep(3)

element = driver.find_element(By.XPATH, '//*[@id="app"]/div/nav/ul/li[2]/div')
# element = driver.find_element(By.XPATH, '//*[@id="app"]/div/nav/ul/li[3]/div')
# element = driver.find_element(By.XPATH, '//*[@id="app"]/div/nav/ul/li[4]/div')
actions = ActionChains(driver)
time.sleep(1)
actions.move_to_element(element).perform()

element2 = driver.find_element(By.XPATH, '//div[3]/div/ul/li[1]/div[1]')
# element2 = driver.find_element(By.XPATH, '//div[4]/div/ul/li[1]/div[1]')
# element2 = driver.find_element(By.XPATH, '//div[6]/div/ul/li[1]/div[1]')
# element2 = driver.find_element(By.XPATH, '//div[5]/div/ul/li[1]/div[1]')
actions2 = ActionChains(driver)
time.sleep(1)
actions2.move_to_element(element2).perform()

driver.find_element(By.LINK_TEXT, "Sektor Gudang").click()

i = 6 

while i <= len(sheetrange['A']):
	Gudang 				= sheetrange['A'+str(i)].value
	NamaSektor 			= sheetrange['B'+str(i)].value
	KapasitasSektor 	= sheetrange['C'+str(i)].value
	Keterangan 			= sheetrange['D'+str(i)].value
	JumlahBaris 	   	= sheetrange['E'+str(i)].value
	Baris  				= sheetrange['F'+str(i)].value
	JumlahNORUT 		= sheetrange['G'+str(i)].value

	N1 			= sheetrange['H'+str(i)].value
	N2			= sheetrange['I'+str(i)].value
	N3 			= sheetrange['J'+str(i)].value
	N4 			= sheetrange['K'+str(i)].value
	N5 			= sheetrange['L'+str(i)].value
	N6 			= sheetrange['M'+str(i)].value
	N7 			= sheetrange['N'+str(i)].value
	N8 			= sheetrange['O'+str(i)].value
	N9 			= sheetrange['P'+str(i)].value
	N10 		= sheetrange['Q'+str(i)].value

	driver.find_element(By.ID, "createButton").click()

	try:
		gd = driver.find_element(By.ID, 'input_kondisi_baran_basan')
		gd.click()
		WebDriverWait(driver, 90).until(EC.element_to_be_clickable((By.ID, 'chooseGudangOption-0')))
		if Gudang == 'Gudang Umum Terbuka':
			driver.find_element(By.ID,'chooseGudangOption-0').click()
		elif Gudang == 'Gudang Umum Tertutup':
			driver.find_element(By.ID,'chooseGudangOption-1').click()
		elif Gudang == 'Gudang Berharga':
			driver.find_element(By.ID,'chooseGudangOption-2').click()
		elif Gudang == 'Gudang Berbahaya':
			driver.find_element(By.ID,'chooseGudangOption-3').click()
		elif Gudang == 'Gudang Hewan dan Tumbuhan':
			driver.find_element(By.ID,'chooseGudangOption-4').click()

		driver.find_element(By.ID, 'deskripsi').send_keys(NamaSektor)
		driver.find_element(By.ID, 'volumeKapasitas').send_keys(KapasitasSektor)
		driver.find_element(By.ID, 'keterangan').send_keys(Keterangan)

		driver.find_element(By.ID, 'inputBaris-0').send_keys(Baris)

		norut = driver.find_element(By.ID, 'action-0')

		for x in range(9):
			x = norut.click()

		time.sleep(2)
		driver.find_element(By.ID, 'inputNoUrut-0').send_keys(N1)
		driver.find_element(By.ID, 'inputNoUrut-1').send_keys(N2)
		driver.find_element(By.ID, 'inputNoUrut-2').send_keys(N3)
		driver.find_element(By.ID, 'inputNoUrut-3').send_keys(N4)
		driver.find_element(By.ID, 'inputNoUrut-4').send_keys(N5)
		driver.find_element(By.ID, 'inputNoUrut-5').send_keys(N6)
		driver.find_element(By.ID, 'inputNoUrut-6').send_keys(N7)
		driver.find_element(By.ID, 'inputNoUrut-7').send_keys(N8)
		driver.find_element(By.ID, 'inputNoUrut-8').send_keys(N9)
		driver.find_element(By.ID, 'inputNoUrut-9').send_keys(N10)

		driver.find_element(By.ID, 'submitButton').click()
		WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/div[2]/form/div[1]/div/div/button')))

	except TimeoutException:
		pass
	i = i + 1
print ("Success Created")

# 