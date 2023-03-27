from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from pytest_html_reporter import attach
import os, platform, time, pytest
from selenium import webdriver
from os import environ, path
from pathlib import Path
from pytest import mark
import pyautogui
import platform
import logging
import sys
from openpyxl import load_workbook
from dotenv import load_dotenv
load_dotenv()

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR")) 
elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))

from Settings.setupbrowser import initDriver, loadDataPath, secondaryinit
from Settings.login import login, oprupbasanbdg
from Settings.Page.Rupbasan import daftarpegawai

Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Rupbasan_pegawairupbasan.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

wb = load_workbook(environ.get("RUPEXEL"))

# init driver by os
@mark.fixture_pegawai
def testconfigandlogin():
	Log.info('Konfigurasi agar berjalan di setiap sistem operasi (mac dan Windos)')
	global driver, pathData
	# driver = initDriver()
	driver = secondaryinit()
	pathData = loadDataPath()
	Log.info('Memasukan User name dan Password di halaman Login')
	# oprupbasanbdg(driver)

@mark.fixture_pegawai
def test_menmabahpetugasrupbasan():
	tambah = wb['Pegawairupbasan']
	inputbaris = 3
	for row in tambah.iter_rows(min_row=inputbaris, max_row=inputbaris, values_only=True):
		nip             = row[0] 
		nama            = row[1] 
		tempatlahir     = row[2] 
		tanggallahir    = row[3] 
		jeniskelamin    = row[4] 
		alamat          = row[5] 
		jabatan         = row[6] 
		pangkat         = row[7] 
		golongan        = row[8] 
		bagian          = row[9] 
		email           = row[10] 
		telepon         = row[11] 
		
		Log.info('Mengakses Halaman')
		daftarpegawai(driver)
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'createButton')))
		driver.find_element(By.ID, 'createButton').click()
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'Nip')))
		driver.find_element(By.ID, "Nip").send_keys(nip)
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'Nama')))
		driver.find_element(By.ID, "Nama").send_keys(nama)
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'tempatLahir')))
		driver.find_element(By.ID, "tempatLahir").send_keys(tempatlahir)
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'tglLahir')))
		tgllhrpgw = driver.find_element(By.ID, "tglLahir")
		tgllhrpgw.send_keys(tanggallahir)
		tgllhrpgw.send_keys(Keys.ENTER)
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'jenisKelamin')))
		driver.find_element(By.ID, "jenisKelamin").click()
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'Laki-laki')))
		driver.find_element(By.XPATH, "//li[contains(.,'"+ jeniskelamin+"')]").click()
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'Alamat')))
		driver.find_element(By.ID, "Alamat").send_keys(alamat)
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'Jabatan')))
		driver.find_element(By.ID, "Jabatan").send_keys(jabatan)
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'Pangkat')))
		driver.find_element(By.ID, "Pangkat").send_keys(pangkat)
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'Golongan')))
		driver.find_element(By.ID, "Golongan").send_keys(golongan)
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'Bagian')))
		driver.find_element(By.ID, "Bagian").send_keys(bagian)
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'Email')))
		driver.find_element(By.ID, "Email").send_keys(email)
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'Telepon')))
		driver.find_element(By.ID, "Telepon").send_keys(telepon)   
		driver.find_element(By.ID, 'submitButton').click()
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'createButton')))

