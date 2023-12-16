from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import os, sys
from os import environ, path
import pyautogui
from pytest import mark
import pytest
import time
import platform
from pathlib import Path
import logging
from faker import Faker
from datetime import datetime
import openpyxl
from faker.providers import date_time
from datetime import datetime, timedelta
import random
from pytest_html_reporter import attach
import pyautogui

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook


fake = Faker('id_ID')
from dotenv import load_dotenv
load_dotenv()

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUAT"))

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))


from Settings.setup import *
from Settings.login import *
from Settings.Page.accessmenu import *


global driver, pathData
driver = initDriver()
pathData = loadDataPath()
