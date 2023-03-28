from distutils.archive_util import make_archive
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from selenium.webdriver.firefox.options import Options

options = Options()
options.add_argument('--disable-blink-features=AutomationControlled')
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUAT"))

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))


from Settings.setup import initDriver, loadDataPath, quit, sleep
from Settings.login import op_keamanan_mp, SpvRutanBdg
from Settings.Page.keamanan import manajemenblokdankamar

import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('LogBlokDanKamar.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

sheetrange = wb['tambahBlokdanKamar']

i = 2

filterColumn                            = sheetrange['B'+str(i)].value
namablok                                = sheetrange['C'+str(i)].value
tipeblok                                = sheetrange['D'+str(i)].value
jeniskelamin                            = sheetrange['E'+str(i)].value
jenisKejahatan                          = sheetrange['F'+str(i)].value
jumlahLantai                            = sheetrange['G'+str(i)].value
formatPenomoranKamar                    = sheetrange['H'+str(i)].value
kelUsia                                 = sheetrange['I'+str(i)].value
jumlahKamarPerlantai                    = sheetrange['J'+str(i)].value
jumlahKolomPerlantai                    = sheetrange['K'+str(i)].value

TambahKamar = wb['tambahBlokdanKamar']


namablok                                = TambahKamar['C'+str(i)].value
lantai                                  = TambahKamar['M'+str(i)].value
nomorKamar                              = TambahKamar['N'+str(i)].value
kelompokJenisKejahatan                  = TambahKamar['O'+str(i)].value
kapasitasInput                          = TambahKamar['P'+str(i)].value
kondisiRuangan                          = TambahKamar['Q'+str(i)].value

VerifikasiBlok = wb['tambahBlokdanKamar']

filterColumn                            = VerifikasiBlok['B'+str(i)].value
namablok                                = VerifikasiBlok['C'+str(i)].value
verifikasi                              = VerifikasiBlok['S'+str(i)].value
keterangan                              = VerifikasiBlok['S'+str(i)].value


    #=============== Halaman Tambah Blok Operaotor =============== 
@mark.fixture_test()
def test_1_setupOS():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')

@mark.fixture_test()
def test_2_login():
    op_keamanan_mp(driver)
    Log.info('Login Operator Manajemen Penempatan')


@mark.fixture_test()
def test_3_AksesMenu():
    manajemenblokdankamar(driver)
    Log.info('Akses halaman Manajemen Blok dan Kamar')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_3_ClickButtonCreate():
    sleep(driver)
    driver.implicitly_wait(10)
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.border:nth-child(1)')))
    driver.find_element(By.XPATH, '//*[@id="createButton"]').click()
    print('.')
    Log.info('Click Button Tambah halaman Pemetaan Block')
    attach(data=driver.get_screenshot_as_png())


@mark.fixture_test()
def test_4_CreateButtonMaster():
    driver.implicitly_wait(10)
    sleep(driver)
    WebDriverWait(driver,30).until(EC.element_to_be_clickable((By.ID, 'createButton')))
    driver.find_element(By.ID, 'createButton').click()
    Log.info('Click Button Tambah Master Blok Dan Kamar')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_5_InputNamaBlok():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, 'blok')))
    driver.find_element(By.ID, 'blok').send_keys(namablok)
    Log.info('Input Nama Blok')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_6_InputTipeBlok():
    print('== NEXT == / input tipe blok')
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, 'tipe_id')))
    driver.find_element(By.ID, 'tipe_id').click()
    Log.info('Input Tipe Blok')

    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,'"+ tipeblok +"')]")))
    driver.find_element(By.XPATH, "//li[contains(.,'"+ tipeblok +"')]").click()
    Log.info(' input tipe blok ')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_7_InputJenisKelamin():
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, 'kel_jenis_kelamin_id')))
    driver.find_element(By.ID, "kel_jenis_kelamin_id").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,'"+ jeniskelamin +"')]")))
    driver.find_element(By.XPATH, "//li[contains(.,'"+ jeniskelamin +"')]").click()
    Log.info(' input jenis kelamin ')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_8_InputJenisKejahatan():
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.el-select__input')))
    driver.find_element(By.CSS_SELECTOR, ".el-select__input").click()

    if jenisKejahatan == 'Korupsi Teroris Kriminal':
        driver.find_element(By.XPATH, "//li[contains(.,\'Korupsi\')]").click()
        driver.find_element(By.XPATH, "//li[contains(.,\'Teroris\')]").click()
        driver.find_element(By.XPATH, "//li[contains(.,\'Kriminal Umum\')]").click()
        time.sleep(1)
        driver.find_element(By.ID, "jenis_kejahatan_values").send_keys(Keys.TAB)
        Log.info('Input untuk Jenis kejahatan Teroris, Kriminal, Korupsi')
        attach(data=driver.get_screenshot_as_png())

    elif jenisKejahatan == 'Korupsi':
        driver.find_element(By.XPATH, "//li[contains(.,\'Korupsi\')]").click()
        time.sleep(1)
        driver.find_element(By.ID, "jenis_kejahatan_values").send_keys(Keys.TAB)
        Log.info('Input untuk Jenis kejahatan Korupsi')
        attach(data=driver.get_screenshot_as_png())

    elif jenisKejahatan == 'Teroris':
        driver.find_element(By.XPATH, "//li[contains(.,\'Teroris\')]").click()
        time.sleep(1)
        driver.find_element(By.ID, "jenis_kejahatan_values").send_keys(Keys.TAB)
        Log.info('Input untuk Jenis kejahatan Teroris')
        attach(data=driver.get_screenshot_as_png())

    elif jenisKejahatan == 'Kriminal':
        driver.find_element(By.XPATH, "//li[contains(.,\'Kriminal\')]").click()
        time.sleep(1)
        driver.find_element(By.ID, "jenis_kejahatan_values").send_keys(Keys.TAB)
        Log.info('Input untuk Jenis kejahatan Kriminal')
        attach(data=driver.get_screenshot_as_png())

    elif jenisKejahatan == 'Korupsi Teroris':
        driver.find_element(By.XPATH, "//li[contains(.,\'Korupsi\')]").click()
        driver.find_element(By.XPATH, "//li[contains(.,\'Teroris\')]").click()
        time.sleep(1)
        driver.find_element(By.ID, "jenis_kejahatan_values").send_keys(Keys.TAB)
        Log.info('Input untuk Jenis kejahatan Korupsi,Teroris')
        attach(data=driver.get_screenshot_as_png())

    elif jenisKejahatan == 'Teroris Kriminal':
        driver.find_element(By.XPATH, "//li[contains(.,\'Teroris\')]").click()
        driver.find_element(By.XPATH, "//li[contains(.,\'Kriminal Umum\')]").click()
        time.sleep(1)
        driver.find_element(By.ID, "jenis_kejahatan_values").send_keys(Keys.TAB)
        Log.info('Input untuk Jenis kejahatan Teroris, Kriminal')
        attach(data=driver.get_screenshot_as_png())
        
@mark.fixture_test()
def test_9_InputKelompokUsia():
    print('== NEXT == / kelompok usia')
    sleep(driver)   
    driver.find_element(By.ID, "kel_usia_id").click()
    driver.find_element(By.XPATH, "//li[contains(.,\'"+ kelUsia +"')]").click()
    Log.info(' input Kelompok Usia ')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_10_InputJumlahLantai():
    print('== NEXT == / masukan jumlah lantai')
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, '// *[ @ id = "Masukkan Jumlah Lantai"]')))
    driver.find_element(By.XPATH, '// *[ @ id = "Masukkan Jumlah Lantai"]').send_keys(jumlahLantai)
    Log.info('input jumlah lantai')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_11_InputFormatPenomoran():
    print('== NEXT == / penomoran kamar')
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, 'penomoran_kamar')))
    driver.find_element(By.ID, "penomoran_kamar").click()
    if formatPenomoranKamar == 'Angka 8 mulai kiri atas':
        driver.find_element(By.XPATH, "//li[contains(.,\'Angka 8 (mulai kiri atas)\')]").click()
        Log.info('Input Format Penomoran Kamar')
        attach(data=driver.get_screenshot_as_png())

    elif formatPenomoranKamar == '8 mulai kiri bawah':
        driver.find_element(By.XPATH, "//li[contains(.,\'Angka 8 (mulai kiri bawah)\')]").click()
        Log.info('Input Format Penomoran Kamar 8 mulai kiri bawah')
        attach(data=driver.get_screenshot_as_png())

    elif formatPenomoranKamar == '8 mulai kanan bawah':
        driver.find_element(By.XPATH, "//li[contains(.,\'Angka 8 (mulai kanan bawah)\')]").click()
        Log.info('Input Format Penomoran Kamar Angka 8 mulai kanan bawah')
        attach(data=driver.get_screenshot_as_png())

    elif formatPenomoranKamar == 'Ke Depan 1 arah':
        driver.find_element(By.XPATH, "//li[contains(.,\'Belakang Ke Depan (1 arah)\')]").click()
        Log.info('Input Format Penomoran Kamar Belakang Ke Depan 1 arah')
        attach(data=driver.get_screenshot_as_png())

    elif formatPenomoranKamar ==  'Ke Depan 2 arah':
        driver.find_element(By.XPATH, "//li[contains(.,\'Belakang Ke Depan (2 arah)\')]").click()
        Log.info('Input Format Penomoran Kamar Belakang Ke Depan 2 arah')
        attach(data=driver.get_screenshot_as_png())

    elif formatPenomoranKamar == 'Depan Ke Belakang 1 arah':
        driver.find_element(By.XPATH, "//li[contains(.,\'Depan Ke Belakang (1 arah)\')]").click()
        Log.info('Input Format Penomoran Kamar Depan Ke Belakang 1 arah')
        attach(data=driver.get_screenshot_as_png())

    elif formatPenomoranKamar == 'Depan Ke Belakang 2 arah':
        driver.find_element(By.XPATH, "//li[contains(.,\'Depan Ke Belakang (2 arah)\')]").click()
        Log.info('Input Format Penomoran Kamar Depan Ke Belakang 2 arah')
        attach(data=driver.get_screenshot_as_png())

    elif formatPenomoranKamar == 'Kanan Kiri 1 arah':
        driver.find_element(By.XPATH, "//li[contains(.,\'Kanan Kiri (1 arah)\')]").click()
        Log.info('Input Format Penomoran Kamar Kanan Kiri 1 arah')
        attach(data=driver.get_screenshot_as_png())
    Log.info('input penomoran kamar')
    driver.find_element(By.CSS_SELECTOR, ".el-dialog__close > svg").click()

@mark.fixture_test()
def test_12_JumlahKamarPerlantai():
    print('== NEXT == / input jumlah kamar per lantai')
    sleep(driver)
    driver.find_element(By.ID, "jumlah_kamar_per_lantai-0").click()

    if jumlahLantai == 1:
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-0").send_keys(jumlahKamarPerlantai)
        driver.find_element(By.ID, "jumlah_kolom_per_lantai-0").send_keys(jumlahKolomPerlantai)
        Log.info(' input 1 lantai')
        attach(data=driver.get_screenshot_as_png())

    elif jumlahLantai == 2:
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-0").click()
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-0").send_keys(jumlahKamarPerlantai)
        driver.find_element(By.ID, "jumlah_kolom_per_lantai-0").send_keys(jumlahKolomPerlantai)
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-1").click()
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-1").send_keys(jumlahKamarPerlantai)
        driver.find_element(By.ID, "jumlah_kolom_per_lantai-1").send_keys(jumlahKolomPerlantai)
        Log.info(' input 2 lantai')
        attach(data=driver.get_screenshot_as_png())

    elif jumlahLantai == 3:
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-0").click()
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-0").send_keys(jumlahKamarPerlantai)
        driver.find_element(By.ID, "jumlah_kolom_per_lantai-0").send_keys(jumlahKolomPerlantai)
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-1").click()
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-1").send_keys(jumlahKamarPerlantai)
        driver.find_element(By.ID, "jumlah_kolom_per_lantai-1").send_keys(jumlahKolomPerlantai)
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-2").click()
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-2").send_keys(jumlahKamarPerlantai)
        driver.find_element(By.ID, "jumlah_kolom_per_lantai-2").send_keys(jumlahKolomPerlantai)
        Log.info(' input 3 lantai')
        attach(data=driver.get_screenshot_as_png())

    elif jumlahLantai == 4:
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-0").click()
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-0").send_keys(jumlahKamarPerlantai)
        driver.find_element(By.ID, "jumlah_kolom_per_lantai-0").send_keys(jumlahKolomPerlantai)
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-1").click()
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-1").send_keys(jumlahKamarPerlantai)
        driver.find_element(By.ID, "jumlah_kolom_per_lantai-1").send_keys(jumlahKolomPerlantai)
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-2").click()
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-2").send_keys(jumlahKamarPerlantai)
        driver.find_element(By.ID, "jumlah_kolom_per_lantai-2").send_keys(jumlahKolomPerlantai)
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-3").click()
        driver.find_element(By.ID, "jumlah_kamar_per_lantai-3").send_keys(jumlahKamarPerlantai)
        driver.find_element(By.ID, "jumlah_kolom_per_lantai-3").send_keys(jumlahKolomPerlantai)
        Log.info(' input 4 lantai')
        attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_13_ClickButtonSubmit():
    print('== NEXT == / Click Button Submit')
    driver.find_element(By.ID, 'submitButton').click()

    Log.info('Input Jumlah Kamar perlantai')
    attach(data=driver.get_screenshot_as_png())

    #=============== Halaman Tambah Kamar =============== 

@mark.fixture_test()
def test_14_loginOperatorMP():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    op_keamanan_mp(driver)
    Log.info('Login Operator Manajemen Penempatan')


@mark.fixture_test()
def test_15_AksesMenuOperatorKeamanan():

    manajemenblokdankamar(driver)
    Log.info('Akses halaman Manajemen Blok dan Kamar')
    attach(data=driver.get_screenshot_as_png())
    Log.info('Akses halaman Daftar Lalu Lintas P2U External')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_16_ClickButtonHalamanTambah():
    sleep(driver)
    driver.implicitly_wait(10)
    sleep(driver)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.border:nth-child(1)')))
    driver.find_element(By.XPATH, '//*[@id="createButton"]').click()
    print('.')
    Log.info('Click Button Tambah halaman Pemetaan Block')
    attach(data=driver.get_screenshot_as_png())
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID,'backButton')))

            
@mark.fixture_test()
def test_17_SearchNamaBlok():
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID,'searchButton')))
    driver.find_element(By.ID, 'kataKunci').send_keys(namablok)
    driver.find_element(By.ID, 'searchButton').click()
    Log.info('Search Nama Blok')

@mark.fixture_test()
def test_18_ClickButtonSearch():
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID,'searchButton')))
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#roomButton0 .h-5')))
    sleep(driver)
    driver.find_element(By.CSS_SELECTOR, '#roomButton0 .h-5').click()
    
@mark.fixture_test()
def test_19_PilihLanta():
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID,'backButton')))
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'lantai_id')))
    driver.execute_script("window.scrollTo(0,244.5)")
    time.sleep(3)
    driver.find_element(By.ID, 'lantai_id').click()
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'lantai-0')))
    driver.find_element(By.ID, "lantai-0").click()
    driver.find_element(By.XPATH, "//li[contains(.,\'"+ lantai +"')]").click()
    Log.info('Pilih Lantai')

@mark.fixture_test()
def test_20_InputNomorkamar():
    sleep(driver)
    driver.find_element(By.ID, "nomorKamar").click()
    driver.find_element(By.ID, "nomorKamar").send_keys(nomorKamar)
    Log.info('Input Nomor Kamar')

@mark.fixture_test()
def test_21_InputKondisiRuangan():
    driver.find_element(By.ID, "kondisiInput").click()
    driver.find_element(By.XPATH, "//li[contains(.,\'"+ kondisiRuangan +"')]").click()
    Log.info('Input Kondisi Ruangan')
    
@mark.fixture_test()
def test_22_InputkelompokJenisKejahatan():
    driver.find_element(By.ID, "kel_jenis_kejahatan_id").click()
    driver.find_element(By.XPATH, "//li[contains(.,\'"+ kelompokJenisKejahatan +"')]").click()
    Log.info('Input Kelompok Jenis Kejahatan')
    
@mark.fixture_test()
def test_23_InputKapasitas():
    driver.find_element(By.ID, "kapasitasInput").click()
    driver.find_element(By.ID, "kapasitasInput").send_keys(kapasitasInput)
    Log.info('Input Kapasitas')

@mark.fixture_test()
def test_24_ClickButtonSubmitKamar():
    driver.find_element(By.ID, 'submitButton').click()
    Log.info('Click Button Submit Kamar')


@mark.fixture_test()
def test_25_ClickButtonSubmitBlok():
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'submitButton')))

    WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, '//div[contains(.,\'Berhasil Ditambahkan\')]')))

    time.sleep(3)
    driver.find_element(By.CSS_SELECTOR, "a > #submitButton > span").click()



    #=============== Halaman Verifikasi Supervisor =============== 

@mark.fixture_test()
def test_26_loginSuperVisor():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    SpvRutanBdg(driver)
    Log.info('Login Operator Manajemen Penempatan')

@mark.fixture_test()
def test_27_AksesMenu():
    manajemenblokdankamar(driver)
    Log.info('Akses halaman Manajemen Blok dan Kamar')
    attach(data=driver.get_screenshot_as_png())
    print('.')
    Log.info('Akses halaman Daftar Lalu Lintas P2U External')
    attach(data=driver.get_screenshot_as_png())

    
@mark.fixture_test()
def test_28_ClickTambahPemetaanBlok():
    sleep(driver)
    driver.implicitly_wait(10)
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.border:nth-child(1)')))
    driver.find_element(By.XPATH, '//*[@id="createButton"]').click()
    print('.')
    Log.info('Click Button Tambah halaman Pemetaan Block')
    attach(data=driver.get_screenshot_as_png())
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID,'backButton')))
    
@mark.fixture_test()
def test_29_SearchNamaBlok():
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID,'searchButton')))
    driver.find_element(By.ID, 'kataKunci').send_keys(namablok)
    driver.find_element(By.ID, 'searchButton').click()
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID,'searchButton')))
    Log.info('Search Nama Blok')

@mark.fixture_test()
def test_30_ClickButtonVerifikasi():
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID,'searchButton')))
    driver.find_element(By.CSS_SELECTOR, '#verifikasiButton-0 .h-5').click()
    Log.info('ClickButtonVerifikasi')

@mark.fixture_test()
def test_31_ClickStatusVerifikasi():
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID,'submitButton')))
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID,'status_verifikasi')))
    driver.find_element(By.ID, 'status_verifikasi').click()
    Log.info('Click status verifikasi')

@mark.fixture_test()
def test_32_Verifikasi():
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH,"//li[contains(.,\'"+ verifikasi +"')]")))
    driver.find_element(By.XPATH, "//li[contains(.,\'"+ verifikasi +"')]").click()
    Log.info('Verifikasi')
    
@mark.fixture_test()
def test_33_InputKeterangan():
    driver.find_element(By.ID, 'keterangan').send_keys(keterangan)
    Log.info('Input Keterangan')
    
@mark.fixture_test()
def test_34_ClickButtonSubmit():
    driver.find_element(By.ID, 'submitButton').click()
    Log.info('Clik Button Submit')

@mark.fixture_test()
def test_35_exit():
    quit(driver)
        

  
        






