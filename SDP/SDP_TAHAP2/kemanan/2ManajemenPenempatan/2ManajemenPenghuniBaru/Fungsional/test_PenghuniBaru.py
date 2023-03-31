from distutils.archive_util import make_archive
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from selenium.webdriver.firefox.options import Options
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUAT"))

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))

from Settings.setup import initDriver, loadDataPath, quit, sleep
from Settings.login import loginOperatorSumedang, Op_Keamanan_p2u, SpvRutanBdg, op_keamanan_mp
from Settings.Page.keamanan import manajemenpenghunibaru

import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('LogManajemenPenghuniBaru.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

wb = load_workbook(environ.get("KeamananUAT"))
sheetrange = wb['TambahManajemenPenghuniBaru']
i = 2

    
Nama                                    = sheetrange['A'+str(i)].value

Kejahatan                               = sheetrange['B'+str(i)].value
blokform                                = sheetrange['C'+str(i)].value
Lantai                                  = sheetrange['D'+str(i)].value
Kamar                                   = sheetrange['E'+str(i)].value
TanggalPenempatan                       = sheetrange['F'+str(i)].value
Keterangan                              = sheetrange['G'+str(i)].value


@mark.fixture_test()
def test_2_login():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    op_keamanan_mp(driver)
    Log.info('Login Operator Manajemen Penempatan')


@mark.fixture_test()
def test_3_AksesMenu():
    
    sleep(driver)
    manajemenpenghunibaru(driver)
    print('.')
    Log.info('Akses halaman Manajemen Penghuni Baru')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_4_CreateButton():
    
    sleep(driver)
    driver.implicitly_wait(10)
    #sleep(driver)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
    driver.find_element(By.CSS_SELECTOR, "#createButton > span").click()
    Log.info('Click Button Tambah halaman Pemetaan Block')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_5_SearchNamaWBP():
    sleep(driver)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'backButton')))
    driver.find_element(By.ID, 'kataKunci').click()
    driver.find_element(By.ID, 'kataKunci').send_keys(Nama)

@mark.fixture_test()
def test_6_ClickButtonSearch():
    sleep(driver)
    driver.find_element(By.ID, 'searchButton').click()
    Log.info('memilih pencarian berdasarkan Nama Lengkap')
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_7_ClickButtonPenempatan():
    sleep(driver)
    driver.find_element(By.ID, "penempatan-0").click()
    Log.info("Click Button Penempatan")
    driver.execute_script("window.scrollTo(0,1.5)")

@mark.fixture_test()
def test_8_SelectBlock():
    sleep(driver)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'submitButton')))
    driver.find_element(By.ID, "blokForm").click()
    driver.find_element(By.XPATH, "//li[contains(.,\'"+ blokform +"')]").click()
    Log.info('Memilih Blok')

@mark.fixture_test()
def test_9_MemilihLantai():
    
    sleep(driver)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'lantaiForm')))
    driver.find_element(By.ID, "lantaiForm").click()
    driver.find_element(By.XPATH, "//li[contains(.,\'"+ Lantai +"')]").click()
    Log.info('Memilih Lantai')

@mark.fixture_test()
def test_10_MemilihKamar():
    
    sleep(driver)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'kamarForm')))
    driver.find_element(By.ID, "kamarForm").click()
    driver.find_element(By.XPATH, "//li[contains(.,\'"+ Kamar +"')]").click()
    Log.info('Memilih Kamar')

@mark.fixture_test()
def test_11_InputTanggalPenempatan():
    sleep(driver)
    driver.find_element(By.ID, "tanggalMutasiForm").send_keys(TanggalPenempatan)
    Log.info('Input Tanggal Penempatan')

@mark.fixture_test()
def test_12_InputKeterangan():
    sleep(driver)
    driver.find_element(By.ID, "keteranganForm").click()
    driver.find_element(By.ID, "keteranganForm").send_keys("OK")
    Log.info('Input Keterangan')

@mark.fixture_test()
def test_13_ClickButtonSubmit():
    sleep(driver)
    driver.find_element(By.ID, 'submitButton').click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
    Log.info('Klik Button Submit')


@mark.fixture_test()
def test_14_loginSpv():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    SpvRutanBdg(driver)
    Log.info('Login Operator Manajemen Penempatan')


@mark.fixture_test()
def test_15_AksesMenu():
    
    sleep(driver)
    manajemenpenghunibaru(driver)
    print('.')
    Log.info('Akses halaman Manajemen Penghuni Baru')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_16_FilterColumnNama():
    
    sleep(driver)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
    driver.find_element(By.ID, 'filterColumn').click()
    driver.find_element(By.XPATH, "//li[contains(.,\'Nama\')]").click()
    Log.info('Pilih Filter Column Nama')

@mark.fixture_test()
def test_17_SelectStatusDalamProses():
    sleep(driver)
    driver.find_element(By.ID, 'statusVerifikasi').click()
    driver.find_element(By.ID, 'DalamProses').click()
    Log.info('Pilih Status Dalam Proses')

@mark.fixture_test()
def test_18_PilihKataKunciNama():
    sleep(driver)
    driver.find_element(By.ID, 'kataKunci').click()
    driver.find_element(By.ID, 'kataKunci').send_keys(Nama)
    Log.info('Pilih Kata Kunci nama')

@mark.fixture_test()
def test_18_ClickButtonSearch():
    sleep(driver)
    driver.find_element(By.ID, 'searchButton').click()
    driver.execute_script("window.scrollTo(0,1.5)")
    Log.info('Click Button Search')

@mark.fixture_test()
def test_19_ClickButtonVerifikasi():
    sleep(driver)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
    attach(data=driver.get_screenshot_as_png())
    time.sleep(3)
    driver.find_element(By.CSS_SELECTOR, ".text-yellow-500").click()
    Log.info("Click Button  Verifikasi")

@mark.fixture_test()
def test_20_ChangeStatusDiizinkan():
    sleep(driver)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#status_verifikasi')))
    driver.find_element(By.CSS_SELECTOR, '#status_verifikasi').click()
    driver.find_element(By.XPATH, "//li[@id='diizinkan']").click()
    Log.info('Status Diizinkan')

@mark.fixture_test()
def test_21_InputKeterangan():
    sleep(driver)
    driver.find_element(By.ID, "keterangan").send_keys(Keterangan)
    Log.info('Input Keterangan')

@mark.fixture_test()
def test_22_Verifikasi():
    sleep(driver)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'submitButton')))
    driver.find_element(By.ID, "submitButton").click()
    Log.info('Click Button Verifikasi')

@mark.fixture_test()
def test_21_Exit():
    sleep(driver)
    quit(driver)


