from distutils.archive_util import make_archive
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from selenium.webdriver.firefox.options import Options

options = Options()
options.add_argument('--disable-blink-features=AutomationControlled')
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUAT"))

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))


from Settings.setup import initDriver, loadDataPath, quit, sleep
from Settings.login import loginOperatorSumedang, Op_Keamanan_p2u, SpvP2U
from Settings.Page.keamanan import p2uinternal

import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Log3_KonfirmasiOperator.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)


sheetrangeIndex = wb['P2U_Internal']

@mark.fixture_test()
def test_1_setupOS():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')

@mark.fixture_test()
def test_2_login():
    Op_Keamanan_p2u(driver)
    Log.info('Login Operator Keamanan')


@mark.fixture_test()
def test_3_Input():
    sleep(driver)
    p2uinternal(driver)
    Log.info('Akses halaman Daftar Lalu Lintas P2U Internal')
    attach(data=driver.get_screenshot_as_png())


    i = 2
        
    while i <= len(sheetrangeIndex['A']):
    
        NamaInput                                 = sheetrangeIndex['A'+str(i)].value
        Keterangan                                = sheetrangeIndex['F'+str(i)].value
            
        try:
            sleep(driver)
            
            print('Pilih Dropdown Nama')
            sleep(driver)
            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="searchButton"]')))

            driver.find_element(By.ID, 'statusColumn').click()
            driver.find_element(By.XPATH, "//li[contains(.,\'Dalam Proses\')]").click()
            Log.info('Filter Status Dalam Proses')

            sleep(driver)
            driver.find_element(By.XPATH, '//*[@id="kataKunci"]').clear()
            driver.find_element(By.XPATH, '//*[@id="kataKunci"]').send_keys(NamaInput)
            Log.info('Search Nama WBP')
            driver.find_element(By.ID, 'searchButton' ).click()
            Log.info('Click Button Search')
            sleep(driver)

            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="searchButton"]')))
            sleep(driver)
            driver.find_element(By.CSS_SELECTOR, "#detailButton-0 .h-5").click()
            Log.info('Click Button Detile')
            
            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID, 'submitButton')))
            driver.find_element(By.ID, 'submitButton').click()
            Log.info('clik button keluar P2U')
            
            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="searchButton"]')))
            driver.find_element(By.ID, 'statusColumn').click()
            driver.find_element(By.XPATH, "//li[contains(.,\'Keluar P2U\')]").click()
            Log.info('filter status Keluar P2U')

            sleep(driver)
            driver.find_element(By.XPATH, '//*[@id="kataKunci"]').clear()
            driver.find_element(By.XPATH, '//*[@id="kataKunci"]').send_keys(NamaInput)
            Log.info('Search Nama WBP')

            driver.find_element(By.ID, 'searchButton' ).click()
            Log.info('click button search')

            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="searchButton"]')))
            sleep(driver)
            driver.find_element(By.CSS_SELECTOR, "#detailButton-0 .h-5").click()
            Log.info('Click Button Detile')

            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID, 'submitButton')))
            driver.find_element(By.ID, 'submitButton').click()
            Log.info('click button masuk P2U')

            sleep(driver)


        except TimeoutException:
            print("ERRROR")
            pass
                
        sleep(driver)
        i = i + 1
    print('DONE')

def teardown():
    quit(driver)
        

  
        






