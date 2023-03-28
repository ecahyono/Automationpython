from distutils.archive_util import make_archive
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from selenium.webdriver.firefox.options import Options

options = Options()
options.add_argument('--disable-blink-features=AutomationControlled')
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUAT"))

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))


from Settings.setup import initDriver, loadDataPath, quit, sleep
from Settings.login import Op_Keamanan_p2u
from Settings.Page.keamanan import p2ueksternal

import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Log1_P2UExternalTambah.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

sheetrangeIndex = wb['P2U_External']

@mark.fixture_test()
def test_1_setupOS():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')

@mark.fixture_test()
def test_2_login():
    Op_Keamanan_p2u(driver)
    Log.info('Login')


@mark.fixture_test()
def test_3_Input():
    sleep(driver)
    p2ueksternal(driver)
    Log.info('Akses halaman Daftar Lalu Lintas P2U External')
    attach(data=driver.get_screenshot_as_png())


    i = 2
        
    while i <= len(sheetrangeIndex['A']):
    
        inputKategori                             = sheetrangeIndex['A'+str(i)].value
        NamaAtauNip                               = sheetrangeIndex['B'+str(i)].value
        inputKeperluan                            = sheetrangeIndex['C'+str(i)].value
        inputNip                                  = sheetrangeIndex['D'+str(i)].value
        inputJabatan                              = sheetrangeIndex['E'+str(i)].value
        InputInstansi                             = sheetrangeIndex['F'+str(i)].value
    

        try:
            sleep(driver)
            
            print('masuk ke halaman daftar lalu lintas external')
            sleep(driver)
            WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="searchButton"]')))
            Log.info('masuk ke halaman daftar lalu lintas external')

            sleep(driver)
            driver.find_element(By.ID, 'createButton').click()
            Log.info('masuk ke halaman tambah')

            WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'submitButton')))
            driver.find_element(By.ID, 'inputKategori').click()


            if inputKategori == 'Pegawai':
                driver.find_element(By.ID, "pegawai").click()
                driver.find_element(By.ID, 'inputSearch').send_keys(NamaAtauNip)
            
                if driver.find_elements(By.XPATH, "//li[@id='searchOption0']/div/div/table/tbody/tr/td"):
                    driver.find_element(By.XPATH, "//li[@id='searchOption0']/div/div/table/tbody/tr/td").click()
                    driver.find_element(By.ID, 'inputKeperluan').send_keys(inputKeperluan)
                else:
                    driver.find_element(By.ID, 'inputNip').click()
                    driver.find_element(By.ID, 'inputNip').send_keys(inputNip)
                    Log.info(' Input NIP ')
                    sleep(driver)
                    attach(data=driver.get_screenshot_as_png())
                    driver.find_element(By.ID, 'inputNama').send_keys(NamaAtauNip)
                    Log.info(' Input Nama ')
                    sleep(driver)
                    attach(data=driver.get_screenshot_as_png())
                    driver.find_element(By.ID, 'inputJabatan').send_keys(inputJabatan)
                    Log.info(' Input Jabatan ')
                    sleep(driver)
                    attach(data=driver.get_screenshot_as_png())
                    driver.find_element(By.ID, 'inputKeperluan').send_keys(inputKeperluan)
                    Log.info(' Input Keperluan ')
                    sleep(driver)
                    attach(data=driver.get_screenshot_as_png())

            elif inputKategori == 'Tamu Dinas':
                driver.find_element(By.ID, "tamuDinas").click()
                Log.info(' Memilih Kategori Tamu Dinas ')
                driver.find_element(By.ID, 'inputSearch').send_keys(NamaAtauNip)
                Log.info(' Cari Nama atau Nip Kategori Tamu Dinas ')
                attach(data=driver.get_screenshot_as_png())

                if driver.find_elements(By.CSS_SELECTOR, 'tr:nth-child(1) > .el-descriptions__label'):
                    driver.find_element(By.CSS_SELECTOR, 'tr:nth-child(1) > .el-descriptions__label').click()
                    driver.find_element(By.ID, 'inputKeperluan').send_keys(inputKeperluan)

                else:
                    driver.find_element(By.ID, 'inputNip').send_keys(inputNip)
                    sleep(driver)
                    driver.find_element(By.ID, 'inputNama').send_keys(NamaAtauNip)
                    sleep(driver)
                    driver.find_element(By.ID, 'inputInstansiId').send_keys(InputInstansi)

                    
                    driver.find_element(By.XPATH, "//li[contains(.,'"+ InputInstansi +"')]").click()
                    sleep(driver)
                    driver.find_element(By.ID, 'inputJabatan').send_keys(inputJabatan)
                    sleep(driver)
                    driver.find_element(By.ID, 'inputKeperluan').send_keys(inputKeperluan)
                    print('.')
                    Log.info(' Input NIP ')
                    attach(data=driver.get_screenshot_as_png())
            elif inputKategori == 'Kunjungan Tatap Muka':
                WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="inputKategori"]')))
                driver.find_element(By.XPATH, '//*[@id="inputKategori"]').click()
                driver.find_element(By.ID, "kunjungan").click()

                driver.find_element(By.ID, "kataKunci").send_keys(NamaAtauNip)
                driver.find_element(By.ID, 'searchButton').click()
                sleep(driver)
                WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
                
                driver.find_element(By.ID, 'pickKunjungan0').click()

            elif inputKategori == 'Kunjungan Online':
                WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="inputKategori"]')))
                driver.find_element(By.XPATH, '//*[@id="inputKategori"]').click()
                driver.find_element(By.ID, "kunjungan").click()

                driver.find_element(By.ID, "kataKunci").send_keys(NamaAtauNip)
                driver.find_element(By.ID, 'searchButton').click()
                sleep(driver)
                WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
                
                driver.find_element(By.ID, 'pickKunjungan0').click()
                
            driver.find_element(By.CSS_SELECTOR, '#submitButton').click()


            sleep(driver)
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="searchButton"]')))
            driver.find_element(By.XPATH, "//input[@type=\'text\']").send_keys('semua')
            sleep(driver)
            # KETIK NAMA
            driver.find_element(By.XPATH, "//li[contains(.,\'Semua\')]").click()
            sleep(driver)
            # PILIH DROPDOWN NAMA LENGKAP
            driver.find_element(By.XPATH, '//*[@id="kataKunci"]').send_keys(NamaAtauNip)
            sleep(driver)
            # KETIK GALIH DI FORM MASUKAN KATA KUNCI
            Log.info('Search Bedasarkan Nama Lengkap')
            attach(data=driver.get_screenshot_as_png())
            driver.find_element(By.ID, "konfirmasiKeluar").click()
            driver.find_element(By.XPATH, '//*[@id="searchButton"]').click()
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="searchButton"]')))

            sleep(driver)
            driver.find_element(By.CSS_SELECTOR, "#detailButton0 .h-5").click()

            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, 'submitButton')))
            time.sleep(5)
            driver.find_element(By.ID, "submitButton").click()
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, '//div[contains(.,\'Berhasil Diperbaharui\')]')))

        except TimeoutException:
            print("ERRROR")
            pass
                
        sleep(driver)
        i = i + 1
    print('DONE')

def teardown():
    quit(driver)
        

  
        






