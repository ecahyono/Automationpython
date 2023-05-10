from distutils.archive_util import make_archive
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUAT"))
    file_path = environ.get("fakermac")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))


from Settings.setupkeamanan import initDriver, loadDataPath, quit, sleep
from Settings.loginkeamanan import SpvRutanBdg, oplapkamtibwaru,kanwiljabar,pusat
from Settings.Page.keamanan import menulaporan6a
import pytest
from openpyxl import Workbook
from faker import Faker
import random

import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('LogLaporan6A.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

i = 2
laporan6a = wb['Laporan6A']

nama                                            = laporan6a['A'+str(i)].value
nosk                                            = laporan6a['B'+str(i)].value
tanggalSkAsimilasi                              = laporan6a['C'+str(i)].value
tanggalAsimilasi                                = laporan6a['D'+str(i)].value
lokasiAsimilasi                                 = laporan6a['E'+str(i)].value
namaPetugas                                     = laporan6a['F'+str(i)].value
namaPenjamin                                    = laporan6a['G'+str(i)].value
keterlibatanLain                                = laporan6a['H'+str(i)].value
keterangan                                      = laporan6a['I'+str(i)].value


workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Laporan7Faker'

fake = Faker('id_ID')
for i in range(5):
    tanggalSkAsimilasiFaker                             = fake.date_between(start_date='today', end_date='today').strftime('%d/%m/%Y')
    tanggalAsimilasiFaker                               = fake.date_between(start_date='today', end_date='today').strftime('%d/%m/%Y')
    rekapitulasiKehadiranFaker                          = fake.text(max_nb_chars=255)
    permasalahanFaker                                   = fake.text(max_nb_chars=255)
    alternatifPenyelesaianFaker                         = fake.text(max_nb_chars=255)
    keteranganFaker                                     = fake.text(max_nb_chars=255)


    worksheet.append([
        tanggalSkAsimilasiFaker,tanggalAsimilasiFaker
        ])
workbook.save(file_path)

global driver, pathData
driver = initDriver()
pathData = loadDataPath()
Log.info('Setup Os')

@pytest.mark.webtest
def test_1loginOperator():
    oplapkamtibwaru(driver)
    Log.info('Login Op Laporan Kamtib 6A')


@pytest.mark.webtest
def test_2_AksesMenu():
    sleep(driver)
    menulaporan6a(driver)
    Log.info('Akses halaman Laporan 6A')
    attach(data=driver.get_screenshot_as_png())


@pytest.mark.webtest
def test_3_CreateButton():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'createButton')))
    driver.find_element(By.ID, "createButton").click()
    Log.info('Klik tombol tambah data')

@pytest.mark.webtest
def test_4_FilterColumn():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonSearch')))
    driver.find_element(By.ID, "filterColumn").click()
    driver.find_element(By.ID, "semua").click()
    Log.info('Filter Column Semua')

@pytest.mark.webtest
def test_5_Katakunci():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonSearch')))
    driver.find_element(By.ID, "kataKunci").click()
    driver.find_element(By.ID, "kataKunci").send_keys(nama)
    Log.info('Search Nama WBP')

@pytest.mark.webtest
def test_6_ButtonSearch():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonSearch')))
    driver.find_element(By.ID, "buttonSearch").click()
    Log.info('Klik tombol search')

@pytest.mark.webtest
def test_7_ClickButtonDaftarkan():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonSearch')))
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'daftarkan0')))
    driver.find_element(By.ID, "daftarkan0").click()
    Log.info('Klik tombol daftarkan')

@pytest.mark.webtest
def test_8_InputnoSK():
    sleep(driver)
    driver.find_element(By.ID, "noSkAsimilasi").click()
    driver.find_element(By.ID, "noSkAsimilasi").send_keys(nosk)
    Log.info('Input No SK Asimilasi')

@pytest.mark.webtest
def test_9_InputTanggalSk():
    sleep(driver)
    driver.find_element(By.ID, "tanggalSkAsimilasi").click()
    driver.find_element(By.ID, "tanggalSkAsimilasi").send_keys(tanggalSkAsimilasiFaker)
    Log.info('Input Tanggal SK Asimilasi')

@pytest.mark.webtest
def test_10_InputTanggalAsimilasi():
    sleep(driver)
    driver.find_element(By.ID, "tanggalAsimilasi").click()
    driver.find_element(By.ID, "tanggalAsimilasi").send_keys(tanggalAsimilasiFaker)
    Log.info('Input Tanggal Asimilasi')

@pytest.mark.webtest
def test_11_Inputlokasi():
    sleep(driver)
    driver.find_element(By.ID, "lokasiAsimilasi").click()
    driver.find_element(By.ID, "lokasiAsimilasi").send_keys(lokasiAsimilasi)
    Log.info('Input Lokasi Asimilasi')

@pytest.mark.webtest
def test_12_InputNamaPetugas():
    sleep(driver)
    driver.find_element(By.ID, "namaPetugas").click()
    driver.find_element(By.ID, "namaPetugas").send_keys(namaPetugas)
    Log.info('Input Nama Petugas')

@pytest.mark.webtest
def test_13_InputNamaPenjamin():
    sleep(driver)
    driver.find_element(By.ID, "namaPenjamin").click()
    driver.find_element(By.ID, "namaPenjamin").send_keys(namaPenjamin)
    Log.info('Input Nama Pengjamin')

@pytest.mark.webtest
def test_14_InputPihakLain():
    sleep(driver)
    driver.find_element(By.ID, "keterlibatanLain").click()
    driver.find_element(By.ID, "keterlibatanLain").send_keys(keterlibatanLain)
    Log.info('Input Keterlibatan Pihak Lain')

@pytest.mark.webtest
def test_15_InputKeterangan():
    sleep(driver)
    driver.find_element(By.ID, "keterangan").click()
    driver.find_element(By.ID, "keterangan").send_keys(keterangan)
    Log.info('Input Keterangan')

@pytest.mark.webtest
def test_16_ClickButtonSubmit():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'submitButton')))
    driver.find_element(By.ID, "submitButton").click()
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'createButton')))
    Log.info('Klik tombol submit')

@pytest.mark.webtest
def test_17_loginSPV():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    SpvRutanBdg(driver)
    Log.info('Login Spv Laporan Kamtib 6A')

@pytest.mark.webtest
def test_18_AksesMenuSpv():
    sleep(driver)
    menulaporan6a(driver)
    Log.info('Akses halaman Laporan 6A')
    attach(data=driver.get_screenshot_as_png())

@pytest.mark.webtest
def test_19_KirimLaporanSpv():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
    time.sleep(2)
    driver.find_element(By.ID, "buttonVerifikasiLaporan").click()
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'kirimLaporan')))
    time.sleep(2)
    driver.find_element(By.CSS_SELECTOR, "#kirimLaporan > span").click()


@pytest.mark.webtest
def test_20_loginKanwil():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    kanwiljabar(driver)
    Log.info('Login Spv Laporan Kamtib 6A')

@pytest.mark.webtest
def test_21_AksesMenuSpv():
    sleep(driver)
    menulaporan6a(driver)
    Log.info('Akses halaman Laporan 6A')
    attach(data=driver.get_screenshot_as_png())

@pytest.mark.webtest
def test_22_VerifikasiLaporanKanwil():
    sleep(driver)
    time.sleep(2)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterUpt2')))
    driver.find_element(By.ID, "formfilterUpt").click()
    time.sleep(1)
    driver.find_element(By.ID, "formfilterUpt2").send_keys("Rutan Kelas I Bandung")
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
    driver.find_element(By.XPATH, "//span[contains(.,\'Rutan Kelas I Bandung\')]").click()
    Log.info('Input UPT')

@pytest.mark.webtest
def test_23_ClickButtonSearch():
    sleep(driver)
    driver.find_element(By.ID, "searchButton").click()
    Log.info('Click Button Search')

@pytest.mark.webtest
def test_24_ClickButtonVerifikasi():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasi')))
    driver.find_element(By.ID, "buttonVerifikasi").click()
    Log.info('Click Button Verifikasi')


@pytest.mark.webtest
def test_25_StatusVerifikasiModal():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'simpanVerifikasi')))
    time.sleep(3)
    driver.find_element(By.ID, "statusVerifikasiModal").click()
    Log.info('Click Verifikasi Modal')

@pytest.mark.webtest
def test_26_UbahStatus():
    sleep(driver)
    driver.find_element(By.ID, "diizinkan").click()
    Log.info('Ubah Status Verifikasi')

@pytest.mark.webtest
def test_27_Inputketerangan():
    sleep(driver)
    driver.find_element(By.ID, "keterangan").send_keys("keterangan")
    Log.info('Input Keterangan')

@pytest.mark.webtest
def test_28_SimpanVerifikasi():
    sleep(driver)
    driver.find_element(By.ID, "simpanVerifikasi").click()
    Log.info('Click Button Simpan Verifikasi')

@pytest.mark.webtest
def test_29_loginPusat():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    pusat(driver)
    Log.info('Login Spv Laporan Kamtib 6A')

@pytest.mark.webtest
def test_30_AksesMenuPusat():
    sleep(driver)
    menulaporan6a(driver)
    Log.info('Akses halaman Laporan 6A')
    attach(data=driver.get_screenshot_as_png())

@pytest.mark.webtest
def test_31_PilihKanwil():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterKanwil')))
    driver.find_element(By.ID, "formfilterKanwil").click()
    driver.find_element(By.ID, "formfilterKanwil").send_keys("jawa barat")
    driver.find_element(By.XPATH, "//li[contains(.,\'Jawa Barat\')]").click()
    Log.info('Pilih Kanwil')

@pytest.mark.webtest
def test_32_PilihUPT():
    sleep(driver)
    driver.find_element(By.ID, "formfilterUpt").click()
    time.sleep(1)
    driver.find_element(By.ID, "formfilterUpt").send_keys("rutan kelas I Bandung")
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
    driver.find_element(By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]").click()
    Log.info('PilihUPT')

@pytest.mark.webtest
def test_33_ClickButtonSearch():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
    driver.find_element(By.CSS_SELECTOR, "#searchButton > span").click()
    Log.info('Click Button Search')


@pytest.mark.webtest
def test_34_exit():
    time.sleep(5)
    quit(driver)















