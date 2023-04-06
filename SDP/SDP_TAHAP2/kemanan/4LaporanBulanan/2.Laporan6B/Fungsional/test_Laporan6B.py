from distutils.archive_util import make_archive
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUAT"))

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))


from Settings.setupkeamanan import initDriver, loadDataPath, quit, sleep
from Settings.loginkeamanan import oplapkamtibwaru,kanwiljabar,pusat, SpvRutanBdg
from Settings.Page.keamanan import menulaporan6b

import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('LogLaporan6B.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)


i = 2
laporan6b = wb['Laporan6B']


UraianSingkatKejadian                       = laporan6b['A'+str(i)].value
WaktuKejadian                               = laporan6b['B'+str(i)].value
Kerusakan                                   = laporan6b['C'+str(i)].value
TotalKerugian                               = laporan6b['D'+str(i)].value
KorbanJiwaPetugas                           = laporan6b['E'+str(i)].value
KorbanJiwaWBP                               = laporan6b['F'+str(i)].value
TindakanSementaraYangDilakukan              = laporan6b['G'+str(i)].value
TindakLanjut                                = laporan6b['H'+str(i)].value
SumberDana                                  = laporan6b['I'+str(i)].value
BantuanPihakLain                            = laporan6b['J'+str(i)].value
keterangan                                  = laporan6b['K'+str(i)].value


@mark.fixture_test()
def test_1loginOperator():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    oplapkamtibwaru(driver)
    Log.info('Login Op Laporan Kamtib 6A')


@mark.fixture_test()
def test_2_AksesMenu():
    sleep(driver)
    menulaporan6b(driver)
    Log.info('Akses halaman Laporan 6B')
    attach(data=driver.get_screenshot_as_png())


@mark.fixture_test()
def test_3_CreateButton():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'createButton')))
    driver.find_element(By.ID, "createButton").click()
    Log.info('Klik tombol tambah data')

@mark.fixture_test()
def test_4_UraianSingkatKejadian():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'submitButton')))
    driver.find_element(By.ID, "uraianKejadian").click()
    driver.find_element(By.ID, "uraianKejadian").send_keys(UraianSingkatKejadian)
    Log.info('Isi Uraian Singkat Kejadian')

@mark.fixture_test()
def test_5_WaktuKejadian():
    sleep(driver)
    driver.find_element(By.ID, "waktuKejadian").send_keys(WaktuKejadian)
    driver.find_element(By.ID, "waktuKejadian").send_keys(Keys.ENTER)
    Log.info('Isi Waktu Kejadian')

@mark.fixture_test()
def test_6_Kerusakan():
    sleep(driver)
    driver.find_element(By.ID, "kerusakan").click()
    driver.find_element(By.ID, "kerusakan").send_keys(Kerusakan)
    Log.info('Isi Kerusakan')
    

@mark.fixture_test()
def test_7_TotalKerugian():
    sleep(driver)
    driver.find_element(By.ID, "totalKerugian").click()
    driver.find_element(By.ID, "totalKerugian").send_keys(TotalKerugian)
    Log.info('Isi Total Kerugian')
    
    
@mark.fixture_test()
def test_8_KorbanJiwa():
    sleep(driver)
    driver.find_element(By.CSS_SELECTOR, "#korbanPetugas .el-input__inner").click()
    driver.find_element(By.CSS_SELECTOR, "#korbanPetugas .el-input__inner").send_keys(KorbanJiwaPetugas)
    driver.find_element(By.CSS_SELECTOR, "#korbanPenghuni .el-input__inner").click()
    driver.find_element(By.CSS_SELECTOR, "#korbanPenghuni .el-input__inner").send_keys(KorbanJiwaWBP)
    Log.info('Isi Korban Jiwa')

@mark.fixture_test()
def test_9_TindakanSementaraYangDilakukan():
    sleep(driver)
    driver.find_element(By.ID, "tindakanSementara").click()
    driver.find_element(By.ID, "tindakanSementara").send_keys(TindakanSementaraYangDilakukan)
    Log.info('Isi Tindakan Sementara Yang Dilakukan')
    

@mark.fixture_test()
def test_10_TindakLanjut():
    sleep(driver)
    driver.find_element(By.ID, "tindakLanjut").click()
    driver.find_element(By.ID, "tindakLanjut").send_keys(TindakLanjut)
    Log.info('Isi Tindak Lanjut')
    

@mark.fixture_test()
def test_11_SumberDana():
    sleep(driver)
    driver.find_element(By.ID, "sumberDana").click()
    driver.find_element(By.ID, "sumberDana").send_keys(SumberDana)
    Log.info('Isi Sumber Dana')
    
@mark.fixture_test()
def test_12_BantuanPihakLain():
    sleep(driver)
    driver.find_element(By.ID, "bantuanPihakLain").click()
    driver.find_element(By.ID, "bantuanPihakLain").send_keys(BantuanPihakLain)
    Log.info('Isi Bantuan Pihak Lain')
    

@mark.fixture_test()
def test_13_Keterangan():
    sleep(driver)
    driver.find_element(By.ID, "keterangan").click()
    driver.find_element(By.ID, "keterangan").send_keys(keterangan)
    Log.info('Isi Keterangan')

@mark.fixture_test()
def test_14_Submit():
    sleep(driver)
    driver.find_element(By.ID, "submitButton").click()
    Log.info('Klik tombol submit')


@mark.fixture_test()
def test_17_loginSPV():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    SpvRutanBdg(driver)
    Log.info('Login Spv Laporan Kamtib 6A')

@mark.fixture_test()
def test_18_AksesMenuSpv():
    sleep(driver)
    menulaporan6b(driver)
    Log.info('Akses halaman Laporan 6A')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_19_KirimLaporanSpv():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
    time.sleep(2)
    driver.find_element(By.ID, "buttonVerifikasiLaporan").click()
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'kirimLaporan')))
    time.sleep(2)
    driver.find_element(By.CSS_SELECTOR, "#kirimLaporan > span").click()


@mark.fixture_test()
def test_20_loginKanwil():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    kanwiljabar(driver)
    Log.info('Login Spv Laporan Kamtib 6A')

@mark.fixture_test()
def test_21_AksesMenuSpv():
    sleep(driver)
    menulaporan6b(driver)
    Log.info('Akses halaman Laporan 6A')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_22_VerifikasiLaporanKanwil():
    sleep(driver)
    time.sleep(2)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterUpt2')))
    driver.find_element(By.ID, "formfilterUpt").click()
    time.sleep(1)
    driver.find_element(By.ID, "formfilterUpt2").send_keys("Rutan Kelas I ")
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
    driver.find_element(By.XPATH, "//span[contains(.,\'Rutan Kelas I Bandung\')]").click()
    Log.info('Input UPT')

@mark.fixture_test()
def test_23_ClickButtonSearch():
    sleep(driver)
    driver.find_element(By.ID, "searchButton").click()
    Log.info('Click Button Search')

@mark.fixture_test()
def test_24_ClickButtonVerifikasi():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasi')))
    driver.find_element(By.ID, "buttonVerifikasi").click()
    Log.info('Click Button Verifikasi')


@mark.fixture_test()
def test_25_StatusVerifikasiModal():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'simpanVerifikasi')))
    time.sleep(3)
    driver.find_element(By.ID, "statusVerifikasiModal").click()
    Log.info('Click Verifikasi Modal')

@mark.fixture_test()
def test_26_UbahStatus():
    sleep(driver)
    driver.find_element(By.ID, "diizinkan").click()
    Log.info('Ubah Status Verifikasi')

@mark.fixture_test()
def test_27_Inputketerangan():
    sleep(driver)
    driver.find_element(By.ID, "keterangan").send_keys("keterangan")
    Log.info('Input Keterangan')

@mark.fixture_test()
def test_28_SimpanVerifikasi():
    sleep(driver)
    driver.find_element(By.ID, "simpanVerifikasi").click()
    Log.info('Click Button Simpan Verifikasi')

@mark.fixture_test()
def test_29_loginPusat():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    pusat(driver)
    Log.info('Login Spv Laporan Kamtib 6A')

@mark.fixture_test()
def test_30_AksesMenuPusat():
    sleep(driver)
    menulaporan6b(driver)
    Log.info('Akses halaman Laporan 6A')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_31_PilihKanwil():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterKanwil')))
    driver.find_element(By.ID, "formfilterKanwil").click()
    driver.find_element(By.ID, "formfilterKanwil").send_keys("jawa barat")
    driver.find_element(By.XPATH, "//li[contains(.,\'Jawa Barat\')]").click()
    Log.info('Pilih Kanwil')

@mark.fixture_test()
def test_32_PilihUPT():
    sleep(driver)
    driver.find_element(By.ID, "formfilterUpt").click()
    time.sleep(1)
    driver.find_element(By.ID, "formfilterUpt").send_keys("rutan kelas I ")
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
    driver.find_element(By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]").click()
    Log.info('PilihUPT')

@mark.fixture_test()
def test_33_ClickButtonSearch():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
    driver.find_element(By.CSS_SELECTOR, "#searchButton > span").click()
    Log.info('Click Button Search')


@mark.fixture_test()
def test_34_exit():
    time.sleep(5)
    quit(driver)





    
    