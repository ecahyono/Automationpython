from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime
import pytest
import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("data"))
    file_path = environ.get("fakermac")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
import random

from Settings.setupkeamanan import initDriver, loadDataPath, quit, sleep, upload
from Settings.loginkeamanan import oplapkamtibwaru,kanwiljabar,pusat, SpvRutanBdg
from Settings.Page.keamanan import menulaporan7e
import random
import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Loglaporan7E.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)



workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Laporan7Faker'

fake = Faker('id_ID')
KondisiFisik = ['pisau','obeng','tali','obat','Kipas','Kacamata','uang','Kunci','Hp','Kertas','Buku']
pemilikbarang = ['IRIANA TAMBa BINTI waloyo','MURSININ HARYANTi BIN sugiono','WULAN SIREGAr BIN waloyo','NILAM WIBOWo BIN sugiono','CUT NILAM HARTATi BINTI doniserigar']
jumlahbarang = ['1','2','3','4', '5', '6' ]
LariDari = ['dalamLapas','luarLapas']
for i in range(5):
    NamaPenangkapfaker = fake.name()
    LariDariFaker = random.choice(LariDari)
    KondisiFisikFaker = random.choice(KondisiFisik)
    keteranganFaker = fake.text()
    
    

    worksheet.append([
        NamaPenangkapfaker, LariDariFaker, KondisiFisikFaker, keteranganFaker
        ])
workbook.save(file_path)

workbook = load_workbook(filename=file_path)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=2, values_only=True):
    NamaFkr                     = row[0]
    LariDariFkr                 = row[1]
    KondisiFisikFkr             = row[2]
    keteranganFkr               = row[3]



@pytest.mark.webtest
def test_1_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      

@pytest.mark.webtest
def test_2_Login():
    Log.info('Setup Os')
    try:
        oplapkamtibwaru(driver)
        Log.info('Login Op Laporan Kamtib 7D')
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)

@pytest.mark.webtest
def test_3_AksesMenu():
    sleep(driver)
    menulaporan7e(driver)
    Log.info('Akses halaman Laporan 7D')
    attach(data=driver.get_screenshot_as_png())

        
@pytest.mark.webtest
def test_4_ClickButtonEdit():
    sleep(driver)
    driver.find_element(By.CSS_SELECTOR, ".text-green-500 path").click()

@pytest.mark.webtest
def test_5_PetugasPenangkap():
    sleep(driver)
    time.sleep(2)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".el-tag--success > .el-tag__content")))
        driver.find_element(By.ID, "petugasPenangkap").click()
        driver.find_element(By.ID, "petugasPenangkap").send_keys(NamaFkr)
        Log.info('Input Petugas penangkap')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        
        Log.info('Gagal Input Petugas penangkap')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@pytest.mark.webtest
def test_6_LariDari():
    sleep(driver)
    try:
        driver.find_element(By.ID, "isLariDalamUpt").click()
        driver.find_element(By.ID, ""+LariDariFkr+"").click()
        driver.execute_script("window.scrollTo(0,254)")
        Log.info('Input Lari dari')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal Input Lari dari')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@pytest.mark.webtest
def test_7_KondisiFisik():
    sleep(driver)
    try:
        driver.find_element(By.ID, "kondisiFisik").click()
        driver.find_element(By.ID, "kondisiFisik").send_keys(KondisiFisikFkr)
        Log.info('Input Kondisi Fisik')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal Input Kondisi Fisik')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@pytest.mark.webtest
def test_8_Keterangan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "keterangan").click()
        driver.find_element(By.ID, "keterangan").send_keys(keteranganFkr)
        Log.info('Input Keterangan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal Input Keterangan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@pytest.mark.webtest
def test_9_ClickButtonSimpan():
    try:
        driver.find_element(By.ID, "submitButton").click()
        Log.info('Click Button Simpan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal Click Button Simpan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@pytest.mark.webtest
def test_10_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      

@pytest.mark.webtest
def test_11_LoginSPV():
    Log.info('Setup Os')
    SpvRutanBdg(driver)
    Log.info('Login Op Laporan Kamtib 7 E')


@pytest.mark.webtest
def test_12_AksesMenuSpv():
    sleep(driver)
    try:
        menulaporan7e(driver)
        Log.info('Akses halaman Laporan 7 E')
        attach(data=driver.get_screenshot_as_png())
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal Akses halaman Laporan 7 E')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@pytest.mark.webtest
def test_13_KirimLaporanSpv():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasiLaporan')))
        time.sleep(2)
        driver.find_element(By.ID, "buttonVerifikasiLaporan").click()
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'kirimLaporan')))
        time.sleep(6)
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#kirimLaporan > span')))
        driver.find_element(By.CSS_SELECTOR, "#kirimLaporan > span").click()
        Log.info('Kirim laporan')
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal Kirim laporan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@pytest.mark.webtest
def test_14_loginKanwil():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    kanwiljabar(driver)
    Log.info('Login Spv Laporan Kamtib 7 E')

@pytest.mark.webtest
def test_15_AksesMenuSpv():
    sleep(driver)
    try:
        menulaporan7e(driver)
        Log.info('Akses halaman Laporan 7 E')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal Akses halaman Laporan 7 E')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@pytest.mark.webtest
def test_16_VerifikasiLaporanKanwil():
    try:
        sleep(driver)
        time.sleep(2)
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterUpt2')))
        driver.find_element(By.ID, "formfilterUpt2").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterUpt2").send_keys("Rutan Kelas")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
        
        driver.find_element(By.XPATH, "//span[contains(.,\'Rutan Kelas I Bandung\')]").click()
        Log.info('Input UPT')
        assert True
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal Input UPT')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@pytest.mark.webtest
def test_17_ClickButtonSearch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        driver.find_element(By.ID, "searchButton").click()
        Log.info('Click Button Search')
        assert True
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal Click Button Search')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@pytest.mark.webtest
def test_18_ClickButtonVerifikasi():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasi')))
        driver.find_element(By.ID, "buttonVerifikasi").click()
        Log.info('Click Button Verifikasi')
        assert True 
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal Click Button Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@pytest.mark.webtest
def test_19_StatusVerifikasiModal():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'simpanVerifikasi')))
        time.sleep(3)
        driver.find_element(By.ID, "statusVerifikasiModal").click()
        Log.info('Click Verifikasi Modal')
        assert True
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal Click Verifikasi Modal')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@pytest.mark.webtest
def test_20_UbahStatus():
    sleep(driver)
    try:
        driver.find_element(By.ID, "diizinkan").click()
        Log.info('Ubah Status Verifikasi')
        assert True
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal Ubah Status Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@pytest.mark.webtest
def test_21_Inputketerangan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "keterangan").send_keys("keterangan")
        Log.info('Input Keterangan')
        assert True
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal Input Keterangan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@pytest.mark.webtest
def test_22_SimpanVerifikasi():
    sleep(driver)
    try:
        driver.find_element(By.ID, "simpanVerifikasi").click()
        Log.info('Click Button Simpan Verifikasi')
        assert True
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal Click Button Simpan Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@pytest.mark.webtest
def test_23_loginPusat():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    pusat(driver)
    Log.info('Login Spv Laporan Kamtib 7 E')

@pytest.mark.webtest
def test_24_AksesMenuPusat():
    sleep(driver)
    try:
        menulaporan7e(driver)
        Log.info('Akses halaman Laporan 7 E')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal Akses halaman Laporan 7 E')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@pytest.mark.webtest
def test_25_PilihKanwil():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterKanwil')))
        driver.find_element(By.ID, "formfilterKanwil").send_keys("jawa barat")
        driver.find_element(By.XPATH, "//li[contains(.,\'Jawa Barat\')]").click()
        Log.info('Pilih Kanwil')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal Pilih Kanwil')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@pytest.mark.webtest
def test_26_PilihUPT():
    sleep(driver)
    try:
        driver.find_element(By.ID, "formfilterUpt").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterUpt").send_keys("rutan kelas")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
        driver.find_element(By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]").click()
        Log.info('PilihUPT')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal PilihUPT')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@pytest.mark.webtest
def test_27_ClickButtonSearch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        driver.find_element(By.CSS_SELECTOR, "#searchButton > span").click()
        Log.info('Click Button Search')
        attach(data=driver.get_screenshot_as_png())
    except Exception as e:
        print(
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        Log.info('Gagal Click Button Search')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@pytest.mark.webtest
def test_28_exit():
    quit(driver)
    Log.info('Exit')







    
    