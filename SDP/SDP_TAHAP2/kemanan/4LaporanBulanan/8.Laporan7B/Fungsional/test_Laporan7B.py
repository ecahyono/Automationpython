from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUAT"))
    file_path = environ.get("fakermac")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))


from Settings.setupkeamanan import initDriver, loadDataPath, quit, sleep
from Settings.loginkeamanan import oplapkamtibwaru,kanwiljabar,pusat, SpvRutanBdg
from Settings.Page.keamanan import menulaporan7b
import random
import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Loglaporan7B.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)


i = 2
laporan7b = wb['Laporan7B']


Blok                                                = laporan7b['A'+str(i)].value
Lantai                                              = laporan7b['B'+str(i)].value
Kamar                                               = laporan7b['C'+str(i)].value

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Laporan7Faker'

fake = Faker('id_ID')
Lokasikegiatan = ['POLRES BANDUNG','POLRES CIMAHI','POLDA JABAR','POSEK CIMAHI','KAMTIBNAS BANDUNG','MABES TNI']
pemilikbarang = ['IRIANA TAMBa BINTI waloyo','MURSININ HARYANTi BIN sugiono','WULAN SIREGAr BIN waloyo','NILAM WIBOWo BIN sugiono','CUT NILAM HARTATi BINTI doniserigar']
jumlahAgen = ['1','2','3','4', '5', '6' ]
TindakLanjutBarang = ['Evaluasi','Agenda Ulang']
for i in range(5):
    tanggalKegiatanIntelejenfaker = fake.date_between(start_date='-7d', end_date='today').strftime('%d/%m/%Y')
    LokasiKegiatanFaker = random.choice(Lokasikegiatan)
    jumlahAgenFaker = random.choice(jumlahAgen)
    HasilKegiatanFaker = fake.paragraph(nb_sentences=2)
    TindakLanjutFaker = random.choice(TindakLanjutBarang)
    keteranganFaker = fake.paragraph(nb_sentences=2)
    

    worksheet.append([
         tanggalKegiatanIntelejenfaker,
         LokasiKegiatanFaker,
         jumlahAgenFaker,
         HasilKegiatanFaker, 
         TindakLanjutFaker,
         keteranganFaker
        ])
workbook.save(file_path)

workbook = load_workbook(filename=file_path)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=2, values_only=True):
    tanggalkegiatanfkr                 = row[0]
    lokasikegiatanfkr                  = row[1]
    jumlahagenfkr                      = row[2]
    hasilkegiatanfkr                   = row[3]
    tindaklanjutfkr                    = row[4]
    keteranganFkr                      = row[5]
   


@mark.fixture_test()
def test_1_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      

@mark.fixture_test()
def test_2_Login():
    Log.info('Setup Os')
    oplapkamtibwaru(driver)
    Log.info('Login Op Laporan Kamtib 7 B')

@mark.fixture_test()
def test_2_AksesMenu():
    sleep(driver)
    try:
        menulaporan7b(driver)
        Log.info('Akses halaman Laporan 7 B')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Akses halaman Laporan 7 B')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_3_CreateButton():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'createButton')))
        driver.find_element(By.ID, "createButton").click()
        Log.info('Klik tombol tambah data')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Klik tombol tambah data')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_4_inputTglKegitatan():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'tglKegiatan')))
        driver.find_element(By.ID, "tglKegiatan").click()
        driver.find_element(By.ID, "tglKegiatan").send_keys(tanggalkegiatanfkr)
        driver.find_element(By.CSS_SELECTOR, ".el-form:nth-child(2) .el-form-item__content").click()
        Log.info('Input tanggal kegiatan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input tanggal kegiatan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_5_InputLokasiKegiatan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "lokasiKegiatan").click()
        driver.find_element(By.ID, "lokasiKegiatan").send_keys(lokasikegiatanfkr)
        Log.info('Input lokasi kegiatan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input lokasi kegiatan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_6_InputJumlahAgen():
    sleep(driver)
    try:
        driver.find_element(By.ID, "jumlahAgen").click()
        driver.find_element(By.ID, "jumlahAgen").send_keys(jumlahagenfkr)
        Log.info('Input jumlah agen')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input jumlah agen')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_7_InputHasilkegiatan():
    try:
        sleep(driver)
        driver.find_element(By.ID, "hasilKegiatan").send_keys(hasilkegiatanfkr)
        Log.info('Input hasil kegiatan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input hasil kegiatan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_8_InputTindaklanjut():
    sleep(driver)
    try:
        driver.find_element(By.ID, "tindakLanjut").send_keys(tindaklanjutfkr)
        Log.info('Input tindak lanjut')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input tindak lanjut')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_9_InputKeterangan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "keterangan").send_keys(keteranganFkr)
        Log.info('Input keterangan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input keterangan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_10_ClickbuttonSubmit():
    sleep(driver)
    try:
        driver.find_element(By.CSS_SELECTOR, "#submitButton > span").click()
        Log.info('Klik tombol submit')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Klik tombol submit')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_11_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      

@mark.fixture_test()
def test_12_LoginSPV():
    Log.info('Setup Os')
    SpvRutanBdg(driver)
    Log.info('Login Op Laporan Kamtib 7 B')


@mark.fixture_test()
def test_13_AksesMenuSpv():
    sleep(driver)
    try:
        menulaporan7b(driver)
        Log.info('Akses halaman Laporan 7 B')
        attach(data=driver.get_screenshot_as_png())
    except:
        Log.info('Gagal Akses halaman Laporan 7 B')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_14_KirimLaporanSpv():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasiLaporan')))
        time.sleep(2)
        driver.find_element(By.ID, "buttonVerifikasiLaporan").click()
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'kirimLaporan')))
        time.sleep(3)
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#kirimLaporan > span')))
        driver.find_element(By.CSS_SELECTOR, "#kirimLaporan > span").click()
        Log.info('Kirim laporan')
    except:
        Log.info('Gagal Kirim laporan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_15_loginKanwil():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    kanwiljabar(driver)
    Log.info('Login Spv Laporan Kamtib 6D')

@mark.fixture_test()
def test_16_AksesMenuSpv():
    sleep(driver)
    try:
        menulaporan7b(driver)
        Log.info('Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_17_VerifikasiLaporanKanwil():
    try:
        sleep(driver)
        time.sleep(2)
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterUpt2')))
        driver.find_element(By.ID, "formfilterUpt2").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterUpt2").send_keys("Rutan Kelas")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
        
        driver.find_element(By.XPATH, "//span[contains(.,\'Rutan Kelas I Bandung\')]").click()
        Log.info('Input UPT')
        assert True
    except:
        Log.info('Gagal Input UPT')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_18_ClickButtonSearch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        driver.find_element(By.ID, "searchButton").click()
        Log.info('Click Button Search')
        assert True
    except:
        Log.info('Gagal Click Button Search')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_19_ClickButtonVerifikasi():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasi')))
        driver.find_element(By.ID, "buttonVerifikasi").click()
        Log.info('Click Button Verifikasi')
        assert True 
    except:
        Log.info('Gagal Click Button Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_20_StatusVerifikasiModal():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'simpanVerifikasi')))
        time.sleep(3)
        driver.find_element(By.ID, "statusVerifikasiModal").click()
        Log.info('Click Verifikasi Modal')
        assert True
    except:
        Log.info('Gagal Click Verifikasi Modal')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_21_UbahStatus():
    sleep(driver)
    try:
        driver.find_element(By.ID, "diizinkan").click()
        Log.info('Ubah Status Verifikasi')
        assert True
    except:
        Log.info('Gagal Ubah Status Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_22_Inputketerangan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "keterangan").send_keys("keterangan")
        Log.info('Input Keterangan')
        assert True
    except:
        Log.info('Gagal Input Keterangan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_23_SimpanVerifikasi():
    sleep(driver)
    try:
        driver.find_element(By.ID, "simpanVerifikasi").click()
        Log.info('Click Button Simpan Verifikasi')
        assert True
    except:
        Log.info('Gagal Click Button Simpan Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_24_loginPusat():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    pusat(driver)
    Log.info('Login Spv Laporan Kamtib 6D')

@mark.fixture_test()
def test_25_AksesMenuPusat():
    sleep(driver)
    try:
        menulaporan7b(driver)
        Log.info('Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_26_PilihKanwil():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterKanwil')))
        driver.find_element(By.ID, "formfilterKanwil").send_keys("jawa barat")
        driver.find_element(By.XPATH, "//li[contains(.,\'Jawa Barat\')]").click()
        Log.info('Pilih Kanwil')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Pilih Kanwil')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_27_PilihUPT():
    sleep(driver)
    try:
        driver.find_element(By.ID, "formfilterUpt").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterUpt").send_keys("rutan kelas")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
        driver.find_element(By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]").click()
        Log.info('PilihUPT')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal PilihUPT')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_28_ClickButtonSearch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        driver.find_element(By.CSS_SELECTOR, "#searchButton > span").click()
        Log.info('Click Button Search')
        attach(data=driver.get_screenshot_as_png())
    except:
        Log.info('Gagal Click Button Search')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_29_exit():
    quit(driver)
    Log.info('Exit')








    
    