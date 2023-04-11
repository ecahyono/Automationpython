from distutils.archive_util import make_archive
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUAT"))

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))


from Settings.setupkeamanan import initDriver, loadDataPath, quit, sleep
from Settings.loginkeamanan import oplapkamtibwaru,kanwiljabar,pusat, SpvRutanBdg
from Settings.Page.keamanan import menulaporan6d

import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Loglaporan6D.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)


i = 2
laporan6D = wb['Laporan6D']


jenisSarana                              = laporan6D['A'+str(i)].value
kondisiBaik                              = laporan6D['B'+str(i)].value
kondisiRusak                             = laporan6D['C'+str(i)].value
keterangan                               = laporan6D['D'+str(i)].value

usulan                                   = laporan6D['F'+str(i)].value
jumlah                                   = laporan6D['G'+str(i)].value

nama                                     = laporan6D['I'+str(i)].value
deskripsi                                = laporan6D['J'+str(i)].value
status                                   = laporan6D['K'+str(i)].value




@mark.fixture_test()
def test_1_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    

@mark.fixture_test()
def test_2_Login():
    Log.info('Setup Os')
    oplapkamtibwaru(driver)
    Log.info('Login Op Laporan Kamtib 6 D')


@mark.fixture_test()
def test_2_AksesMenu():
    sleep(driver)
    try:    
        menulaporan6d(driver)
        Log.info('Akses halaman Laporan 6 D')
        attach(data=driver.get_screenshot_as_png())
    except:
        Log.info('Gagal Akses halaman Laporan 6 D')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_3_CreateButton():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'createButton')))
        driver.find_element(By.ID, "createButton").click()
        Log.info('Klik tombol tambah data')
    except:
        Log.info('Gagal Klik tombol tambah data')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_4_InputJenisSarana():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'submitButton')))
        driver.find_element(By.ID, "jenisSarana").click()
        driver.execute_script("window.scrollTo(0,0)")
        driver.find_element(By.ID, "jenisSarana").send_keys(jenisSarana)
        driver.find_element(By.XPATH, "//li[contains(.,\'"+ jenisSarana +"')]").click()
        Log.info('Input Jenis Sarana')
    except:
        Log.info('Gagal Input Jenis Sarana')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_5_InputKondisi():
    sleep(driver)
    try:
        driver.find_element(By.CSS_SELECTOR, "#kondisiBaik .el-input__inner").send_keys(kondisiBaik)
        driver.find_element(By.CSS_SELECTOR, "#kondisiRusak .el-input__inner").send_keys(kondisiRusak)
        Log.info('Input Kondisi')
    except:
        Log.info('Gagal Input Kondisi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_6_InputKeterangan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "keterangan").click()
        driver.find_element(By.ID, "keterangan").send_keys(keterangan)
        Log.info('Input Keterangan')
    except:
        Log.info('Gagal Input Keterangan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_7_ButtonSubmit():
    sleep(driver)
    try:
        driver.find_element(By.CSS_SELECTOR, "#submitButton > span").click()
        WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, '//div[contains(.,\'Berhasil Ditambahkan\')]')))
        Log.info('Klik tombol submit')
    except:
        Log.info('Gagal Klik tombol submit')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_8_TambahUsulan():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.invisibility_of_element((By.CSS_SELECTOR, '.el-loading-mask')))
        driver.find_element(By.CSS_SELECTOR, "#tambahUsulan > span").click()
        driver.find_element(By.ID, "usulan").send_keys(usulan)
        driver.find_element(By.ID, "usulan5").click()
    except:
        Log.info('Gagal Klik tombol tambah usulan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False
    

@mark.fixture_test()
def test_9_InputJumlah():
    sleep(driver)
    try:
        driver.find_element(By.ID, "jumlah").click()
        driver.find_element(By.ID, "jumlah").send_keys(jumlah)
        Log.info('Input Jumlah')
    except:
        Log.info('Gagal Input Jumlah')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_10_ClickButtonSUbmit():
    sleep(driver)
    try:
        driver.find_element(By.CSS_SELECTOR, "#submitButton > span").click()
        WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, '//div[contains(.,\'Berhasil Ditambahkan\')]')))
    except:
        Log.info('Gagal Klik tombol submit')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_11_SetupOS():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')


@mark.fixture_test()
def test_12_LoginSpv():
    sleep(driver)
    SpvRutanBdg(driver)
    Log.info('Login Spv Laporan Kamtib 6D')

@mark.fixture_test()
def test_13_AksesMenuSpv():
    sleep(driver)
    try:
        menulaporan6d(driver)
        Log.info('Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
    except:
        Log.info('Gagal Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_14_KirimLaporanSpv():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasiLaporan')))
        time.sleep(2)
        driver.find_element(By.ID, "buttonVerifikasiLaporan").click()
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'kirimLaporan')))
        time.sleep(3)
        driver.find_element(By.CSS_SELECTOR, "#kirimLaporan > span").click()
    except:
        Log.info('Gagal Kirim Laporan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_15_loginKanwil():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    kanwiljabar(driver)
    Log.info('Login Spv Laporan Kamtib 6D')

@mark.fixture_test()
def test_16_AksesMenuSpv():
    sleep(driver)
    menulaporan6d(driver)
    Log.info('Akses halaman Laporan 6D')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_17_VerifikasiLaporanKanwil():
    sleep(driver)
    try:
        time.sleep(2)
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterUpt2')))
        driver.find_element(By.ID, "formfilterUpt2").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterUpt2").send_keys("Rutan Kelas")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))

        driver.find_element(By.XPATH, "//span[contains(.,\'Rutan Kelas I Bandung\')]").click()
        Log.info('Input UPT')
    except:
        Log.info('Gagal Input UPT')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_18_ClickButtonSearch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        driver.find_element(By.ID, "searchButton").click()
        Log.info('Click Button Search')
    except:
        Log.info('Gagal Click Button Search')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_19_ClickButtonVerifikasi():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasi')))
        driver.find_element(By.ID, "buttonVerifikasi").click()
        Log.info('Click Button Verifikasi')
    except:
        Log.info('Gagal Click Button Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_20_StatusVerifikasiModal():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'simpanVerifikasi')))
        time.sleep(3)
        driver.find_element(By.ID, "statusVerifikasiModal").click()
        Log.info('Click Verifikasi Modal')
    except:
        Log.info('Gagal Click Verifikasi Modal')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_21_UbahStatus():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'diizinkan')))
        driver.find_element(By.ID, "diizinkan").click()
        Log.info('Ubah Status Verifikasi')
    except:
        Log.info('Gagal Ubah Status Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_22_Inputketerangan():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'keterangan')))
        driver.find_element(By.ID, "keterangan").click()
        Log.info('Click Keterangan')
    except:
        Log.info('Gagal Click Keterangan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_23_SimpanVerifikasi():
    try:
        sleep(driver)
        driver.find_element(By.ID, "simpanVerifikasi").click()
        Log.info('Click Button Simpan Verifikasi')
    except:
        Log.info('Gagal Click Button Simpan Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_24_loginPusat():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    pusat(driver)
    Log.info('Login Spv Laporan Kamtib 6D')

@mark.fixture_test()
def test_25_AksesMenuPusat():
    sleep(driver)
    menulaporan6d(driver)
    Log.info('Akses halaman Laporan 6D')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_26_PilihKanwil():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterKanwil')))
        driver.find_element(By.ID, "formfilterKanwil").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterKanwil").send_keys("jawa barat")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Jawa Barat\')]")))
        driver.find_element(By.XPATH, "//li[contains(.,\'Jawa Barat\')]").click()
        Log.info('Pilih Kanwil')
    except:
        Log.info('Gagal Pilih Kanwil')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_27_PilihUPT():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterUpt')))
        driver.find_element(By.ID, "formfilterUpt").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterUpt").send_keys("rutan kelas")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
        driver.find_element(By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]").click()
        Log.info('PilihUPT')
    except:
        Log.info('Gagal Pilih UPT')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_28_ClickButtonSearch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        driver.find_element(By.CSS_SELECTOR, "#searchButton > span").click()
        Log.info('Click Button Search')
    except:
        Log.info('Gagal Click Button Search')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_29_ClickButtonMaster():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonMaster')))
        driver.find_element(By.ID, "buttonMaster").click()
        Log.info('Click Button Master')
    except:
        Log.info('Gagal Click Button Master')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_30_CreateButton():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'createButton')))
        driver.find_element(By.ID, "createButton").click()
        Log.info('Click Button Create')
    except:
        Log.info('Gagal Click Button Create')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_31_InputNama():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'nama')))
        driver.find_element(By.ID, "nama").click()
        Log.info('Click Nama barang')
    except:
        Log.info('Gagal Click Nama barang')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_32_InputDeskripsi():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'deskripsi')))
        driver.find_element(By.ID, "deskripsi").click()
        Log.info('Click Deskripsi')
    except:
        Log.info('Gagal Click Deskripsi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_33_Switch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//div[@id='app']/div/div[2]/div/div[2]/div/div/div[2]/form/div[3]/div/div/div/span/span")))
        driver.find_element(By.XPATH, "//div[@id='app']/div/div[2]/div/div[2]/div/div/div[2]/form/div[3]/div/div/div/span/span").click()
        Log.info('Click Switch')
    except:
        Log.info('Gagal Click Switch')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_34_Tambah():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'submitButton')))
        driver.find_element(By.ID, "submitButton").click()
        Log.info('Click Button Tambah')
    except:
        Log.info('Gagal Click Button Tambah')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_35_exit():
    quit(driver)
    Log.info('Exit')






    
    