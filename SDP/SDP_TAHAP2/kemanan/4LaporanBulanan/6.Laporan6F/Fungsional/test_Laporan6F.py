from distutils.archive_util import make_archive
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUAT"))

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))


from Settings.setupkeamanan import initDriver, loadDataPath, quit, sleep
from Settings.loginkeamanan import oplapkamtibwaru,kanwiljabar,pusat, SpvRutanBdg
from Settings.Page.keamanan import menulaporan6f

import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Loglaporan6f.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)


i = 2
laporan6e = wb['Laporan6F']


luasUpt                                             = laporan6e['A'+str(i)].value
kapasitasHunian                                     = laporan6e['B'+str(i)].value
LuasBlok                                            = laporan6e['C'+str(i)].value
jumlahPenghuni                                      = laporan6e['D'+str(i)].value
KamarLayak                                          = laporan6e['E'+str(i)].value
KamarTidakLayak                                     = laporan6e['F'+str(i)].value
ventilasiLayak                                      = laporan6e['G'+str(i)].value
ventilasiTidakLayak                                 = laporan6e['H'+str(i)].value
Pencahayaanlayak                                    = laporan6e['I'+str(i)].value

Pencahayaantidaklayak                               = laporan6e['J'+str(i)].value
KamarMandiLayak                                     = laporan6e['K'+str(i)].value
KamarMandiTidakLayak                                = laporan6e['L'+str(i)].value
WcLayak                                             = laporan6e['M'+str(i)].value
WcTidakLayak                                        = laporan6e['N'+str(i)].value
AirLayak                                            = laporan6e['O'+str(i)].value
AirTidakLayak                                       = laporan6e['P'+str(i)].value

TempatTidurLayak                                    = laporan6e['Q'+str(i)].value
TempatTidurTidakLayak                               = laporan6e['R'+str(i)].value
KasurLayak                                          = laporan6e['S'+str(i)].value
KasurTidakLayak                                     = laporan6e['T'+str(i)].value
LemariLayak                                         = laporan6e['U'+str(i)].value
LemariTidakLayak                                    = laporan6e['V'+str(i)].value

@mark.fixture_test()
def test_1_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      

@mark.fixture_test()
def test_2_Login():
    Log.info('Setup Os')
    oplapkamtibwaru(driver)
    Log.info('Login Op Laporan Kamtib 6 F')

@mark.fixture_test()
def test_2_AksesMenu():
    sleep(driver)
    try:
        menulaporan6f(driver)
        Log.info('Akses halaman Laporan 6 F')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Akses halaman Laporan 6 F')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_3_CreateButton():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'createButton')))
        driver.find_element(By.ID, "createButton").click()
        Log.info('Klik tombol tambah data')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Klik tombol tambah data')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_4_InputLuasUpt():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'backButton')))
        driver.find_element(By.ID, "luasUpt").send_keys(luasUpt)
        Log.info('Input Luas UPT')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Luas UPT')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_5_InputKapasitasHunian():
    sleep(driver)
    try:
        driver.find_element(By.ID, "kapasitasHunian").send_keys(kapasitasHunian)
        Log.info('Input Kapasitas Hunian')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Kapasitas Hunian')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_6_InputLuasBlok():
    sleep(driver)
    try:
        driver.find_element(By.ID, "luasBlok").send_keys(LuasBlok)
        Log.info('Input Luas Blok')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Luas Blok')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_7_InputJumlahPenghuni():
    sleep(driver)
    try:
        driver.find_element(By.ID, "jumlahPenghuni").send_keys(jumlahPenghuni)
        Log.info('Input Jumlah Penghuni')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Jumlah Penghuni')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_8_InputKamarLayak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "layak0").send_keys(KamarLayak)
        Log.info('Input kamar layak')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input kamar layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_9_InputKamarTidakLayak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "tidakLayar0").click()
        driver.find_element(By.ID, "tidakLayar0").send_keys(KamarTidakLayak)
        Log.info('Input kamar tidak layak')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input kamar tidak layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_10_InputVentilasiLayak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "layak1").click()
        driver.find_element(By.ID, "layak1").send_keys(ventilasiLayak)
        Log.info('Input ventilasi')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input ventilasi Layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_11_InputVentilasiTidakLayak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "tidakLayar1").click()
        driver.find_element(By.ID, "tidakLayar1").send_keys(ventilasiTidakLayak)
        Log.info('Input ventilasi')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input ventilasi Tidak Layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False
        
@mark.fixture_test()
def test_12_InputPencahayaanLayak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "layak2").click()
        driver.find_element(By.ID, "layak2").send_keys(Pencahayaanlayak)
        Log.info('Input Pencahayaan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Pencahayaan Layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_13_InputPencahayaanLayak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "tidakLayar2").click()
        driver.find_element(By.ID, "tidakLayar2").send_keys(Pencahayaantidaklayak)
        Log.info('Input Pencahayaan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Pencahayaan Layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_14_InputKamarMandiLayak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "layak3").click()
        driver.find_element(By.ID, "layak3").send_keys(KamarMandiLayak)
        Log.info('Input Kamar Mandi Layak')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Kamar Mandi Layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_15_InputKamarMandiTidakLayal():
    sleep(driver)
    try:
        driver.find_element(By.ID, "tidakLayar3").click()
        driver.find_element(By.ID, "tidakLayar3").send_keys(KamarMandiTidakLayak)
        Log.info('Input Kamar Mandi Tidak Layak')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Kamar Mandi Tidak Layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_16_InputWcLayak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "layak4").click()
        driver.find_element(By.ID, "layak4").send_keys(WcLayak)
        Log.info('input Wc Layak')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('gagal input Wc Layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_17_InpuWcTidakLayak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "tidakLayar4").click()
        driver.find_element(By.ID, "tidakLayar4").send_keys(WcTidakLayak)
        Log.info('input Wc Tidak Layak')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('gagal input Wc Tidak Layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_18_InputAirLayak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "layak5").click()
        driver.find_element(By.ID, "layak5").send_keys(AirLayak)
        Log.info('input Air Layak')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info(' gagal input Air Layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_19_InputAirLayak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "tidakLayar5").click()
        driver.find_element(By.ID, "tidakLayar5").send_keys(AirTidakLayak)
        Log.info('input Air Tidak Layak')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('gagal input Air Tidak Layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_20_InputTempatTidurLayak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "layak6").click()
        driver.find_element(By.ID, "layak6").send_keys(TempatTidurLayak)
        Log.info('input Tempat Tidur Layak')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('gagal input Tempat Tidur Layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_21_InputTempatTidurTidakLayak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "tidakLayar6").click()
        driver.find_element(By.ID, "tidakLayar6").send_keys(TempatTidurTidakLayak)
        Log.info('input Tempat Tidur Tidak Layak')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('gaagal input Tempat Tidur Tidak Layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_22_kasurLayak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "layak7").click()
        driver.find_element(By.ID, "layak7").send_keys(KasurLayak)
        Log.info('kasur Layak')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('gagal kasur Layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_23_KasurTidakLayak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "tidakLayar7").click()
        driver.find_element(By.ID, "tidakLayar7").send_keys(KasurTidakLayak)
        Log.info('input Kasur Tidak Layak')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('gagal input Kasur Tidak Layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_24_LemariLayak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "layak8").click()
        driver.find_element(By.ID, "layak8").send_keys(LemariLayak)
        Log.info('Lemari Layak')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('gagal Lemari Layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_25_LemariTidakLayak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "tidakLayar8").click()
        driver.find_element(By.ID, "tidakLayar8").send_keys(LemariTidakLayak)
        Log.info('input Lemari Tidak Layak')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('gagal input Lemari Tidak Layak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_26_ClickButtonSubmit():
    sleep(driver)
    try:
        driver.find_element(By.CSS_SELECTOR, "#submitButton > span").click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "searchButton")))
        Log.info('click button submit')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('gagal click button submit')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False



@mark.fixture_test()
def test_27_SetupOS():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')


@mark.fixture_test()
def test_28_LoginSpv():
    sleep(driver)
    try:
        SpvRutanBdg(driver)
        Log.info('Login Spv Laporan Kamtib 6F')
    except:
        Log.info('Gagal Login Spv Laporan Kamtib 6F')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_29_AksesMenuSpv():
    sleep(driver)
    try:
        menulaporan6f(driver)
        Log.info('Akses halaman Laporan 6F')
        attach(data=driver.get_screenshot_as_png())
    except:
        Log.info('Gagal Akses halaman Laporan 6F')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_30_KirimLaporanSpv():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasiLaporan')))
        time.sleep(2)
        driver.find_element(By.ID, "buttonVerifikasiLaporan").click()
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'kirimLaporan')))
        time.sleep(3)
        driver.find_element(By.CSS_SELECTOR, "#kirimLaporan > span").click()
        Log.info('Kirim laporan')
    except:
        Log.info('Gagal Kirim laporan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False



@mark.fixture_test()
def test_31_loginKanwil():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    kanwiljabar(driver)
    Log.info('Login Spv Laporan Kamtib 6D')

@mark.fixture_test()
def test_32_AksesMenuSpv():
    sleep(driver)
    try:
        menulaporan6f(driver)
        Log.info('Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_33_VerifikasiLaporanKanwil():
    try:
        sleep(driver)
        time.sleep(2)
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterUpt2')))
        driver.find_element(By.ID, "formfilterUpt2").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterUpt2").send_keys("Rutan Kelas")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
        
        driver.find_element(By.XPATH, "//span[contains(.,\'Rutan Kelas I Bandung\')]").click()
        Log.info('Input UPT')
        assert True
    except:
        Log.info('Gagal Input UPT')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_34_ClickButtonSearch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        driver.find_element(By.ID, "searchButton").click()
        Log.info('Click Button Search')
        assert True
    except:
        Log.info('Gagal Click Button Search')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_35_ClickButtonVerifikasi():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasi')))
        driver.find_element(By.ID, "buttonVerifikasi").click()
        Log.info('Click Button Verifikasi')
        assert True 
    except:
        Log.info('Gagal Click Button Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_36_StatusVerifikasiModal():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'simpanVerifikasi')))
        time.sleep(3)
        driver.find_element(By.ID, "statusVerifikasiModal").click()
        Log.info('Click Verifikasi Modal')
        assert True
    except:
        Log.info('Gagal Click Verifikasi Modal')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_37_UbahStatus():
    sleep(driver)
    try:
        driver.find_element(By.ID, "diizinkan").click()
        Log.info('Ubah Status Verifikasi')
        assert True
    except:
        Log.info('Gagal Ubah Status Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_38_Inputketerangan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "keterangan").send_keys("keterangan")
        Log.info('Input Keterangan')
        assert True
    except:
        Log.info('Gagal Input Keterangan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_39_SimpanVerifikasi():
    sleep(driver)
    try:
        driver.find_element(By.ID, "simpanVerifikasi").click()
        Log.info('Click Button Simpan Verifikasi')
        assert True
    except:
        Log.info('Gagal Click Button Simpan Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False



@mark.fixture_test()
def test_39_loginPusat():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    pusat(driver)
    Log.info('Login Spv Laporan Kamtib 6D')

@mark.fixture_test()
def test_40_AksesMenuPusat():
    sleep(driver)
    try:
        menulaporan6f(driver)
        Log.info('Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_41_PilihKanwil():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterKanwil')))
        driver.find_element(By.ID, "formfilterKanwil").send_keys("jawa barat")
        driver.find_element(By.XPATH, "//li[contains(.,\'Jawa Barat\')]").click()
        Log.info('Pilih Kanwil')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Pilih Kanwil')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_42_PilihUPT():
    sleep(driver)
    try:
        driver.find_element(By.ID, "formfilterUpt").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterUpt").send_keys("rutan kelas")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
        driver.find_element(By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]").click()
        Log.info('PilihUPT')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal PilihUPT')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_43_ClickButtonSearch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        driver.find_element(By.CSS_SELECTOR, "#searchButton > span").click()
        Log.info('Click Button Search')
        attach(data=driver.get_screenshot_as_png())
    except:
        Log.info('Gagal Click Button Search')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_44_exit():
    quit(driver)
    Log.info('Exit')






    
    