from distutils.archive_util import make_archive
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUAT"))

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))


from Settings.setupkeamanan import initDriver, loadDataPath, quit, sleep
from Settings.loginkeamanan import oplapkamtibwaru,kanwiljabar,pusat, SpvRutanBdg
from Settings.Page.keamanan import menulaporan6c

import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Loglaporan6c.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)


i = 2
laporan6c = wb['Laporan6C']


JenisKerusuhan                              = laporan6c['A'+str(i)].value
WaktuKejadian                               = laporan6c['B'+str(i)].value
IndikasiPenyebab                            = laporan6c['C'+str(i)].value
TotalKerugian                               = laporan6c['D'+str(i)].value
KorbanJiwaPetugas                           = laporan6c['E'+str(i)].value
KorbanJiwaWBP                               = laporan6c['F'+str(i)].value
TindakanSementaraYangDilakukan              = laporan6c['G'+str(i)].value
BantuanPihakLain                            = laporan6c['H'+str(i)].value
KondisiAkhir                                = laporan6c['I'+str(i)].value
keterangan                                  = laporan6c['J'+str(i)].value


@mark.fixture_test()
def test_1loginOperator():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    oplapkamtibwaru(driver)
    Log.info('Login Op Laporan Kamtib 6C')


@mark.fixture_test()
def test_2_AksesMenu():
    sleep(driver)
    menulaporan6c(driver)
    Log.info('Akses halaman Laporan 6 C')
    attach(data=driver.get_screenshot_as_png())


@mark.fixture_test()
def test_3_CreateButton():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'createButton')))
    driver.find_element(By.ID, "createButton").click()
    Log.info('Klik tombol tambah data')

@mark.fixture_test()
def test_4_JenisKerusuhan():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'submitButton')))
    driver.find_element(By.ID, "jenisKerusuhan").click()
    driver.find_element(By.ID, "jenisKerusuhan").send_keys(JenisKerusuhan)
    Log.info('Isi Uraian Jenis Kerusuhan')

@mark.fixture_test()
def test_5_WaktuKejadian():
    sleep(driver)
    driver.find_element(By.ID, "waktuKejadian").send_keys(WaktuKejadian)
    driver.find_element(By.ID, "waktuKejadian").send_keys(Keys.ENTER)
    Log.info('Isi Waktu Kejadian')

@mark.fixture_test()
def test_6_IndikasiPenyebab():
    sleep(driver)
    driver.find_element(By.ID, "indikasiPenyebab").send_keys(IndikasiPenyebab)
    Log.info('Isi IndikasiPenyebab')
    

@mark.fixture_test()
def test_7_TotalKerugian():
    sleep(driver)
    driver.find_element(By.ID, "kerugianMateri").send_keys(TotalKerugian)
    Log.info('Isi Total Kerugian')
    
    
@mark.fixture_test()
def test_8_KorbanJiwa():
    sleep(driver)
    driver.find_element(By.CSS_SELECTOR, "#korbanPetugas .el-input__inner").click()
    driver.find_element(By.CSS_SELECTOR, "#korbanPetugas .el-input__inner").send_keys(KorbanJiwaPetugas)
    driver.find_element(By.CSS_SELECTOR, "#korbanPenghuni .el-input__inner").click()
    driver.find_element(By.CSS_SELECTOR, "#korbanPenghuni .el-input__inner").send_keys(KorbanJiwaWBP)
    Log.info('Isi Korban Jiwa')

@mark.fixture_test()
def test_9_TindakanSementaraYangDilakukan():
    sleep(driver)
    driver.find_element(By.ID, "tindakanSementara").click()
    driver.find_element(By.ID, "tindakanSementara").send_keys(TindakanSementaraYangDilakukan)
    Log.info('Isi Tindakan Sementara Yang Dilakukan')
    
@mark.fixture_test()
def test_10_BantuanPihakLain():
    sleep(driver)
    driver.find_element(By.ID, "bantuanPihakLain").click()
    driver.find_element(By.ID, "bantuanPihakLain").send_keys(BantuanPihakLain)
    Log.info('Isi Bantuan Pihak Lain')

@mark.fixture_test()
def test_11_KondisiAkhir():
    sleep(driver)
    driver.find_element(By.ID, "kondisiAkhir").click()
    driver.find_element(By.ID, "kondisiAkhir").send_keys(KondisiAkhir)
    Log.info('Isi Kondisi Akhir')

@mark.fixture_test()
def test_12_Keterangan():
    sleep(driver)
    driver.find_element(By.ID, "keterangan").click()
    driver.find_element(By.ID, "keterangan").send_keys(keterangan)
    Log.info('Isi Keterangan')

@mark.fixture_test()
def test_13_Submit():
    sleep(driver)
    driver.find_element(By.ID, "submitButton").click()
    Log.info('Klik tombol submit')


@mark.fixture_test()
def test_14_loginSPV():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    SpvRutanBdg(driver)
    Log.info('Login Spv Laporan Kamtib 6C')

@mark.fixture_test()
def test_15_AksesMenuSpv():
    sleep(driver)
    menulaporan6c(driver)
    Log.info('Akses halaman Laporan 6C')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_16_KirimLaporanSpv():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasiLaporan')))
    time.sleep(2)
    driver.find_element(By.ID, "buttonVerifikasiLaporan").click()
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'kirimLaporan')))
    time.sleep(2)
    driver.find_element(By.CSS_SELECTOR, "#kirimLaporan > span").click()


@mark.fixture_test()
def test_17_loginKanwil():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    kanwiljabar(driver)
    Log.info('Login Spv Laporan Kamtib 6C')

@mark.fixture_test()
def test_18_AksesMenuSpv():
    sleep(driver)
    menulaporan6c(driver)
    Log.info('Akses halaman Laporan 6C')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_19_VerifikasiLaporanKanwil():
    sleep(driver)
    time.sleep(2)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterUpt2')))
    driver.find_element(By.ID, "formfilterUpt2").click()
    time.sleep(1)
    driver.find_element(By.ID, "formfilterUpt2").send_keys("Rutan Kelas")
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
    
    driver.find_element(By.XPATH, "//span[contains(.,\'Rutan Kelas I Bandung\')]").click()
    Log.info('Input UPT')

@mark.fixture_test()
def test_20_ClickButtonSearch():
    sleep(driver)
    driver.find_element(By.ID, "searchButton").click()
    Log.info('Click Button Search')

@mark.fixture_test()
def test_21_ClickButtonVerifikasi():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasi')))
    driver.find_element(By.ID, "buttonVerifikasi").click()
    Log.info('Click Button Verifikasi')


@mark.fixture_test()
def test_22_StatusVerifikasiModal():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'simpanVerifikasi')))
    time.sleep(3)
    driver.find_element(By.ID, "statusVerifikasiModal").click()
    Log.info('Click Verifikasi Modal')

@mark.fixture_test()
def test_23_UbahStatus():
    sleep(driver)
    driver.find_element(By.ID, "diizinkan").click()
    Log.info('Ubah Status Verifikasi')

@mark.fixture_test()
def test_24_Inputketerangan():
    sleep(driver)
    driver.find_element(By.ID, "keterangan").send_keys("keterangan")
    Log.info('Input Keterangan')

@mark.fixture_test()
def test_25_SimpanVerifikasi():
    sleep(driver)
    driver.find_element(By.ID, "simpanVerifikasi").click()
    Log.info('Click Button Simpan Verifikasi')

@mark.fixture_test()
def test_26_loginPusat():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    pusat(driver)
    Log.info('Login Spv Laporan Kamtib 6C')

@mark.fixture_test()
def test_27_AksesMenuPusat():
    sleep(driver)
    menulaporan6c(driver)
    Log.info('Akses halaman Laporan 6C')
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_28_PilihKanwil():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterKanwil')))
    driver.find_element(By.ID, "formfilterKanwil").send_keys("jawa barat")
    driver.find_element(By.XPATH, "//li[contains(.,\'Jawa Barat\')]").click()
    Log.info('Pilih Kanwil')

@mark.fixture_test()
def test_29_PilihUPT():
    sleep(driver)
    driver.find_element(By.ID, "formfilterUpt").click()
    time.sleep(1)
    driver.find_element(By.ID, "formfilterUpt").send_keys("rutan kelas")
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
    driver.find_element(By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]").click()
    Log.info('PilihUPT')

@mark.fixture_test()
def test_30_ClickButtonSearch():
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
    driver.find_element(By.CSS_SELECTOR, "#searchButton > span").click()
    Log.info('Click Button Search')


@mark.fixture_test()
def test_31_exit():
    time.sleep(5)
    quit(driver)
    Log.info('Exit')






    
    