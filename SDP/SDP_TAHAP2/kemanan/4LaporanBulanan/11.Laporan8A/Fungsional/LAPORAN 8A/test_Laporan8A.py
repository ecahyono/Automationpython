from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime
import pytest

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("data"))
    file_path = environ.get("fakermac")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
import random

from Settings.setupkeamanan import initDriver, loadDataPath, quit, sleep, upload
from Settings.loginkeamanan import oplapkamtibwaru,kanwiljabar,pusat, SpvRutanBdg
from Settings.Page.keamanan import menulaporan8a
import random
import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Loglaporan8A.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

sheetrangeIndex = wb['ListPetugasRutan1Bandung']

i = random.randint(1,40)

NamaInput                                 = sheetrangeIndex['A'+str(i)].value
print('nama input adalah', NamaInput)
time.sleep(3)

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Laporan7Faker'

fake = Faker('id_ID')
for i in range(5):
    KronologisKejadianFaker                          = fake.text()
    TanggalPemberianBantuanHukumFaker                = fake.date_between(start_date='today', end_date='today').strftime('%d/%m/%Y')
    MelanggarPeraturanPerundangundanganFaker         = fake.text()
    PemberiBantuanHukumFaker                         = fake.name()
    PertimbanganHukumFaker                           = fake.text()
    PutusanPengadilanFaker                           = fake.text()
    KeteranganFaker                                  = fake.text()

    worksheet.append([
        KronologisKejadianFaker,TanggalPemberianBantuanHukumFaker,MelanggarPeraturanPerundangundanganFaker,PemberiBantuanHukumFaker,PertimbanganHukumFaker,PutusanPengadilanFaker,KeteranganFaker
        ])
workbook.save(file_path)

workbook = load_workbook(filename=file_path)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=2, values_only=True):
    KronologisKejadianFkr                     = row[0]
    TanggalPemberianBantuanHukumFkr           = row[1]
    MelanggarPeraturanPerundangundanganFkr    = row[2]
    PemberiBantuanHukumFkr                    = row[3]
    PertimbanganHukumFkr                      = row[4]
    PutusanPengadilanFkr                      = row[5]
    KeteranganFkr                             = row[6]
  

@pytest.mark.webtest
def test_1_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      

@pytest.mark.webtest
def test_2_Login():
    Log.info('Setup Os')
    oplapkamtibwaru(driver)
    Log.info('Login Op Laporan Kamtib 8 A')

@pytest.mark.webtest
def test_3_AksesMenu():
    sleep(driver)
    menulaporan8a(driver)
    Log.info('Akses halaman Laporan 8 A')
    attach(data=driver.get_screenshot_as_png())

@pytest.mark.webtest
def test_4_ClickButtonTambah():
    sleep(driver)
    try:
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#searchButton > span")))
        time.sleep(2)
        driver.find_element(By.CSS_SELECTOR, "#createButton").click()
        Log.info('Klik Button Tambah')
    except Exception as e:
        Log.info('Button Tambah Tidak Ditemukan' )
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)
        assert False
        quit(driver)


@pytest.mark.webtest
def test_5_NamaInput():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#backButton > span")))
        WebDriverWait(driver,10).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#pegawai")))
        driver.find_element(By.XPATH, "//input[@id=\'pegawai\']").click()
        sleep(driver)
        driver.find_element(By.CSS_SELECTOR, "#pegawai").send_keys(NamaInput)
        driver.find_element(By.XPATH, "//li[contains(.,\'"+ NamaInput +"')]").click()
        Log.info('Input Nama')
    except Exception as e:
        Log.info('Nama Input Tidak Ditemukan')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)
        assert False

@pytest.mark.webtest
def test_6_KronologisKejadianFkr():
    sleep(driver)
    try:
        driver.find_element(By.ID, "kronologis").send_keys(KronologisKejadianFkr)
        Log.info('Input Kronologis Kejadian ')
    except Exception as e:
        Log.info('Kronologis Kejadian Tidak Ditemukan')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)
        assert False

@pytest.mark.webtest
def test_7_TanggalPemberianBantuanHukumFkr():
    sleep(driver)
    try:
        driver.find_element(By.ID, "tanggal").send_keys(TanggalPemberianBantuanHukumFkr)
        Log.info('Input Tanggal Pemberian Bantuan Hukum ')
    except Exception as e:
        Log.info('Tanggal Pemberian Bantuan Hukum Tidak Ditemukan')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)
        assert False


@pytest.mark.webtest
def test_8_MelanggarPeraturanPerundangundanganFkr():
    sleep(driver)
    try:
        driver.find_element(By.CSS_SELECTOR, ".el-form:nth-child(2)").click()
        driver.find_element(By.ID, "melanggar").send_keys(MelanggarPeraturanPerundangundanganFkr)
        Log.info('Input Melanggar Peraturan Perundangundangan')
    except Exception as e:
        Log.info('Melanggar Peraturan Perundangundangan Tidak Ditemukan')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)
        assert False
    


@pytest.mark.webtest
def test_9_PemberiBantuanHukumFkr():
    sleep(driver)
    try:    
        driver.find_element(By.ID, "pemberiBantuan").send_keys(PemberiBantuanHukumFkr)
        Log.info('Input Pemberi Bantuan Hukum ')
    except Exception as e:
        Log.info('Pemberi Bantuan Hukum Tidak Ditemukan')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)
        assert False

@pytest.mark.webtest
def test_10_PertimbanganHukumFkr():
    sleep(driver)
    try:    
        driver.find_element(By.ID, "pertimbanganHukum").send_keys(PertimbanganHukumFkr)
        Log.info('Pertimbangan Hukum Ditemukan')
    except Exception as e:
        Log.info('Pertimbangan Hukum Tidak Ditemukan')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)
        assert False

@pytest.mark.webtest
def test_11_PutusanPengadilanFkr():
    sleep(driver)
    try:
        driver.find_element(By.ID, "putusanPengadilan").click()
        driver.find_element(By.ID, "putusanPengadilan").send_keys(PutusanPengadilanFkr)
        Log.info('Input Putusan Pengadilan')
    except Exception as e:  
        Log.info('Putusan Pengadilan Tidak Ditemukan')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)
        assert False

@pytest.mark.webtest
def test_12_KeteranganFkr():
    sleep(driver)
    try:
        driver.find_element(By.ID, "keterangan").click()
        driver.find_element(By.ID, "keterangan").send_keys(KeteranganFkr)
        Log.info('Input Keterangan')
    except Exception as e:
        Log.info('Keterangan Tidak Ditemukan')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)
        assert False

@pytest.mark.webtest
def test_13_ClickButtonSimpan():
    sleep(driver)
    try:
        driver.find_element(By.CSS_SELECTOR, "#submitButton > span").click()
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#searchButton > span")))
        Log.info('Click Button Simpan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)

    except Exception as e:
        Log.info('Button Simpan Tidak Ditemukan')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False




@pytest.mark.webtest
def test_14_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      

@pytest.mark.webtest
def test_15_LoginSPV():
    Log.info('Setup Os')
    SpvRutanBdg(driver)
    Log.info('Login Op Laporan Kamtib 8 A')


@pytest.mark.webtest
def test_16_AksesMenuSpv():
    sleep(driver)
    try:
        menulaporan8a(driver)
        Log.info('Akses halaman Laporan 8 A')
        attach(data=driver.get_screenshot_as_png())
    except Exception as e:
        Log.info('Akses halaman Laporan 8 A Gagal')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)
        assert False


@pytest.mark.webtest
def test_17_KirimLaporanSpv():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasiLaporan')))
        driver.find_element(By.ID, "buttonVerifikasiLaporan").click()
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'kirimLaporan')))
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#kirimLaporan > span')))
        WebDriverWait(driver,30).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, '.circular')))
        driver.find_element(By.CSS_SELECTOR, "#kirimLaporan > span").click()
        WebDriverWait(driver,10).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, '#kirimLaporan > span')))
        
        Log.info('Kirim laporan')
        quit(driver)
    except Exception as e:
        Log.info('Kirim laporan Gagal')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@pytest.mark.webtest
def test_18_loginKanwil():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    kanwiljabar(driver)
    Log.info('Login Spv Laporan Kamtib 8 A')

@pytest.mark.webtest
def test_19_AksesMenuKanwil():
    sleep(driver)
    try:
        menulaporan8a(driver)
        Log.info('Akses halaman Laporan 8 A')
        attach(data=driver.get_screenshot_as_png())
        
    except Exception as e:
        Log.info('Akses halaman Laporan 8 A Gagal')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)
        assert False

@pytest.mark.webtest
def test_20_VerifikasiLaporanKanwil():
    try:
        sleep(driver)
        time.sleep(2)
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterUpt2')))
        driver.find_element(By.ID, "formfilterUpt2").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterUpt2").send_keys("Rutan Kelas")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
        
        driver.find_element(By.XPATH, "//span[contains(.,\'Rutan Kelas I Bandung\')]").click()
        Log.info('Input UPT')

    except Exception as e:
        Log.info('Input UPT Gagal')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False      

@pytest.mark.webtest
def test_21_ClickButtonSearch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        driver.find_element(By.ID, "searchButton").click()
        Log.info('Click Button Search')
    except Exception as e:
        Log.info('Click Button Search Gagal')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@pytest.mark.webtest
def test_22_ClickButtonVerifikasi():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasi')))
        driver.find_element(By.ID, "buttonVerifikasi").click()
        Log.info('Click Button Verifikasi')
    except Exception as e:
        Log.info('Click Button Verifikasi Gagal')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False



@pytest.mark.webtest
def test_23_StatusVerifikasiModal():
    sleep(driver)
    try:
        WebDriverWait(driver,30).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, '.circular')))
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'simpanVerifikasi')))
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'statusVerifikasiModal')))
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.justify-end > .w-8')))
        time.sleep(5)
        driver.find_element(By.ID, "statusVerifikasiModal").click()
        Log.info('Click Verifikasi Modal')
    except Exception as e:
        Log.info('Click Verifikasi Modal Gagal')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@pytest.mark.webtest
def test_24_UbahStatus():
    sleep(driver)
    try:
        driver.find_element(By.ID, "diizinkan").click()
        Log.info('Ubah Status Verifikasi')
    except Exception as e:
        Log.info('Ubah Status Verifikasi Gagal')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@pytest.mark.webtest
def test_25_Inputketerangan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "keterangan").send_keys("keterangan")
        Log.info('Input Keterangan')
    except Exception as e:
        Log.info('Input Keterangan Gagal')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@pytest.mark.webtest
def test_26_SimpanVerifikasi():
    sleep(driver)
    try:
        driver.find_element(By.ID, "simpanVerifikasi").click()
        WebDriverWait(driver,30).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, '.circular')))
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'simpanVerifikasi')))
        Log.info('Click Button Simpan Verifikasi')
        quit(driver)
    except Exception as e:
        Log.info('Click Button Simpan Verifikasi Gagal')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False



@pytest.mark.webtest
def test_27_loginPusat():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    pusat(driver)
    Log.info('Login Spv Laporan Kamtib 8 A')

@pytest.mark.webtest
def test_28_AksesMenuPusat():
    sleep(driver)
    try:
        menulaporan8a(driver)
        Log.info('Akses halaman Laporan 8 A')
        attach(data=driver.get_screenshot_as_png())
    except Exception as e:
        Log.info('Akses halaman Laporan 8 A Gagal')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@pytest.mark.webtest
def test_29_PilihKanwil():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterKanwil')))
        driver.find_element(By.ID, "formfilterKanwil").send_keys("jawa barat")
        driver.find_element(By.XPATH, "//li[contains(.,\'Jawa Barat\')]").click()
        Log.info('Pilih Kanwil')
        attach(data=driver.get_screenshot_as_png())
    except Exception as e:
        Log.info('Pilih Kanwil Gagal')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@pytest.mark.webtest
def test_30_PilihUPT():
    sleep(driver)
    try:
        driver.find_element(By.ID, "formfilterUpt").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterUpt").send_keys("rutan kelas")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
        driver.find_element(By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]").click()
        Log.info('PilihUPT')
        attach(data=driver.get_screenshot_as_png())
    except Exception as e:
        Log.info('PilihUPT Gagal')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@pytest.mark.webtest
def test_31_ClickButtonSearch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        driver.find_element(By.CSS_SELECTOR, "#searchButton > span").click()
        Log.info('Click Button Search')
        attach(data=driver.get_screenshot_as_png())
    except Exception as e:
        Log.info('Click Button Search Gagal')
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@pytest.mark.webtest
def test_32_exit():
    quit(driver)
    Log.info('Exit')







    
    