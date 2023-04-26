from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("data"))
    file_path = environ.get("fakermac")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
import random

from Settings.setupkeamanan import initDriver, loadDataPath, quit, sleep, upload
from Settings.loginkeamanan import oplapkamtibwaru,kanwiljabar,pusat, SpvRutanBdg
from Settings.Page.keamanan import menulaporan7d
import random
import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Loglaporan7D.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)


sheetrangeIndex = wb['listWbpRutan1Bandung']

i = random.randint(0,50)

NamaInput                                 = sheetrangeIndex['A'+str(i)].value


workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'LaporanFaker'

fake = Faker('id_ID')
Wbp = ['NASRULLAH SIRAIt BIN doniserigar','PRAYITNA SALAHUDIn BINTI waloyo','NAJIB NAINGGOLAn BIN sugiono','PRAYOGA KURNIAWAn BINTI sugiono','PUSPA SUWARNo BIN zulahmad','KAREN ZULKARNAIn BINTI doniserigar','BAHUWIRYA UTAMa BIN waloyo','H CENGKIR HARDIANSYAh BINTI sugiono','DALIONO PRADIPTa BIN sugiono','HUMAIRA NASYIAH S.KEd BIN zulahmad','R. VICKY USAMAH S.E.I BINTI doniserigar','IR. AISYAH HASANAh BIN waloyo','QUEEN PURNAWATi BINTI sugiono','DR. JABAL SUSANTi BIN sugiono','VANESA RIYANTI S.Gz BINTI zulahmad']
Gangguan = ['Berat','Sedang','Ringan']
JenisGangguanSedang = ['jenisGangguanOption-0','jenisGangguanOption-1','jenisGangguanOption-2','jenisGangguanOption-3', 'jenisGangguanOption-4' ]
JenisGangguanBerat = ['jenisGangguanOption-0','jenisGangguanOption-1','jenisGangguanOption-2','jenisGangguanOption-3', 'jenisGangguanOption-4', 'jenisGangguanOption-5','jenisGangguanOption-6' ]
JenisGangguanRingan = ['jenisGangguanOption-0','jenisGangguanOption-1','jenisGangguanOption-2','jenisGangguanOption-3', 'jenisGangguanOption-4', 'jenisGangguanOption-5','jenisGangguanOption-6' ]
followup = ['0','1']
korban = ['1','2','3','4', '5', '6' ]
JenisGangguanBerat = ['Tidak Mengikuti Program Pembinaan Yang Telah Ditetapkan','Mengancam, Melawan, Atau Melakukan Penyerangan Terhadap Petugas','Membuat Atau Menyimpan Senjata API, Senjata Tajam, Atau Sejenisnya','Merusak Fasilitas Lapas Atau Rutan','Mengancam, Memprovokasi, Atau Perbuatan Lain Yang Menimbulkan Gangguan Keamanan Dan Ketertiban','Memiliki, Membawa, Atau Menggunakan Alat Komunikasi Atau Alat Elektronik','Membuat, Membawa, Menyimpan, Mengedarkan Atau Mengkonsumsi Minuman Yang Mengandung Alkohol','Membuat, Membawa, Menyimpan, Mengedarkan, Atau Mengkonsumsi Narkotika Dan Obat Terlarang Serta Zat Adiktif Lainnya','Melakukan Upaya Melarikan Diri Atau Membantu Narapidana Atau Tahanan Lain Untuk Melarikan Diri']

for i in range(50):
    WbpFaker                    = random.randint(0,50)
    GangguanFaker               = random.choice(Gangguan)
    JenisGangguanSedangFaker    = random.choice(JenisGangguanSedang)
    JenisGangguanBeratFaker     = random.choice(JenisGangguanBerat)
    JenisGangguanRinganFaker    = random.choice(JenisGangguanRingan)
    waktukejadianFaker          = datetime.now().strftime('%d/%m/%Y')
    KronologiFaker              = fake.paragraph(nb_sentences=2)
    TglSuratPtsFaker            = datetime.now().strftime('%d/%m/%Y')
    NamaPimpinanFaker           = fake.name()
    PengawasFaker               = fake.name()
    NoSuratPutusanFaker         = fake.numerify('SRT/'+'#################')
    NipPimpinanFaker            = fake.year()+"1"f"{i+1:05}"
    NipPengawasFaker            = fake.year()+"1"f"{i+1:05}"
    tglMulaiHukumanFaker        = datetime.now().strftime('%d/%m/%Y')
    tglAkhirHukumanFaker        = datetime.now().strftime('%d/%m/%Y')
    KeteranganFaker             = fake.paragraph(nb_sentences=2)
    NoSkFaker                   = fake.numerify('SK/'+'#################')
    NoBAPFaker                  = fake.numerify('BAP/'+'#################')
    followupFaker               = random.choice(followup)




    worksheet.append([
        WbpFaker, 
        GangguanFaker, 
        JenisGangguanSedangFaker,
        JenisGangguanBeratFaker,
        JenisGangguanRinganFaker, 
        waktukejadianFaker, 
        KronologiFaker,
        TglSuratPtsFaker, 
        NamaPimpinanFaker, 
        PengawasFaker,
        NoSuratPutusanFaker,
        NipPimpinanFaker,
        NipPengawasFaker,
        tglMulaiHukumanFaker,
        tglAkhirHukumanFaker,
        KeteranganFaker,
        NoSkFaker,
        NoBAPFaker,
        followupFaker
         
        ])
workbook.save(file_path)

workbook = load_workbook(filename=file_path)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=2, values_only=True):

    WbpXC                                   = row[0]
    GangguanXC                              = row[1]
    JenisGangguanSedangXC                   = row[2]
    JenisGangguanBeratXC                    = row[3]
    JenisGangguanRinganXC                   = row[4]
    waktukejadianXC                         = row[5]
    KronologiXC                             = row[6]
    TglSuratPtsXC                           = row[7]
    NamaPimpinanXC                          = row[8]
    PengawasXC                              = row[9]
    NoSuratPutusanXC                        = row[10]
    NipPimpinanXC                           = row[11]
    NipPengawasXC                           = row[12]
    tglMulaiHukumanXC                       = row[13]
    tglAkhirHukumanXC                       = row[14]
    KeteranganXC                            = row[15]
    NoSkXC                                  = row[16]
    NoBAPXC                                 = row[17]
    followupXC                              = row[18]


@mark.fixture_test()
def test_1_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      

@mark.fixture_test()
def test_2_Login():
    Log.info('Setup Os')
    oplapkamtibwaru(driver)
    Log.info('Login Op Laporan Kamtib 7 D')

@mark.fixture_test()
def test_2_AksesMenu():
    sleep(driver)
    try:
        menulaporan7d(driver)
        Log.info('Akses halaman Laporan 7 D')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Akses halaman Laporan 7 D')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        

@mark.fixture_test()
def test_3_CreateButton():
    sleep(driver)
    
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'createButton')))
    driver.find_element(By.ID, "createButton").click()
    Log.info('Klik tombol tambah data')
    attach(data=driver.get_screenshot_as_png())
    
@mark.fixture_test()
def test_3_SearchWBP():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, "backButton")))
        driver.find_element(By.ID, "kataKunci").click()
        driver.find_element(By.ID, "kataKunci").send_keys(NamaInput)
        print(NamaInput)
        driver.find_element(By.CSS_SELECTOR, "#searchButton > span").click()
        
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, "searchButton")))
        Log.info('Cari WBP')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except: 
        Log.info('Gagal Cari WBP')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        

@mark.fixture_test()
def test_4_PilihJenisGangguan():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, "searchButton")))
        driver.find_element(By.CSS_SELECTOR, ".cell svg").click()
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'"+ GangguanXC +"')]")))
        driver.find_element(By.XPATH, "//li[contains(.,\'"+ GangguanXC +"')]").click()
        Log.info('Pilih Jenis Gangguan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Pilih Jenis Gangguan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        

@mark.fixture_test()
def test_5_ClickButtonDaftarkan():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#daftarkan-0 > span')))
        driver.find_element(By.CSS_SELECTOR, "#daftarkan-0 > span").click()
        Log.info('Klik tombol daftarkan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Klik tombol daftarkan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        

@mark.fixture_test()
def test_6_InputData():
    Log.info('Iput Data Bersadarkan Jenis Gangguan')
    sleep(driver)
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//td[contains(.,'Status Verifikasi')]")))
    if driver.find_elements(By.XPATH, "//h1[contains(.,\'Catat Pelanggaran Sedang\')]"):
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//td[contains(.,'Status Verifikasi')]")))
        driver.find_element(By.ID, "jenisGangguan").click()
        time.sleep(2)
        driver.find_element(By.XPATH, "//li[@id=\'"+JenisGangguanSedangXC+"\']").click()

        sleep(driver)
        driver.find_element(By.ID, "tglTerjadi").click()
        driver.find_element(By.ID, "tglTerjadi").send_keys(waktukejadianXC)

        driver.find_element(By.ID, "kronologiSingkat").click()
        driver.find_element(By.ID, "kronologiSingkat").send_keys(KronologiXC)

        driver.find_element(By.ID, "tglSuratPutusan").click()
        driver.find_element(By.ID, "tglSuratPutusan").send_keys(TglSuratPtsXC)

        driver.find_element(By.ID, "namaPimpinan").click()
        driver.find_element(By.ID, "namaPimpinan").send_keys(NamaPimpinanXC)

        driver.find_element(By.ID, "namaPengawas").click()
        driver.find_element(By.ID, "namaPengawas").send_keys(PengawasXC)

        driver.find_element(By.ID, "noSuratPutusan").click()
        driver.find_element(By.ID, "noSuratPutusan").send_keys(NoSuratPutusanXC)

        driver.find_element(By.ID, "nipPimpinan").click()
        driver.find_element(By.ID, "nipPimpinan").send_keys(NipPimpinanXC)

        driver.find_element(By.ID, "nipPengawas").click()
        driver.find_element(By.ID, "nipPengawas").send_keys(NipPengawasXC)

        driver.find_element(By.ID, "tglMulaiHukuman").click()
        driver.find_element(By.ID, "tglMulaiHukuman").send_keys(tglMulaiHukumanXC)

        driver.find_element(By.ID, "tglAkhirHukuman").click()
        driver.find_element(By.ID, "tglAkhirHukuman").send_keys(tglMulaiHukumanXC)

        driver.find_element(By.ID, "keterangan").click()
        driver.find_element(By.ID, "keterangan").send_keys(KeteranganXC)

        driver.find_element(By.CSS_SELECTOR, "#chooseFile > span").click()
        upload(driver)
        

        driver.find_element(By.ID, "noSk").click()
        driver.find_element(By.ID, "noSk").send_keys(NoSkXC)
        
        # driver.find_element(By.ID, "generateSk").click()
        # driver.find_element(By.CSS_SELECTOR, ".el-form-item:nth-child(1) #noSk").click()
        # driver.find_element(By.CSS_SELECTOR, ".el-form-item:nth-child(1) #noSk").send_keys(NoSkXC)
        # driver.find_element(By.ID, "noBap").send_keys(NoBAPXC)
        # driver.find_element(By.ID, "submitGenerateSk").click()
        # WebDriverWait(driver,10).until(EC.invisibility_of_element((By.ID, 'submitGenerateSk')))

        Log.info('Input Jenis Gangguan sedang')
        print("Catat Pelanggaran Sedang")
        attach(data=driver.get_screenshot_as_png())

    elif driver.find_elements(By.XPATH, "//h1[contains(.,\'Catat Pelanggaran Ringan\')]"):
        driver.find_element(By.ID, "jenisGangguan").click()
        driver.find_element(By.XPATH, "//li[@id=\'"+JenisGangguanRinganXC+"\']").click()

        driver.find_element(By.ID, "tglTerjadi").click()
        driver.find_element(By.ID, "tglTerjadi").send_keys(waktukejadianXC)

        driver.find_element(By.ID, "kronologi").click()
        driver.find_element(By.ID, "kronologi").send_keys(KronologiXC)

        driver.find_element(By.ID, "motif").click()
        driver.find_element(By.ID, "motif").send_keys(KronologiXC)

        driver.find_element(By.ID, "sanksi").click()
        driver.find_element(By.ID, "sanksi").send_keys(PengawasXC)

        driver.find_element(By.ID, "tglPenyelesaian").click()
        driver.find_element(By.ID, "tglPenyelesaian").send_keys(tglMulaiHukumanXC)

        driver.find_element(By.ID, "penyelesaian").click()
        driver.find_element(By.ID, "penyelesaian").send_keys(KeteranganXC)

        driver.find_element(By.ID, "uraianTambahan").click()
        driver.find_element(By.ID, "uraianTambahan").send_keys(KeteranganXC)

        Log.info('Input Jenis Gangguan Ringan')
        print("Catat Pelanggaran Ringan")
        attach(data=driver.get_screenshot_as_png())

    elif driver.find_elements(By.XPATH, "//h1[contains(.,\'Catat Pelanggaran Berat\')]"):
        sleep(driver)
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, "tanggalTerjadi")))
        print("Catat Pelanggaran Berat")
        driver.find_element(By.ID, "jenisPelanggaran").click()
        time.sleep(1)
        sleep(driver)
        driver.find_element(By.XPATH, "//li[contains(.,\'"+JenisGangguanBeratXC+"\')]").click()
        sleep(driver)
        driver.find_element(By.ID, "tanggalTerjadi").click()
        driver.find_element(By.ID, "tanggalTerjadi").send_keys(waktukejadianXC)

        driver.find_element(By.ID, "kronologisSingkat").click()
        driver.find_element(By.ID, "kronologisSingkat").send_keys(KronologiXC)

        driver.find_element(By.ID, "tanggalSuratPutusan").click()
        driver.find_element(By.ID, "tanggalSuratPutusan").send_keys(TglSuratPtsXC)

        driver.find_element(By.ID, "namaPimpinan").click()
        driver.find_element(By.ID, "namaPimpinan").send_keys(NamaPimpinanXC)

        driver.find_element(By.ID, "namaPengawas").click()
        driver.find_element(By.ID, "namaPengawas").send_keys(PengawasXC)

        driver.find_element(By.ID, "noSuratPutusan").click()
        driver.find_element(By.ID, "noSuratPutusan").send_keys(NoSuratPutusanXC)

        driver.find_element(By.ID, "nipPimpinan").click()
        driver.find_element(By.ID, "nipPimpinan").send_keys(NipPimpinanXC)

        driver.find_element(By.ID, "nipPengawas").click()
        driver.find_element(By.ID, "nipPengawas").send_keys(NipPengawasXC)

        driver.find_element(By.CSS_SELECTOR, "#chooseFile > span").click()
        upload(driver)

        if followup == 0:
            driver.find_element(By.CLASS_NAME, '.el-switch__core').click()
        else:
            pass    

        driver.find_element(By.ID, "tglMulaiHukuman-0").click()
        driver.find_element(By.ID, "tglMulaiHukuman-0").send_keys(tglMulaiHukumanXC)
        driver.find_element(By.ID, "tglAkhirHukuman-0").click()
        driver.find_element(By.ID, "tglAkhirHukuman-0").send_keys(tglMulaiHukumanXC)

        driver.find_element(By.ID, "tglMulaiHukuman-1").click()
        driver.find_element(By.ID, "tglMulaiHukuman-1").send_keys(tglMulaiHukumanXC)
        driver.find_element(By.ID, "tglAkhirHukuman-1").click()
        driver.find_element(By.ID, "tglAkhirHukuman-1").send_keys(tglMulaiHukumanXC)

        driver.find_element(By.ID, "keterangan").click()
        driver.find_element(By.ID, "keterangan").send_keys(KeteranganXC)
        
        driver.find_element(By.ID, "keterangan").click()
        driver.find_element(By.XPATH, "//tr[2]/td[5]/div/div/div/div/input").send_keys(KeteranganXC)


        driver.find_element(By.ID, "noSk").click()
        driver.find_element(By.ID, "noSk").send_keys(NoSkXC)
        
        # driver.find_element(By.ID, "generateSk").click()
        # driver.find_element(By.ID, "noSurat").click()
        # driver.find_element(By.ID, "noSurat").send_keys(NoSkXC)
        # driver.find_element(By.ID, "noBap").send_keys(NoBAPXC)
        # driver.find_element(By.ID, "submitGenerateSk").click()
  
        # WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'submitGenerateSk')))
        Log.info('Input Jenis Gangguan Berat')
        print("Catat Pelanggaran Berat")

@mark.fixture_test()
def test_7_submit():
    driver.find_element(By.ID, "submitButton").click()
    sleep(driver)
    print("Submit")
    attach(data=driver.get_screenshot_as_png())

@mark.fixture_test()
def test_8_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      

@mark.fixture_test()
def test_9_LoginSPV():
    Log.info('Setup Os')
    SpvRutanBdg(driver)
    Log.info('Login Op Laporan Kamtib 7 B')


@mark.fixture_test()
def test_10_AksesMenuSpv():
    sleep(driver)
    try:
        menulaporan7d(driver)
        Log.info('Akses halaman Laporan 7 B')
        attach(data=driver.get_screenshot_as_png())
    except:
        Log.info('Gagal Akses halaman Laporan 7 B')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_11_KirimLaporanSpv():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasiLaporan')))
        time.sleep(2)
        driver.find_element(By.ID, "buttonVerifikasiLaporan").click()
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'kirimLaporan')))
        time.sleep(3)
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#kirimLaporan > span')))
        driver.find_element(By.CSS_SELECTOR, "#kirimLaporan > span").click()
        Log.info('Kirim laporan')
    except:
        Log.info('Gagal Kirim laporan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_12_loginPusat():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    pusat(driver)
    Log.info('Login Spv Laporan Kamtib 6D')

@mark.fixture_test()
def test_13_AksesMenuPusat():
    sleep(driver)
    try:
        menulaporan7d(driver)
        Log.info('Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_14_PilihKanwil():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterKanwil')))
        driver.find_element(By.ID, "formfilterKanwil").send_keys("jawa barat")
        driver.find_element(By.XPATH, "//li[contains(.,\'Jawa Barat\')]").click()
        Log.info('Pilih Kanwil')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Pilih Kanwil')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_15_PilihUPT():
    sleep(driver)
    try:
        driver.find_element(By.ID, "formfilterUpt").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterUpt").send_keys("rutan kelas")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
        driver.find_element(By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]").click()
        Log.info('PilihUPT')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal PilihUPT')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_16_ClickButtonSearch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        driver.find_element(By.CSS_SELECTOR, "#searchButton > span").click()
        Log.info('Click Button Search')
        attach(data=driver.get_screenshot_as_png())
    except:
        Log.info('Gagal Click Button Search')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_17_ClickButtonVerifikasi():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasi')))
        driver.find_element(By.ID, "buttonVerifikasi").click()
        Log.info('Click Button Verifikasi')
        attach(data=driver.get_screenshot_as_png())
    except:
        Log.info('Gagal Click Button Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_18_StatusVerifikasiModal():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'simpanVerifikasi')))
        time.sleep(3)
        driver.find_element(By.ID, "statusVerifikasiModal").click()
        Log.info('Click Verifikasi Modal')
        assert True
    except:
        Log.info('Gagal Click Verifikasi Modal')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_19_UbahStatus():
    sleep(driver)
    try:
        driver.find_element(By.ID, "diizinkan").click()
        Log.info('Ubah Status Verifikasi')
        assert True
    except:
        Log.info('Gagal Ubah Status Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_20_Inputketerangan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "keterangan").send_keys("keterangan")
        Log.info('Input Keterangan')
        assert True
    except:
        Log.info('Gagal Input Keterangan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_21_SimpanVerifikasi():
    sleep(driver)
    try:
        driver.find_element(By.ID, "simpanVerifikasi").click()
        Log.info('Click Button Simpan Verifikasi')
        assert True
    except:
        Log.info('Gagal Click Button Simpan Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_22_exit():
    quit(driver)
    Log.info('Exit')







    
    