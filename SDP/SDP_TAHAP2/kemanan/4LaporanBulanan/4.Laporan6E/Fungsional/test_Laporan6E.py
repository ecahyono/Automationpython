from distutils.archive_util import make_archive
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUAT"))

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))


from Settings.setupkeamanan import initDriver, loadDataPath, quit, sleep
from Settings.loginkeamanan import oplapkamtibwaru,kanwiljabar,pusat, SpvRutanBdg
from Settings.Page.keamanan import menulaporan6e

import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Loglaporan6e.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)


i = 2
laporan6e = wb['Laporan6E']


tipeSenjata                                     = laporan6e['A'+str(i)].value
jenisSenjata                                    = laporan6e['B'+str(i)].value
nomorSenjata                                    = laporan6e['C'+str(i)].value
masaBerlaku                                     = laporan6e['D'+str(i)].value
kepemilikan                                     = laporan6e['E'+str(i)].value
kondisiBaik                                     = laporan6e['F'+str(i)].value
kondisiRusak                                    = laporan6e['G'+str(i)].value
jumlahPeluru                                    = laporan6e['H'+str(i)].value
deskripsi                                       = laporan6e['I'+str(i)].value

tipeSenjata1                                     = laporan6e['k'+str(i)].value
jenisSenjata1                                    = laporan6e['L'+str(i)].value
jenispeluru1                                     = laporan6e['M'+str(i)].value
kepemilikan1                                     = laporan6e['N'+str(i)].value
deskripsi1                                       = laporan6e['O'+str(i)].value


@mark.fixture_test()
def test_1_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      

@mark.fixture_test()
def test_2_Login():
    Log.info('Setup Os')
    oplapkamtibwaru(driver)
    Log.info('Login Op Laporan Kamtib 6 E')

@mark.fixture_test()
def test_2_AksesMenu():
    sleep(driver)
    try:
        menulaporan6e(driver)
        Log.info('Akses halaman Laporan 6 E')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Akses halaman Laporan 6 E')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_3_CreateButton():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'createButton')))
        driver.find_element(By.ID, "createButton").click()
        Log.info('Klik tombol tambah data')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Klik tombol tambah data')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False
    

@mark.fixture_test()
def test_4_TipeSenjata():
    sleep(driver)
    try:
        driver.find_element(By.ID, "tipeSenjata").click()
        time.sleep(1)
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'submitButton')))
        driver.find_element(By.ID, "tipeSenjata").send_keys(tipeSenjata)
        time.sleep(1)
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'tipeSenjata0')))
        driver.find_element(By.ID, "tipeSenjata0").click()
        Log.info('Pilih tipe senjata')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Pilih tipe senjata')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_5_PilihJenisSenjata():
    sleep(driver)
    try:
        driver.find_element(By.ID, "jenisSenjata").click()
        time.sleep(1)
        driver.find_element(By.ID, "jenisSenjata").send_keys(jenisSenjata)
        time.sleep(1)
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'jenisSenjata1')))
        driver.find_element(By.ID, "jenisSenjata1").click()
        Log.info('Pilih jenis senjata')
        assert True
    except:
        Log.info('Gagal Pilih jenis senjata')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_6_InputNomorSenjata():
    sleep(driver)
    try:
        driver.find_element(By.ID, "nomorSenjata").click()
        driver.find_element(By.ID, "nomorSenjata").send_keys(nomorSenjata)
        Log.info('Input nomor senjata')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input nomor senjata')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_7_MasaBerlaku():
    sleep(driver)
    try:
        driver.find_element(By.ID, "masaBerlaku").click()
        time.sleep(0.5)
        driver.find_element(By.ID, "masaBerlaku").send_keys(masaBerlaku)
        Log.info('Input masa berlaku')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input masa berlaku')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_8_kepemilikan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "kepemilikan").click()
        time.sleep(0.5)
        driver.find_element(By.ID, "kepemilikan0").click()
        Log.info('Pilih kepemilikan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Pilih kepemilikan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_9_kondisibaik():
    sleep(driver)
    try:
        driver.find_element(By.ID, "kondisiBaik").click()
        time.sleep(0.5)
        driver.find_element(By.ID, "kondisiBaik").send_keys(kondisiBaik)
        Log.info('Input kondisi baik')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input kondisi baik')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_10_kondisirusak():
    sleep(driver)
    try:
        driver.find_element(By.ID, "kondisiRusak").click()
        time.sleep(0.5)
        driver.find_element(By.ID, "kondisiRusak").send_keys(kondisiRusak)
        Log.info('Input kondisi rusak')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input kondisi rusak')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_11_jumlahpeluru():
    sleep(driver)
    try:
        driver.find_element(By.ID, "jumlahPeluru").click()
        time.sleep(0.5)
        driver.find_element(By.ID, "jumlahPeluru").send_keys(jumlahPeluru)
        Log.info('Input jumlah peluru')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input jumlah peluru')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_12_deskripsi():
    sleep(driver)
    try:
        driver.find_element(By.ID, "deskripsi").click()
        driver.find_element(By.ID, "deskripsi").send_keys(deskripsi)
        Log.info('Input deskripsi')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input deskripsi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_13_ClickButtonSubmit():
    sleep(driver)
    try:
        driver.find_element(By.CSS_SELECTOR, "#submitButton > span").click()
        Log.info('Klik tombol simpan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Klik tombol simpan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_11_SetupOS():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')


@mark.fixture_test()
def test_12_LoginSpv():
    sleep(driver)
    try:
        SpvRutanBdg(driver)
        Log.info('Login Spv Laporan Kamtib 6D')
    except:
        Log.info('Gagal Login Spv Laporan Kamtib 6D')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_13_AksesMenuSpv():
    sleep(driver)
    try:
        menulaporan6e(driver)
        Log.info('Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
    except:
        Log.info('Gagal Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_14_KirimLaporanSpv():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasiLaporan')))
        time.sleep(2)
        driver.find_element(By.ID, "buttonVerifikasiLaporan").click()
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'kirimLaporan')))
        time.sleep(3)
        driver.find_element(By.CSS_SELECTOR, "#kirimLaporan > span").click()
        Log.info('Kirim laporan')
    except:
        Log.info('Gagal Kirim laporan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_15_loginKanwil():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    kanwiljabar(driver)
    Log.info('Login Spv Laporan Kamtib 6D')

@mark.fixture_test()
def test_16_AksesMenuSpv():
    sleep(driver)
    try:
        menulaporan6e(driver)
        Log.info('Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_17_VerifikasiLaporanKanwil():
    try:
        sleep(driver)
        time.sleep(2)
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterUpt2')))
        driver.find_element(By.ID, "formfilterUpt2").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterUpt2").send_keys("Rutan Kelas")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
        
        driver.find_element(By.XPATH, "//span[contains(.,\'Rutan Kelas I Bandung\')]").click()
        Log.info('Input UPT')
        assert True
    except:
        Log.info('Gagal Input UPT')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_18_ClickButtonSearch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        driver.find_element(By.ID, "searchButton").click()
        Log.info('Click Button Search')
        assert True
    except:
        Log.info('Gagal Click Button Search')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_19_ClickButtonVerifikasi():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasi')))
        driver.find_element(By.ID, "buttonVerifikasi").click()
        Log.info('Click Button Verifikasi')
        assert True 
    except:
        Log.info('Gagal Click Button Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_20_StatusVerifikasiModal():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'simpanVerifikasi')))
        time.sleep(3)
        driver.find_element(By.ID, "statusVerifikasiModal").click()
        Log.info('Click Verifikasi Modal')
        assert True
    except:
        Log.info('Gagal Click Verifikasi Modal')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_21_UbahStatus():
    sleep(driver)
    try:
        driver.find_element(By.ID, "diizinkan").click()
        Log.info('Ubah Status Verifikasi')
        assert True
    except:
        Log.info('Gagal Ubah Status Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_22_Inputketerangan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "keterangan").send_keys("keterangan")
        Log.info('Input Keterangan')
        assert True
    except:
        Log.info('Gagal Input Keterangan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_23_SimpanVerifikasi():
    sleep(driver)
    try:
        driver.find_element(By.ID, "simpanVerifikasi").click()
        Log.info('Click Button Simpan Verifikasi')
        assert True
    except:
        Log.info('Gagal Click Button Simpan Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_24_loginPusat():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    pusat(driver)
    Log.info('Login Spv Laporan Kamtib 6D')

@mark.fixture_test()
def test_25_AksesMenuPusat():
    sleep(driver)
    try:
        menulaporan6e(driver)
        Log.info('Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_26_PilihKanwil():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterKanwil')))
        driver.find_element(By.ID, "formfilterKanwil").send_keys("jawa barat")
        driver.find_element(By.XPATH, "//li[contains(.,\'Jawa Barat\')]").click()
        Log.info('Pilih Kanwil')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Pilih Kanwil')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_27_PilihUPT():
    sleep(driver)
    try:
        driver.find_element(By.ID, "formfilterUpt").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterUpt").send_keys("rutan kelas")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
        driver.find_element(By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]").click()
        Log.info('PilihUPT')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal PilihUPT')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_28_ClickButtonSearch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        driver.find_element(By.CSS_SELECTOR, "#searchButton > span").click()
        Log.info('Click Button Search')
        attach(data=driver.get_screenshot_as_png())
    except:
        Log.info('Gagal Click Button Search')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_29_ClickButtonMaster():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonMaster')))
        time.sleep(3)
        driver.find_element(By.ID, "buttonMaster").click()
        Log.info('Click Button Master')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Click Button Master')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_30_CreateButton():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'backButton')))
        driver.find_element(By.ID, "createButton").click()
        Log.info('Click Button Create')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Click Button Create')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_31_TipeSenjata():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'submitButton')))
        driver.find_element(By.ID, "tipeSenjata").click()
        driver.find_element(By.XPATH, "//li[contains(.,\'"+ tipeSenjata1 +"')]").click()
        Log.info('Input Tipe Senjata')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Tipe Senjata')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_32_JenisSenjata():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'submitButton')))
        driver.find_element(By.ID, "jenisSenjata").click()
        driver.find_element(By.ID, "jenisSenjata").send_keys(jenisSenjata1)
        Log.info('Input Jenis Senjata')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Jenis Senjata')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_33_JenisPeluru():
    sleep(driver)
    try:
        driver.find_element(By.ID, "jenisPeluru").click()
        driver.find_element(By.XPATH, "//li[contains(.,\'"+ jenispeluru1 +"')]").click()
        Log.info('Input Jenis Peluru')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Jenis Peluru')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_34_Kepemilikan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "kepemilikan").click()
        driver.find_element(By.XPATH, "//li[contains(.,\'"+ kepemilikan1 +"')]").click()
        Log.info('Input Kepemilikan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Kepemilikan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_35_Deskripsi():
    sleep(driver)
    try:
        driver.find_element(By.ID, "deskripsi").click()
        driver.find_element(By.ID, "deskripsi").send_keys(deskripsi1)
        Log.info('Input Deskripsi')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Deskripsi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_36_CreateButton():
    sleep(driver)
    try:
        driver.find_element(By.CSS_SELECTOR, "#submitButton > span").click()
        Log.info('Click Button Create')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Click Button Create')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

        
@mark.fixture_test()
def test_37_exit():
    quit(driver)
    Log.info('Exit')






    
    