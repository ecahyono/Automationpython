from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("data"))
    file_path = environ.get("fakermac")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))


from Settings.setupkeamanan import initDriver, loadDataPath, quit, sleep
from Settings.loginkeamanan import oplapkamtibwaru,kanwiljabar,pusat, SpvRutanBdg
from Settings.Page.keamanan import menulaporan7c
import random
import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Loglaporan7C.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

sheetrangeIndex = wb['ListPetugasRutan1Bandung']

i = random.randint(0,50)

IdentitasPelakuData                                 = sheetrangeIndex['A'+str(i)].value

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'LaporanFaker'

fake = Faker('id_ID')
JenisGangguan = ['Kerusuhan Kunjungan','Kerusuhan Petugas dengan WBP ','Kerusuhan WBP dengan WBP ','Kerusuhan Kunjungan']
identitaspelaku = ['Drs. Tina Uyainah','Paris Mandasari, S.Ked','Okto Irawan','dr. Nilam Siregar, M.Pd','Kayla Prastuti']
tingkatgangguan = ['1','2','3','4', '5', '6' ]
korban = ['1','2','3','4', '5', '6' ]
TindakLanjutBarang = ['Evaluasi','Pengawasan','Disiplin']
for i in range(50):
    JenisGangguanFaker          = random.choice(JenisGangguan)
    waktukejadianFaker          = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    tingkatgangguanFaker        = random.choice(tingkatgangguan)
    korbanptgsFaker             = random.choice(korban)
    korbanwbpFaker              = random.choice(korban)
    korbanMasyarakatFaker       = random.choice(korban)
    KerugianFaker               = fake.pyint(min_value=1000, max_value=1000000)
    TindaLanjutFaker            = random.choice(TindakLanjutBarang)
    keteranganFaker             = fake.paragraph(nb_sentences=2)
    

    worksheet.append([
        JenisGangguanFaker, 
        waktukejadianFaker, 
        tingkatgangguanFaker, 
        korbanptgsFaker,
        korbanwbpFaker, 
        korbanMasyarakatFaker, 
        KerugianFaker,
        TindaLanjutFaker,
        keteranganFaker
        ])
workbook.save(file_path)

workbook = load_workbook(filename=file_path)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=2, values_only=True):
    jenisgangguanfkr                        = row[0]
    waktukejadianfkr                        = row[1]
    tingkatgangguanfkr                      = row[2]
    korbanptgsfkr                           = row[3]
    korbanwbpfkr                            = row[4]
    korbanmasyarakanfkr                     = row[5]
    KerugianFkr                             = row[6]
    TindakLanjutFkr                         = row[7]
    keteranganFkr                           = row[8]



   


@mark.fixture_test()
def test_1_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      

@mark.fixture_test()
def test_2_Login():
    Log.info('Setup Os')
    oplapkamtibwaru(driver)
    Log.info('Login Op Laporan Kamtib 7 C')

@mark.fixture_test()
def test_2_AksesMenu():
    sleep(driver)
    try:
        menulaporan7c(driver)
        Log.info('Akses halaman Laporan 7 C')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Akses halaman Laporan 7 C')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_3_CreateButton():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'createButton')))
        driver.find_element(By.ID, "createButton").click()
        time.sleep(3)
        Log.info('Klik tombol tambah data')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Klik tombol tambah data')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False
    
@mark.fixture_test()
def test_4_JenisGanggguan():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'submitButton')))
        driver.find_element(By.ID, "jenisGangguan").send_keys(jenisgangguanfkr)
        Log.info('Pilih Jenis Gangguan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Pilih Jenis Gangguan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)

@mark.fixture_test()
def test_5_IdentitasPelaku():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'identitasPelaku')))
        driver.find_element(By.ID, "identitasPelaku").send_keys(IdentitasPelakuData)
        driver.find_element(By.XPATH, "//li[contains(.,\'"+ IdentitasPelakuData +"')]").click()
        Log.info('Pilih Identitas Pelaku')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Pilih Identitas Pelaku')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)

@mark.fixture_test()
def test_6_WaktuKejadian():
    sleep(driver)
    try:
        driver.find_element(By.ID, "waktuKejadian").click()
        driver.find_element(By.ID, "waktuKejadian").send_keys(waktukejadianfkr)
        driver.find_element(By.CSS_SELECTOR, ".el-form:nth-child(3)").click()
        Log.info('Pilih Waktu Kejadian')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Pilih Waktu Kejadian')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)

@mark.fixture_test()
def test_7_TingkatGangguan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "tingkatGangguan").click()
        driver.find_element(By.ID, "tingkatGangguan").send_keys(tingkatgangguanfkr)
        Log.info('Pilih Tingkat Gangguan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Pilih Tingkat Gangguan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)

@mark.fixture_test()    
def test_8_Korban():
    sleep(driver)
    try:
        driver.find_element(By.ID, "jumlahPetugas").send_keys(korbanptgsfkr)
        driver.find_element(By.ID, "jumlahWbp").send_keys(korbanwbpfkr)
        driver.find_element(By.ID, "jumlahMasyarakat").send_keys(korbanmasyarakanfkr)
        Log.info('Input Korban Petugas,WBP,Masyarakat')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Pilih Korban')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
    
@mark.fixture_test()
def test_9_Kerugian():
    sleep(driver)
    try:
        driver.find_element(By.ID, "kerugianMateril").send_keys(KerugianFkr)
        Log.info('Input Kerugian')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Kerugian')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)

@mark.fixture_test()
def test_10_TindakLanjut():
    sleep(driver)
    try:
        driver.find_element(By.ID, "tindakLanjut").send_keys(TindakLanjutFkr)
        Log.info('Input Tindak Lanjut')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Tindak Lanjut')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)

@mark.fixture_test()
def test_11_Keterangan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "keterangan").send_keys(keteranganFkr)
        Log.info('Input Keterangan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Keterangan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)

@mark.fixture_test()
def test_12_SubmitButton():
    sleep(driver)
    try:
        driver.find_element(By.ID, "submitButton").click()
        Log.info('Klik tombol submit')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Klik tombol submit')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)


@mark.fixture_test()
def test_13_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      

@mark.fixture_test()
def test_14_LoginSPV():
    Log.info('Setup Os')
    SpvRutanBdg(driver)
    Log.info('Login Op Laporan Kamtib 7 B')


@mark.fixture_test()
def test_15_AksesMenuSpv():
    sleep(driver)
    try:
        menulaporan7c(driver)
        Log.info('Akses halaman Laporan 7 B')
        attach(data=driver.get_screenshot_as_png())
    except:
        Log.info('Gagal Akses halaman Laporan 7 B')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_16_KirimLaporanSpv():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasiLaporan')))
        time.sleep(2)
        driver.find_element(By.ID, "buttonVerifikasiLaporan").click()
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'kirimLaporan')))
        time.sleep(3)
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#kirimLaporan > span')))
        driver.find_element(By.CSS_SELECTOR, "#kirimLaporan > span").click()
        Log.info('Kirim laporan')
    except:
        Log.info('Gagal Kirim laporan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_17_loginKanwil():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    kanwiljabar(driver)
    Log.info('Login Spv Laporan Kamtib 6D')

@mark.fixture_test()
def test_18_AksesMenuSpv():
    sleep(driver)
    try:
        menulaporan7c(driver)
        Log.info('Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_19_VerifikasiLaporanKanwil():
    try:
        sleep(driver)
        time.sleep(2)
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterUpt2')))
        driver.find_element(By.ID, "formfilterUpt2").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterUpt2").send_keys("Rutan Kelas")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
        
        driver.find_element(By.XPATH, "//span[contains(.,\'Rutan Kelas I Bandung\')]").click()
        Log.info('Input UPT')
        assert True
    except:
        Log.info('Gagal Input UPT')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_20_ClickButtonSearch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        driver.find_element(By.ID, "searchButton").click()
        Log.info('Click Button Search')
        assert True
    except:
        Log.info('Gagal Click Button Search')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_21_ClickButtonVerifikasi():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasi')))
        driver.find_element(By.ID, "buttonVerifikasi").click()
        Log.info('Click Button Verifikasi')
        assert True 
    except:
        Log.info('Gagal Click Button Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_22_StatusVerifikasiModal():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'simpanVerifikasi')))
        time.sleep(3)
        driver.find_element(By.ID, "statusVerifikasiModal").click()
        Log.info('Click Verifikasi Modal')
        assert True
    except:
        Log.info('Gagal Click Verifikasi Modal')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_23_UbahStatus():
    sleep(driver)
    try:
        driver.find_element(By.ID, "diizinkan").click()
        Log.info('Ubah Status Verifikasi')
        assert True
    except:
        Log.info('Gagal Ubah Status Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_24_Inputketerangan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "keterangan").send_keys("keterangan")
        Log.info('Input Keterangan')
        assert True
    except:
        Log.info('Gagal Input Keterangan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_25_SimpanVerifikasi():
    sleep(driver)
    try:
        driver.find_element(By.ID, "simpanVerifikasi").click()
        Log.info('Click Button Simpan Verifikasi')
        assert True
    except:
        Log.info('Gagal Click Button Simpan Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_26_loginPusat():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    pusat(driver)
    Log.info('Login Spv Laporan Kamtib 6D')

@mark.fixture_test()
def test_27_AksesMenuPusat():
    sleep(driver)
    try:
        menulaporan7c(driver)
        Log.info('Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_28_PilihKanwil():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterKanwil')))
        driver.find_element(By.ID, "formfilterKanwil").send_keys("jawa barat")
        driver.find_element(By.XPATH, "//li[contains(.,\'Jawa Barat\')]").click()
        Log.info('Pilih Kanwil')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Pilih Kanwil')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_29_PilihUPT():
    sleep(driver)
    try:
        driver.find_element(By.ID, "formfilterUpt").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterUpt").send_keys("rutan kelas")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
        driver.find_element(By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]").click()
        Log.info('PilihUPT')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal PilihUPT')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_30_ClickButtonSearch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        driver.find_element(By.CSS_SELECTOR, "#searchButton > span").click()
        Log.info('Click Button Search')
        attach(data=driver.get_screenshot_as_png())
    except:
        Log.info('Gagal Click Button Search')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_31_exit():
    quit(driver)
    Log.info('Exit')








    
    