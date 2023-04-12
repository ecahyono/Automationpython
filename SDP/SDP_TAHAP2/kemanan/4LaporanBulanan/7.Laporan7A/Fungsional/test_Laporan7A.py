from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUAT"))
    file_path = environ.get("fakermac")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))


from Settings.setupkeamanan import initDriver, loadDataPath, quit, sleep
from Settings.loginkeamanan import oplapkamtibwaru,kanwiljabar,pusat, SpvRutanBdg
from Settings.Page.keamanan import menulaporan7a
import random
import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Loglaporan7A.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)


i = 2
laporan7a = wb['Laporan7A']


Blok                                                = laporan7a['A'+str(i)].value
Lantai                                              = laporan7a['B'+str(i)].value
Kamar                                               = laporan7a['C'+str(i)].value

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Laporan7Faker'

fake = Faker('id_ID')
NamaBarangTemuan = ['pisau','obeng','tali','obat','Kipas','Kacamata','uang','Kunci','Hp','Kertas','Buku']
pemilikbarang = ['IRIANA TAMBa BINTI waloyo','MURSININ HARYANTi BIN sugiono','WULAN SIREGAr BIN waloyo','NILAM WIBOWo BIN sugiono','CUT NILAM HARTATi BINTI doniserigar']
jumlahbarang = ['1','2','3','4', '5', '6' ]
TindakLanjutBarang = ['Musnahkan','Amankan']
for i in range(5):
    NoSuratPerintahfaker = fake.numerify('SRT/'+'#################')
    tanggalpelaksanaanfaker = fake.date_between(start_date='-1d', end_date='today').strftime('%d/%m/%Y')
    keteranganFaker = fake.numerify('############')
    NamaBarangTemuanFaker = random.choice(NamaBarangTemuan)
    pemilikFaker = random.choice(pemilikbarang)
    jumlahFaker = random.choice(jumlahbarang)
    TindakLanjutFaker = random.choice(TindakLanjutBarang)
    

    worksheet.append([
        NoSuratPerintahfaker, tanggalpelaksanaanfaker, keteranganFaker,NamaBarangTemuanFaker, pemilikFaker, jumlahFaker, TindakLanjutFaker
        ])
workbook.save(file_path)

workbook = load_workbook(filename=file_path)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=2, values_only=True):
    nosrtfkr                 = row[0]
    tglpelaksanaanfkr        = row[1]
    keteranganFkr            = row[2]
    namabrgfkr               = row[3]
    Pemilikfkr               = row[4]
    jumlahfkr                = row[5]
    Tindaklanjutfkr          = row[6]


@mark.fixture_test()
def test_1_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      

@mark.fixture_test()
def test_2_Login():
    Log.info('Setup Os')
    oplapkamtibwaru(driver)
    Log.info('Login Op Laporan Kamtib 7 A')

@mark.fixture_test()
def test_2_AksesMenu():
    sleep(driver)
    try:
        menulaporan7a(driver)
        Log.info('Akses halaman Laporan 7 A')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Akses halaman Laporan 7 A')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_3_CreateButton():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'createButton')))
        driver.find_element(By.ID, "createButton").click()
        Log.info('Klik tombol tambah data')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Klik tombol tambah data')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

    
@mark.fixture_test()
def test_4_InpuNoSuratPerintah():
    sleep(driver)
    try:
        driver.find_element(By.ID, "nmrSuratPerintah").click()
        driver.find_element(By.ID, "nmrSuratPerintah").send_keys(nosrtfkr)
        Log.info('Input No Surat Perintah')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input No Surat Perintah')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_5_InputTanggalPelaksanaan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "tglPelaksanaanKegiatan").click()
        driver.find_element(By.ID, "tglPelaksanaanKegiatan").send_keys(tglpelaksanaanfkr)
        Log.info('Input Tanggal Pelaksanaan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Tanggal Pelaksanaan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_6_InputBlok():
    sleep(driver)
    try:
        driver.find_element(By.ID, "blokForm").click()
        driver.find_element(By.ID, "blokForm").send_keys(Blok)
        driver.find_element(By.XPATH, "//li[contains(.,\'"+ Blok +"')]").click()
        Log.info('Input Blok')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Blok')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_7_lantai():
    sleep(driver)
    try:
        driver.find_element(By.ID, "lantaiForm").click()
        driver.find_element(By.ID, "lantaiForm").send_keys(Lantai)
        driver.find_element(By.XPATH, "//li[contains(.,\'"+ Lantai +"')]").click()
        Log.info('Input Lantai')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Lantai')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_8_InputKamar():
    sleep(driver)
    try:
        driver.find_element(By.ID, "kamarForm").click()
        driver.find_element(By.ID, "kamarForm").send_keys(Kamar)
        driver.find_element(By.XPATH, "//li[contains(.,\'"+ Kamar +"')]").click()
        Log.info('Input Kamar')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Kamar')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_9_InputKeterangan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "keterangan").click()
        driver.find_element(By.ID, "keterangan").send_keys(keteranganFkr)
        Log.info('Input Keterangan')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Keterangan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_10_InputNamaBarang():
    sleep(driver)
    try:
        driver.find_element(By.ID, "namaBarang-0").click()
        driver.find_element(By.ID, "namaBarang-0").send_keys(namabrgfkr)
        Log.info('Input Nama Barang')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Nama Barang')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_11_InputJumlah():
    sleep(driver)
    try:
        driver.find_element(By.ID, "jumlahBarang-0").click()
        driver.find_element(By.ID, "jumlahBarang-0").send_keys(jumlahfkr)
        Log.info('Input Jumlah')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Jumlah')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_12_namaPemilik():
    sleep(driver)
    try:
        driver.find_element(By.ID, "namaPemilikBarang").click()
        driver.find_element(By.ID, "namaPemilikBarang").send_keys(Pemilikfkr)
        driver.find_element(By.XPATH, "//li[contains(.,\'"+ Pemilikfkr +"')]").click()
        Log.info('Input Nama Pemilik')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Nama Pemilik')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_13_InputtindakLanjut():
    sleep(driver)
    try:
        driver.find_element(By.ID, "tindakLanjut").click()
        driver.find_element(By.ID, "tindakLanjut").send_keys(Tindaklanjutfkr)
        driver.find_element(By.XPATH, "//li[contains(.,\'"+ Tindaklanjutfkr +"')]").click()
        Log.info('Input Tindak Lanjut')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Input Tindak Lanjut')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_14_submit():
    sleep(driver)
    try:
        driver.find_element(By.ID, "submitButton").click()
        Log.info('Submit')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Submit')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False



@mark.fixture_test()
def test_15_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      

@mark.fixture_test()
def test_16_LoginSPV():
    Log.info('Setup Os')
    SpvRutanBdg(driver)
    Log.info('Login Op Laporan Kamtib 7 A')


@mark.fixture_test()
def test_17_AksesMenuSpv():
    sleep(driver)
    try:
        menulaporan7a(driver)
        Log.info('Akses halaman Laporan 7 A')
        attach(data=driver.get_screenshot_as_png())
    except:
        Log.info('Gagal Akses halaman Laporan 7 A')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_18_KirimLaporanSpv():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasiLaporan')))
        time.sleep(2)
        driver.find_element(By.ID, "buttonVerifikasiLaporan").click()
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'kirimLaporan')))
        time.sleep(3)
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#kirimLaporan > span')))
        driver.find_element(By.CSS_SELECTOR, "#kirimLaporan > span").click()
        Log.info('Kirim laporan')
    except:
        Log.info('Gagal Kirim laporan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_19_loginKanwil():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    kanwiljabar(driver)
    Log.info('Login Spv Laporan Kamtib 6D')

@mark.fixture_test()
def test_20_AksesMenuSpv():
    sleep(driver)
    try:
        menulaporan7a(driver)
        Log.info('Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_21_VerifikasiLaporanKanwil():
    try:
        sleep(driver)
        time.sleep(2)
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterUpt2')))
        driver.find_element(By.ID, "formfilterUpt2").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterUpt2").send_keys("Rutan Kelas")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
        
        driver.find_element(By.XPATH, "//span[contains(.,\'Rutan Kelas I Bandung\')]").click()
        Log.info('Input UPT')
        assert True
    except:
        Log.info('Gagal Input UPT')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_22_ClickButtonSearch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        driver.find_element(By.ID, "searchButton").click()
        Log.info('Click Button Search')
        assert True
    except:
        Log.info('Gagal Click Button Search')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_23_ClickButtonVerifikasi():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'buttonVerifikasi')))
        driver.find_element(By.ID, "buttonVerifikasi").click()
        Log.info('Click Button Verifikasi')
        assert True 
    except:
        Log.info('Gagal Click Button Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False


@mark.fixture_test()
def test_24_StatusVerifikasiModal():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'simpanVerifikasi')))
        time.sleep(3)
        driver.find_element(By.ID, "statusVerifikasiModal").click()
        Log.info('Click Verifikasi Modal')
        assert True
    except:
        Log.info('Gagal Click Verifikasi Modal')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_25_UbahStatus():
    sleep(driver)
    try:
        driver.find_element(By.ID, "diizinkan").click()
        Log.info('Ubah Status Verifikasi')
        assert True
    except:
        Log.info('Gagal Ubah Status Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_26_Inputketerangan():
    sleep(driver)
    try:
        driver.find_element(By.ID, "keterangan").send_keys("keterangan")
        Log.info('Input Keterangan')
        assert True
    except:
        Log.info('Gagal Input Keterangan')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_27_SimpanVerifikasi():
    sleep(driver)
    try:
        driver.find_element(By.ID, "simpanVerifikasi").click()
        Log.info('Click Button Simpan Verifikasi')
        assert True
    except:
        Log.info('Gagal Click Button Simpan Verifikasi')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False



@mark.fixture_test()
def test_28_loginPusat():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    pusat(driver)
    Log.info('Login Spv Laporan Kamtib 6D')

@mark.fixture_test()
def test_29_AksesMenuPusat():
    sleep(driver)
    try:
        menulaporan7a(driver)
        Log.info('Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Akses halaman Laporan 6D')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_30_PilihKanwil():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'formfilterKanwil')))
        driver.find_element(By.ID, "formfilterKanwil").send_keys("jawa barat")
        driver.find_element(By.XPATH, "//li[contains(.,\'Jawa Barat\')]").click()
        Log.info('Pilih Kanwil')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal Pilih Kanwil')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_31_PilihUPT():
    sleep(driver)
    try:
        driver.find_element(By.ID, "formfilterUpt").click()
        time.sleep(1)
        driver.find_element(By.ID, "formfilterUpt").send_keys("rutan kelas")
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]")))
        driver.find_element(By.XPATH, "//li[contains(.,\'Rutan Kelas I Bandung\')]").click()
        Log.info('PilihUPT')
        attach(data=driver.get_screenshot_as_png())
        assert True
    except:
        Log.info('Gagal PilihUPT')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_32_ClickButtonSearch():
    sleep(driver)
    try:
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
        driver.find_element(By.CSS_SELECTOR, "#searchButton > span").click()
        Log.info('Click Button Search')
        attach(data=driver.get_screenshot_as_png())
    except:
        Log.info('Gagal Click Button Search')
        attach(data=driver.get_screenshot_as_png())
        quit(driver)
        assert False

@mark.fixture_test()
def test_33_exit():
    quit(driver)
    Log.info('Exit')








    
    