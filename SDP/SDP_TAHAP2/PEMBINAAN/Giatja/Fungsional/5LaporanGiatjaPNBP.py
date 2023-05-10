from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime
import pytest

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("data"))
    file_path = environ.get("fakerGiatja")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
import random

from Settings.setupGiatja import initDriver, loadDataPath, quit, sleep, upload, uploadGambar
from Settings.loginGiatja import Op_Giatja
from Settings.Page.Giatja import MenuLaporan
import random
import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('logLaporan.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Pemasaran'

fake = Faker('id_ID')
akunSetor                                               = ['pendapatan','nihil']
pemasaran                                               = ['pemasaran0opt1']

for i in range(5):
    TanggalSetorFaker                                   = fake.date_between(start_date='today', end_date='today').strftime('%m')
    akunSetorFaker                                      = random.choice(akunSetor)
    keteranganFaker                                     = fake.text(max_nb_chars=255)
    pemasaranFaker                                      = random.choice(pemasaran)
    NilaiFaker                                          = fake.random_int(min=100000, max=1000000)
    worksheet.append([
        TanggalSetorFaker,
        akunSetorFaker,
        keteranganFaker,
        pemasaranFaker,
        NilaiFaker
        ])
workbook.save(file_path)

workbook = load_workbook(filename=file_path)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=2, values_only=True):
    TanggalSetorFaker                                   = row[0]
    akunSetorFaker                                      = row[1]
    keteranganFaker                                     = row[2]
    pemasaranFaker                                      = row[3]
    NilaiFaker                                          = row[4]

@pytest.mark.webtest
def test_1_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      
    Op_Giatja(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Setup Os')

@pytest.mark.webtest
def test_TC_GIATJA_022():
    sleep(driver)
    MenuLaporan(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Operator mengakses halaman Laporan Kegiatan Kerja dan Produksi')

@pytest.mark.webtest
def test_TC_GIATJA_016():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "Aprilsarana3")))
    if TanggalSetorFaker == "01":
        driver.find_element(By.ID, "Januaripemasaran-pnbp1").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Januari pnbp')
    elif TanggalSetorFaker == "02":
        driver.find_element(By.ID, "Febuaripemasaran-pnbp1").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Febuari pnbp')
    elif TanggalSetorFaker == "03":
        driver.find_element(By.ID, "Maretpemasaran-pnbp1").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Maret pnbp')
    elif TanggalSetorFaker == "04":
        driver.find_element(By.ID, "Aprilpemasaran-pnbp1").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('April pnbp')
    elif TanggalSetorFaker == "05":
        driver.find_element(By.ID, "Meipemasaran-pnbp1").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Mei pnbp')
    elif TanggalSetorFaker == "06":
        driver.find_element(By.ID, "Junipemasaran-pnbp1").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Juni pnbp')
    elif TanggalSetorFaker == "07":
        driver.find_element(By.ID, "Julipemasaran-pnbp1").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Juli pnbp')
    elif TanggalSetorFaker == "08":
        driver.find_element(By.ID, "Agustuspemasaran-pnbp1").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Agustus pnbp')
    elif TanggalSetorFaker == "09":
        driver.find_element(By.ID, "Septemberpemasaran-pnbp1").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('September pnbp')
    elif TanggalSetorFaker == "10":
        driver.find_element(By.ID, "Oktoberpemasaran-pnbp1").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Oktober pnbp')
    elif TanggalSetorFaker == "11":
        driver.find_element(By.ID, "Novemberpemasaran-pnbp1").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('November pnbp')
    elif TanggalSetorFaker == "12":
        driver.find_element(By.ID, "Desemberpemasaran-pnbp1").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Desember pnbp')
    driver.find_element(By.XPATH, "//span[contains(.,'OK')]").click()
    WebDriverWait(driver, 50).until(EC.invisibility_of_element((By.ID, ".circular")))
    time.sleep(2)
    WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, "Aprilpemasaran-pnbp1")))
     
    Log.info('Operator membuat Laporan Kegiatan Kerja dan Produksi')
     

   

@pytest.mark.webtest
def test_exit():
    quit(driver)
    Log.info('Exit')







    
    