from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime
import pytest

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("data"))
    file_path = environ.get("fakerGiatja")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
import random

from Settings.setupGiatja import initDriver, loadDataPath, quit, sleep, upload, uploadGambar
from Settings.loginGiatja import Op_Giatja
from Settings.Page.Giatja import MenuPNBP
import random
import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('LogOpPNBP.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Pemasaran'

fake = Faker('id_ID')
akunSetor                                               = ['pendapatan','nihil']
pemasaran                                               = ['pemasaran0opt1']

for i in range(5):
    TanggalSetorFaker                                   = fake.date_between(start_date='today', end_date='today').strftime('%d/%m/%Y')
    akunSetorFaker                                      = random.choice(akunSetor)
    keteranganFaker                                     = fake.text(max_nb_chars=255)
    pemasaranFaker                                      = random.choice(pemasaran)
    NilaiFaker                                          = fake.random_int(min=100000, max=1000000)
    worksheet.append([
        TanggalSetorFaker,
        akunSetorFaker,
        keteranganFaker,
        pemasaranFaker,
        NilaiFaker
        ])
workbook.save(file_path)

workbook = load_workbook(filename=file_path)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=2, values_only=True):
    TanggalSetorFaker                                   = row[0]
    akunSetorFaker                                      = row[1]
    keteranganFaker                                     = row[2]
    pemasaranFaker                                      = row[3]
    NilaiFaker                                          = row[4]

@pytest.mark.webtest
def test_1_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      
    Op_Giatja(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Setup Os')

@pytest.mark.webtest
def test_TC_GIATJA_015():
    sleep(driver)
    MenuPNBP(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Operator mengakses halaman PNBP')

@pytest.mark.webtest
def test_TC_GIATJA_016():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "createButton")))
    driver.find_element(By.ID, "createButton").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "backButton")))
    attach(data=driver.get_screenshot_as_png())

    WebDriverWait(driver, 30).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".circular")))
    driver.find_element(By.ID, "tanggalSetor").send_keys(TanggalSetorFaker)
    driver.find_element(By.ID, "tanggalSetor").send_keys(Keys.ENTER)

    driver.find_element(By.ID, "akunSetor").click()
    driver.find_element(By.ID, ""+akunSetorFaker+"").click()

    driver.find_element(By.ID, "uploadButton").click()
    uploadGambar(driver)
    driver.find_element(By.ID, "keterangan").send_keys(keteranganFaker)

    driver.find_element(By.CSS_SELECTOR, "#addButton .h-5").click()

    driver.find_element(By.ID, "pemasaran0").click()
    driver.find_element(By.ID, ""+pemasaranFaker+"").click()

    driver.find_element(By.ID, "jumlah0").send_keys(NilaiFaker)

    driver.find_element(By.ID, "submitButton").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "createButton")))

    Log.info('Operator menambahkan data PNPB')

@pytest.mark.webtest
def test_TC_GIATJA_017():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "createButton")))
    driver.find_element(By.ID, "filterColumn").click()
    driver.find_element(By.XPATH, "//li[@id='semua']").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    driver.find_element(By.CSS_SELECTOR, "#lihat0 > span").click()
    time.sleep(2)
    driver.find_element(By.CSS_SELECTOR, ".el-dialog__close > svg").click()
    Log.info('Operator melihat bukti setor yang telah di upload')
    attach(data=driver.get_screenshot_as_png())

    Log.info('Operator melihat bukti setor yang telah di upload')
    
@pytest.mark.webtest
def test_TC_GIATJA_018():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "createButton")))
    driver.find_element(By.ID, "filterColumn").click()
    driver.find_element(By.XPATH, "//li[@id='semua']").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))

    driver.find_element(By.XPATH, "//tr[1]/td[8]/div/div/div/a/button").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".el-image__inner")))
    driver.find_element(By.CSS_SELECTOR, "#backButton > span").click()
    Log.info('Operator mengakses halaman Detail PNPB')

@pytest.mark.webtest
def test_TC_GIATJA_019():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "createButton")))
    driver.find_element(By.ID, "filterColumn").click()
    driver.find_element(By.XPATH, "//li[@id='semua']").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))

    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#updateButton0 .h-5")))
    driver.find_element(By.CSS_SELECTOR, "#updateButton0 .h-5").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#backButton > span")))

    Log.info('Operator mengakses halaman Ubah PNPB')

@pytest.mark.webtest
def test_TC_GIATJA_020():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".el-loading-spinner")))
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "keterangan")))
    driver.find_element(By.ID, "keterangan").send_keys(keteranganFaker)
    driver.find_element(By.CSS_SELECTOR, "#submitButton > span").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "createButton")))

    Log.info('Operator mengubah data PNPB')

@pytest.mark.webtest
def test_TC_GIATJA_021():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "createButton")))
    driver.find_element(By.ID, "filterColumn").click()
    driver.find_element(By.XPATH, "//li[@id='semua']").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    driver.find_element(By.CSS_SELECTOR, ".el-table__row:nth-child(1) .el-space__item > .el-button path").click()
    driver.find_element(By.CSS_SELECTOR, ".el-button--primary > span").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//div[contains(.,'Berhasil Menghapus')]"))).click()
    

    Log.info('Operator menghapus data PNPB')

@pytest.mark.webtest
def test_exit():
    quit(driver)
    Log.info('Exit')







    
    