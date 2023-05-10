from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime
import pytest

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("data"))
    file_path = environ.get("fakerGiatja")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
import random

from Settings.setupGiatja import initDriver, loadDataPath, quit, sleep, upload, uploadGambar
from Settings.loginGiatja import Op_Giatja
from Settings.Page.Giatja import MenuPemasaran
import random
import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('LogOpPemasaran.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

sheetrangeIndex = wb['Negara']

i = random.randint(1,200)

NamaNegara                                   = sheetrangeIndex['A'+str(i)].value
KotaTujuan                                   = sheetrangeIndex['B'+str(i)].value
print('input Negara adalah', NamaNegara,KotaTujuan)
time.sleep(3)

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Pemasaran'

fake = Faker('id_ID')
JenisPemasaran                                          = ['jenis0','jenis1','jenis2','jenis3','jenis4']
mitra                                                   = ['mitra0','mitra1','mitra2','mitra3','mitra4']
produk                                                  = ['Meja Kayu','Sablon Baju','Kursi Kayu','mobil mini','Sweater Anak']
satuan                                                  = ['pasang','Buah','Lusin','Kotak','Kotak','Unit']
for i in range(5):
    JenisPemasaranFaker                                 = random.choice(JenisPemasaran)
    NamaKegiatanFaker                                   = fake.text(max_nb_chars=7)
    TanggalKegiatanFaker                                = fake.date_between(start_date='today', end_date='today').strftime('%d/%m/%Y')
    MitraFaker                                          = random.choice(mitra)
    UraianFaker                                         = fake.text(max_nb_chars=255)
    LokasiFaker                                         = fake.address()
    produkFaker                                         = random.choice(produk)
    JumlahBarangFaker                                   = fake.random_int(min=1, max=10)
    satuanFaker                                         = random.choice(satuan)
    NilaiFaker                                          = fake.random_int(min=100000, max=1000000)
    worksheet.append([
        JenisPemasaranFaker,
        NamaKegiatanFaker,
        TanggalKegiatanFaker,
        MitraFaker,
        UraianFaker,
        LokasiFaker,
        produkFaker,
        JumlahBarangFaker,
        satuanFaker,
        NilaiFaker
        ])
workbook.save(file_path)

workbook = load_workbook(filename=file_path)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=2, values_only=True):
    JenisPemasaranExcell                         = row[0]
    NamaKegiatanExcell                           = row[1]
    TanggalKegiatanExcell                        = row[2]
    MitraExcell                                  = row[3]
    UraianExcell                                 = row[4]
    LokasiExcell                                 = row[5]
    produkExcell                                 = row[6]
    JumlahBarangExcell                           = row[7]
    satuanExcell                                 = row[8]
    NilaiExcell                                  = row[9]



@pytest.mark.webtest
def test_1_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      
    Op_Giatja(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Setup Os')

@pytest.mark.webtest
def test_TC_GIATJA_008():
    sleep(driver)
    MenuPemasaran(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Operator mengakses halaman Pemasaran')

@pytest.mark.webtest
def test_TC_GIATJA_009():
    sleep(driver)

    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "createButton")))
    driver.find_element(By.ID, "createButton").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "backButton")))
    attach(data=driver.get_screenshot_as_png())
    Log.info('Klik Button Tambah')

    driver.find_element(By.ID, "jenisPemasaran").click()
    driver.find_element(By.ID, ""+JenisPemasaranFaker+"").click()

    driver.find_element(By.ID, "namaKegiatan").click()
    driver.find_element(By.ID, "namaKegiatan").send_keys(NamaKegiatanFaker)

    driver.find_element(By.ID, "tanggalPemasaran").send_keys(TanggalKegiatanFaker)
    driver.find_element(By.ID, "tanggalPemasaran").send_keys(Keys.ENTER)
    

    driver.find_element(By.ID, "mitra").click()
    driver.find_element(By.ID, ""+MitraFaker+"").click()

    if JenisPemasaran == 'jenis4':
        driver.find_element(By.ID, "negaraTujuan").click()
        driver.find_element(By.ID, ""+NamaNegara+"").click()
        driver.find_element(By.ID, "negaraTujuan").send_keys(Keys.ENTER)

        driver.find_element(By.ID, "kotaTujuan").send_keys(KotaTujuan)
        driver.find_element(By.XPATH, "//li[contains(.,\'"+ KotaTujuan +"')]").click()
        driver.find_element(By.ID, "kotaTujuan").send_keys(Keys.ENTER)

    else:

        driver.find_element(By.ID, "kota").send_keys(KotaTujuan)
    
        driver.find_element(By.XPATH, "//li[contains(.,\'"+ KotaTujuan +"')]").click()
        driver.find_element(By.ID, "kota").send_keys(Keys.ENTER)

        driver.find_element(By.ID, "lokasi").click()
        driver.find_element(By.ID, "lokasi").send_keys(LokasiFaker)

    driver.find_element(By.ID, "uploadButton").click()
    uploadGambar(pathData)

    driver.execute_script("window.scrollTo(0,432)")
    driver.find_element(By.ID, "uraianKegiatan").click()
    driver.find_element(By.ID, "uraianKegiatan").send_keys(UraianFaker)


    driver.find_element(By.CSS_SELECTOR, ".h-5").click()

    driver.find_element(By.ID, "produk0").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'"+ produkFaker +"')]")))
    driver.find_element(By.XPATH, "//li[contains(.,\'"+ produkFaker +"')]").click()

    driver.find_element(By.ID, "jumlah0").send_keys(JumlahBarangFaker)

    driver.find_element(By.ID, "satuan0").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'"+ satuanFaker +"')]")))
    driver.find_element(By.XPATH, "//li[contains(.,\'"+ satuanFaker +"')]").click()

    driver.find_element(By.XPATH, "(//input[@type=\'text\'])[10]").send_keys(NilaiFaker)
    driver.find_element(By.CSS_SELECTOR, "#submitButton > span").click()

    Log.info('Operator mengisi form Pemasaran')
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, "createButton")))
    Log.info('Operator berhasil mengisi form Pemasaran')
    attach(data=driver.get_screenshot_as_png())
   

@pytest.mark.webtest
def test_TC_GIATJA_010():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "createButton")))
    driver.find_element(By.ID, "filterColumn").click()
    driver.find_element(By.XPATH, "//li[@id='semua']").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    driver.find_element(By.CSS_SELECTOR, "#lihat0 > span").click()
    driver.find_element(By.CSS_SELECTOR, ".el-dialog__close > svg").click()
    Log.info('Operator melihat bukti setor yang telah di upload')
    attach(data=driver.get_screenshot_as_png())

@pytest.mark.webtest
def test_TC_GIATJA_011():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "createButton")))
    driver.find_element(By.ID, "filterColumn").click()
    driver.find_element(By.XPATH, "//li[@id='semua']").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))

    driver.find_element(By.XPATH, "//tr[1]/td[8]/div/div/div/a/button").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".el-image__inner")))
    driver.find_element(By.CSS_SELECTOR, "#backButton > span").click()
    Log.info("Operator mengakses halaman Detail Pemasaran")
    attach(data=driver.get_screenshot_as_png())

@pytest.mark.webtest
def test_TC_GIATJA_012():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "createButton")))
    driver.find_element(By.ID, "filterColumn").click()
    driver.find_element(By.XPATH, "//li[@id='semua']").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))

    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#updateButton0 .h-5")))
    driver.find_element(By.CSS_SELECTOR, "#updateButton0 .h-5").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#backButton > span")))

    Log.info('Operator mengakses halaman Edit Pemasaran')
    attach(data=driver.get_screenshot_as_png())

@pytest.mark.webtest
def test_TC_GIATJA_013():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".el-loading-spinner")))
    WebDriverWait(driver, 30).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".circular")))
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "uraianKegiatan")))
    driver.find_element(By.ID, "uraianKegiatan").send_keys(UraianFaker)
    driver.find_element(By.CSS_SELECTOR, "#submitButton > span").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "createButton")))


    Log.info('Operator mengubah data Pemasaran')
    attach(data=driver.get_screenshot_as_png())

@pytest.mark.webtest
def test_TC_GIATJA_014():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "createButton")))
    driver.find_element(By.ID, "filterColumn").click()
    driver.find_element(By.XPATH, "//li[@id='semua']").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    driver.find_element(By.CSS_SELECTOR, ".el-table__row:nth-child(1) .el-space__item > .el-button path").click()
    driver.find_element(By.CSS_SELECTOR, ".el-button--primary > span").click()

    Log.info('Operator menghapus data Pemasaran')
    attach(data=driver.get_screenshot_as_png())


@pytest.mark.webtest
def test2_exit():
    quit(driver)
    Log.info('Exit')







    
    