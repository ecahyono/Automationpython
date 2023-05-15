from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime
import pytest

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("data"))
    file_path = environ.get("fakerGiatja")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
import random

from Settings.setupGiatja import initDriver, loadDataPath, quit, sleep, upload, uploadGambar
from Settings.loginGiatja import Op_Giatja
from Settings.Page.Giatja import MenuGiatja
import random
import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('LogGiatja.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

sheetrangeIndex = wb['listWbpSumedang']

random1 = random.randint(1,170)
random2= random.randint(1,170)
random3 = random.randint(1,170)

NamaInput1                                 = sheetrangeIndex['A'+str(random1)].value
NamaInput2                                 = sheetrangeIndex['B'+str(random2)].value
NamaInput3                                 = sheetrangeIndex['C'+str(random3)].value
print('nama input WBP adalah', NamaInput1, NamaInput2, NamaInput3)
time.sleep(3)

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'GiatjaTambah'

fake = Faker('id_ID')
KegiatanKerja   = ['Manufaktur','Jasa','Agribisnis']
Jeniskegiatan   = ['jenisKegiatan1','jenisKegiatan2','jenisKegiatan3']
Skalakegiatan   = ['skala0','skala1','skala2']
area            = ['area0','area1','area2']
sarana          = ['sarana0','sarana1','sarana2']
prasarana       = ['prasaranaundefined-0','prasaranaundefined-1','prasaranaundefined-2','prasaranaundefined-3']
mitra           = ['mitra0','mitra1','mitra2']
checkbox        = ['.el-table__row:nth-child(1) .el-checkbox__inner','.el-table__row:nth-child(2) .el-checkbox__inner','.el-table__row:nth-child(3) .el-checkbox__inner','.el-table__row:nth-child(4) .el-checkbox__inner','.el-table__row:nth-child(6) .el-checkbox__inner','.el-table__row:nth-child(7) .el-checkbox__inner','.el-table__row:nth-child(8) .el-checkbox__inner','.el-table__row:nth-child(9) .el-checkbox__inner','.el-table__row:nth-child(10) .el-checkbox__inner','.el-table__row:nth-child(5) .el-checkbox__inner']
idjenis         = ['#jenis0 > span','#jenis1 > span']
satuan          = ['satuan0','satuan1','satuan3','satuan4','satuan5']


for i in range(1):
    KegiatanKerjaFaker                = random.choice(KegiatanKerja)
    JeniskegiatanFaker                = random.choice(Jeniskegiatan)
    namaKegiatanFaker                 = fake.text(max_nb_chars=7)
    SkalakegiatanFaker                = random.choice(Skalakegiatan)
    tanggalAwalKegiatanFaker          = fake.date_between(start_date='-30d', end_date='-1d').strftime('%d/%m/%Y')
    tanggalAkhirKegiatanFaker         = fake.date_between(start_date='today', end_date='today').strftime('%d/%m/%Y')
    areaFaker                         = random.choice(area)
    lokasiKegiatanFaker               = fake.address()
    luasLokasiKegiatanFaker           = fake.random_int(min=10, max=1000)
    jumlahRuangKegiatanFaker          = fake.random_int(min=1, max=10)
    saranaFaker                       = random.choice(sarana)
    prasaranaFaker                    = random.choice(prasarana)
    mitraFaker                        = random.choice(mitra)
    ketereranganFaker                 = fake.text(max_nb_chars=100)
    jumlahPesertaFaker                = fake.random_int(min=1, max=3)
    PesertaRandomFaker1               = random.choice(checkbox)
    PesertaRandomFaker2               = random.choice(checkbox)
    PesertaRandomFaker3               = random.choice(checkbox)
    idjenisFaker                      = random.choice(idjenis)
    namaProdukFaker                   = fake.text(max_nb_chars=7)
    jumlahprodukFaker                 = fake.random_int(min=1, max=3)
    satuanFaker                       = random.choice(satuan)
    HargaFaker                        = fake.random_int(min=10000, max=1000000)
    lamaPengerjaanFaker               = fake.random_int(min=1, max=7)



    worksheet.append([
        KegiatanKerjaFaker,
        JeniskegiatanFaker,
        namaKegiatanFaker,
        SkalakegiatanFaker,
        tanggalAwalKegiatanFaker,
        tanggalAkhirKegiatanFaker,
        areaFaker,
        lokasiKegiatanFaker,
        luasLokasiKegiatanFaker,
        jumlahRuangKegiatanFaker,
        saranaFaker,
        prasaranaFaker,
        mitraFaker,
        ketereranganFaker,
        jumlahPesertaFaker,
        PesertaRandomFaker1,
        PesertaRandomFaker2,
        PesertaRandomFaker3,
        idjenisFaker,
        namaProdukFaker,
        jumlahprodukFaker,
        satuanFaker,
        HargaFaker,
        lamaPengerjaanFaker
        ])
workbook.save(file_path)

workbook = load_workbook(filename=file_path)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=1, values_only=True):
    KegiatanKerjaFkr                            = row[0]
    JeniskegiatanFkr                            = row[1]
    namaKegiatanFkr                             = row[2]
    SkalakegiatanFkr                            = row[3]
    tanggalAwalKegiatanFkr                      = row[4]
    tanggalAkhirKegiatanFkr                     = row[5]
    areaFkr                                     = row[6]
    lokasiKegiatanFkr                           = row[7]
    luasLokasiKegiatanFkr                       = row[8]
    jumlahRuangKegiatanFkr                      = row[9]
    saranaFkr                                   = row[10]
    prasaranaFkr                                = row[11]
    mitraFkr                                    = row[12]
    ketereranganFkr                             = row[13]
    jumlahPesertaFkr                            = row[14]
    PesertaRandomFkr                            = row[15]
    PesertaRandomFkr2                           = row[16]
    PesertaRandomFkr3                           = row[17]
    idjenisFkr                                  = row[18]
    namaProdukFkr                               = row[19]
    jumlahprodukFkr                             = row[20]
    satuanFkr                                   = row[21]
    HargaFkr                                    = row[22]
    lamaPengerjaanFkr                           = row[23]




@pytest.mark.webtest
def test_SetupOsOpen():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      

@pytest.mark.webtest
def test_TC_GIATJA_001():
    Log.info('Setup Os')
    Op_Giatja(driver)
    Log.info('Login Op Giatja')
    attach(data=driver.get_screenshot_as_png())

@pytest.mark.webtest
def test_TC_GIATJA_002():
    sleep(driver)
    MenuGiatja(driver)
    Log.info('Operator mengakses halaman Kegiatan Kerja dan Produksi')
    attach(data=driver.get_screenshot_as_png())

@pytest.mark.webtest
def test_TC_GIATJA_003():
    sleep(driver)
    try:
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "createButton")))
        driver.find_element(By.ID, "createButton").click()
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "backButton")))
        attach(data=driver.get_screenshot_as_png())
        Log.info('Klik Button Tambah')
    except Exception as e:
        Log.info('Button Tambah Tidak Ditemukan' )
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)
        assert False

    sleep(driver)
    try:
        driver.find_element(By.ID, "bidangKegiatan").click()
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'"+ KegiatanKerjaFaker +"')]/span")))
        driver.find_element(By.XPATH, "//li[contains(.,\'"+ KegiatanKerjaFaker +"')]").click()
        Log.info('Input Bidang Kegiatan')
        attach(data=driver.get_screenshot_as_png())
    except Exception as e:
        Log.info('Input Bidang Kegiatan Tidak Ditemukan' )
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)
        assert False

    sleep(driver)
    driver.find_element(By.ID, "jenisKegiatan").click()
    driver.find_element(By.ID, ""+JeniskegiatanFaker+"").click()
    attach(data=driver.get_screenshot_as_png())
    Log.info('Input Jenis Kegiatan')
   
    sleep(driver)
    try:
        driver.find_element(By.ID, "namaKegiatan").send_keys(namaKegiatanFaker)
        Log.info('Input Nama Kegiatan')
    except Exception as e:
        Log.info('Input Nama Kegiatan Tidak Ditemukan' )
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)
        assert False

    sleep(driver)
    try:
        driver.find_element(By.ID, "skalaKegiatan").click()
        driver.find_element(By.ID, ""+SkalakegiatanFaker+"").click()
        Log.info('Input Skala Kegiatan')
    except Exception as e:
        Log.info('Input Skala Kegiatan Tidak Ditemukan' )
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)
        assert False

    sleep(driver)
    try:
        driver.find_element(By.ID, "tanggalAwalKegiatan").send_keys(tanggalAwalKegiatanFaker)
        Log.info('Input Tanggal Awal Kegiatan')
    except Exception as e:
        Log.info('Input Tanggal Awal Kegiatan Tidak Ditemukan' )
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)
        assert False

    sleep(driver)
    try:
        driver.find_element(By.ID, "tanggalAkhirKegiatan").send_keys(tanggalAkhirKegiatanFaker)
        Log.info('Input Tanggal Akhir Kegiatan')
    except Exception as e:
        Log.info('Input Tanggal Akhir Kegiatan Tidak Ditemukan' )
        print('ERROR FOUND',
            type(e).__name__,          # TypeError
            __file__,                  # /tmp/example.py
            e.__traceback__.tb_lineno  # 2
        )
        quit(driver)
        assert False

    sleep(driver)
    try:
        driver.find_element(By.ID, "lokasiKegiatan").send_keys(lokasiKegiatanFaker)
        Log.info('Input Lokasi Kegiatan')
    except Exception as e:
        Log.info('Input Lokasi Kegiatan Tidak Ditemukan' )
        print('ERROR FOUND',
            type(e).__name__,          
            __file__,                  
            e.__traceback__.tb_lineno  
        )
        quit(driver)
        assert False

    sleep(driver)
    driver.find_element(By.ID, "areaKegiatan").click()
    driver.find_element(By.ID, ""+ areaFaker+"").click()
    Log.info('Input Area Kegiatan')


    sleep(driver)
    try:
        driver.find_element(By.ID, "luas").send_keys(luasLokasiKegiatanFaker)
        Log.info('Input Luas')
    except Exception as e:
        Log.info('Input Luas Tidak Ditemukan' )
        print('ERROR FOUND',
            type(e).__name__,          
            __file__,                  
            e.__traceback__.tb_lineno  
        )
        time.sleep(5)
        quit(driver)
        assert False

    sleep(driver)
    try:
        driver.find_element(By.ID, "jumlahRuang").send_keys(jumlahRuangKegiatanFaker)
        Log.info('Input Jumlah Ruang')
    except Exception as e:
        Log.info('Input Jumlah Ruang Tidak Ditemukan' )
        print('ERROR FOUND',
            type(e).__name__,          
            __file__,                  
            e.__traceback__.tb_lineno  
        )
        quit(driver)
        assert False

    sleep(driver)
    driver.find_element(By.CSS_SELECTOR, ".el-form-item:nth-child(7) .el-select__input").click()
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, ""+saranaFaker+"")))
    driver.find_element(By.ID, ""+saranaFaker+"").click()
    driver.find_element(By.CSS_SELECTOR, " .el-form-item:nth-child(7) > .el-form-item__label").click()
    Log.info('Input Sarana')

    sleep(driver)
    try:
        driver.find_element(By.CSS_SELECTOR, ".el-form-item:nth-child(8) .el-select__input").click()
        WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, ""+prasaranaFaker+"")))
        driver.find_element(By.ID, ""+prasaranaFaker+"").click()
        Log.info('Input Prasarana')
    except Exception as e:
        Log.info('Input Prasarana Tidak Ditemukan' )
        print('ERROR FOUND',
            type(e).__name__,          
            __file__,                  
            e.__traceback__.tb_lineno  
        )
        time.sleep(5)
        quit(driver)
        assert False

    sleep(driver)
    try:
        driver.find_element(By.CSS_SELECTOR, ".el-form-item:nth-child(9) .el-select__input").click()
        WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, ""+mitraFaker+"")))
        driver.find_element(By.ID, ""+mitraFaker+"").click()
        Log.info('Input Mitra')
    except Exception as e:
        Log.info('Input Mitra Tidak Ditemukan' )
        print('ERROR FOUND',
            type(e).__name__,          
            __file__,                  
            e.__traceback__.tb_lineno  
        )
        time.sleep(5)
        quit(driver)
        assert False

    sleep(driver)
    try:
        driver.find_element(By.ID, "uploadButton").click()
        uploadGambar(driver)
        Log.info('Input gambar')
    except Exception as e:
        Log.info('Input gambar Tidak Ditemukan' )
        print('ERROR FOUND',
            type(e).__name__,          
            __file__,                  
            e.__traceback__.tb_lineno  
        )
        time.sleep(5)
        quit(driver)
        assert False

    sleep(driver)
    try:
        driver.find_element(By.ID, "keterangan").send_keys(ketereranganFaker)
        Log.info('Input Keterangan')
    except Exception as e:
        Log.info('Input Keterangan Tidak Ditemukan' )
        print('ERROR FOUND',
            type(e).__name__,          
            __file__,                  
            e.__traceback__.tb_lineno  
        )
        time.sleep(5)
        quit(driver)
        assert False

    sleep(driver)
    driver.find_element(By.ID, "createPeserta").click()
    Log.info('Button Tambah Peserta')
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, "buttonFilter")))
    time.sleep(1)
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, "buttonFilter")))
    time.sleep(1)
    driver.find_element(By.CSS_SELECTOR, ""+ PesertaRandomFaker1 +"").click()
    Log.info('Tambah Peserta 1')
    driver.find_element(By.CSS_SELECTOR, ""+ PesertaRandomFaker2 +"").click()
    Log.info('Tambah Peserta 2')
    driver.find_element(By.CSS_SELECTOR, ""+ PesertaRandomFaker3 +"").click()
    Log.info('Tambah Peserta 3')
    driver.find_element(By.ID, "buttonDaftarkan").click()
    Log.info('Click Button Daftarkan')

    sleep(driver)
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, "createProduk")))
    driver.find_element(By.ID, "createProduk").click()
    Log.info('Click Button Create Produk')

    sleep(driver)
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, "idJenisProduk")))
    driver.find_element(By.ID, "idJenisProduk").click()
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ""+idjenisFaker+"")))
    driver.find_element(By.CSS_SELECTOR, ""+idjenisFaker+"").click()
    Log.info('Click Button Jenis Produk')

    sleep(driver)
    driver.find_element(By.ID, "namaProduk").click()
    driver.find_element(By.ID, "namaProduk").send_keys(namaProdukFaker)
    driver.find_element(By.XPATH, "//li[contains(.,\'"+ namaProdukFaker +"')]").click()
    Log.info('Input Nama Produk')

    sleep(driver)
    driver.find_element(By.ID, "jumlah").click()
    driver.find_element(By.ID, "jumlah").send_keys(jumlahprodukFaker)
    Log.info('Input Jumlah Produk')

    sleep(driver)
    driver.find_element(By.ID, "jenisSatuan").click()
    WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, ""+satuanFaker+"")))
    driver.find_element(By.ID, ""+satuanFaker+"").click()
    Log.info('Input Satuan Produk')

    sleep(driver)
    driver.find_element(By.ID, "harga").click()
    driver.find_element(By.ID, "harga").send_keys(HargaFaker)
    Log.info('Input Harga Produk')

    sleep(driver)
    driver.find_element(By.ID, "lamaPengerjaan").click()
    pyautogui.press('backspace')
    driver.find_element(By.ID, "lamaPengerjaan").send_keys(lamaPengerjaanFaker)
    Log.info('Input Lama Pengerjaan Produk')

    sleep(driver)
    driver.find_element(By.CSS_SELECTOR, "#uploadButtonProduk > span").click()
    uploadGambar(driver)
    Log.info('Input Gambar Produk')

    sleep(driver)
    driver.find_element(By.ID, "keteranganProduk").click()
    driver.find_element(By.ID, "keteranganProduk").send_keys(ketereranganFaker)
    Log.info('Input Keterangan Produk')

    sleep(driver)
    driver.find_element(By.CSS_SELECTOR, "#submitButtonProduk > span").click()
    Log.info('Click Button Submit Produk')

    sleep(driver)
    driver.execute_script("window.scrollTo(0,1160)")
    time.sleep(2)
    driver.find_element(By.CSS_SELECTOR, ".el-form").click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#submitButton > span")))
    time.sleep(2)
    driver.find_element(By.ID, "submitButton").click()
    WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, "createButton")))

@pytest.mark.webtest
def test_TC_GIATJA_004():
    sleep(driver)
  
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "filterColumn")))
    driver.find_element(By.ID, "filterColumn").click()
    driver.find_element(By.ID, "semua").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#detail0 .h-5")))
    driver.find_element(By.CSS_SELECTOR, "#detail0 .h-5").click()
    time.sleep(10)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//span[4]/span")))
    driver.find_element(By.XPATH, "//span[4]/span").click()
    Log.info('Click Button Detail')
    
@pytest.mark.webtest
def test_TC_GIATJA_005():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#edit0 .h-5")))
    driver.find_element(By.CSS_SELECTOR, "#edit0 .h-5").click()
    Log.info('Click Button Edit')
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "keterangan")))
    driver.find_element(By.ID, "keterangan").send_keys(ketereranganFaker)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))

@pytest.mark.webtest
def test_TC_GIATJA_006():
    sleep(driver)
    try:

        sleep(driver)
        driver.find_element(By.CSS_SELECTOR, ".el-form-item:nth-child(7) .el-select__input").click()
        WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.ID, ""+saranaFaker+"")))
        driver.find_element(By.ID, ""+saranaFaker+"").click()
        driver.find_element(By.CSS_SELECTOR, " .el-form-item:nth-child(7) > .el-form-item__label").click()

        driver.execute_script("window.scrollTo(0,1160)")
        time.sleep(2)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        time.sleep(2)
        driver.find_element(By.CSS_SELECTOR, "#submitButton svg").click()
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "createButton")))
        Log.info("2.Klik button Ubah")
    except Exception as e:
        driver.execute_script("window.scrollTo(0,0)")
        driver.find_element(By.CSS_SELECTOR, "#backButton").click()
        Log.info('Input Keterangan Tidak Ditemukan' )
        print('ERROR FOUND',
            type(e).__name__,          
            __file__,                  
            e.__traceback__.tb_lineno  
        )
        time.sleep(5)
        assert False


@pytest.mark.webtest
def test_TC_GIATJA_007():
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".el-table__row:nth-child(1) .el-space__item > .el-button")))
    driver.find_element(By.CSS_SELECTOR, ".el-table__row:nth-child(1) .el-space__item > .el-button").click()
    Log.info('Click Button delete')
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(.,'Batalkan')]")))
    driver.find_element(By.XPATH, "/html/body/div[3]/div/div/div[3]/button[2]").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[3]/div/div/div[3]/button[2]")))

@pytest.mark.webtest
def test_EXIT():
    quit(driver)
    Log.info('Exit')







    
    