from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime
import pytest

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("data"))
    file_path = environ.get("fakerGiatja")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
import random

from Settings.setupGiatja import initDriver, loadDataPath, quit, sleep, upload, uploadGambar
from Settings.loginGiatja import KasieGiatja,Op_Giatja,KalapasGiatja
from Settings.Page.Giatja import MenuLaporan
import random
import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('logLaporan.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Pemasaran'

fake = Faker('id_ID')
akunSetor                                               = ['pendapatan','nihil']
pemasaran                                               = ['pemasaran0opt1']
fakeMonth                                               = ['07']

for i in range(5):
    TanggalSetorFaker                                   = random.choice(fakeMonth)
    # TanggalSetorFaker                                   = fake.date_between(start_date='today', end_date='today').strftime('%m')
    akunSetorFaker                                      = random.choice(akunSetor)
    keteranganFaker                                     = fake.text(max_nb_chars=255)
    pemasaranFaker                                      = random.choice(pemasaran)
    NilaiFaker                                          = fake.random_int(min=100000, max=1000000)
    worksheet.append([
        TanggalSetorFaker,
        akunSetorFaker,
        keteranganFaker,
        pemasaranFaker,
        NilaiFaker
        ])
workbook.save(file_path)

workbook = load_workbook(filename=file_path)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=2, values_only=True):
    TanggalSetorFaker                                   = row[0]
    akunSetorFaker                                      = row[1]
    keteranganFaker                                     = row[2]
    pemasaranFaker                                      = row[3]
    NilaiFaker                                          = row[4]

@pytest.mark.webtest
def test_1_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()      
    Op_Giatja(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Setup Os Akses aplikasi SDP')

@pytest.mark.webtest
def test_TC_GIATJA_022():
    sleep(driver)
    MenuLaporan(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Operator mengakses halaman Laporan Kegiatan Kerja dan Produksi')

@pytest.mark.webtest
def test_TC_GIATJA_023():
    sleep(driver)
    driver.implicitly_wait(60)
    WebDriverWait(driver, 70).until(EC.element_to_be_clickable((By.ID, "Aprilsarana3")))
    if TanggalSetorFaker == "01":
        driver.find_element(By.ID, "Januarikegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Januari Kegiatan Kerja')
    elif TanggalSetorFaker == "02":
        driver.find_element(By.ID, "Febuarikegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Febuari Kegiatan Kerja')
    elif TanggalSetorFaker == "03":
        driver.find_element(By.ID, "Maretkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Maret Kegiatan Kerja')
    elif TanggalSetorFaker == "04":
        driver.find_element(By.ID, "Aprilkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('April Kegiatan Kerja')
    elif TanggalSetorFaker == "05":
        driver.find_element(By.ID, "Meikegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Mei Kegiatan Kerja')
     
    elif TanggalSetorFaker == "06":
        driver.find_element(By.ID, "Junikegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Juni Kegiatan Kerja')
     
    elif TanggalSetorFaker == "07":
        driver.find_element(By.ID, "Julikegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Juli Kegiatan Kerja')
     
    elif TanggalSetorFaker == "08":
        driver.find_element(By.ID, "Agustuskegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Agustus Kegiatan Kerja')
     
    elif TanggalSetorFaker == "09":
        driver.find_element(By.ID, "Septemberkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('September Kegiatan Kerja')
     
    elif TanggalSetorFaker == "10":
        driver.find_element(By.ID, "Oktoberkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Oktober Kegiatan Kerja')
     
    elif TanggalSetorFaker == "11":
        driver.find_element(By.ID, "Novemberkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('November Kegiatan Kerja')
     
    elif TanggalSetorFaker == "12":
        driver.find_element(By.ID, "Desemberkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Desember Kegiatan Kerja')

    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//button[contains(.,'OK')]")))
    driver.find_element(By.XPATH, "//button[contains(.,'OK')]").click()
    WebDriverWait(driver, 30).until(EC.invisibility_of_element((By.ID, ".circular")))
     
    Log.info('membuat Laporan Kegiatan Kerja dan Produksi')
    quit(driver)
     

@pytest.mark.webtest
def test_SetupOs_Kasie():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()  
    Log.info('Setup Os Akses aplikasi SDP')


@pytest.mark.webtest
def test_TC_GIATJA_024():    
    KasieGiatja(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Login Kasie')

@pytest.mark.webtest
def test_TC_GIATJA_025():
    sleep(driver)
    MenuLaporan(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Kasie mengakses halaman Laporan Kegiatan Kerja dan Produksi')

@pytest.mark.webtest
def test_TC_GIATJA_026():
    sleep(driver)
    driver.implicitly_wait(60)
    WebDriverWait(driver, 70).until(EC.element_to_be_clickable((By.ID, "Aprilsarana3")))
    if TanggalSetorFaker == "01":
        driver.find_element(By.ID, "Januarikegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Januari Kegiatan Kerja')
    elif TanggalSetorFaker == "02":
        driver.find_element(By.ID, "Febuarikegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Febuari Kegiatan Kerja')
    elif TanggalSetorFaker == "03":
        driver.find_element(By.ID, "Maretkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Maret Kegiatan Kerja')
    elif TanggalSetorFaker == "04":
        driver.find_element(By.ID, "Aprilkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('April Kegiatan Kerja')
    elif TanggalSetorFaker == "05":
        driver.find_element(By.ID, "Meikegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Mei Kegiatan Kerja')
     
    elif TanggalSetorFaker == "06":
        driver.find_element(By.ID, "Junikegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Juni Kegiatan Kerja')
     
    elif TanggalSetorFaker == "07":
        driver.find_element(By.ID, "Julikegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Juli Kegiatan Kerja')
     
    elif TanggalSetorFaker == "08":
        driver.find_element(By.ID, "Agustuskegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Agustus Kegiatan Kerja')
     
    elif TanggalSetorFaker == "09":
        driver.find_element(By.ID, "Septemberkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('September Kegiatan Kerja')
     
    elif TanggalSetorFaker == "10":
        driver.find_element(By.ID, "Oktoberkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Oktober Kegiatan Kerja')
     
    elif TanggalSetorFaker == "11":
        driver.find_element(By.ID, "Novemberkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('November Kegiatan Kerja')
     
    elif TanggalSetorFaker == "12":
        driver.find_element(By.ID, "Desemberkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Desember Kegiatan Kerja')

    Log.info('1. Klik status pada bulan / tahun yang akan di verifikasi')

    

    WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, "//div[2]/span[2]/span")))
    driver.find_element(By.ID, "verifikasi").click()
    Log.info('2. Pada halaman Laporan Kegiatan Kerja dan Produk, klik button "Verifikasi"')

    WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, "cancel")))
    driver.find_element(By.ID, "statuVerifikasi").click()
    WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, "div:nth-child(1) > #verifikasi")))
    driver.find_element(By.XPATH, "div:nth-child(1) > #verifikasi").click()
    Log.info('Pilih status verifikasi"')

    driver.find_element(By.CSS_SELECTOR, ".el-textarea__inner").send_keys("Verifikasi")
    Log.info('input Keterangan ')

    driver.find_element(By.ID, "submitButton").click()
    Log.info('4. Klik button Simpan')

    time.sleep(2)
    Log.info('Kasie melakukan verifikasi pada Laporan Kegiatan Kerja dan Produksi yang telah di generate oleh operator')
    quit(driver)
    

@pytest.mark.webtest
def test_SetupOsKalapas():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()  
    Log.info('Setup Os Akses aplikasi SDP')

@pytest.mark.webtest
def test_TC_GIATJA_027():
    KalapasGiatja(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Login Kalapas')

@pytest.mark.webtest
def test_TC_GIATJA_028():
    sleep(driver)
    MenuLaporan(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Kalapas mengakses halaman Laporan Kegiatan Kerja dan Produksi')


@pytest.mark.webtest
def test_TC_GIATJA_029():
    sleep(driver)
    driver.implicitly_wait(60)
    WebDriverWait(driver, 70).until(EC.element_to_be_clickable((By.ID, "Aprilsarana3")))
    if TanggalSetorFaker == "01":
        driver.find_element(By.ID, "Januarikegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Januari Kegiatan Kerja')
    elif TanggalSetorFaker == "02":
        driver.find_element(By.ID, "Febuarikegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Febuari Kegiatan Kerja')
    elif TanggalSetorFaker == "03":
        driver.find_element(By.ID, "Maretkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Maret Kegiatan Kerja')
    elif TanggalSetorFaker == "04":
        driver.find_element(By.ID, "Aprilkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('April Kegiatan Kerja')
    elif TanggalSetorFaker == "05":
        driver.find_element(By.ID, "Meikegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Mei Kegiatan Kerja')
     
    elif TanggalSetorFaker == "06":
        driver.find_element(By.ID, "Junikegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Juni Kegiatan Kerja')
     
    elif TanggalSetorFaker == "07":
        driver.find_element(By.ID, "Julikegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Juli Kegiatan Kerja')
     
    elif TanggalSetorFaker == "08":
        driver.find_element(By.ID, "Agustuskegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Agustus Kegiatan Kerja')
     
    elif TanggalSetorFaker == "09":
        driver.find_element(By.ID, "Septemberkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('September Kegiatan Kerja')
     
    elif TanggalSetorFaker == "10":
        driver.find_element(By.ID, "Oktoberkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Oktober Kegiatan Kerja')
     
    elif TanggalSetorFaker == "11":
        driver.find_element(By.ID, "Novemberkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('November Kegiatan Kerja')
     
    elif TanggalSetorFaker == "12":
        driver.find_element(By.ID, "Desemberkegiatan-kerja0").click()
        sleep(driver)
        attach(data=driver.get_screenshot_as_png())
        Log.info('Desember Kegiatan Kerja')

    Log.info('1. Klik status pada bulan / tahun yang akan di verifikasi')
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//div[2]/span[2]")))

    driver.find_element(By.CSS_SELECTOR, "#otorisasi > span").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#submitButton > span")))

    Log.info('Kalapas melakukan otorisasi pada Laporan Kegiatan Kerja dan Produksi yang telah di verifikasi oleh kasie')

@pytest.mark.webtest
def test_exit_Operator():
    quit(driver)
    Log.info('Exit')







    
    