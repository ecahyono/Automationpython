from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime
import pytest
from selenium import webdriver

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("data"))
    file_path = environ.get("fakersppn")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
import random

from Settings.SetupAll import *
from Settings.loginPembinaan import *
from Settings.Page.Pembinaan import *
import random
import logging

sheetrangeIndex = wb['listWbpSumedang']

random1 = random.randint(1,9)


PetugasSheet                                 = sheetrangeIndex['E'+str(random1)].value
# print(PetugasSheet)

time.sleep(3)

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'SheetSppn'

fake = Faker('id_ID')

KegiatanKerja   = ['Manufaktur','Jasa']
nums = '01234567891011121415161716'
SkPerwalian = ['opt0','opt2','opt3','opt4','opt5','opt6']
checkboxPeseeta1 = ['#check0 .el-checkbox__inner','#check1 .el-checkbox__inner','#check2 .el-checkbox__inner','#check3 .el-checkbox__inner','#check4 .el-checkbox__inner','#check5 .el-checkbox__inner']
checkboxPeseeta2 = ['#check6 .el-checkbox__inner','#check7 .el-checkbox__inner','#check8 .el-checkbox__inner','#check9 .el-checkbox__inner']
fakeMonth               = ['04']




for i in range(1):
    NoSkFaker                = "Wi"+ fake.isbn10() + ".PASS" + ".PASS" + random.choice(nums) +".PK." + fake.date_between(start_date='today', end_date='today').strftime('%d.%m.%Y') + "-" + random.choice(nums)
    SkPerwalianFaker         = random.choice(SkPerwalian)
    NamaFaker                = fake.name()
    PegawaiFaker             = PetugasSheet
    TempatPenetapanFaker     = fake.city()
    TanggalPenetapanFaker    = fake.date_between(start_date='-30d', end_date='today').strftime('%d.%m.%Y')
    checkboxPeseeta1Faker    = random.choice(checkboxPeseeta1)
    checkboxPeseeta2Faker    = random.choice(checkboxPeseeta2)
    # BulanWbpWaliFaker      = random.choice(fakeMonth)
    BulanWbpWaliFaker                                   = fake.date_between(start_date='today', end_date='today').strftime('%m')
    rekomendasiFaker        = fake.text()

    worksheet.append([
        NoSkFaker,
        SkPerwalianFaker,
        NamaFaker,
        PegawaiFaker,
        TempatPenetapanFaker,
        TanggalPenetapanFaker,
        checkboxPeseeta1Faker,
        checkboxPeseeta2Faker,
        BulanWbpWaliFaker,
        rekomendasiFaker

       
     
        ])
        
workbook.save(file_path)

workbook = load_workbook(filename=file_path)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=1, values_only=True):
    NoSkExcell                            = row[0]
    SkPerwalianExcell                     = row[1]
    NamaExcell                            = row[2]
    PegawaiExcell                         = row[3]
    TempatPenetapanExcell                 = row[4]
    TanggalPenetapanExcell                = row[5]
    checkboxPeseeta1Excell                = row[6]
    checkboxPeseeta2Excell                = row[7]
    BulanWbpWaliExcell                    = row[8]
    rekomendasiExcell                     = row[9]

    


   



global driver, pathData
driver = initDriver()
pathData = loadDataPath()   