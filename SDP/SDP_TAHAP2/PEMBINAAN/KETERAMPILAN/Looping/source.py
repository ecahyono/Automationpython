from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime
import pytest

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    file_path = environ.get("fakerKTR")
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("fakerKTR"))
    wbd = load_workbook(environ.get("data"))

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
from Settings.setupKeterampilan import initDriver, loadDataPath, quit, sleep, upload, uploadGambar
from Settings.loginKeterampilan import * 
from Settings.Page.Keterampilan import *
import random
import logging




workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Keterampilan'

fake = Faker('id_ID')
#test_1
PelatihanKeterampilan   = ['Manufaktur','Jasa','Agribisnis']
tingkatPelatihan        = ['Pemula','Mahir','Lanjutan']
sarana                  = ['saranaOption-0','saranaOption-1','saranaOption-2','saranaOption-3']
prasarana               = ['prasaranaOption-0','prasaranaOption-1','prasaranaOption-2','prasaranaOption-3','prasaranaOption-4']
mitra                   = ['mitra-0','mitra-1','mitra-2','mitra-3','mitra-4']
instruktur              = ['#petugas-0','#instansiLain-0','#mitra-0']
InstrukturOption        = ['instrukturOption-0','instrukturOption-1']
penanggungJawabOption   = ['penanggungJawabOption-0','penanggungJawabOption-1','penanggungJawabOption-2']

#test_6
Predikat                = ['Sangat Baik', 'Baik', 'Cukup', 'Kurang']

fakeMonth                                               = ['05']
       



for i in range(1):
    PelatihanKeterampilanFaker                      = random.choice(PelatihanKeterampilan)
    tingkatPelatihanFaker                           = random.choice(tingkatPelatihan)
    NamaProgramPelatihanFaker                       = fake.text(max_nb_chars=20)
    tempatPelatihanFaker                            = fake.address()
    tanggalPelatihanFaker                           = fake.date_between(start_date='today', end_date='today').strftime('%d/%m/%Y')
    saranaFaker                                     = random.choice(sarana)
    prasaranaFaker                                  = random.choice(prasarana)
    mitraFaker                                      = random.choice(mitra)
    instrukturFaker                                 = random.choice(instruktur)
    InstrukturOptionFaker                           = random.choice(InstrukturOption)
    penanggungJawabOptionFaker                      = random.choice(penanggungJawabOption)
    materiPelatihanFaker                            = fake.text(max_nb_chars=20)
    keteranganFaker                                 = fake.text(max_nb_chars=20)
    # jumlahpesertaFaker                              = random.randint(1,6)
    jumlahpesertaFaker                              = 2
    no_sertifikatFaker0                             = fake.license_plate()
    no_sertifikatFaker1                             = fake.license_plate()
    no_sertifikatFaker2                             = fake.license_plate()
    no_sertifikatFaker3                             = fake.license_plate()
    no_sertifikatFaker4                             = fake.license_plate()
    NilaiFaker0                                     = random.randint(1,10)
    NilaiFaker1                                     = random.randint(1,10)
    NilaiFaker2                                     = random.randint(1,10)
    NilaiFaker3                                     = random.randint(1,10)
    NilaiFaker4                                     = random.randint(1,10)
    PredikatFaker0                                  = random.choice(Predikat)
    PredikatFaker1                                  = random.choice(Predikat)
    PredikatFaker2                                  = random.choice(Predikat)
    PredikatFaker3                                  = random.choice(Predikat)
    PredikatFaker4                                  = random.choice(Predikat)


    no_sertifikatFaker0                             = fake.license_plate()

    BulanFaker                                      = random.choice(fakeMonth)
    KeteranganFaker                                 = fake.text()
# BulanFaker                                   = fake.date_between(start_date='today', end_date='today').strftime('%m')

    


    worksheet.append([
        
        PelatihanKeterampilanFaker,
        tingkatPelatihanFaker,
        NamaProgramPelatihanFaker,
        tempatPelatihanFaker,
        tanggalPelatihanFaker,
        saranaFaker,
        prasaranaFaker,
        mitraFaker,
        instrukturFaker,
        InstrukturOptionFaker,
        penanggungJawabOptionFaker,
        materiPelatihanFaker,
        keteranganFaker,
        jumlahpesertaFaker,
        no_sertifikatFaker0,
        no_sertifikatFaker1,
        no_sertifikatFaker2,
        no_sertifikatFaker3,
        no_sertifikatFaker4,
        NilaiFaker0,
        NilaiFaker1,
        NilaiFaker2,
        NilaiFaker3,
        NilaiFaker4,
        PredikatFaker0,
        PredikatFaker1,
        PredikatFaker2,
        PredikatFaker3,
        PredikatFaker4,
        BulanFaker,
        KeteranganFaker



        
        ])
workbook.save(file_path)

workbook = load_workbook(filename=file_path)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=1, values_only=True):
    PelatihanKeterampilanExcel                                   = row[0]
    tingkatPelatihanExcel                                        = row[1]
    NamaProgramPelatihanExcel                                    = row[2]
    tempatPelatihanExcel                                         = row[3]
    tanggalPelatihanExcel                                        = row[4]
    saranaExcel                                                  = row[5]
    prasaranaExcel                                               = row[6]
    mitraExcel                                                   = row[7]
    instrukturExcel                                              = row[8]
    InstrukturOptionExcel                                        = row[9]
    penanggungJawabOptionExcel                                   = row[10]
    materiPelatihanExcel                                         = row[11]
    keteranganExcel                                              = row[12]
    jumlahpesertaExcel                                           = row[13]
    no_sertifikatExcel0                                          = row[14]
    no_sertifikatExcel1                                          = row[15]
    no_sertifikatExcel2                                          = row[16]
    no_sertifikatExcel3                                          = row[17]
    no_sertifikatExcel4                                          = row[18]
    NilaiExcel0                                                  = row[19]
    NilaiExcel1                                                  = row[20]
    NilaiExcel2                                                  = row[21]
    NilaiExcel3                                                  = row[22]
    NilaiExcel4                                                  = row[23]
    PredikatExcel0                                               = row[24]
    PredikatExcel1                                               = row[25]
    PredikatExcel2                                               = row[26]
    PredikatExcel3                                               = row[27]
    PredikatExcel4                                               = row[28]
    BulanExcel                                                   = row[29]
    KeteranganExcel                                              = row[30]
    

x = 1

sheetrangeIndex = wb['Keterampilan']

NamaKegiatan                                  = sheetrangeIndex['C'+str(x)].value
JumlahPeserta                                 = sheetrangeIndex['N'+str(x)].value
