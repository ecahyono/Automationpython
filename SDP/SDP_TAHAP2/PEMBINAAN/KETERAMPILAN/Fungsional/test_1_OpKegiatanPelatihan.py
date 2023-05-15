from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime
import pytest

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("data"))
    file_path = environ.get("fakerKTR")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
import random

from Settings.setupKeterampilan import initDriver, loadDataPath, quit, sleep, upload, uploadGambar
from Settings.loginKeterampilan import Op_Keterampilan
from Settings.Page.Keterampilan import MenuKegiatanPelatihan,MenuLaporanPelatihanKeterampilan,MenuPresensiKegiatan
import random
import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Log1OpKegiatanPelatihan.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)



sheetrangeIndex = wb['KeterampilanData']

random1 = random.randint(1,9)
random2= random.randint(1,6)
random3 = random.randint(1,5)
peserta1 = random.randint(1,2)
peserta2 = random.randint(1,2)
peserta3 = random.randint(1,2)
peserta4 = random.randint(1,2)
peserta5 = random.randint(1,2)

Manufaktur                                  = sheetrangeIndex['A'+str(random1)].value
jasa                                        = sheetrangeIndex['B'+str(random2)].value
Agribisnis                                  = sheetrangeIndex['C'+str(random3)].value
peserta0                                    = sheetrangeIndex['D'+str(peserta1)].value
peserta1                                    = sheetrangeIndex['E'+str(peserta2)].value
peserta2                                    = sheetrangeIndex['F'+str(peserta3)].value
peserta3                                    = sheetrangeIndex['G'+str(peserta4)].value
peserta4                                    = sheetrangeIndex['H'+str(peserta5)].value
peserta5                                    = sheetrangeIndex['I'+str(peserta5)].value

print('nama input WBP adalah', Manufaktur, jasa, Agribisnis)
time.sleep(3)


workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Keterampilan'

fake = Faker('id_ID')

PelatihanKeterampilan   = ['Manufaktur','Jasa','Agribisnis']
tingkatPelatihan        = ['Pemula','Mahir','Lanjutan']
sarana                  = ['saranaOption-0','saranaOption-1','saranaOption-2','saranaOption-3']
prasarana               = ['prasaranaOption-0','prasaranaOption-1','prasaranaOption-2','prasaranaOption-3','prasaranaOption-4']
mitra                   = ['mitra-0','mitra-1','mitra-2','mitra-3','mitra-4']
instruktur              = ['#petugas-0','#instansiLain-0','#mitra-0']
InstrukturOption        = ['instrukturOption-0','instrukturOption-1']
penanggungJawabOption   = ['penanggungJawabOption-0','penanggungJawabOption-1','penanggungJawabOption-2']
       



for i in range(5):
    PelatihanKeterampilanFaker                      = random.choice(PelatihanKeterampilan)
    tingkatPelatihanFaker                           = random.choice(tingkatPelatihan)
    NamaProgramPelatihanFaker                       = fake.text(max_nb_chars=20)
    tempatPelatihanFaker                            = fake.address()
    tanggalPelatihanFaker                           = fake.date_between(start_date='today', end_date='today').strftime('%d/%m/%Y')
    saranaFaker                                     = random.choice(sarana)
    prasaranaFaker                                  = random.choice(prasarana)
    mitraFaker                                      = random.choice(mitra)
    instrukturFaker                                 = random.choice(instruktur)
    InstrukturOptionFaker                           = random.choice(InstrukturOption)
    penanggungJawabOptionFaker                      = random.choice(penanggungJawabOption)
    materiPelatihanFaker                            = fake.text(max_nb_chars=20)
    keteranganFaker                                 = fake.text(max_nb_chars=20)
    # jumlahpesertaFaker                              = random.randint(1,6)
    jumlahpesertaFaker                              = 2


    worksheet.append([
        
        PelatihanKeterampilanFaker,
        tingkatPelatihanFaker,
        NamaProgramPelatihanFaker,
        tempatPelatihanFaker,
        tanggalPelatihanFaker,
        saranaFaker,
        prasaranaFaker,
        mitraFaker,
        instrukturFaker,
        InstrukturOptionFaker,
        penanggungJawabOptionFaker,
        materiPelatihanFaker,
        keteranganFaker,
        jumlahpesertaFaker


        
        ])
workbook.save(file_path)

workbook = load_workbook(filename=file_path)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=2, values_only=True):
    PelatihanKeterampilanExcel                                   = row[0]
    tingkatPelatihanExcel                                        = row[1]
    NamaProgramPelatihanExcel                                    = row[2]
    tempatPelatihanExcel                                         = row[3]
    tanggalPelatihanExcel                                        = row[4]
    saranaExcel                                                  = row[5]
    prasaranaExcel                                               = row[6]
    mitraExcel                                                   = row[7]
    instrukturExcel                                              = row[8]
    InstrukturOptionExcel                                        = row[9]
    penanggungJawabOptionExcel                                   = row[10]
    materiPelatihanExcel                                         = row[11]




@pytest.mark.webtest
def test7_SetupOs_OpKegiatanPelatihan():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')

@pytest.mark.webtest
def test_TC_KTR_001():      
    Op_Keterampilan(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Login aplikasi menggunakan akun dengan role operator')

@pytest.mark.webtest
def test_TC_KTR_002():
    sleep(driver)
    MenuKegiatanPelatihan(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Operator mengakses halaman Program Pelatihan Keterampilan')

@pytest.mark.webtest
def test_TC_KTR_003():
    driver.implicitly_wait(60)
    d = driver.find_element
    wait = WebDriverWait(driver, 50)
    sleep(driver)
    wait.until(EC.element_to_be_clickable((By.ID, "createButton")))
    d(By.ID, "createButton").click()
    Log.info('Operator mengklik tombol Tambah Program Pelatihan Keterampilan')

    wait.until(EC.element_to_be_clickable((By.ID, "backButton")))
    attach(data=driver.get_screenshot_as_png())

    wait.until(EC.element_to_be_clickable((By.ID, "jenisPelatihan")))
    d(By.ID, "jenisPelatihan").click()
    d(By.XPATH, "//li[contains(.,\'"+ PelatihanKeterampilanFaker +"')]").click()
    d(By.ID, "bidangPelatihan").click()
    if PelatihanKeterampilanExcel == 'Manufaktur':
        wait.until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'"+ Manufaktur +"')]")))
        d(By.XPATH, "//li[contains(.,\'"+ Manufaktur +"')]").click()
        Log.info('Operator memilih jenis pelatihan keterampilan Manufaktur')
    elif PelatihanKeterampilanExcel == 'Jasa':
        sleep(driver)
        wait.until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'"+ jasa +"')]")))
        d(By.XPATH, "//li[contains(.,\'"+ jasa +"')]").click()
        Log.info('Operator memilih jenis pelatihan keterampilan Jasa')
    elif PelatihanKeterampilanExcel == 'Agribisnis':
        sleep(driver)
        wait.until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'"+ Agribisnis +"')]")))
        d(By.XPATH, "//li[contains(.,\'"+ Agribisnis +"')]").click()
        Log.info('Operator memilih jenis pelatihan keterampilan Agribisnis')

    d(By.ID, "tingkatPelatihan").click()
    wait.until(EC.element_to_be_clickable((By.XPATH, "//li[contains(.,\'"+ tingkatPelatihanFaker +"')]")))
    d(By.XPATH, "//li[contains(.,\'"+ tingkatPelatihanFaker +"')]").click()
    Log.info('Operator memilih tingkat pelatihan keterampilan ')

    d(By.ID, "namaProgram").send_keys(NamaProgramPelatihanFaker)
    Log.info('Operator mengisi nama program pelatihan keterampilan ')

    d(By.ID, "tempatPelaksanaan").send_keys(tempatPelatihanFaker)
    Log.info('Operator mengisi tempat pelatihan keterampilan ')

    d(By.ID, "waktuPelaksanaan").click()
    d(By.ID, "tanggalEmpty").send_keys(tanggalPelatihanFaker)
    Log.info('Operator mengisi waktu pelatihan keterampilan ')

    d(By.ID, "j").click()
    Log.info('Operator mengisi jam pelatihan keterampilan ')

    d(By.CSS_SELECTOR, ".flex:nth-child(5) > #submitButton > span").click()

    d(By.CSS_SELECTOR, ".el-form-item:nth-child(8) .el-select__input").click()
    time.sleep(1)
    wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@id=\'"+saranaFaker+"\']")))
    d(By.XPATH, "//li[@id=\'"+saranaFaker+"\']").click()
    Log.info('Operator memilih sarana pelatihan keterampilan ')

    d(By.CSS_SELECTOR, ".el-form-item:nth-child(9) .el-select__input").click()
    time.sleep(1)
    wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@id=\'"+prasaranaFaker+"\']")))
    d(By.XPATH, "//li[@id=\'"+prasaranaFaker+"\']").click()
    Log.info('Operator memilih prasarana pelatihan keterampilan ')

    d(By.CSS_SELECTOR, "#mitra").click()
    time.sleep(1)
    wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@id=\'"+mitraFaker+"\']")))
    d(By.XPATH, "//li[@id=\'"+mitraFaker+"\']").click()
    Log.info('Operator memilih mitra pelatihan keterampilan ')

    driver.execute_script("window.scrollTo(0,132)")

    d(By.ID, "instruktur-0").click()
    time.sleep(2)
    d(By.ID, "petugas-0").click()
    Log.info('Operator memilih instruktur pelatihan keterampilan ')

    d(By.ID, "instrukturSelect").click()
    time.sleep(2)
    wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@id=\'"+InstrukturOptionFaker+"\']")))
    d(By.XPATH, "//li[@id=\'"+InstrukturOptionFaker+"\']").click()
    Log.info('Operator memilih instruktur pelatihan keterampilan ')

    driver.execute_script("window.scrollTo(0,132)")

    d(By.ID, "penanggungJawan").click()
    wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@id=\'"+penanggungJawabOptionFaker+"\']")))
    d(By.XPATH, "//li[@id=\'"+penanggungJawabOptionFaker+"\']").click()
    Log.info('Operator memilih penanggung jawab pelatihan keterampilan ')

    d(By.ID, "materi").send_keys(materiPelatihanFaker)
    Log.info('Operator mengisi materi pelatihan keterampilan ')

    d(By.ID, "keterangan").send_keys(keteranganFaker)
    Log.info('Operator mengisi keterangan pelatihan keterampilan ')
    
    time.sleep(2)
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#createButtonPeserta > span")))
    d(By.CSS_SELECTOR, "#createButtonPeserta > span").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(.,'Pilih')]")))
    if jumlahpesertaFaker == 1:
        d(By.CSS_SELECTOR, ""+peserta0+"").click()
        Log.info('Operator memilih 1 peserta pelatihan keterampilan ')
    elif jumlahpesertaFaker == 2:
        d(By.CSS_SELECTOR, ""+peserta0+"").click()
        d(By.CSS_SELECTOR, ""+peserta1+"").click()
        Log.info('Operator memilih 2 peserta pelatihan keterampilan ')
    elif jumlahpesertaFaker == 3:
        d(By.CSS_SELECTOR, ""+peserta0+"").click()
        d(By.CSS_SELECTOR, ""+peserta1+"").click()
        d(By.CSS_SELECTOR, ""+peserta2+"").click()
        Log.info('Operator memilih 3 peserta pelatihan keterampilan ')
    elif jumlahpesertaFaker == 4:
        d(By.CSS_SELECTOR, ""+peserta0+"").click()
        d(By.CSS_SELECTOR, ""+peserta1+"").click()
        d(By.CSS_SELECTOR, ""+peserta2+"").click()
        d(By.CSS_SELECTOR, ""+peserta3+"").click()
        Log.info('Operator memilih 4 peserta pelatihan keterampilan ')
    elif jumlahpesertaFaker == 5:
        d(By.CSS_SELECTOR, ""+peserta0+"").click()
        d(By.CSS_SELECTOR, ""+peserta1+"").click()
        d(By.CSS_SELECTOR, ""+peserta2+"").click()
        d(By.CSS_SELECTOR, ""+peserta3+"").click()
        d(By.CSS_SELECTOR, ""+peserta4+"").click()
        Log.info('Operator memilih 5 peserta pelatihan keterampilan ')
    d(By.XPATH, "//span[contains(.,'Pilih')]").click()
    
    d(By.ID, "submitButton").click()
    Log.info('Operator menekan tombol simpan pelatihan keterampilan ')

    wait.until(EC.element_to_be_clickable((By.ID, "createButton")))
    Log.info('Operator menambahkan Program Pelatihan Keterampilan')

@pytest.mark.webtest
def test_TC_KTR_004():
    driver.implicitly_wait(60)
    
    sleep(driver)
    d = driver.find_element

    wait = WebDriverWait(driver, 20) 
    wait.until(EC.element_to_be_clickable((By.ID, "createButton"))) 
    wait.until(EC.element_to_be_clickable((By.ID, "buttonSearch")))

    # d(By.ID, 'detail-0').click()
    # wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".circular")))
    # time.sleep(5)
    # d(By.ID, "backButton").click()

    attach(data=driver.get_screenshot_as_png())

    Log.info('Operator mengakses halaman Detail Program Pelatihan Keterampilan')

@pytest.mark.webtest
def test_TC_KTR_005():
    driver.implicitly_wait(60)
    sleep(driver)
    d = driver.find_element
    wait = WebDriverWait(driver, 20) 
    wait.until(EC.element_to_be_clickable((By.ID, "createButton"))) 
    wait.until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    time.sleep(1)
    d(By.ID, "update-0").click()
    Log.info('Operator mengakses halaman Ubah Program Pelatihan Keterampilan')


@pytest.mark.webtest
def test_TC_KTR_006():
    driver.implicitly_wait(60)
    sleep(driver)
    d = driver.find_element
    wait = WebDriverWait(driver, 60)
    wait.until(EC.element_to_be_clickable((By.ID, "keterangan")))
    driver.execute_script("window.scrollTo(0,972)")
    d(By.CSS_SELECTOR, '#keterangan').send_keys(keteranganFaker)
    wait.until(EC.element_to_be_clickable((By.ID, "submitButton")))
    d(By.ID, "submitButton").click()
    wait.until(EC.element_to_be_clickable((By.ID, "createButton"))) 
    Log.info('Operator mengubah data Program Pelatihan Keterampilan')

@pytest.mark.webtest
def test_TC_KTR_007():
    sleep(driver)
    # d = driver.find_element   
    # wait = WebDriverWait(driver, 20)
    # wait.until(EC.element_to_be_clickable((By.ID, "createButton"))) 
    # wait.until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    # wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "h-5 w-5")))
    Log.info('Operator menghapus data Program Pelatihan Keterampilan')




@pytest.mark.webtest
def test1_exit_OpKegiatanPelatihan():
    quit(driver)
    Log.info('Exit')







    
    