from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime
import pytest

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("fakerKTR"))
    file_path = environ.get("fakerKTR")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
import random

from Settings.setupKeterampilan import initDriver, loadDataPath, quit, sleep, upload, uploadGambar
from Settings.loginKeterampilan import KalapasKeterampilan
from Settings.Page.Keterampilan import VerifikasiKelulusanPesertaKegiatan
import random
import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Log7KasieVerifikasiKelulusanPeserta.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

sheetrangeIndex = wb['Keterampilan']

i = 5
NamaKegiatan                                  = sheetrangeIndex['C'+str(i)].value
JumlahPeserta                                 = sheetrangeIndex['N'+str(i)].value
print(NamaKegiatan)



workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Keterampilan'

fake = Faker('id_ID')
Predikat = ['Sangat Baik', 'Baik', 'Cukup', 'Kurang']
       

for i in range(5):
    no_sertifikatFaker0                    = fake.license_plate()
    PredikatFaker0                         = random.choice(Predikat)

    
@pytest.mark.webtest
def test1_KasieVerifikasiKelulusanPeserta():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')

@pytest.mark.webtest
def test_TC_KTR_027():      
    KalapasKeterampilan(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Login aplikasi menggunakan akun dengan role Kalapas')

@pytest.mark.webtest
def test_TC_KTR_028():
    sleep(driver)
    VerifikasiKelulusanPesertaKegiatan(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Kalapas mengakses halaman Verifikasi Kelulusan Peserta Kegiatan')

    