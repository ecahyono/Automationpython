from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime
import pytest

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("fakerKTR"))
    file_path = environ.get("fakerKTR")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
import random

from Settings.setupKeterampilan import initDriver, loadDataPath, quit, sleep, upload, uploadGambar
from Settings.loginKeterampilan import Op_Keterampilan
from Settings.Page.Keterampilan import MenuPresensiKegiatan
import random
import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Log6OpkelulusanPeserta.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

sheetrangeIndex = wb['Keterampilan']

i = 5
NamaKegiatan                                  = sheetrangeIndex['C'+str(i)].value
JumlahPeserta                                 = sheetrangeIndex['N'+str(i)].value
print(NamaKegiatan)



workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Keterampilan'

fake = Faker('id_ID')
Predikat = ['Sangat Baik', 'Baik', 'Cukup', 'Kurang']
       



for i in range(5):
    no_sertifikatFaker0                    = fake.license_plate()
    no_sertifikatFaker1                    = fake.license_plate()
    no_sertifikatFaker2                    = fake.license_plate()
    no_sertifikatFaker3                    = fake.license_plate()
    no_sertifikatFaker4                    = fake.license_plate()
    NilaiFaker0                             = random.randint(1,10)
    NilaiFaker1                            = random.randint(1,10)
    NilaiFaker2                             = random.randint(1,10)
    NilaiFaker3                             = random.randint(1,10)
    NilaiFaker4                             = random.randint(1,10)
    PredikatFaker0                          = random.choice(Predikat)
    PredikatFaker1                          = random.choice(Predikat)
    PredikatFaker2                          = random.choice(Predikat)
    PredikatFaker3                          = random.choice(Predikat)
    PredikatFaker4                          = random.choice(Predikat)


    
@pytest.mark.webtest
def test1_OpKegiatanPelatihan_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')

@pytest.mark.webtest
def test_TC_KTR_024():      
    Op_Keterampilan(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Login aplikasi menggunakan akun dengan role operator')

@pytest.mark.webtest
def test_TC_KTR_025():
    sleep(driver)
    MenuPresensiKegiatan(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Operator mengakses halaman Program Pelatihan Keterampilan')

@pytest.mark.webtest
def test_TC_KTR_026():
    sleep(driver)
    driver.implicitly_wait(30)
    attach(data=driver.get_screenshot_as_png())
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    driver.find_element(By.ID, "filterColumn").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "nama_program")))
    driver.find_element(By.ID, "nama_program").click()
    driver.find_element(By.ID, "kataKunci").send_keys(NamaKegiatan)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    driver.find_element(By.ID, "buttonSearch").click()

    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".h-5")))
    time.sleep(5)
    driver.find_element(By.CSS_SELECTOR, ".h-5").click()
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//span[contains(.,'Catat Kelulusan')]")))

    if JumlahPeserta == 1:  
        driver.find_element(By.ID, "catatKelulusan-0").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "no_sertifikat")))
        driver.find_element(By.ID, "no_sertifikat").clear()
        driver.find_element(By.ID, "no_sertifikat").send_keys(no_sertifikatFaker0)
        driver.find_element(By.ID, 'pilihFile').click()
        uploadGambar(driver)
        driver.find_element(By.ID, 'nilai').send_keys(NilaiFaker0)
        driver.find_element(By.ID, 'predikat').send_keys(PredikatFaker0)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        driver.find_element(By.ID, 'submitButton').click()
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        time.sleep(2)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#detailKelulusan-0 > span")))
        
    elif JumlahPeserta == 2:
        driver.find_element(By.ID, "catatKelulusan-0").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "no_sertifikat")))
        driver.find_element(By.ID, "no_sertifikat").clear()
        driver.find_element(By.ID, "no_sertifikat").send_keys(no_sertifikatFaker0)
        driver.find_element(By.ID, 'pilihFile').click()
        uploadGambar(driver)
        driver.find_element(By.ID, 'nilai').send_keys(NilaiFaker0)
        driver.find_element(By.ID, 'predikat').send_keys(PredikatFaker0)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        driver.find_element(By.ID, 'submitButton').click()
        time.sleep(10)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(.,'Catat Kelulusan')]")))
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#detailKelulusan-0 > span")))
        
        driver.find_element(By.ID, "catatKelulusan-1").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "no_sertifikat")))
        driver.find_element(By.ID, "no_sertifikat").clear()
        driver.find_element(By.ID, "no_sertifikat").send_keys(no_sertifikatFaker1)
        driver.find_element(By.ID, 'pilihFile').click()
        uploadGambar(driver)
        driver.find_element(By.ID, 'nilai').send_keys(NilaiFaker1)
        driver.find_element(By.ID, 'predikat').send_keys(PredikatFaker1)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        driver.find_element(By.ID, 'submitButton').click()
        time.sleep(10)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#detailKelulusan-1 > span")))
        
    elif JumlahPeserta == 3:
        driver.find_element(By.ID, "catatKelulusan-0").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "no_sertifikat")))
        driver.find_element(By.ID, "no_sertifikat").clear()
        driver.find_element(By.ID, "no_sertifikat").send_keys(no_sertifikatFaker0)
        driver.find_element(By.ID, 'pilihFile').click()
        uploadGambar(driver)
        driver.find_element(By.ID, 'nilai').send_keys(NilaiFaker0)
        driver.find_element(By.ID, 'predikat').send_keys(PredikatFaker0)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        driver.find_element(By.ID, 'submitButton').click()
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))

    
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(.,'Catat Kelulusan')]")))
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#detailKelulusan-0 > span")))
        

        driver.find_element(By.ID, "catatKelulusan-1").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "no_sertifikat")))
        driver.find_element(By.ID, "no_sertifikat").clear()
        driver.find_element(By.ID, "no_sertifikat").send_keys(no_sertifikatFaker1)
        driver.find_element(By.ID, 'pilihFile').click()
        uploadGambar(driver)
        driver.find_element(By.ID, 'nilai').send_keys(NilaiFaker1)
        driver.find_element(By.ID, 'predikat').send_keys(PredikatFaker1)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        driver.find_element(By.ID, 'submitButton').click()
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        time.sleep(2)
    
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(.,'Catat Kelulusan')]")))
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#detailKelulusan-1 > span")))
        
        driver.find_element(By.ID, "catatKelulusan-2").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "no_sertifikat")))
        driver.find_element(By.ID, "no_sertifikat").clear()
        driver.find_element(By.ID, "no_sertifikat").send_keys(no_sertifikatFaker2)
        driver.find_element(By.ID, 'pilihFile').click()
        uploadGambar(driver)
        driver.find_element(By.ID, 'nilai').send_keys(NilaiFaker2)
        driver.find_element(By.ID, 'predikat').send_keys(PredikatFaker2)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        driver.find_element(By.ID, 'submitButton').click()
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        time.sleep(2)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#detailKelulusan-2 > span")))
        
    elif JumlahPeserta == 4:
        driver.find_element(By.ID, "catatKelulusan-0").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "no_sertifikat")))
        driver.find_element(By.ID, "no_sertifikat").clear()
        driver.find_element(By.ID, "no_sertifikat").send_keys(no_sertifikatFaker0)
        driver.find_element(By.ID, 'pilihFile').click()
        uploadGambar(driver)
        driver.find_element(By.ID, 'nilai').send_keys(NilaiFaker0)
        driver.find_element(By.ID, 'predikat').send_keys(PredikatFaker0)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        driver.find_element(By.ID, 'submitButton').click()
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        time.sleep(2)
    
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(.,'Catat Kelulusan')]")))
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#detailKelulusan-0 > span")))
        
        
        driver.find_element(By.ID, "catatKelulusan-1").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "no_sertifikat")))
        driver.find_element(By.ID, "no_sertifikat").clear()
        driver.find_element(By.ID, "no_sertifikat").send_keys(no_sertifikatFaker1)
        driver.find_element(By.ID, 'pilihFile').click()
        uploadGambar(driver)
        driver.find_element(By.ID, 'nilai').send_keys(NilaiFaker1)
        driver.find_element(By.ID, 'predikat').send_keys(PredikatFaker1)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        driver.find_element(By.ID, 'submitButton').click()
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        time.sleep(2)
    
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(.,'Catat Kelulusan')]")))
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#detailKelulusan-1 > span")))
        
        
        driver.find_element(By.ID, "catatKelulusan-2").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "no_sertifikat")))
        driver.find_element(By.ID, "no_sertifikat").clear()
        driver.find_element(By.ID, "no_sertifikat").send_keys(no_sertifikatFaker2)
        driver.find_element(By.ID, 'pilihFile').click()
        uploadGambar(driver)
        driver.find_element(By.ID, 'nilai').send_keys(NilaiFaker2)
        driver.find_element(By.ID, 'predikat').send_keys(PredikatFaker2)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        driver.find_element(By.ID, 'submitButton').click()
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        time.sleep(2)
    
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(.,'Catat Kelulusan')]")))
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#detailKelulusan-2 > span")))
        
        
        driver.find_element(By.ID, "catatKelulusan-3").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "no_sertifikat")))
        driver.find_element(By.ID, "no_sertifikat").clear()
        driver.find_element(By.ID, "no_sertifikat").send_keys(no_sertifikatFaker3)
        driver.find_element(By.ID, 'pilihFile').click()
        uploadGambar(driver)
        driver.find_element(By.ID, 'nilai').send_keys(NilaiFaker3)
        driver.find_element(By.ID, 'predikat').send_keys(PredikatFaker3)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        driver.find_element(By.ID, 'submitButton').click()
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        time.sleep(2)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#detailKelulusan-0 > span")))
        

    elif JumlahPeserta == 5:
        driver.find_element(By.ID, "catatKelulusan-0").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "no_sertifikat")))
        driver.find_element(By.ID, "no_sertifikat").clear()
        driver.find_element(By.ID, "no_sertifikat").send_keys(no_sertifikatFaker0)
        driver.find_element(By.ID, 'pilihFile').click()
        uploadGambar(driver)
        driver.find_element(By.ID, 'nilai').send_keys(NilaiFaker0)
        driver.find_element(By.ID, 'predikat').send_keys(PredikatFaker0)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        driver.find_element(By.ID, 'submitButton').click()
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        time.sleep(2)
    
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(.,'Catat Kelulusan')]")))
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#detailKelulusan-0 > span")))
        
        
        driver.find_element(By.ID, "catatKelulusan-1").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "no_sertifikat")))
        driver.find_element(By.ID, "no_sertifikat").clear()
        driver.find_element(By.ID, "no_sertifikat").send_keys(no_sertifikatFaker1)
        driver.find_element(By.ID, 'pilihFile').click()
        uploadGambar(driver)
        driver.find_element(By.ID, 'nilai').send_keys(NilaiFaker1)
        driver.find_element(By.ID, 'predikat').send_keys(PredikatFaker1)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        driver.find_element(By.ID, 'submitButton').click()
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        time.sleep(2)
    
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(.,'Catat Kelulusan')]")))
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#detailKelulusan-1 > span")))
        
        
        driver.find_element(By.ID, "catatKelulusan-2").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "no_sertifikat")))
        driver.find_element(By.ID, "no_sertifikat").clear()
        driver.find_element(By.ID, "no_sertifikat").send_keys(no_sertifikatFaker2)
        driver.find_element(By.ID, 'pilihFile').click()
        uploadGambar(driver)
        driver.find_element(By.ID, 'nilai').send_keys(NilaiFaker2)
        driver.find_element(By.ID, 'predikat').send_keys(PredikatFaker2)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        driver.find_element(By.ID, 'submitButton').click()
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        time.sleep(2)
    
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(.,'Catat Kelulusan')]")))
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#detailKelulusan-2 > span")))
        
        
        driver.find_element(By.ID, "catatKelulusan-3").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "no_sertifikat")))
        driver.find_element(By.ID, "no_sertifikat").clear()
        driver.find_element(By.ID, "no_sertifikat").send_keys(no_sertifikatFaker3)
        driver.find_element(By.ID, 'pilihFile').click()
        uploadGambar(driver)
        driver.find_element(By.ID, 'nilai').send_keys(NilaiFaker3)
        driver.find_element(By.ID, 'predikat').send_keys(PredikatFaker3)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        driver.find_element(By.ID, 'submitButton').click()
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        time.sleep(2)
    
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(.,'Catat Kelulusan')]")))
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#detailKelulusan-3 > span")))
        
        
        driver.find_element(By.ID, "catatKelulusan-4").click()
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "no_sertifikat")))
        driver.find_element(By.ID, "no_sertifikat").clear()
        driver.find_element(By.ID, "no_sertifikat").send_keys(no_sertifikatFaker4)
        driver.find_element(By.ID, 'pilihFile').click()
        uploadGambar(driver)
        driver.find_element(By.ID, 'nilai').send_keys(NilaiFaker4)
        driver.find_element(By.ID, 'predikat').send_keys(PredikatFaker4)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        driver.find_element(By.ID, 'submitButton').click()
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "submitButton")))
        time.sleep(2)
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#detailKelulusan-4 > span")))
        
        
    
    Log.info('Operator mengklik tab Kegiatan Pelatihan')

@pytest.mark.webtest
def test_exit_1():
    sleep(driver)
    quit(driver)
    Log.info('Operator mengklik tombol Keluar')