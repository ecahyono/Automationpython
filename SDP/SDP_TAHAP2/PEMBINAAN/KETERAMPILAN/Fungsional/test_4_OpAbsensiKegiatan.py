from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime
import pytest

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("fakerKTR"))
    file_path = environ.get("fakerKTR")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
import random

from Settings.setupKeterampilan import initDriver, loadDataPath, quit, sleep, upload, uploadGambar
from Settings.loginKeterampilan import Op_Keterampilan
from Settings.Page.Keterampilan import MenuPresensiKegiatan
import random
import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Log4OpAbsensiKegiatan.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

sheetrangeIndex = wb['Keterampilan']

i = 5
NamaKegiatan                                  = sheetrangeIndex['C'+str(i)].value
JumlahPeserta                                 = sheetrangeIndex['N'+str(i)].value
print(NamaKegiatan)

@pytest.mark.webtest
def test4_SetupOs__OpAbsensiKegiatan():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')

@pytest.mark.webtest
def test_TC_KTR_016():      
    Op_Keterampilan(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Login aplikasi menggunakan akun dengan role operator')

@pytest.mark.webtest
def test_TC_KTR_017():
    sleep(driver)
    MenuPresensiKegiatan(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Operator mengakses halaman Program Pelatihan Keterampilan')

@pytest.mark.webtest
def test_TC_KTR_018():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "buttonSearch")))
    driver.find_element(By.ID, "filterColumn").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "nama_program")))
    driver.find_element(By.ID, "nama_program").click()

    driver.find_element(By.ID, "kataKunci").send_keys(NamaKegiatan)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    driver.find_element(By.ID, "buttonSearch").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    time.sleep(1)

    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#absensi-0 > span")))
    time.sleep(2)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    driver.find_element(By.CSS_SELECTOR, "#absensi-0 > span").click()
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "backButton")))

    driver.find_element(By.CSS_SELECTOR, ".el-select__input").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "pesertaOption-0")))

    if JumlahPeserta == 1:  
        driver.find_element(By.ID, "pesertaOption-0").click()
    elif JumlahPeserta == 2:
        driver.find_element(By.ID, "pesertaOption-0").click()
        driver.find_element(By.ID, "pesertaOption-1").click()
    elif JumlahPeserta == 3:
        driver.find_element(By.ID, "pesertaOption-0").click()
        driver.find_element(By.ID, "pesertaOption-1").click()
        driver.find_element(By.ID, "pesertaOption-2").click()
    elif JumlahPeserta == 4:
        driver.find_element(By.ID, "pesertaOption-0").click()
        driver.find_element(By.ID, "pesertaOption-1").click()
        driver.find_element(By.ID, "pesertaOption-2").click()
        driver.find_element(By.ID, "pesertaOption-3").click()
    elif JumlahPeserta == 5:
        driver.find_element(By.ID, "pesertaOption-0").click()
        driver.find_element(By.ID, "pesertaOption-1").click()
        driver.find_element(By.ID, "pesertaOption-2").click()
        driver.find_element(By.ID, "pesertaOption-3").click()
        driver.find_element(By.ID, "pesertaOption-4").click()

    driver.find_element(By.CSS_SELECTOR, ".flex:nth-child(2) > #submitButton").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//p[contains(.,'Berhasil Ditambahkan')]")))
    driver.find_element(By.CSS_SELECTOR, ".mt-6 > #submitButton").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))


    attach(data=driver.get_screenshot_as_png())
    Log.info('Operator mencatat absensi peserta kegiatan')


@pytest.mark.webtest
def test_TC_KTR_019():
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "buttonSearch")))
    driver.find_element(By.ID, "filterColumn").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "nama_program")))
    driver.find_element(By.ID, "nama_program").click()

    driver.find_element(By.ID, "kataKunci").send_keys(NamaKegiatan)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    driver.find_element(By.ID, "buttonSearch").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    time.sleep(1)

    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#absensi-0 > span")))
    time.sleep(2)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#detail-0 .h-5")))
    driver.find_element(By.CSS_SELECTOR, "#detail-0 .h-5").click()

    driver.find_element(By.CSS_SELECTOR, ".el-breadcrumb__item:nth-child(4) > .el-breadcrumb__inner").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "buttonSearch")))

    Log.info('Operator mengakses halaman Detail Absensi Kegiatan')
@pytest.mark.webtest
def test_4_exit_OpAbsensiKegiatan():
    quit(driver)
    Log.info('Exit')







    
    