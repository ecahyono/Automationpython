from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime
import pytest

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("fakerKTR"))
    file_path = environ.get("fakerKTR")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
import random

from Settings.setupKeterampilan import initDriver, loadDataPath, quit, sleep, upload, uploadGambar
from Settings.loginKeterampilan import kasieKeterampilan
from Settings.Page.Keterampilan import MenuPersetujuanPesertaKegiatan
import random
import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Log2_KasieVerifikasiPesertaKegiatan.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)



sheetrangeIndex = wb['Keterampilan']

i = 5
NamaKegiatan                                  = sheetrangeIndex['C'+str(i)].value
print(NamaKegiatan)

@pytest.mark.webtest
def test7_SetupOs_KasieVerifikasiPesertaKegiatan():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')

@pytest.mark.webtest
def test_TC_KTR_008():      
    kasieKeterampilan(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('"Berhasil login dan menu yang ditampilkan sesuai hak akses role kasie"')

@pytest.mark.webtest
def test_TC_KTR_009():
    sleep(driver)
    MenuPersetujuanPesertaKegiatan(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Kasie mengakses halaman Persetujuan Peserta Kegiatan')

@pytest.mark.webtest
def test_TC_KTR_011():
    sleep(driver)
    driver.implicitly_wait(50)
    d = driver.find_element
    wait = WebDriverWait(driver, 30)
    wait.until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    d(By.ID, "filterColumn").click()
    d(By.CSS_SELECTOR, "#nama_program > span").click()
    
    d(By.ID, "kataunci").send_keys(NamaKegiatan)
    d(By.ID, "buttonSearch").click()
    time.sleep(1)
    wait.until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#verifikasi-0 path:nth-child(2)")))

    d(By.CSS_SELECTOR, "#verifikasi-0 path:nth-child(2)").click()
    Log.info('Click Action Verifikasi')

    wait.until(EC.element_to_be_clickable((By.ID, "chooseVerifikasi")))
    d(By.ID, "chooseVerifikasi").click()
    wait.until(EC.element_to_be_clickable((By.ID, "verifikasi")))
    time.sleep(1)
    d(By.ID, "verifikasi").click()
    Log.info ('Click Verifikasi')

    d(By.ID, "keterangan").send_keys("Verifikasi Peserta Kegiatan")
    Log.info('Input Keterangan')
    wait.until(EC.element_to_be_clickable((By.ID, "submitButton")))
    d(By.ID, "submitButton").click()
    Log.info('Click Submit')
    wait.until(EC.element_to_be_clickable((By.XPATH, "//p[contains(.,'Berhasil Memverifikasi Data')]")))
    Log.info('Kasie melakukan verifikasi peserta kegiatan pada kegiatan yang telah ditambahkan oleh operator')

@pytest.mark.webtest
def test_TC_KTR_010():
    sleep(driver)
    d=driver.find_element
    wait = WebDriverWait(driver, 30)
    wait.until(EC.element_to_be_clickable((By.ID, "buttonSearch")))
    d(By.CSS_SELECTOR, ".text-blue-500 .h-5").click()
    wait.until(EC.element_to_be_clickable((By.ID, "backButton")))
    
    Log.info('Kasie mengakses halaman Detail Program Pelatihan Keterampilan')

@pytest.mark.webtest
def test_exit2__KasieVerifikasiPesertaKegiatan():
    quit(driver)
    Log.info('Exit')







    
    