from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime
import pytest

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("fakerKTR"))
    file_path = environ.get("fakerKTR")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
import random

from Settings.setupKeterampilan import initDriver, loadDataPath, quit, sleep, upload, uploadGambar
from Settings.loginKeterampilan import kasieKeterampilan
from Settings.Page.Keterampilan import MenuPersetujuanPresensiPesertaKegiatan
import random
import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Log4OpAbsensiKegiatan.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

sheetrangeIndex = wb['Keterampilan']

i = 5
NamaKegiatan                                  = sheetrangeIndex['C'+str(i)].value
JumlahPeserta                                 = sheetrangeIndex['N'+str(i)].value
print(NamaKegiatan)


@pytest.mark.webtest
def test1_OpKegiatanPelatihan_SetupOs():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')

@pytest.mark.webtest
def test_TC_KTR_020():
    attach(data=driver.get_screenshot_as_png())      
    kasieKeterampilan(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('"Login aplikasi menggunakan akun dengan role kasie"')

@pytest.mark.webtest
def test_TC_KTR_021():
    sleep(driver)
    MenuPersetujuanPresensiPesertaKegiatan(driver)
    attach(data=driver.get_screenshot_as_png())
    Log.info('Kasie mengakses halaman Persetujuan Presensi Peserta Kegiatan')


@pytest.mark.webtest
def test_TC_KTR_022():
    attach(data=driver.get_screenshot_as_png())
    driver.implicitly_wait(60)
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "buttonSearch")))
    driver.find_element(By.ID, "filterColumn").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "nama_program")))
    driver.find_element(By.ID, "nama_program").click()

    driver.find_element(By.ID, "kataKunci").send_keys(NamaKegiatan)
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "buttonSearch")))

    
    driver.find_element(By.CSS_SELECTOR, "#verifikasi-0 .h-5").click()

    driver.find_element(By.CSS_SELECTOR, ".el-form-item:nth-child(1) > .el-form-item__content > .el-select .el-input__inner").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//div[@id='el-popper-container-1804']/div[8]/div/div/div/ul/li")))
    driver.find_element(By.XPATH, "//div[@id='el-popper-container-1804']/div[8]/div/div/div/ul/li").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".el-textarea__inner")))
    driver.find_element(By.CSS_SELECTOR, ".el-textarea__inner").send_keys("Verifikasi")
    driver.find_element(By.CSS_SELECTOR, "#submitButton > span").click()

    Log.info("Kasie melakukan v erifikasi pada presensi peserta kegiatan")

@pytest.mark.webtest
def test_TC_KTR_023():
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "buttonSearch")))
    driver.find_element(By.CSS_SELECTOR, "#detail-0 .h-5").click()
    Log.info('Kasie mengakses halaman Detail Absensi Kegiatan')

@pytest.mark.webtest
def test1_exit_OpKegiatanPelatihan():
    quit(driver)
    Log.info('Exit')







    
    