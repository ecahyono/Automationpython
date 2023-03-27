# import imp
from json import load

import string
from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys

from openpyxl import load_workbook
import time

#target halaman excel ada dimana , wb = variablenya
# wb = load_workbook(filename="C:\chromedriver\Data.xlsx")
wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Filexel\lainlain.xlsx")

# jadi ini bisa read sheet yang dibawah itu yang di excel
sheetrange = wb['Lain-lain']

# ini web driver disimpen dimana, kalo disimpen di path kosongin aja
driver = webdriver.Chrome(r'C:\Users\user\Documents\TRCH\chromedriver.exe')

# link nya ini dimana
driver.get("http://192.168.2.11:32400/")
# seting windows nya jadi max   
driver.maximize_window()
# script gakan di eksekusi kalo web ga muncul. kalo lebih dari 10 detik ga muncul error
driver.implicitly_wait(6)
# ini letak xpath icon login
driver.find_element(By.XPATH, "//div/span").click()
# ini masuk ke form input username
driver.find_element(By.ID, "username").click()
# masukin input username
driver.find_element(By.ID, "username").send_keys("test-user")
# masukin input password
driver.find_element(By.ID, "password").send_keys("password")
# click button login
driver.find_element(By.ID, "kc-login").click()
time.sleep(3)

element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[9]")
actions = ActionChains(driver)
actions.move_to_element(element).perform()
time.sleep(2)

element2 = driver.find_element(By.XPATH, "//div[11]/div/ul/li[1]/div[1]")
actions2 = ActionChains(driver)
actions2.move_to_element(element2).perform()
driver.find_element(By.LINK_TEXT, "Lain-Lain").click()

i =3 

while i <= len(sheetrange['A']):
    Id_Lookup = sheetrange['A'+str(i)].value
    Groups = sheetrange['B'+str(i)].value
    Deskripsi = sheetrange['C'+str(i)].value
    Catatan = sheetrange['D'+str(i)].value

    time.sleep(1)

    # driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/button").click()
    driver.find_element(By.ID, "createButton").click()

    try:
        WebDriverWait(driver,6).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/div/div/h1')))
        time.sleep(2)
        driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div/div[1]/input').send_keys(Id_Lookup)

        driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[2]/div/div/div/div/input').click()
        driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[2]/div/div/div/div/input').send_keys(Groups)
        driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[2]/div/div/div/div/input').send_keys(Keys.DOWN)
        driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[2]/div/div/div/div/input').send_keys(Keys.ENTER)

        driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[3]/div/div[1]/input').send_keys(Deskripsi)

        driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[4]/div/div/textarea').send_keys(Catatan)
        
        driver.find_element(By.ID, 'submitButton').click()
    except TimeoutException:
        pass
    i = i + 1
print ("Success Created")