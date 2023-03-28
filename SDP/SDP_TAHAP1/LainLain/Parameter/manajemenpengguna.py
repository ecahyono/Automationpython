# ==================================================================================================
#===================================================================================================
from json import load
from lib2to3.pgen2 import driver
import string
from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
import pyautogui

from openpyxl import load_workbook
import time

#target halaman excel ada dimana , wb = variablenya
# wb = load_workbook(filename="C:\chromedriver\Data.xlsx")
wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\SDP\Filexel\lainlain.xlsx")

# jadi ini bisa read sheet yang dibawah itu yang di excel
sheetrange = wb['manajemenpengguna']

# ini web driver disimpen dimana, kalo disimpen di path kosongin aja
driver = webdriver.Chrome(r'C:\Users\user\Documents\TRCH\chromedriver.exe')

# link nya ini dimana
driver.get("http://kumbang.torche.id:32400/")
# seting windows nya jadi max   
driver.maximize_window()
# script gakan di eksekusi kalo web ga muncul. kalo lebih dari 10 detik ga muncul error
WebDriverWait(driver,60).until(EC.element_to_be_clickable((By.ID, 'login')))
# ini letak xpath icon login
driver.find_element(By.ID, "login").click()
# ini masuk ke form input username
driver.find_element(By.ID, "username").click()
# masukin input username
driver.find_element(By.ID, "username").send_keys("test-user")
# masukin input password
driver.find_element(By.ID, "password").send_keys("password")
# click button login
driver.find_element(By.ID, "kc-login").click()
WebDriverWait(driver,60).until(EC.element_to_be_clickable((By.ID, '601')))
element = driver.find_element(By.ID, '601')
ActionChains(driver).move_to_element(element).perform()
WebDriverWait(driver,60).until(EC.element_to_be_clickable((By.ID, '80')))
element2 = driver.find_element(By.ID, "80")
ActionChains(driver).move_to_element(element2).perform()
WebDriverWait(driver,60).until(EC.element_to_be_clickable((By.ID, '/lain-lain/manajemen-pengguna')))
driver.find_element(By.LINK_TEXT, "Manajemen Pengguna").click()

i = 16

while i <= len(sheetrange['A']):
	Username		= sheetrange['A'+str(i)].value
	Password		= sheetrange['B'+str(i)].value
	Nip			 	= sheetrange['C'+str(i)].value
	Email		   	= sheetrange['D'+str(i)].value
	No_Hp		   	= sheetrange['E'+str(i)].value
	Unit_Kerja	  	= sheetrange['F'+str(i)].value
	Tipe			= sheetrange['G'+str(i)].value
	Kanwil		  	= sheetrange['H'+str(i)].value
	UPT			 	= sheetrange['I'+str(i)].value
	Level		   	= sheetrange['J'+str(i)].value
	Aktif_Level		= sheetrange['K'+str(i)].value

	WebDriverWait(driver,60).until(EC.element_to_be_clickable((By.ID, 'createButton')))
	driver.find_element(By.ID, "createButton").click()

	try:
		WebDriverWait(driver,6).until(EC.visibility_of_element_located((By.ID, 'Username')))
		driver.find_element(By.ID, "Username").send_keys(Username)
		driver.find_element(By.ID, "Password").send_keys(Password)
		driver.find_element(By.ID, "nama").send_keys('Akun-'+ Username)
		driver.find_element(By.ID, "Nip").send_keys(Nip)
		driver.find_element(By.ID, "Email").send_keys(Email)
		driver.find_element(By.ID, "NoHp").send_keys(No_Hp)
		driver.find_element(By.ID, "UnitKerja").send_keys(Unit_Kerja)
		if Tipe  == 'Pusat':
			driver.find_element(By.ID, "tipePusat").click()
			levelnya = driver.find_element(By.XPATH, "//*[@id='app']/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[2]/div/div/div/div[1]/input")
			levelnya.click()
			levelnya.send_keys(Level)
			driver.find_element(By.XPATH, "//li[contains(.,'"+ Level+"')]").click()
		elif Tipe  == 'Kanwil':
			driver.find_element(By.ID, "tipeKanwil").click()
			driver.find_element(By.ID, 'Kanwil').click()
			WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'Kanwil-0')))
			driver.find_element(By.XPATH, "//li[contains(.,'"+ Kanwil+"')]").click()
			levelnya = driver.find_element(By.XPATH, "//*[@id='app']/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[3]/div/div/div/div[1]/input")
			levelnya.click()
			levelnya.send_keys(Level)
			driver.find_element(By.XPATH, "//li[contains(.,'"+ Level+"')]").click()
		elif Tipe == "UPT":
			driver.find_element(By.ID, "tipeUpt").click()
			driver.find_element(By.ID, 'Kanwil').click()
			WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'Kanwil-0')))
			driver.find_element(By.ID, 'Kanwil').send_keys(Kanwil)
			driver.find_element(By.XPATH, "//li[contains(.,'"+ Kanwil+"')]").click()
			uptnya = driver.find_element(By.ID, 'Upt').click()
			WebDriverWait(driver,60).until(EC.element_to_be_clickable((By.ID, 'Upt-0')))
			driver.find_element(By.XPATH, "//li[contains(.,'"+ UPT+"')]").click()
			levelnya = driver.find_element(By.XPATH, "//div[@id='app']/div/div[2]/div/div[2]/div/div/form/div/div[2]/div[4]/div/div/div/div/input")
			levelnya.click()
			levelnya.send_keys(Level)
			driver.find_element(By.XPATH, "//li[contains(.,'"+ Level+"')]").click()
				
		driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div').click()
		WebDriverWait(driver,60).until(EC.element_to_be_clickable((By.ID, 'AktifLevel-0')))
		driver.find_element(By.ID, 'AktifLevel-0').click()
		driver.find_element(By.ID, "submitButton").click()
	except TimeoutException:
		# print("")
		pass
	WebDriverWait(driver,90).until(EC.element_to_be_clickable((By.ID, 'createButton')))
	i = i + 1
print ("Success Created")