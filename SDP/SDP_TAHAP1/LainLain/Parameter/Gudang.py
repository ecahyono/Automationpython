# import imp
from json import load
import pyautogui
import string
from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys

from openpyxl import load_workbook
import time

#target halaman excel ada dimana , wb = variablenya
# wb = load_workbook(filename="C:\chromedriver\Data.xlsx")
wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Filexel\lainlain.xlsx")

# jadi ini bisa read sheet yang dibawah itu yang di excel
sheetrange = wb['Lain-lain']

# ini web driver disimpen dimana, kalo disimpen di path kosongin aja
driver = webdriver.Chrome(r'C:\Users\user\Documents\TRCH\chromedriver.exe')

# link nya ini dimana
driver.get("http://192.168.2.11:32400/")
# seting windows nya jadi max   
driver.maximize_window()
# script gakan di eksekusi kalo web ga muncul. kalo lebih dari 10 detik ga muncul error
driver.implicitly_wait(6)
# ini letak xpath icon login
driver.find_element(By.XPATH, "//div/span").click()
# ini masuk ke form input username
driver.find_element(By.ID, "username").click()
# masukin input username
driver.find_element(By.ID, "username").send_keys("test-user")
# masukin input password
driver.find_element(By.ID, "password").send_keys("password")
# click button login
driver.find_element(By.ID, "kc-login").click()
time.sleep(3)

element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[9]")
actions = ActionChains(driver)
time.sleep(2)
actions.move_to_element(element).perform()

element2 = driver.find_element(By.XPATH, "//div[11]/div/ul/li[1]/div[1]")
actions2 = ActionChains(driver)
time.sleep(2)
actions2.move_to_element(element2).perform()

driver.find_element(By.LINK_TEXT, "Gudang").click()

driver.find_element(By.ID, "createButton").click()

Jenis_Gudang = driver.find_element(By.ID, 'input_kondisi_baran_basan')
Jenis_Gudang.click()
Jenis_Gudang.send_keys('Gudang Berharga')
Jenis_Gudang.send_keys(Keys.DOWN)
Jenis_Gudang.send_keys(Keys.ENTER)

Alamat= driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[2]/div/div/textarea')
Alamat.send_keys('Alamt Gudang')

Provinsi = driver.find_element(By. XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[3]/div/div/div/div/input')
Provinsi.click()
Provinsi.send_keys('Jawa Barat')
Provinsi.send_keys(Keys.DOWN)
Provinsi.send_keys(Keys.ENTER)

KotaKabupaten = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[4]/div/div/div/div/input')
KotaKabupaten.click()
time.sleep(2)
KotaKabupaten.send_keys('Kuningan')
KotaKabupaten.send_keys(Keys.DOWN)
KotaKabupaten.send_keys(Keys.ENTER)

Luas_Target = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[1]/div/div/input')
Luas_Target.send_keys("10.000")

Kapasitas_Gudang = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[2]/div/div/input')
Kapasitas_Gudang.send_keys('10000')

Foto = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[3]/div/div/div/button')
Foto.click()
time.sleep(3)
pyautogui.typewrite(r'C:\Users\user\Pictures\Saved Pictures\limvo.png')
pyautogui.press('enter')
time.sleep(3)

Keterangan = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[4]/div/div/textarea')
Keterangan.send_keys("Keterangan Tambah Gudang Baru Untuk Gudang Berharga")

driver.find_element(By.ID, 'submitButton').click()

# i =3 

# while i <= len(sheetrange['A']):
#     Id_Lookup = sheetrange['A'+str(i)].value
#     Groups = sheetrange['B'+str(i)].value
#     Deskripsi = sheetrange['C'+str(i)].value
#     Catatan = sheetrange['D'+str(i)].value

#     time.sleep(1)

#     # driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/button").click()
#     

#     try:
#         WebDriverWait(driver,6).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/div/div/h1')))
#         time.sleep(2)
#         driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div/div[1]/input').send_keys(Id_Lookup)

#         driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[2]/div/div/div/div/input').click()
#         driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[2]/div/div/div/div/input').send_keys(Groups)
#         driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[2]/div/div/div/div/input').send_keys(Keys.DOWN)
#         driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[2]/div/div/div/div/input').send_keys(Keys.ENTER)

#         driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[3]/div/div[1]/input').send_keys(Deskripsi)

#         driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[4]/div/div/textarea').send_keys(Catatan)
        
#         driver.find_element(By.ID, 'submitButton').click()
#     except TimeoutException:
#         pass
#     i = i + 1
# print ("Success Created")