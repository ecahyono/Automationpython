from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
import platform
import sys

from openpyxl import load_workbook
import time

if platform.system() == 'Darwin':
    wb = load_workbook(filename=r"/Users/will/Documents/Automationpython/SDP/Filexel/lainlain.xlsx")
elif platform.system() == 'Windows':
    wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\SDP\Filexel\lainlain.xlsx")

# jadi ini bisa read sheet yang dibawah itu yang di excel
sheetrange = wb['Pegawai']

# ini web driver disimpen dimana, kalo disimpen di path kosongin aja
driver = webdriver.Chrome()

#mb driver = webdriver.Chrome(r"/Users/will/Downloads/chromedriver")

# link nya ini dimana
# driver.get("http://kumbang.torche.id:32400/")
driver.get("http://kumbang.torche.id:32400/")
# seting windows nya jadi max   
driver.maximize_window()
# script gakan di eksekusi kalo web ga muncul. kalo lebih dari 10 detik ga muncul error
driver.implicitly_wait(10)
# ini letak xpath icon login
driver.find_element(By.XPATH, "//div/span").click()
# ini masuk ke form input username
driver.find_element(By.ID, "username").click()
# masukin input username
driver.find_element(By.ID, "username").send_keys("oprupbasaninternal")
# masukin input password
driver.find_element(By.ID, "password").send_keys("password")
# click button login
driver.find_element(By.ID, "kc-login").click()
time.sleep(2)

# variable element nyari dimana letak menu lain lain
WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, '601')))
element = driver.find_element(By.ID, '601')
ActionChains(driver).move_to_element(element).perform()
WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.LINK_TEXT, 'Daftar Pegawai')))
driver.find_element(By. LINK_TEXT, 'Daftar Pegawai').click()

# deklarasi variable dimana i adalah 2 ( jadi dia read data excel dari baris ke 2)
i = 2

# nge baca mulai dari tabel A
while i <= len(sheetrange['A']):
    nip             = sheetrange['A'+str(i)].value
    nama            = sheetrange['B'+str(i)].value
    tempatlahir     = sheetrange['C'+str(i)].value
    tanggallahir    = sheetrange['D'+str(i)].value
    jeniskelamin    = sheetrange['E'+str(i)].value
    alamat          = sheetrange['F'+str(i)].value
    jabatan         = sheetrange['G'+str(i)].value
    pangkat         = sheetrange['H'+str(i)].value
    golongan        = sheetrange['I'+str(i)].value
    bagian          = sheetrange['J'+str(i)].value
    email           = sheetrange['K'+str(i)].value
    telepon         = sheetrange['L'+str(i)].value


    time.sleep(1)
    # Click button tambah
    driver.find_element(By.ID, 'createButton').click()

    try:
        
        driver.find_element(By.ID, "Nip").send_keys(nip)
        driver.find_element(By.ID, "Nama").send_keys(nama)
        driver.find_element(By.ID, "tempatLahir").send_keys(tempatlahir)
        tgllhrpgw = driver.find_element(By.ID, "tglLahir")
        tgllhrpgw.send_keys(tanggallahir)
        tgllhrpgw.send_keys(Keys.ENTER)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div/div[1]/div[5]/div/div/div/div/input").click()
        if jeniskelamin == 'L': 
            driver.find_element(By.XPATH, "//li[contains(.,\'Laki-laki\')]").click()
            print(jeniskelamin)
        elif jeniskelamin == 'P':
            driver.find_element(By.XPATH, "//li[contains(.,\'Perempuan\')]").click()
            print(jeniskelamin)
        time.sleep(2)
        driver.find_element(By.ID, "Alamat").send_keys(alamat)
        driver.find_element(By.ID, "Jabatan").send_keys(jabatan)
        driver.find_element(By.ID, "Pangkat").send_keys(pangkat)
        driver.find_element(By.ID, "Golongan").send_keys(golongan)
        driver.find_element(By.ID, "Bagian").send_keys(bagian)
        driver.find_element(By.ID, "Email").send_keys(email)
        driver.find_element(By.ID, "Telepon").send_keys(telepon) 

        # nunggu 4 detik baru pencet button submit, kumbang lambat jadi harus nunggu
        # time.sleep(10)
        driver.find_element(By.ID, 'submitButton').click()

        
# kalo script gagal keterangan di cmd muncul ini
    except TimeoutException:
        print("ga muncul")
        pass

#looping i + 1, jadi dia bakal masukin data berdasarkan urutan dan baka stop kalo data urutan udah abis ( no di excel )
    time.sleep(5)
    i = i + 1
print("done")

