from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
import os
import pyautogui
import pytest
import time
import platform
import subprocess
from pathlib import Path

@pytest.fixture()
def test_setup():
    global driver
    global wb
    swin = Service(r'C:/Users/user/Documents/TRCH/chromedriver.exe')
    smac = Service('/Users/wildanwijaksana/Documents/chromedriver')

    if platform.system() == 'Darwin':
        driver = webdriver.Chrome(service=smac)
        url = "http://kumbang.torche.id:32400/"
       
        driver.get(url)
        # seting windows nya jadi max
        wb = load_workbook(filename=r"/Users/will/Documents/work/Automationpython/Filexel/Kunjungan.xlsx")   
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
    elif platform.system() == 'Windows':
        driver = webdriver.Chrome(service=swin)
        # url = "http://kumbang.torche.id:32400/"
        url = "http://192.168.2.11:32400/"
        
        driver.get(url)
        # seting windows nya jadi max   
        wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Filexel\Kunjungan.xlsx")
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()

def test_DafkunOnline(test_setup):
    driver.implicitly_wait(10)
    sheetrange = wb ['Onsitedaftar']
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div/div[2]/div/div/div[1]/canvas")))
        
    driver.find_element(By.XPATH, "//div/span").click()
        
    driver.find_element(By.ID, "username").click()
    driver.find_element(By.ID, "username").send_keys("wildan")
    driver.find_element(By.ID, "password").send_keys("wildan")
    driver.find_element(By.ID, "kc-login").click()
        
    nav1 = driver.find_element(By.XPATH, '//*[@id="app"]/div/nav/ul/li[4]/div')
    actions = ActionChains(driver)
    actions.move_to_element(nav1).perform()

    element2 = driver.find_element(By.XPATH, "//div[6]/div/ul/li[2]/div[1]")
    time.sleep(1)
    actions2 = ActionChains(driver)
    actions2.move_to_element(element2).perform()
    time.sleep(1)

    driver.find_element(By.LINK_TEXT, 'Manajemen Kunjungan Online').click()

    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="createButton"]')))
    driver.find_element(By.XPATH, '//*[@id="createButton"]').click()
    
    
    i = 2

    
    while i <= len(sheetrange['A']):
        q                        = sheetrange['A'+str(i)].value
        r                        = sheetrange['B'+str(i)].value
        r                        = sheetrange['C'+str(i)].value
        r                        = sheetrange['D'+str(i)].value
        r                        = sheetrange['E'+str(i)].value
        r                        = sheetrange['F'+str(i)].value

        try:
            time.sleep(1)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div/div[1]/label/div/div/div/input').click()         
            
            #submmit
            driver.find_element(By.XPATH, '//*[@id="submitButton"]').click()
            time.sleep(5)

    
        except TimeoutException:
            print("Mohon Cek Kembali")
            pass
        i = i + 1
    print("Done")