from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
import os
import pyautogui
import pytest
import time
import platform
import subprocess

from pathlib import Path

@pytest.fixture()
def test_setup():
    global driver
    global wb
    swin = Service(r'C:/Users/user/Documents/TRCH/chromedriver.exe')
    smac = Service('/Users/will/Downloads/chromedriver')

    if platform.system() == 'Darwin':
        driver = webdriver.Chrome(service=smac)
        # url = "http://kumbang.torche.id:32400/"
        url = "http://192.168.2.11:32400/"
        
        driver.get(url)
        # seting windows nya jadi max
        wb = load_workbook(filename=r"/Users/will/Documents/work/Automationpython/Filexel/Registrasi.xlsx")   
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
    elif platform.system() == 'Windows':
        driver = webdriver.Chrome(service=swin)
        # url = "http://kumbang.torche.id:32400/"
        url = "http://192.168.2.11:32400/"
        
        driver.get(url)
        # seting windows nya jadi max   
        wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Filexel\Registrasi.xlsx")
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
        
    
def test_Web(test_setup):
    driver.implicitly_wait(10)

    # read excel
    sheetrange = wb['Banhum']
    # Menuju login
    driver.find_element(By.XPATH, "//div/span").click()
    driver.find_element(By.ID, "username").click()
    driver.find_element(By.ID, "username").send_keys("rono")
    driver.find_element(By.ID, "password").send_keys("rene")
    # click button login
    driver.find_element(By.ID, "kc-login").click()
    time.sleep(5) 
    #Registrasi
    element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[1]/div")                                   
    time.sleep(1)
    actions = ActionChains(driver)
    time.sleep(1)
    actions.move_to_element(element).perform()
    time.sleep(1)
    #Identitas
    driver.find_element(By.LINK_TEXT, "Bantuan Hukum").click()
    time.sleep(1)
    i = 2
    while i <= len(sheetrange['A']):
        #deklarasi per colom pada sheet
        noreg           = sheetrange['A'+str(i)].value

        Noajuan         = sheetrange['C'+str(i)].value
        tglajuan        = sheetrange['D'+str(i)].value
        nosrtpgntr      = sheetrange['E'+str(i)].value
        tglsrtpngtr     = sheetrange['F'+str(i)].value
        sbanhum         = sheetrange['G'+str(i)].value
        ket             = sheetrange['H'+str(i)].value
        jmlhdoc         = sheetrange['I'+str(i)].value
        namadoc         = sheetrange['J'+str(i)].value
        ketdoc          = sheetrange['N'+str(i)].value

        #button +Tambah
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/button").click()                                 
        try:
            WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/div/div/h1')))
            #======================== Halaman Cari ============================
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div/div/div/div/input').click()
            time.sleep(2)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div/div/div/div/input').send_keys(noreg)
            time.sleep(3)
            driver.find_element(By.CSS_SELECTOR, ".el-select-dropdown__item:nth-child(1) tr:nth-child(2) > .el-descriptions__cell:nth-child(1)").click()
            time.sleep(1)
            #======================== Button Cari ============================
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div[1]/div/form/button[1]").click()
            #======================== masuk ke halaman input ============================
            time.sleep(5)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/form/div/div/div[2]/div/div[1]/input').click()
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/form/div/div/div[2]/div/div[1]/input').send_keys(Noajuan)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[3]/div/div[1]/input').click()
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[3]/div/div[1]/input').send_keys(tglajuan)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[4]/div/div[1]/input').click()
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[4]/div/div[1]/input').send_keys(nosrtpgntr)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[5]/div/div[1]/input').click()
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[5]/div/div[1]/input').send_keys(tglsrtpngtr)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[5]/div/div[1]/input').send_keys(Keys.ENTER)

            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[6]/div/div/div/div/input').click()
            if sbanhum == 'Diajukan':
                driver.find_element(By.XPATH, '//li[contains(.,\'Diajukan\')]').click()
            elif sbanhum == 'Ditolak':
                driver.find_element(By.XPATH, '//li[contains(.,\'Ditolak\')]').click()
            elif sbanhum == 'Disetujui':
                driver.find_element(By.XPATH, '//span[contains(.,\'Disetujui\')]').click()
            print('sesuai data di seheet')
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[7]/div/div/textarea').click()
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[7]/div/div/textarea').send_keys(ket)

            if  jmlhdoc == 1:
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div/div/input').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div/div/input').send_keys(namadoc)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[3]/div/div/div/button').click()
                time.sleep(2)
                pyautogui.typewrite(r"C:\Users\user\Documents\TRCH\Automationpython\FilePDF\conoth.pdf")
                pyautogui.press("enter")
                time.sleep(2)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[4]/div/div/textarea').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[4]/div/div/textarea').send_keys(ketdoc)
            elif jmlhdoc == 2:  
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div/div/input').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div/div/input').send_keys(namadoc)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[3]/div/div/div/button').click()
                time.sleep(2)
                pyautogui.typewrite(r"C:\Users\user\Documents\TRCH\Automationpython\FilePDF\conoth.pdf")
                pyautogui.press("enter")
                time.sleep(2)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[4]/div/div/textarea').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[4]/div/div/textarea').send_keys(ketdoc)
                #tambah baris =====================================================================================================================================
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[2]/table/thead/tr/th[5]/div/div/div').click()
                # =================================================================================================================================================
                time.sleep(1)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[2]/td[2]/div/div/input').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[2]/td[2]/div/div/input').send_keys(namadoc)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[2]/td[3]/div/div/div/button').click()
                time.sleep(2)
                pyautogui.typewrite(r"C:\Users\user\Documents\TRCH\Automationpython\FilePDF\conoth.pdf")
                pyautogui.press("enter")
                time.sleep(2)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[2]/td[4]/div/div/textarea').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[2]/td[4]/div/div/textarea').send_keys(ketdoc)
            elif jmlhdoc == 3:
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div/div/input').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div/div/input').send_keys(namadoc)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[3]/div/div/div/button').click()
                time.sleep(2)
                pyautogui.typewrite(r"C:\Users\user\Documents\TRCH\Automationpython\FilePDF\conoth.pdf")
                pyautogui.press("enter")
                time.sleep(2)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[4]/div/div/textarea').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[4]/div/div/textarea').send_keys(ketdoc)
                #tambah baris =====================================================================================================================================
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[2]/table/thead/tr/th[5]/div/div/div').click()
                # =================================================================================================================================================
                time.sleep(1)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[2]/td[2]/div/div/input').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[2]/td[2]/div/div/input').send_keys(namadoc)

                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[2]/td[3]/div/div/div/button').click()
                time.sleep(2)
                pyautogui.typewrite(r"C:\Users\user\Documents\TRCH\Automationpython\FilePDF\conoth.pdf")
                pyautogui.press("enter")
                time.sleep(2)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[2]/td[4]/div/div/textarea').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[2]/td[4]/div/div/textarea').send_keys(ketdoc)
                #tambah baris =====================================================================================================================================
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[2]/table/thead/tr/th[5]/div/div/div').click()
                # =================================================================================================================================================
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[3]/td[2]/div/div/input').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[3]/td[2]/div/div/input').send_keys(namadoc)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[3]/td[3]/div/div/div/button').click()
                time.sleep(2)
                pyautogui.typewrite(r"C:\Users\user\Documents\TRCH\Automationpython\FilePDF\conoth.pdf")
                pyautogui.press("enter")
                time.sleep(2)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[3]/td[4]/div/div/textarea').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[3]/td[4]/div/div/textarea').send_keys(ketdoc)
            elif jmlhdoc == 4:
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div/div/input').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div/div/input').send_keys(namadoc)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[3]/div/div/div/button').click()
                time.sleep(2)
                pyautogui.typewrite(r"C:\Users\user\Documents\TRCH\Automationpython\FilePDF\conoth.pdf")
                pyautogui.press("enter")
                time.sleep(2)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[4]/div/div/textarea').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[4]/div/div/textarea').send_keys(ketdoc)
                #tambah baris =====================================================================================================================================
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[2]/table/thead/tr/th[5]/div/div/div').click()
                # =================================================================================================================================================
                time.sleep(1)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[2]/td[2]/div/div/input').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[2]/td[2]/div/div/input').send_keys(namadoc)

                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[2]/td[3]/div/div/div/button').click()
                time.sleep(2)
                pyautogui.typewrite(r"C:\Users\user\Documents\TRCH\Automationpython\FilePDF\conoth.pdf")
                pyautogui.press("enter")
                time.sleep(2)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[2]/td[4]/div/div/textarea').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[2]/td[4]/div/div/textarea').send_keys(ketdoc)
                #tambah baris =====================================================================================================================================
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[2]/table/thead/tr/th[5]/div/div/div').click()
                # =================================================================================================================================================
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[3]/td[2]/div/div/input').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[3]/td[2]/div/div/input').send_keys(namadoc)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[3]/td[3]/div/div/div/button').click()
                time.sleep(2)
                pyautogui.typewrite(r"C:\Users\user\Documents\TRCH\Automationpython\FilePDF\conoth.pdf")
                pyautogui.press("enter")
                time.sleep(2)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[3]/td[4]/div/div/textarea').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[3]/td[4]/div/div/textarea').send_keys(ketdoc)
                #tambah baris =====================================================================================================================================
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[2]/table/thead/tr/th[5]/div/div/div').click()
                # =================================================================================================================================================
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[4]/td[2]/div/div/input').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[4]/td[2]/div/div/input').send_keys(namadoc)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[4]/td[3]/div/div/div/button').click()
                time.sleep(2)
                pyautogui.typewrite(r"C:\Users\user\Documents\TRCH\Automationpython\FilePDF\conoth.pdf")
                pyautogui.press("enter")
                time.sleep(2)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[4]/td[4]/div/div/textarea').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[10]/div[1]/div[3]/div/div[1]/div/table/tbody/tr[4]/td[4]/div/div/textarea').send_keys(ketdoc)
            print('done')
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/form/div/div/div[11]/button[2]').click()
            time.sleep(4)
            WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/div[1]/div')))
        except TimeoutException:
            print("MASIH ADA ERROR, CEK LAGI PAK WIL")
            pass
        time.sleep(5)
        i = i + 1
    print("DONE PAK")

