from pydoc import describe
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
import pytest
from openpyxl import load_workbook
import time

from selenium.webdriver.common.keys import Keys 
@pytest.fixture()
def test_setup():
    global driver
    LokasiChromeDriver = Service('/Users/will/Downloads/chromedriver')
    #LokasiChromeDriver = Service()
    driver = webdriver.Chrome(service=LokasiChromeDriver)
    #url = "http://kumbang.torche.id:32400/"
    url = "http://192.168.2.11:32400/"
    driver.get(url)
    driver.maximize_window()
    driver.implicitly_wait(10)
    yield
    #driver.close()
    #driver.quit()
def test_Jenis_Kejahatan(test_setup):
    
    wb = load_workbook(filename=r"/Users/will/Documents/work/Automationpython/Filexel/lainlain.xlsx")
    #wb = load_workbook(filename=r"C:\Users\wilda\Documents\Automationpython\Filexel/Registrasi.xlsx")
    
    sheetrange = wb['JenisKejahatan']
    WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, '//div/span')))
        
    driver.find_element(By.XPATH, "//div/span").click()
    # ini masuk ke form input username
    driver.find_element(By.ID, "username").click()
    # masukin input username
    driver.find_element(By.ID, "username").send_keys("test-user")
    # masukin input password
    driver.find_element(By.ID, "password").send_keys("password")
    # click button login
    driver.find_element(By.ID, "kc-login").click()
    time.sleep(1)
    WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div/div[1]/div/div[1]")))
    
    #======================== Menu Registrasi ============================
    element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[8]/div")
    time.sleep(0.5)
    actions = ActionChains(driver)
    time.sleep(0.5)
    actions.move_to_element(element).perform()
    time.sleep(0.5)

    element2 = driver.find_element(By.XPATH, "//div[10]/div/ul/li/div")
    time.sleep(0.5)
    
    actions2 = ActionChains(driver)
    time.sleep(0.5)
    actions2.move_to_element(element2).perform()
    time.sleep(0.5)
    driver.find_element(By.LINK_TEXT, "Jenis Kejahatan")
    time.sleep(0.5)
    driver.find_element(By.LINK_TEXT, "Jenis Kejahatan").click()
    time.sleep(0.5)

    #======================== Halaman Create ============================
    i = 2

    # nge baca mulai dari tabel A
    while i <= len(sheetrange['A']):
        # deklarasi bahwa NIP itu ada di A 
        JenisKejahatan          = sheetrange['A'+str(i)].value
        Deskripsi               = sheetrange['A'+str(i)].value
        
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".is-plain:nth-child(2)")))
        driver.find_element(By.CSS_SELECTOR, ".is-plain:nth-child(2)").click()
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[2]/div/div/button[2]")))

        try:    
            time.sleep(1)
            #driver.find_element(By.XPATH, "//*[@id=\"input_terminologi\"]").click()
            driver.find_element(By.XPATH, "//*[@id=\"input_terminologi\"]").send_keys(JenisKejahatan)
            
            #driver.find_element(By.XPATH, "//*[@id=\"input_deskripsi\"]").click()
            driver.find_element(By.XPATH, "//*[@id=\"input_deskripsi\"]").send_keys(Deskripsi)
            WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[2]/div/div/button[2]")))
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[2]/div/div/button[2]").click()

        except TimeoutException:
            print("MASIH ADA ERROR, CEK LAGI PAK WIL")
            pass
        time.sleep(5)
        i = i + 1
    print("DONE PAK WILDAN, SEBATS DULU")