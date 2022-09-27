# ==================================================================================================
#===================================================================================================
from json import load
from lib2to3.pgen2 import driver
import string
from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
import pyautogui

from openpyxl import load_workbook
import time

#target halaman excel ada dimana , wb = variablenya
# wb = load_workbook(filename="C:\chromedriver\Data.xlsx")
wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Filexel\lainlain.xlsx")

# jadi ini bisa read sheet yang dibawah itu yang di excel
sheetrange = wb['manajemenpengguna']

# ini web driver disimpen dimana, kalo disimpen di path kosongin aja
driver = webdriver.Chrome(r'C:\Users\user\Documents\TRCH\chromedriver.exe')

# link nya ini dimana
driver.get("http://kumbang.torche.id:32400/")
# seting windows nya jadi max   
driver.maximize_window()
# script gakan di eksekusi kalo web ga muncul. kalo lebih dari 10 detik ga muncul error
driver.implicitly_wait(6)
# ini letak xpath icon login
driver.find_element(By.XPATH, "//div/span").click()
# ini masuk ke form input username
driver.find_element(By.ID, "username").click()
# masukin input username
driver.find_element(By.ID, "username").send_keys("rono")
# masukin input password
driver.find_element(By.ID, "password").send_keys("rene")
# click button login
driver.find_element(By.ID, "kc-login").click()
time.sleep(2)

element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[8]/div")
actions = ActionChains(driver)
actions.move_to_element(element).perform()
time.sleep(3)

# WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "XPATH LAIN LAIN PARAMETER "))) ## OPSIONAL

element2 = driver.find_element(By.XPATH, "//div[10]/div/ul/li/div")
actions2 = ActionChains(driver)
actions2.move_to_element(element2).perform()
driver.find_element(By.LINK_TEXT, "Manajemen Pengguna").click()



i = 2

while i <= len(sheetrange['A']):
    Username    = sheetrange['A'+str(i)].value
    Password    = sheetrange['B'+str(i)].value
    Nip         = sheetrange['C'+str(i)].value
    Email       = sheetrange['D'+str(i)].value
    No_Hp       = sheetrange['E'+str(i)].value
    Unit_Kerja  = sheetrange['F'+str(i)].value
    Tipe        = sheetrange['G'+str(i)].value
    Kanwil      = sheetrange['H'+str(i)].value
    UPT         = sheetrange['I'+str(i)].value
    Level       = sheetrange['J'+str(i)].value
    Aktif_Level = sheetrange['K'+str(i)].value

    time.sleep(1)

    driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/button").click()

    try:
        WebDriverWait(driver,6).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/div/div/h1')))
        time.sleep(2)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[1]/div[1]/div/div[1]/input").send_keys(Username)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[1]/div[2]/div/div[1]/input").send_keys(Password)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[1]/div[3]/div/div[1]/input").send_keys(Nip)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[1]/div[4]/div/div/input").send_keys(Email)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[1]/div[5]/div/div[1]/input").send_keys(No_Hp)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[1]/div[6]/div/div[1]/input").send_keys(Unit_Kerja)
        
        time.sleep(2)
        if Tipe  == 'Pusat':
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]/div[1]/div/div/label[1]/span[1]/span").click()
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]/div[2]/div/div[1]/div/div[1]/input").click()
            pyautogui.typewrite(Level)
            time.sleep(2)
            pyautogui.keyDown('down')
            pyautogui.press('enter')
            time.sleep(2)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]").click()
            if Aktif_Level == 'UAT':
                driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]/div[3]/div/div/label/span[1]/span").click() 
            elif Aktif_Level == 'Kepala Seksi':
                driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[4]/div/div/label/span[1]/span").click()
            elif Aktif_Level == 'master': 
                driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]/div[5]/div/div/label/span[1]/span").click()
            elif Aktif_Level == 'Administrasi': 
                driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]/div[3]/div/div/label[1]").click()
                print('noting')
            print (Tipe)
        elif Tipe  == 'Kanwil':
                driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]/div[1]/div/div/label[2]/span[1]/span").click()
                time.sleep(2)
                driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[2]/div/div/div/div/input").click()
                pyautogui.typewrite(Kanwil)
                pyautogui.keyDown('down')
                pyautogui.press('enter')
                time.sleep(2)
                driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[3]/div/div/div/div[1]/input").click()                
                pyautogui.typewrite(Level)
                pyautogui.keyDown('down')
                pyautogui.press('enter')
                driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]/div[4]/label").click()
                if Aktif_Level == 'UAT':
                    driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]/div[3]/div/div/label/span[1]/span").click() 
                elif Aktif_Level == 'Kepala Seksi':
                    driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[4]/div/div/label/span[1]/span").click()
                elif Aktif_Level == 'master': 
                    driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]/div[5]/div/div/label/span[1]/span").click()
                elif Aktif_Level == 'Administrasi': 
                    driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]/div[3]/div/div/label[1]").click()
                    print('noting')
                print (Tipe)
        elif Tipe == 'UPT':
                driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]/div[1]/div/div/label[3]/span[1]/span").click()
                time.sleep(2)
                driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[2]/div/div/div/div/input").click()
                pyautogui.typewrite(Kanwil)
                pyautogui.keyDown('down')
                pyautogui.press('enter')
                time.sleep(2)
                driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]/div[3]/div/div/div/div/input").click()
                pyautogui.typewrite(UPT)
                pyautogui.keyDown('down')
                pyautogui.press('enter')
                time.sleep(2)
                driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]/div[4]/div/div[1]/div/div[1]/input").click()
                pyautogui.typewrite(Level)
                pyautogui.keyDown('down')
                pyautogui.press('enter')
                driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]/div[4]/label").click()
                time.sleep(2)
                if Aktif_Level == 'UAT':
                    driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]/div[3]/div/div/label/span[1]/span").click() 
                elif Aktif_Level == 'Kepala Seksi':
                    driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[4]/div/div/label/span[1]/span").click()
                elif Aktif_Level == 'master': 
                    driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]/div[5]/div/div/label/span[1]/span").click()
                elif Aktif_Level == 'Administrasi': 
                    driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div[2]/div[3]/div/div/label[1]").click()
                    print('noting')
                print (Tipe)
        # Submit
        time.sleep(2)
        
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[2]/div/div/button[2]").click()
        
        
    except TimeoutException:
        # print("")
        pass
    time.sleep(5)   
    i = i + 1
print ("Success Created")