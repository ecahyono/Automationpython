from json import load
from lib2to3.pgen2 import driver
import string
from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys


from openpyxl import load_workbook
import time

#target halaman excel ada dimana , wb = variablenya
#macbook
#wb = load_workbook(filename=r"/Users/will/Documents/Automationpython/Filexel/lainlain.xlsx")
wb = load_workbook(filename=r"C:\Users\wilda\Documents\Automationpython\Filexel/lainlain.xlsx")

# jadi ini bisa read sheet yang dibawah itu yang di excel
sheetrange = wb['Pegawai']

# ini web driver disimpen dimana, kalo disimpen di path kosongin aja
driver = webdriver.Chrome()

#mb driver = webdriver.Chrome(r"/Users/will/Downloads/chromedriver")

# link nya ini dimana
driver.get("http://kumbang.torche.id:32400/")
# seting windows nya jadi max   
driver.maximize_window()
# script gakan di eksekusi kalo web ga muncul. kalo lebih dari 10 detik ga muncul error
driver.implicitly_wait(10)
# ini letak xpath icon login
driver.find_element(By.XPATH, "//div/span").click()
# ini masuk ke form input username
driver.find_element(By.ID, "username").click()
# masukin input username
driver.find_element(By.ID, "username").send_keys("wildan")
# masukin input password
driver.find_element(By.ID, "password").send_keys("wildan")
# click button login
driver.find_element(By.ID, "kc-login").click()
time.sleep(5)
# variable element nyari dimana letak menu lain lain
element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[7]/div")
                                        
time.sleep(5)
# lib buat deklarasi element
actions = ActionChains(driver)
# mouse pindah ke variable element yaitu lain lain
actions.move_to_element(element).perform()
#actions.move_to_element(lain).move_to_element(tools).move_to_element(pegawai).click()
# ini bisa dipake buat nunggu menu kalo telat muncul
#WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[7]/div")))
# click menu pegawai
driver.find_element(By.LINK_TEXT, "Daftar Pegawai").click()

# deklarasi variable dimana i adalah 2 ( jadi dia read data excel dari baris ke 2)
i = 2

# nge baca mulai dari tabel A
while i <= len(sheetrange['A']):
    # deklarasi bahwa NIP itu ada di A 
    nip = sheetrange['A'+str(i)].value
    # deklarasi bahwa NAMA itu ada di B 
    nama = sheetrange['B'+str(i)].value
    # deklarasi bahwa NAMA itu ada di B 
    tempatlahir = sheetrange['C'+str(i)].value
    tanggallahir = sheetrange['D'+str(i)].value
    jeniskelamin = sheetrange['E'+str(i)].value
    alamat = sheetrange['F'+str(i)].value
    jabatan = sheetrange['G'+str(i)].value
    pangkat = sheetrange['H'+str(i)].value
    golongan = sheetrange['I'+str(i)].value
    bagian = sheetrange['J'+str(i)].value
    email = sheetrange['K'+str(i)].value
    telepon = sheetrange['L'+str(i)].value




    time.sleep(1)
    # Click button tambah
    driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/button").click()

    try:
        # menunggu form halaman input data pegawai muncul
        #WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/div[2]/div[1]/div/div')))
        
        #memasukan data nip,nama dll yang ada di excel
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div/div[1]/div[1]/div/div[1]/input").send_keys(nip)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div/div[1]/div[2]/div/div[1]/input").send_keys(nama)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div/div[1]/div[3]/div/div/input").send_keys(tempatlahir)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div/div[1]/div[4]/div/div/input").send_keys(tanggallahir)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div/div[1]/div[4]/div/div/input").send_keys(Keys.ENTER)
        # driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div/div[1]/div[3]/div/div/input").send_keys(jeniskelamin)
        
        #click dropdown laki laki
        driver.find_element(By.CSS_SELECTOR, ".select-trigger .el-input__inner").click()
        driver.find_element(By.XPATH, "//li[contains(.,\'Laki-laki\')]").click()

        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div/div[1]/div[6]/div/div/textarea").send_keys(alamat)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div/div[2]/div[1]/div/div/input").send_keys(jabatan)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div/div[2]/div[2]/div/div/input").send_keys(pangkat)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div/div[2]/div[3]/div/div/input").send_keys(golongan)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div/div[2]/div[4]/div/div/input").send_keys(bagian)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div/div[2]/div[5]/div/div/input").send_keys(email)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[1]/div/div[2]/div[6]/div/div/input").send_keys(telepon)

        # nunggu 4 detik baru pencet button submit, kumbang lambat jadi harus nunggu
        time.sleep(7)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[4]/div/div/button[2]").click()

        
# kalo script gagal keterangan di cmd muncul ini
    except TimeoutException:
        print("ga muncul")
        pass

#looping i + 1, jadi dia bakal masukin data berdasarkan urutan dan baka stop kalo data urutan udah abis ( no di excel )
    time.sleep(5)
    i = i + 1
print("done")

