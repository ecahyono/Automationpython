

from json import load
from lib2to3.pgen2 import driver
import string
from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys 

from openpyxl import load_workbook
import time
import pyautogui

wb = load_workbook(filename=r"C:\Users\wilda\Documents\Automationpython\Filexel/lainlain.xlsx")
#wb = load_workbook(filename=r"/Users/will/Documents/Automationpython/Filexel/lainlain.xlsx")

# jadi ini bisa read sheet yang dibawah itu yang di excel
sheetrange = wb['DaerahTingkatII']

# ini web driver disimpen dimana, kalo disimpen di path kosongin aja
driver = webdriver.Chrome()
#driver = webdriver.Chrome('/Users/will/Downloads/chromedriver') 


driver.get("http://kumbang.torche.id:32400/")
# seting windows nya jadi max   
driver.maximize_window()
# script gakan di eksekusi kalo web ga muncul. kalo lebih dari 10 detik ga muncul error
driver.implicitly_wait(10)
# ini letak xpath icon login
driver.find_element(By.XPATH, "//div/span").click()
# ini masuk ke form input username
driver.find_element(By.ID, "username").click()
# masukin input username
driver.find_element(By.ID, "username").send_keys("wildan")
# masukin input password
driver.find_element(By.ID, "password").send_keys("wildan")
# click button login
driver.find_element(By.ID, "kc-login").click()
time.sleep(2)
element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[7]/div")
time.sleep(2)
actions = ActionChains(driver)
actions.move_to_element(element).perform()
time.sleep(2)

element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[7]/div")

actions = ActionChains(driver)
actions.move_to_element(element).perform()


element2 = driver.find_element(By.XPATH, "//div[9]/div/ul/li/div")

actions2 = ActionChains(driver)
actions2.move_to_element(element2).perform()
driver.find_element(By.LINK_TEXT, "Daerah Tingkat II").click()

i = 2

# nge baca mulai dari tabel A
while i <= len(sheetrange['A']):
    # deklarasi bahwa NIP itu ada di A 
    Provinsi = sheetrange['A'+str(i)].value
    Deskripsi = sheetrange['B'+str(i)].value

    try:
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/button").click()
        
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div/div/div/div/input").click()
        
        time.sleep(5)
        pyautogui.typewrite(Provinsi)
        time.sleep(2)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div/div/div/div/input").send_keys(Keys.DOWN)
        time.sleep(2)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div/div/div/div/input").send_keys(Keys.ENTER)
        

        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[2]/div/div/input").click()
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[2]/div/div/input").send_keys(Deskripsi)

        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[3]/div/div/button[2]").click()
        

    except TimeoutException:
        print("MASIH ADA ERROR, CEK LAGI PAK WIL")
        pass
    time.sleep(5)
    i = i + 1
print("DONE PAK WILDAN, SEBATS DULU")
