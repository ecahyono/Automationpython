from json import load
from lib2to3.pgen2 import driver
import string
from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
import pytest
from selenium.webdriver.chrome.service import Service
import platform


from openpyxl import load_workbook
import time

@pytest.fixture()
def test_setup():
	global driver
	global wb
	swin = Service(r'C:/Users/user/Documents/TRCH/chromedriver.exe')
	smac = Service('/Users/will/Downloads/chromedriver')
	if platform.system() == 'Darwin':
		driver = webdriver.Chrome(service=smac)
		# url = 'http://kumbang.torche.id:32400/'
		url = 'http://192.168.2.11:32400/'
				
		driver.get(url)
		# seting windows nya jadi max
		wb = load_workbook(filename='/Users/will/Documents/work/Automationpython/Filexel/Registrasi.xlsx')   
		driver.maximize_window()
		driver.implicitly_wait(5)
		yield
		driver.close()
		driver.quit()
	elif platform.system() == 'Windows':
		driver = webdriver.Chrome(service=swin)
		url = 'http://kumbang.torche.id:32400/'
		# url = 'http://192.168.2.11:32400/'
				
		driver.get(url)
		# seting windows nya jadi max   
		wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Filexel\lainlain.xlsx")
		driver.maximize_window()
		driver.implicitly_wait(5)
		yield
		driver.close()
		driver.quit()
def test_Identitas(test_setup):
	# read excel
	sheetrange = wb ['Pegawai']
	# Menuju login
	WebDriverWait(driver, 90).until(EC.element_to_be_clickable((By.ID , 'login')))
	driver.find_element(By.ID, 'login').click()
	driver.find_element(By.ID, 'username').click()
	driver.find_element(By.ID, "username").send_keys("oprupbasanbdg")
	driver.find_element(By.ID, "password").send_keys("password")
	# click button login
	driver.find_element(By.ID, 'kc-login').click()
	WebDriverWait(driver, 90).until(EC.element_to_be_clickable((By.ID , '601')))
# variable element nyari dimana letak menu lain lain
	element = driver.find_element(By.ID, "601")
	ActionChains(driver).move_to_element(element).perform()
	driver.find_element(By.LINK_TEXT, "Daftar Pegawai").click()

	# deklarasi variable dimana i adalah 2 ( jadi dia read data excel dari baris ke 2)
	i = 3

	# nge baca mulai dari tabel A
	while i <= len(sheetrange['A']):
		nip			 = sheetrange['A'+str(i)].value
		nama			= sheetrange['B'+str(i)].value
		tempatlahir	 = sheetrange['C'+str(i)].value
		tanggallahir	= sheetrange['D'+str(i)].value
		jeniskelamin	= sheetrange['E'+str(i)].value
		alamat		  = sheetrange['F'+str(i)].value
		jabatan		 = sheetrange['G'+str(i)].value
		pangkat		 = sheetrange['H'+str(i)].value
		golongan		= sheetrange['I'+str(i)].value
		bagian		  = sheetrange['J'+str(i)].value
		email		   = sheetrange['K'+str(i)].value
		telepon		 = sheetrange['L'+str(i)].value


		WebDriverWait(driver, 90).until(EC.element_to_be_clickable((By.ID , 'createButton')))
		# Click button tambah
		driver.find_element(By.ID, 'createButton').click()

		try:
					
			driver.find_element(By.ID, 'Nip').send_keys(nip)
			driver.find_element(By.ID, 'Nama').send_keys(nama)
			# driver.find_element(By.ID, '').send_keys(tempatlahir)
			# driver.find_element(By.ID, '').send_keys(tanggallahir)
			# driver.find_element(By.ID, '').send_keys(Keys.ENTER)
			#Jenis Kelamin
			driver.find_element(By.ID, 'jenisKelamin').click()
			if jeniskelamin == 'L': 
				driver.find_element(By.ID, 'Laki-laki').click()
			elif jeniskelamin == 'P':
				driver.find_element(By.ID, 'Perempuan').click()

			driver.find_element(By.ID, 'Alamat').send_keys(alamat)
			driver.find_element(By.ID, 'Jabatan').send_keys(jabatan)
			driver.find_element(By.ID, 'Pangkat').send_keys(pangkat)
			driver.find_element(By.ID, 'Golongan').send_keys(golongan)
			# driver.find_element(By.ID, '').send_keys(bagian)
			# driver.find_element(By.ID, '').send_keys(email)
			# driver.find_element(By.ID, '').send_keys(telepon)

			# nunggu 4 detik baru pencet button submit, kumbang lambat jadi harus nunggu
			# time.sleep(10)
			driver.find_element(By.ID, 'submitButton').click()
			WebDriverWait(driver, 90).until(EC.element_to_be_clickable((By.ID , 'createButton')))				
	# kalo script gagal keterangan di cmd muncul ini
		except TimeoutException:
			print("Loading")
			pass
	#looping i + 1, jadi dia bakal masukin data berdasarkan urutan dan baka stop kalo data urutan udah abis ( no di excel )
		i = i + 1
	print("done")

