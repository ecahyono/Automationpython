from json import load
from lib2to3.pgen2 import driver
import string
from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys


from openpyxl import load_workbook
import time

#target halaman excel ada dimana , wb = variablenya
#macbook
#wb = load_workbook(filename=r"/Users/will/Documents/Automationpython/Filexel/lainlain.xlsx")
wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Filexel\lainlain.xlsx")

# jadi ini bisa read sheet yang dibawah itu yang di excel
sheetrange = wb['TamuDinas']

# ini web driver disimpen dimana, kalo disimpen di path kosongin aja
driver = webdriver.Chrome()

#mb driver = webdriver.Chrome(r"/Users/will/Downloads/chromedriver")

# link nya ini dimana
# driver.get("http://kumbang.torche.id:32400/")
driver.get("http://192.168.2.11:32400/")
# seting windows nya jadi max   
driver.maximize_window()
# script gakan di eksekusi kalo web ga muncul. kalo lebih dari 10 detik ga muncul error
driver.implicitly_wait(10)
# ini letak xpath icon login
driver.find_element(By.XPATH, "//div/span").click()
# ini masuk ke form input username
driver.find_element(By.ID, "username").click()
# masukin input username
driver.find_element(By.ID, "username").send_keys("Ruptest")
# masukin input password
driver.find_element(By.ID, "password").send_keys("Ruptest")
# click button login
driver.find_element(By.ID, "kc-login").click()
time.sleep(2)

# variable element nyari dimana letak menu lain lain
element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[9]/div")
time.sleep(2)
# lib buat deklarasi element
actions = ActionChains(driver)
# mouse pindah ke variable element yaitu lain lain
actions.move_to_element(element).perform()
driver.find_element(By.LINK_TEXT, "Tamu Dinas").click()

# deklarasi variable dimana i adalah 2 ( jadi dia read data excel dari baris ke 2)
i = 6

# nge baca mulai dari tabel A
while i <= len(sheetrange['A']):
    nip             = sheetrange['A'+str(i)].value
    nama            = sheetrange['B'+str(i)].value
    pangkat         = sheetrange['C'+str(i)].value
    golongan        = sheetrange['D'+str(i)].value
    jabatan         = sheetrange['E'+str(i)].value
    Instansi        = sheetrange['F'+str(i)].value
    alminstansi     = sheetrange['G'+str(i)].value
    Kota            = sheetrange['H'+str(i)].value
    Provinsi        = sheetrange['I'+str(i)].value
    Keterangan      = sheetrange['J'+str(i)].value


    time.sleep(1)
    # Click button tambah
    driver.find_element(By.ID, 'createButton').click()

    try:
        
        driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div/div[1]/input').send_keys(nip)
        driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[2]/div/div[1]/input').send_keys(nama)
        driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[3]/div/div[1]/input').send_keys(pangkat)
        driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[4]/div/div[1]/input').send_keys(golongan)
        driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[5]/div/div[1]/input').send_keys(jabatan)

        #Instansi
        driver.find_element(By.ID, 'inputInstansiId').click()
        if Instansi == 'POLRES BANDUNG': 
            driver.find_element(By.ID, 'optionInstansi3').click()

        driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[7]/div/div/textarea').send_keys(alminstansi)

        #Kota
        driver.find_element(By.ID, 'inputKotaId').click()
        time.sleep(1)
        if Kota == 'Kuningan': 
            driver.find_element(By.ID, 'optionKota119').click()
        #Provinsi
        driver.find_element(By.ID, 'inputProvinsiId').click()
        time.sleep(1)
        if Provinsi == 'Jawa Barat': 
            driver.find_element(By.ID, 'optionProvinsi4').click()

        driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[10]/div/div/textarea').send_keys(Keterangan)

        driver.find_element(By.ID, 'submitButton').click()

        
# kalo script gagal keterangan di cmd muncul ini
    except TimeoutException:
        print("ga muncul")
        pass

#looping i + 1, jadi dia bakal masukin data berdasarkan urutan dan baka stop kalo data urutan udah abis ( no di excel )
    time.sleep(3)
    i = i + 1
print("done")

