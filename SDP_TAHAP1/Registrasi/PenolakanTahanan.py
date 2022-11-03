from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
import os
import pyautogui
import pytest
import time
import platform
import subprocess

from pathlib import Path

@pytest.fixture()
def test_setup():
    global driver
    global wb
    swin = Service(r'C:/Users/user/Documents/TRCH/chromedriver.exe')
    smac = Service('/Users/will/Downloads/chromedriver')

    if platform.system() == 'Darwin':
        driver = webdriver.Chrome(service=smac)
        # url = "http://kumbang.torche.id:32400/"
        url = "http://192.168.2.11:32400/"
        
        driver.get(url)
        # seting windows nya jadi max
        wb = load_workbook(filename=r"/Users/will/Documents/work/Automationpython/Filexel/Registrasi.xlsx")   
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
    elif platform.system() == 'Windows':
        driver = webdriver.Chrome(service=swin)
        # url = "http://kumbang.torche.id:32400/"
        url = "http://192.168.2.11:32400/"
        
        driver.get(url)
        # seting windows nya jadi max   
        wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Filexel\Registrasi.xlsx")
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
        
def test_Penolakan(test_setup):
    
    driver.implicitly_wait(5)
    sheetrange = wb['PenolakannTahanan']
    WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, '//div/span')))
        
    driver.find_element(By.XPATH, "//div/span").click()
    # ini masuk ke form input username
    driver.find_element(By.ID, "username").click()
    # masukin input username
    driver.find_element(By.ID, "username").send_keys("test-user")
    # masukin input password
    driver.find_element(By.ID, "password").send_keys("password")
    # click button login
    driver.find_element(By.ID, "kc-login").click()
    time.sleep(1)
    WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div/div[1]/div/div[1]")))
    
    #======================== Menu Registrasi ============================
    element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[1]/div")
    time.sleep(1)
    actions = ActionChains(driver)
    time.sleep(1)
    actions.move_to_element(element).perform()
    time.sleep(1)
    #======================== menu Penerimaan dan Penolakan ============================
    element2 = driver.find_element(By.XPATH, "//div/ul/li/div")
    time.sleep(0.5)
    actions2 = ActionChains(driver)
    time.sleep(0.5)
    actions2.move_to_element(element2).perform()
    time.sleep(0.5)
    driver.find_element(By.LINK_TEXT, "Penerimaan").click()

    #======================== Halaman Create ============================

    WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/div[1]/h1")))
    
    i = 2

    # nge baca mulai dari tabel A
    while i <= len(sheetrange['A']):
        # deklarasi bahwa NIP itu ada di A 
        TanggalPenerimaan = sheetrange['A'+str(i)].value
        AsalInstansi = sheetrange['B'+str(i)].value
        NoSurat = sheetrange['C'+str(i)].value
        TanggalSurat = sheetrange['D'+str(i)].value
        Keterangan = sheetrange['E'+str(i)].value
        Nama = sheetrange['F'+str(i)].value
        JenisKelamin = sheetrange['G'+str(i)].value
        Umur = sheetrange['H'+str(i)].value
        TempatLahir = sheetrange['I'+str(i)].value
        TempatTinggal = sheetrange['J'+str(i)].value
        Agama = sheetrange['K'+str(i)].value
        Pekerjaan = sheetrange['L'+str(i)].value
        PerkaraPasal = sheetrange['M'+str(i)].value
        DasarPenahananPidana = sheetrange['N'+str(i)].value
        LamaPenahananPidanaHari = sheetrange['O'+str(i)].value
        LamaPenahananPidanaBulan = sheetrange['P'+str(i)].value
        LamaPenahananPidanaTahun = sheetrange['Q'+str(i)].value
        Tempat = sheetrange['R'+str(i)].value
        DasarPenahananPidana = sheetrange['S'+str(i)].value
        LamaPenahananPidanaHari = sheetrange['T'+str(i)].value
        LamaPenahananPidanaBulan = sheetrange['U'+str(i)].value
        LamaPenahananPidanaTahun = sheetrange['V'+str(i)].value
        HasilPemeriksaanKesehatan = sheetrange['W'+str(i)].value
        Keterangan1 = sheetrange['X'+str(i)].value
        

        #======================== Halaman Index ============================
        time.sleep(2)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/div[2]/button[2]").click()
        time.sleep(2)
        WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/h2")))
       
        try:     
            #======================== INPUT TANGGAL PENERIMAAN ============================
            driver.find_element(By.XPATH, "//div[@id='app']/div/div[2]/div/div[2]/div/div/form/div/div/div/div/div/input").click()
            driver.find_element(By.XPATH, "//div[@id='app']/div/div[2]/div/div[2]/div/div/form/div/div/div/div/div/input").send_keys(TanggalPenerimaan)
            time.sleep(2)
            driver.find_element(By.XPATH, "//div[@id='app']/div/div[2]/div/div[2]/div/div/form/div/div/div/div/div/input").send_keys(Keys.ENTER)
            #======================== INPUT ASAL AsalInstansi ============================
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[2]/div/div/input").send_keys(AsalInstansi)
            #======================== INPUT ASAL NoSurat ============================
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[3]/div/div/input").send_keys(NoSurat)
            #======================== INPUT TanggalPenerimaan ============================
            driver.find_element(By.XPATH, "//div[@id='app']/div/div[2]/div/div[2]/div/div/form/div/div[2]/div/div/div/input").click()
            driver.find_element(By.XPATH, "//div[@id='app']/div/div[2]/div/div[2]/div/div/form/div/div[2]/div/div/div/input").send_keys(TanggalSurat)
            driver.find_element(By.XPATH, "//div[@id='app']/div/div[2]/div/div[2]/div/div/form/div/div[2]/div/div/div/input").send_keys(Keys.ENTER)
            #======================== INPUT Keterangan ============================
            driver.find_element(By.XPATH, "//div[@id='app']/div/div[2]/div/div[2]/div/div/form/div/div[2]/div[2]/div/div/textarea").send_keys(Keterangan)
            
            #======================== INPUT Nama ============================
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[2]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div/div/div[1]/div/div/input").send_keys(Nama)
            #======================== INPUT Nama ============================
            #======================== INPUT JENIS KELAMIN ============================
            driver.find_element(By.XPATH, "//div[2]/div/div/div/div/input").click()
            time.sleep(1)
            
            if JenisKelamin == 'Perempuan' :
                time.sleep(1)
                dropdown = driver.find_element(By.XPATH, "//span[contains(.,\'Perempuan\')]")
                dropdown.find_element(By.XPATH, "//span[contains(.,\'Perempuan\')]").click()
                time.sleep(1)
            elif JenisKelamin == 'Laki-Laki' :
                time.sleep(1)
                dropdown = driver.find_element(By.XPATH, "//span[contains(.,\'Laki-Laki\')]")
                dropdown.find_element(By.XPATH, "//span[contains(.,\'Laki-Laki\')]").click()

            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[2]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div/div/div[3]/div/div/div/input").send_keys(Keys.DELETE)
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[2]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div/div/div[3]/div/div/div/input").send_keys(Umur)
            time.sleep(2)
            driver.find_element(By.XPATH,  "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[3]/div/div/button[2]").click()

        except TimeoutException:
            print("MASIH ADA ERROR, CEK LAGI PAK WIL")
            pass
        time.sleep(5)
        i = i + 1
    print("DONE PAK WILDAN, SEBATS DULU")
