from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
import os
import pyautogui
import pytest
import time
import platform
import subprocess

from pathlib import Path

@pytest.fixture()
def test_setup():
    global driver
    global wb
    swin = Service(r'C:/Users/user/Documents/TRCH/chromedriver.exe')
    smac = Service('/Users/will/Downloads/chromedriver')

    if platform.system() == 'Darwin':
        driver = webdriver.Chrome(service=smac)
        # url = "http://kumbang.torche.id:32400/"
        url = "http://192.168.2.11:32400/"
        
        driver.get(url)
        # seting windows nya jadi max
        wb = load_workbook(filename=r"/Users/will/Documents/work/Automationpython/Filexel/Registrasi.xlsx")   
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
    elif platform.system() == 'Windows':
        driver = webdriver.Chrome(service=swin)
        # url = "http://kumbang.torche.id:32400/"
        url = "http://192.168.2.11:32400/"
        
        driver.get(url)
        # seting windows nya jadi max   
        wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Filexel\Registrasi.xlsx")
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
        
def test_Meningal_Dunia(test_setup):
    driver.implicitly_wait(10)
    
    sheetrange = wb['MeninggalDunia']
    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div/div[2]/div/div/div[1]/canvas")))
        
    driver.find_element(By.XPATH, "//div/span").click()
    # ini masuk ke form input username
    driver.find_element(By.ID, "username").click()
    # masukin input username
    driver.find_element(By.ID, "username").send_keys("wildan")
    # masukin input password
    driver.find_element(By.ID, "password").send_keys("wildan")
    # click button login
    driver.find_element(By.ID, "kc-login").click()
    time.sleep(1)
    element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[1]/div")
    time.sleep(1)
    actions = ActionChains(driver)
    actions.move_to_element(element).perform()
    time.sleep(1)
    element2 = driver.find_element(By.XPATH, "//div/ul/li[2]/div")
    time.sleep(1)
    actions2 = ActionChains(driver)
    actions2.move_to_element(element2).perform()
    time.sleep(1)
    driver.find_element(By.LINK_TEXT, "Meninggal Dunia").click()
    WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/div[1]/h1")))
    #WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/div[1]/h1")))
       
    i = 2

    # nge baca mulai dari tabel A
    while i <= len(sheetrange['A']):
        # deklarasi bahwa NIP itu ada di A 
        NoInduk                                     = sheetrange['A'+str(i)].value
        NomorSuratMeninggal                         = sheetrange['B'+str(i)].value
        TempatMeninggal                             = sheetrange['C'+str(i)].value
        SebabMeninggal                              = sheetrange['D'+str(i)].value
        TanggalSuratMeninggal                       = sheetrange['E'+str(i)].value
        TanggalMeninggal                            = sheetrange['F'+str(i)].value
        Keterangan                                  = sheetrange['G'+str(i)].value
        NamaDokumen                                 = sheetrange['H'+str(i)].value
        DirFile                                     = sheetrange['I'+str(i)].value
        KeteranganF                                 = sheetrange['J'+str(i)].value
        
        #======================== Halaman Index ============================
    
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/button')))
            
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/button").click()
        
        try:      
            #======================== Halaman Cari ===========================
            WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/button')))
            driver.find_element(By.CSS_SELECTOR, ".el-input__inner").click()
            time.sleep(1)
            driver.find_element(By.CSS_SELECTOR, ".el-input__inner").send_keys(NoInduk)
            time.sleep(3)
            
            driver.find_element(By.CSS_SELECTOR, ".el-select-dropdown__item:nth-child(1) tr:nth-child(2) > .el-descriptions__cell:nth-child(1)").click()
            time.sleep(1)
            #======================== Button Cari ============================
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div[1]/div/form/button[1]").click()
            #======================== Button Cari ============================
            

            driver.find_element(By.CSS_SELECTOR, ".el-col:nth-child(1) .el-input__inner").send_keys(NomorSuratMeninggal)
            driver.find_element(By.CSS_SELECTOR, ".el-form-item:nth-child(2) .el-textarea__inner").send_keys(TempatMeninggal)
            driver.find_element(By.CSS_SELECTOR, ".el-col:nth-child(1) > .el-form-item:nth-child(3) .el-textarea__inner").send_keys(SebabMeninggal)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[2]/div/div/div[2]/div[1]/div/div/input').send_keys(TanggalSuratMeninggal)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[2]/div/div/div[2]/div[1]/div/div/input').send_keys(Keys.ENTER)

            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[2]/div/div/div[2]/div[2]/div/div/input').send_keys(TanggalMeninggal)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[2]/div/div/div[2]/div[2]/div/div/input').send_keys(Keys.ENTER)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[2]/div/div/div[2]/div[3]/div/div/textarea').send_keys(Keterangan)

            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/div[2]/div[2]/button[2]').click()

            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[3]/div/div[2]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[2]/div/div/div/div[1]/input').send_keys(NamaDokumen)
            time.sleep(3)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[3]/div/div[2]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[3]/div/div/div/div/div/div/button').click()
            time.sleep(3)
            pyautogui.write(r"///////users/will/Downloads/pdf/terminologi.pdf")
            time.sleep(2)
            pyautogui.press('enter')
            pyautogui.write(r"///////users/will/Downloads/pdf/terminologi.pdf")
            time.sleep(3)
            pyautogui.press('enter')
            pyautogui.hotkey("backspace")
            pyautogui.write(r"///////users/will/Downloads/pdf/terminologi.pdf")
            pyautogui.press('enter')
            time.sleep(1)
            pyautogui.press('enter')
            time.sleep(3)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[3]/div/div[2]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[4]/div/div/textarea').send_keys(KeteranganF)
            time.sleep(1)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/div[2]/div[2]/button[2]').click()

         
    


        except TimeoutException:
            print("MASIH ADA ERROR, CEK LAGI PAK WIL")
            pass
        time.sleep(5)
        i = i + 1
    print("DONE PAK WILDAN, SEBATS DULU")
