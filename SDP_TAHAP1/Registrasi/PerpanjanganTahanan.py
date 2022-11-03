from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
import os
import pyautogui
import pytest
import time
import platform
import subprocess

from pathlib import Path

@pytest.fixture()
def test_setup():
    global driver
    global wb
    swin = Service(r'C:/Users/user/Documents/TRCH/chromedriver.exe')
    smac = Service('/Users/will/Downloads/chromedriver')
    if platform.system() == 'Darwin':
        driver = webdriver.Chrome(service=smac)
        # url = "http://kumbang.torche.id:32400/"
        url = "http://192.168.2.11:32400/"
        
        driver.get(url)
        # seting windows nya jadi max
        wb = load_workbook(filename=r"/Users/will/Documents/work/Automationpython/Filexel/Registrasi.xlsx")   
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
    elif platform.system() == 'Windows':
        driver = webdriver.Chrome(service=swin)
        # url = "http://kumbang.torche.id:32400/"
        url = "http://192.168.2.11:32400/"
        
        driver.get(url)
        # seting windows nya jadi max   
        wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Filexel\Registrasi.xlsx")
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
        
def test_Web(test_setup):
        # global driver
        # jadi ini bisa read sheet yang dibawah itu yang di excel
        sheetrange = wb ['PerpanjanganTahanan']
        # Menuju login
        driver.find_element(By.XPATH, "//div/span").click()
        driver.find_element(By.ID, "username").click()
        driver.find_element(By.ID, "username").send_keys("rono")
        driver.find_element(By.ID, "password").send_keys("rene")
        # click button login
        driver.find_element(By.ID, "kc-login").click()
        time.sleep(3)
        #Registrasi
        element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[1]/div")                                   
        time.sleep(2)
        actions = ActionChains(driver)
        actions.move_to_element(element).perform()
        #xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx anak ke 2
        element2 = driver.find_element(By.XPATH, "//div/ul/li[2]/div")
        #xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx anak ke 2
        time.sleep(2)
        actions2 = ActionChains(driver)
        actions2.move_to_element(element2).perform()
        #xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx anak ke 2
        element3 = driver.find_element(By.XPATH, "//div[3]/div/ul/li[2]/div[2]/div/ul/li[2]/div[1]")
        #xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx anak ke 2
        time.sleep(2)
        actions2 = ActionChains(driver)
        actions2.move_to_element(element3).perform()
        #Surat Sidang
        driver.find_element(By.LINK_TEXT, "Perpanjangan Tahanan").click()
        time.sleep(1)

        i = 4
        while i <= len(sheetrange['A']):
            noreg               = sheetrange['A'+str(i)].value
            
            Nosrt               = sheetrange['C'+str(i)].value
            tglsrt              = sheetrange['D'+str(i)].value
            Jperpnjngn          = sheetrange['E'+str(i)].value
            tglmulaiperpanjngn  = sheetrange['F'+str(i)].value
            jmlhdoc             = sheetrange['G'+str(i)].value
            dokumen             = sheetrange['H'+str(i)].value
            dokumen1            = sheetrange['I'+str(i)].value
            dokumen2            = sheetrange['J'+str(i)].value
            dokumen3            = sheetrange['K'+str(i)].value
            Keterangan          = sheetrange['L'+str(i)].value
            Keterangan1         = sheetrange['M'+str(i)].value
            Keterangan2         = sheetrange['N'+str(i)].value
            Keterangan3         = sheetrange['O'+str(i)].value

            time.sleep(2) 
            #button +Tambah
            driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/button").click()                                 
            try :
                WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div/div/div/div/input')))
                #======================== Halaman Cari ============================
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div/div/div/div/input').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div/div/div/div/input').send_keys(noreg)
                time.sleep(3)
                WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".el-select-dropdown__item:nth-child(1) tr:nth-child(2) > .el-descriptions__cell:nth-child(1)")))
                driver.find_element(By.CSS_SELECTOR, ".el-select-dropdown__item:nth-child(1) tr:nth-child(2) > .el-descriptions__cell:nth-child(1)").click()
                time.sleep(1)
                #======================== Button Cari ============================
                driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div[1]/div/form/button[1]").click()
                #======================== Button Cari ============================

                WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[1]/div[1]/div[1]/div/div[1]/input')))

                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[1]/div[1]/div[1]/div/div[1]/input').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[1]/div[1]/div[1]/div/div[1]/input').send_keys(Nosrt)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[1]/div[1]/div[2]/div/div[1]/input').click()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[1]/div[1]/div[2]/div/div[1]/input').send_keys(Nosrt)

                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[1]/div[1]/div[2]/div/div[1]/input').click()
                time.sleep(2)
                if Jperpnjngn == 'Perpanjangan Tahanan Kepolisian':
                    driver.find_element(By.XPATH, "//li[contains(.,'Perpanjangan Tahanan Kepolisian')]").click()
                elif Jperpnjngn == 'Perpanjangan Tahanan Kejaksaan':
                    driver.find_element(By.XPATH, "//li[contains(.,'Perpanjangan Tahanan Kejaksaan')]").click()
                elif Jperpnjngn == 'Perpanjangan Tahanan Pengadilan Negeri':
                    driver.find_element(By.XPATH, "//li[contains(.,'Perpanjangan Tahanan Pengadilan Negeri')]").click()
                elif Jperpnjngn == 'Perpanjangan Tahanan Pengadilan Tinggi':
                    driver.find_element(By.XPATH, "//li[contains(.,'Perpanjangan Tahanan Pengadilan Tinggi')]").click()
                elif Jperpnjngn == 'Perpanjangan Tahanan MA':
                    driver.find_element(By.XPATH, "//li[contains(.,'Perpanjangan Tahanan MA')]").click()
                elif Jperpnjngn == 'Perpanjangan Sandera':
                    driver.find_element(By.XPATH, "//li[contains(.,'Perpanjangan Sandera')]").click()
                elif Jperpnjngn == 'Perpanjangan Tahanan Militer':
                    driver.find_element(By.XPATH, "//li[contains(.,'Perpanjangan Tahanan Militer')]").click()
                time.sleep(4)
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[1]/div[2]/div[1]/div/div/input').clear()
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div[2]/div/form/div[1]/div[1]/div[2]/div/div[1]/input').send_keys(Nosrt)

            except TimeoutException:
                    print('ERROR')
                    pass
            time.sleep(4)   
            i = i + 1
        print ("Success Created")