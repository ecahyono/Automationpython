#======================================================================
    #CONTOH CHECKBOX, INPUT NUMBER
#========================I==============================================
#======================================================================
    #CONTOH CAMERA DAN UPLOAD FILE MENYUSUL 
#========================I==============================================
from json import load
from lib2to3.pgen2 import driver
import string
from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.chrome.options import Options

import time 
from openpyxl import load_workbook
import pyautogui



#target halaman excel ada dimana , wb = variablenya
# wb = load_workbook(filename="C:\chromedriver\Data.xlsx")
wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Backup\contohkasus.xlsx")
# jadi ini bisa read sheet yang dibawah itu yang di excel
sheetrange = wb['Identitas']

driver = webdriver.Chrome(chrome_options=options)  


# ini web driver disimpen dimana, kalo disimpen di path kosongin aja
driver = webdriver.Chrome(chrome_options=options, executable_path=r'C:\Users\user\Documents\TRCH\chromedriver.exe')
# driver = webdriver.Chrome(executable_path=r'C:\Users\user\Documents\TRCH\chromedriver.exe')

# link nya ini dimana
driver.get("http://kumbang.torche.id:32400/")
# seting windows nya jadi max   
driver.maximize_window()
# script gakan di eksekusi kalo web ga muncul. kalo lebih dari 10 detik ga muncul error
driver.implicitly_wait(10)
# ini letak xpath icon login
driver.find_element(By.XPATH, "//div/span").click()
# ini masuk ke form input username
driver.find_element(By.ID, "username").click()
# masukin input username
driver.find_element(By.ID, "username").send_keys("rono")
# masukin input password
driver.find_element(By.ID, "password").send_keys("rene")
# click button login
driver.find_element(By.ID, "kc-login").click()
time.sleep(3)

#Registrasi
element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[1]/div")                                   
actions = ActionChains(driver)
actions.move_to_element(element).perform()
time.sleep(1)

#Identitas
driver.find_element(By.LINK_TEXT, "Daftar Identitas").click()
time.sleep(1)

i = 2 

while i <= len(sheetrange['A']):
    # deklarasi per colom pada sheet
    # #--------------------------------------------------------------
    chcktab1                = sheetrange['BN'+str(i)].value
    #--------------------------------------------------------------
    chcktab2                = sheetrange['BL'+str(i)].value
    #--------------------------------------------------------------
    time.sleep(3)
    #button +Tambah
    driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/button").click()                                 

    try:
        WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/div/div/h1')))
        #======================================================================
        driver.find_element(By.ID, "tab-1").click()
        #========================Input Tab Biodata ============================
        cekbocktab1 = driver.find_element(By.ID, "btn_is_pengaruh_terhadap_masyarakat").click()
        if chcktab1 == 'tidak' :
            driver.find_element(By.ID, "btn_is_pengaruh_terhadap_masyarakat").click()
            print ("uncheckh")
        elif chcktab1 == 'ya': 
            print ("Masih default")
        #======================================================================
        driver.find_element(By.ID, "tab-2").click()
        #========================Input Tab Pekerjaan===========================
        #------------------------------------------------------------------------------
        if chcktab2 == 'ya' :
            driver.find_element(By.XPATH, "//*[@id=\"is_baca_quran\"]/span[1]/span").click()
            print (chcktab2)
        elif chcktab2 == 'Tidak':
            print ("tidak di check")
        # #--------------------------------------------------------------
        # #======================================================================
        # driver.find_element(By.ID, "tab-3").click()
        # #========================Input Tab Keluarga============================ 
        # #-----------------------------------------------------------------------
        # MASIH MENCARI UNTUK TRIGER FIELD INPUT ANAK KE -- DARI --
        # #======================================================================
        driver.find_element(By.ID, "tab-4").click()
        # #========================Input Tab Data Fisik========================== 
        # #------------------------------------------------------------------------------
        driver.find_element(By.ID, "tinggi").click()
        pyautogui.hotkey("backspace")
        time.sleep(2)
        driver.find_element(By.XPATH, "//*[@id=\"tinggi\"]/div/input" ).send_keys(Tinggi_badan) 
        #--------------------------------------------------------------
        driver.find_element(By.ID, "berat").click()
        pyautogui.hotkey("backspace")
        time.sleep(2)
        driver.find_element(By.ID, "//*[@id=\"berat\"]/div/input").send_keys(Berat_badan) 
        #--------------------------------------------------------------
        #======================================================================
        driver.find_element(By.ID, "tab-6").click()
        # driver.find_element(By.XPATH, "//*[@id=\"pane-6\"]/form/div/div[1]/div/div/div/button[1]").click()
        
        #Submit
        time.sleep(2)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[4]/div/template/button[2]").click()
        
    except TimeoutException:
        # print("d")
        pass
    time.sleep(5)   
    i = i + 1
print ("Success Created")