from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from pytest_html_reporter import attach
import os, platform, time, pytest
from selenium import webdriver
from os import environ, path
from pathlib import Path
from pytest import mark
import pyautogui
import platform
import logging
import subprocess
import sys
from openpyxl import load_workbook
from dotenv import load_dotenv
load_dotenv()

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR")) 
elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))

from Settings.setupbrowser import initDriver, loadDataPath
from Settings.login import login, bapasbdg
from Settings.Page.bapas import daftarklien

Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Rupbasan_Newrecod.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

# init driver by os
@mark.fixture_penerimaan
def testconfigandlogin():
	Log.info('Konfigurasi agar berjalan di setiap sistem operasi (mac dan Windos)')
	global driver, pathData
	driver = initDriver()
	pathData = loadDataPath()
	Log.info('Memasukan User name dan Password di halaman Login')
	bapasbdg(driver)

@mark.fixture_penerimaan
def test_daftarklien():
    Log.info('Membuka Halaman Penerimaan Klien APH')
    daftarklien(driver)
    driver.find_element(By.ID, 'createButton').click()
    # Create a new workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'PenerimaanAPH'
    # Generate fake data and write it to the worksheet
    fake = Faker('id_ID')
    for i in range(5):
        name = fake.name()
        tgllahir = fake.date_of_birth().strftime('%d/%m/%Y')
        phone = fake.numerify('############')
        worksheet.append([
            name, tgllahir, phone, 
            ])
    # Save the workbook to a file
    file_path = environ.get("BPSEXEL")
    workbook.save(file_path)

    workbook = load_workbook(filename=file_path)
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        name            = row[0]
        tgllahir        = row[1]
        phone           = row[2]

        driver.find_element(By. ID, 'namKlien').send_keys(name)
        driver.find_element(By. ID, 'tanggalLahir').send_keys(tgllahir)
        driver.find_element(By. ID, 'phone').send_keys(phone)





        # driver.find_element_by_css_selector('input[type="submit"]').click()

