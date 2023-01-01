from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from pytest_html_reporter import attach
import os, platform, time, pytest
from selenium import webdriver
from os import environ, path
from pathlib import Path
from pytest import mark
import platform
import logging
from openpyxl import load_workbook
import sys
import pyautogui

from dotenv import load_dotenv
load_dotenv()

if platform.system() == 'Darwin':
	sys.path.append(environ.get("MACPARENTDIR")) 
elif platform.system() == 'Windows':
	sys.path.append(environ.get("WINPARENTDIR"))

from Settings.setup import initDriver, loadDataPath, sleep, hold
from Settings.login import login, oprupbasanbdg


Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Test_Rupbasan_Pemeliharaan.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

wb = load_workbook(environ.get("RUPEXEL"))
sheetrange = wb['Pemeliharan']
i = 4

tglPemeliharaan		= sheetrange['A'+str(i)].value #Tanggal Pemeliharaan
jnspemeliharaan		= sheetrange['B'+str(i)].value #Jenis Pemeliharaan
KgiPemeliharaan		= sheetrange['C'+str(i)].value #Kegiatan Pemeliharaan
Keterangan			= sheetrange['D'+str(i)].value #Keterangan

ketpelaksana	 = sheetrange['E'+str(i)].value #Ketua Pelaksana
ptgsinternal	 = sheetrange['F'+str(i)].value #Petugas Pemeliharaan (Internal)
ptgsexternal	 = sheetrange['G'+str(i)].value #Petugas Pemeliharaan (Eksternal)

#tab Detail Pemeliharan
jumlhdetail		= sheetrange['H'+str(i)].value #jumlah
namabarang		= sheetrange['I'+str(i)].value #Nama Barang
kndbrng			= sheetrange['J'+str(i)].value #Kondisi Barang
subkonbrg		= sheetrange['K'+str(i)].value #Sub Kondisi Barang
jmlhbaik		= sheetrange['L'+str(i)].value #Jumlah Baik
jmlhRpar		= sheetrange['M'+str(i)].value #Jumlah Rusak Parah
jmlhRring		= sheetrange['N'+str(i)].value #Jumlah Rusak Ringan
ketdetail		= sheetrange['O'+str(i)].value #Keterangan

#tab identitas
jmlhbahan		= sheetrange['P'+str(i)].value #jumlah
namabahan		= sheetrange['Q'+str(i)].value #Nama Bahan
jmlhpakai		= sheetrange['R'+str(i)].value #Jumlah Pakai
satuan			= sheetrange['S'+str(i)].value #Satuan
satlain			= sheetrange['T'+str(i)].value #Satuan Lain
ketbahan		= sheetrange['U'+str(i)].value #Keterangan

# init driver by os
@mark.fixture_pemeliharaan
def test_Ossetup_1():
	global driver, pathData
	driver = initDriver()
	pathData = loadDataPath()
	Log.info('Konfigurasi agar berjalan di setiap sistem operasi (mac dan Windos)')

@mark.fixture_pemeliharaan
def test_loggin_2():
	# login(driver)
	oprupbasanbdg(driver)
	Log.info('Memasukan User name dan Password di halaman Login)')

@mark.fixture_pemeliharaan
def test_PMN_001():
	Log.info('Mengakses menu Pengamanan dengan memilih modul Rupbasan kemudian pilih menu Pengamanan')
	hold(driver)
	nav1 = driver.find_element(By.XPATH, pathData['AksesMenu']['Rupbasan']['menu']['Rupbasan'])
	ActionChains(driver).move_to_element(nav1).perform()
	time.sleep(1)
	driver.find_element(By.LINK_TEXT, 'Pengamanan').click()
	driver.find_element(By.ID, 'kataKunci').click()
	attach(data=driver.get_screenshot_as_png())

@mark.fixture_pemeliharaan
def test_PMN_002():
	WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID , 'createButton')))
	Log.info('Menampilkan tanggal sesuai dengan hari saat mengakses dengan menekan button Hari Ini')
	hold(driver)
	pass
wb = load_workbook(environ.get("RUPEXEL"))
sheetrange1 = wb['Barangbasan']
j = 9

Tanggal_Shift = sheetrange1['A'+str(j)].value 
Komandan_Jaga = sheetrange1['B'+str(j)].value 
jenis_barang  = sheetrange1['C'+str(j)].value 
satuan 		  = sheetrange1['D'+str(j)].value 

@mark.fixture_pemeliharaan
def test_PMN_003():
	Log.info('Menambahkan data Shift dengan menekan button Tambah lalu menginputkan data Shift pada form Tambah #kemudian submit data')
	hold(driver)
	driver.find_element(By.ID, 'createButton').click()
	# driver.find_element(By.ID, 'backButton').click()
	# WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID , 'createButton')))

	tgl = driver.find_element(By.ID,'').click()

#Navigasi button
# //*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/div[3]/div[1]/div[3]/button
# //*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/div[3]/div[1]/div[3]/div/button[1]
# //*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/div[3]/div[1]/div[3]/div/button[2]