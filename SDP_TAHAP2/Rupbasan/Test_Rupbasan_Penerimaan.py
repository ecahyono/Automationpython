from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from pytest_html_reporter import attach
import os, platform, time, pytest
from selenium import webdriver
from os import environ, path
from pathlib import Path
from pytest import mark
import pyautogui
import platform
import logging
import subprocess
import sys
from openpyxl import load_workbook
from dotenv import load_dotenv
load_dotenv()

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR")) 
elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))

from Settings.setupbrowser import initDriver, loadDataPath
from Settings.login import login, oprupbasanbdg
from Settings.Page.Rupbasan import Penerimaan, Gudang, SektorGudang

Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Rupbasan_Newrecod.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

wb = load_workbook(environ.get("RUPEXEL"))

# init driver by os
@mark.fixture_penerimaan
def testconfigandlogin():
	Log.info('Konfigurasi agar berjalan di setiap sistem operasi (mac dan Windos)')
	global driver, pathData
	driver = initDriver()
	pathData = loadDataPath()
	Log.info('Memasukan User name dan Password di halaman Login')
	oprupbasanbdg(driver)

@mark.fixture_penerimaan
def test_menmabhapenerimaan():
	tambah = wb['TambahubahPenerimaan']
	for row in tambah.iter_rows(min_row=2, values_only=True):
		JenisRegistrasi    				= row[0] #Jenis Registrasi
		tglPenerimaan      				= row[1] #Tanggal Penerimaan
		instansi           				= row[2] #Instansi
		NomorSuratPengantar           	= row[3] #Nomor Surat Pengantar
		NomorSuratPerintahPenyitaan		= row[4] #Nomor Surat Perintah Penyitaan
		TanggalSuratPerintahPenyitaan   = row[5] #Tanggal Surat Perintah Penyitaan
		PengadilanPenyita           	= row[6] #Pengadilan Penyita
		NomorSuratPenetapanPengadilan	= row[7] #Nomor Surat Penetapan Pengadilan
		tglSuratPenetapanPengadilan     = row[8] #Tanggal Surat Penetapan Pengadilan
		InstitusiPenempatan    			= row[9] #Institusi Penempatan
		Pasal         					= row[10] #Pasal
		KategoriPidana        			= row[11] #Kategori Pidana
		JenisKejahatan           		= row[12] #JenisKejahatan
		NoBAPenyitaan    				= row[13] #NoBAPenyitaan
		NoSuratPernyataan   			= row[14] #No. Surat Pernyataan
		Keterangan					    = row[15] #Keterangan
		Petugaspenerima             	= row[16] #Petugas Penerima
		Namapetugasyangmenyerahkan	    = row[17] #Nama Petugas yang Menyerahkan
		NIPNRPPetugasyangMenyerahkan  	= row[18] #NIP/NRP Petugas yang Menyerahkan
		PangkatGolonganPetugas     		= row[19] #Pangkat/Golongan Petugas yang Menyerahkan
		JabatanPetugas            		= row[20] #Jabatan Petugas yang Menyerahkan
		IdentitasWBP         			= row[21] #Identitas WBP
		Namasaksipenyerah            	= row[22] #nama Petugas Instansi yang Menyerahkan
		NipNrpsaksiinstansipenyerh 		= row[23] #nip Petugas Instansi yang Menyerahkan
		SaksiPenerimaan        			= row[24] #Saksi Penerimaan
		Keterangandkumen      			= row[25] #Keterangan Dokumen

		Log.info('Mengakses Halaman Tambah')
		Penerimaan(driver)
		driver.find_element(By.ID, 'createButton').click()
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'dropdownJenisRegistrasi')))
		Log.info('Memeilih Dropdown Jenis Registrasi')

		driver.find_element(By.ID, 'dropdownJenisRegistrasi').click()
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'jenisRegistrasi0')))
		driver.find_element(By.XPATH, "//li[contains(.,'"+ JenisRegistrasi+"')]").click()

		Log.info('Menginput Tanggal Penerimaan')
		Tanggal_Penerimaan = driver.find_element(By.ID, 'inputTglPenerimaan')  
		Tanggal_Penerimaan.send_keys(tglPenerimaan)
		Tanggal_Penerimaan.send_keys(Keys.ENTER)

		Log.info('Memilih Opsi instansi')
		driver.find_element(By.ID, 'dropdownInstansi').click()  
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'instansi0')))
		driver.find_element(By.XPATH, "//li[contains(.,'"+ instansi+"')]").click()

		Log.info('Menginput Nomor Surat Perintah Penyitaan')
		driver.find_element(By.ID, 'inputNoRegInstansi').send_keys(NomorSuratPengantar)  
		Log.info('Menginput Tgl Surat Perintah Penyitaan')
		driver.find_element(By.ID, 'inputNoSuratIzinPenyitaan').send_keys(NomorSuratPerintahPenyitaan) 

		Log.info('Menginput anggal Surat Perintah Penyitaan')
		Tanggal_Penerimaan = driver.find_element(By.ID, 'inputTglSuratIzinPenyitaan')  
		Tanggal_Penerimaan.send_keys(TanggalSuratPerintahPenyitaan)
		Tanggal_Penerimaan.send_keys(Keys.ENTER)

		Log.info('Memilih Opsi Pengadilan Penyita')
		driver.find_element(By.ID, 'dropdownPengadilanPenyita').click()
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'pengadilanNegeri0')))
		driver.find_element(By.XPATH, "//li[contains(.,'"+ PengadilanPenyita+"')]").click()

		Log.info('menginput nomor surat Penyitaan')
		driver.find_element(By.ID, 'inputNoSuratPenyitaan').send_keys(NomorSuratPenetapanPengadilan)   

		Log.info('menginput tanggal surat Penyitaan')
		Tanggal_Surat_Izin_Penyitaan = driver.find_element(By.ID, 'inputTglSuratPenyitaan')  
		Tanggal_Surat_Izin_Penyitaan.send_keys(tglSuratPenetapanPengadilan)
		Tanggal_Surat_Izin_Penyitaan.send_keys(Keys.ENTER)

		Log.info('Memilih Opsi Institusi Penempatan')
		driver.find_element(By.ID, 'dropdownInstitusiPenempatan').click()
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'institusiPenempatan0')))
		driver.find_element(By.XPATH, "//li[contains(.,'"+ InstitusiPenempatan+"')]").click()

		Log.info('input pasal')
		driver.find_element(By.ID, 'inputPasal').send_keys(Pasal)  

		Log.info('Memilih Opsi Institusi Penempatan')
		driver.find_element(By.ID, 'dropdownKategoriPidana').click()
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'Pidana Umum')))
		driver.find_element(By.XPATH, "//li[contains(.,'"+ KategoriPidana+"')]").click()

		Log.info('Memilih Opsi Institusi Penempatan')
		driver.find_element(By.ID, 'dropdownJenisKejahatan').click()
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'jenisKejahatan0')))
		driver.find_element(By.XPATH, "//li[contains(.,'"+ JenisKejahatan+"')]").click()

		Log.info('input NoBAPenyitaan')
		driver.find_element(By.ID, 'inputNoBaPenyitaan').send_keys(NoBAPenyitaan)  

		Log.info('input No. Surat Pernyataan')
		driver.find_element(By.ID, 'inputNoSuratPernyataan').send_keys(NoSuratPernyataan)  

		Log.info('Input field text area')
		driver.find_element(By.ID, 'inputKeterangan').send_keys(Keterangan)

		Log.info('Melakukan Pencarian data Identitas petugas Penerima')
		try :
			ppenerima = driver.find_element(By.ID, 'searchPetugasPenerima')
			ppenerima.click()
			ppenerima.send_keys(Petugaspenerima)
			WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID, 'searchPetugasPenerima0')))
			driver.find_element(By.XPATH, "//li[contains(.,'"+ Petugaspenerima +"')]").click()
		except NoSuchElementException: 
			Log.info("Element not found, Data petugas belum di tambahkan")
			is_running = True
			while is_running:
				try:
					subprocess.run(["pytest", "Pegawairupbasan.py"], check=True)
					# setelah pytest selesai dijalankan, ulangi proses mencari elemen
					handles = driver.window_handles
					# alihkan fokus kembali ke tab awal
					driver.switch_to.window(handles[0])
					# jalankan script pada tab awal
					# ...
					is_running = False
				except:
					WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'searchPetugasPenerima')))
					ppenerima = driver.find_element(By.ID, 'searchPetugasPenerima')
					ppenerima.click()
					ppenerima.send_keys(Petugaspenerima)
					WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'searchPetugasPenerima0')))
					driver.find_element(By.XPATH, "//li[contains(.,'"+ Petugaspenerima +"')]").click()

		Log.info('Melakukan Pencarian data Petugas yang menyerahkan')
		driver.find_element(By.ID, 'nama_penyerah'	 ).send_keys(Namapetugasyangmenyerahkan) 
		driver.find_element(By.ID, 'nip_penyerah'	 ).send_keys(NIPNRPPetugasyangMenyerahkan) 
		driver.find_element(By.ID, 'pangkat_penyerah').send_keys(PangkatGolonganPetugas) 
		driver.find_element(By.ID, 'jabatan_penyerah').send_keys(JabatanPetugas) 

		Log.info('Melakukan Pencarian data Identitas WBP')
		try : 
			tahanan = driver.find_element(By.ID, 'searchIdentitas-0')
			tahanan.click()
			WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'searchIdentitas-00')))
			driver.find_element(By.XPATH, "//li[contains(.,'"+ IdentitasWBP +"')]").click()
		except NoSuchElementException:
			print("Element not found, Data Identitas Tidak ditemukan")

			
		Log.info('klik tab Petugas Instansi yang Menyerahkan')
		driver.find_element(By.ID, "tab-petugas_instansi").click()
		Log.info('input data  Petugas Instansi Yang menyerahkan')
		driver.find_element(By.ID, 'nama_menyerahkan-0').send_keys(Namasaksipenyerah) 
		driver.find_element(By.ID, 'nip_menyerahkan-0' ).send_keys(NipNrpsaksiinstansipenyerh) 
		
		Log.info('Klik tab Saksi Penerima')
		driver.find_element(By.ID, "tab-saksi_penerimaan").click()
		Log.info('melakukan pencarian data petugas saksi internal')
		try :
			ppenerima = driver.find_element(By.ID, 'searchSaksi-0')
			ppenerima.click()
			ppenerima.send_keys(SaksiPenerimaan)
			WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'searchSaksi-00')))
			driver.find_element(By.XPATH, "//li[contains(.,'"+ SaksiPenerimaan +"')]").click()
		except NoSuchElementException: 
			print("Element not found, Data Petugas belum ditambahkan")

		Log.info('Klik tab Dokument')
		driver.find_element(By.ID, "tab-dokumen_files").click()
		Log.info('melakukan Verifikasi dokumen yang diupload')
		driver.find_element(By.ID, "verifikasi").click()
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'simpanVerifikasi')))
		inputpasswordverif = driver.find_element(By.ID, 'passwordVerifikasi')
		is_filled = inputpasswordverif.get_attribute("value") !=""
		if is_filled :
			pass
		else:
			inputpasswordverif.send_keys(environ.get("SPVLAMPG"))
		driver.find_element(By.ID, "simpanVerifikasi").click()
		WebDriverWait(driver, 50).until_not(EC.presence_of_element_located((By.ID, "simpanVerifikasi")))
		WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'pilihDokumenFile0')))
		driver.find_element(By.ID, "pilihDokumenFile0").click()
		time.sleep(3)
		pyautogui.write(environ.get("FILEPDF"))
		pyautogui.press('enter')
		driver.find_element(By.ID, 'keteranganDokumen-0').send_keys(Keterangandkumen) 

		# Log.info('menekan button submit')
		# driver.find_element(By.ID, "submitButton").click()
		# WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'searchButton')))

#############################################################################################################################
#############################################################################################################################
