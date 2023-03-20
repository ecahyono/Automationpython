import openpyxl
from faker import Faker
from selenium import webdriver
from selenium.webdriver.common.by import By

# Create a Faker object to generate fake data
fake = Faker()

# Create an Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = 'Fake Data'

# Write headers to the worksheet
worksheet.cell(row=1, column=1, value='Name')
worksheet.cell(row=1, column=2, value='Email')
worksheet.cell(row=1, column=3, value='Phone')

# Generate and write fake data to a new worksheet
new_worksheet = workbook.create_sheet('Fake Data2')
for row in range(2, 11):
    name = fake.name()
    email = fake.email()
    phone = fake.phone_number()
    new_worksheet.cell(row=row, column=1, value=name)
    new_worksheet.cell(row=row, column=2, value=email)
    new_worksheet.cell(row=row, column=3, value=phone)


# Create a WebDriver object to automate the browser
driver = webdriver.Chrome()

# Open the Excel file

workbook = openpyxl.load_workbook('fake_data.xlsx', read_only=True)
worksheet = workbook['Fake Data']

# Read data from the worksheet and use it in Selenium
for row in worksheet.iter_rows(min_row=2, max_row=11, values_only=True):
    name = row[0]
    email = row[1]
    phone = row[2]
    
    # Open a website
    driver.get("https://demoqa.com/automation-practice-form")
    driver.maximize_window()
    
    # Find an input element and fill it with the fake name
    element = driver.find_element(By.ID,"firstName")
    element.send_keys(name)

    # Find another input element and fill it with the fake email
    element = driver.find_element(By.ID,"userEmail")
    element.send_keys(email)

    # Find a third input element and fill it with the fake phone number
    element = driver.find_element(By.ID,"userNumber")
    element.send_keys(phone)

    # Find a button and click it
    button = driver.find_element(By. ID, "submit")
    button.click()

# Close the browser



