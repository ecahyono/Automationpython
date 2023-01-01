# import imp
from json import load
import pyautogui
import string
from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
import pyautogui

from openpyxl import load_workbook
import time

#target halaman excel ada dimana , wb = variablenya
# wb = load_workbook(filename="C:\chromedriver\Data.xlsx")
wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Filexel\Rupbasan.xlsx")

# jadi ini bisa read sheet yang dibawah itu yang di excel
sheetrange = wb['Gudang']

# ini web driver disimpen dimana, kalo disimpen di path kosongin aja
driver = webdriver.Chrome(r'C:\Users\user\Documents\TRCH\chromedriver.exe')

# link nya ini dimana
# driver.get("http://192.168.2.11:32400/")
driver.get("http://kumbang.torche.id:32400/")
# seting windows nya jadi max   
driver.maximize_window()
# script gakan di eksekusi kalo web ga muncul. kalo lebih dari 10 detik ga muncul error
driver.implicitly_wait(6)
# ini letak xpath icon login
driver.find_element(By.XPATH, "//div/span").click()
# ini masuk ke form input username
driver.find_element(By.ID, "username").click()
# masukin input username
driver.find_element(By.ID, "username").send_keys("oprupbasanbdg")
# masukin input password
driver.find_element(By.ID, "password").send_keys("password")
# click button login
driver.find_element(By.ID, "kc-login").click()
time.sleep(3)

element = driver.find_element(By.XPATH, '//*[@id="app"]/div/nav/ul/li[2]/div')
# element = driver.find_element(By.XPATH, '//*[@id="app"]/div/nav/ul/li[3]/div')
# element = driver.find_element(By.XPATH, '//*[@id="app"]/div/nav/ul/li[4]/div')

actions = ActionChains(driver)
time.sleep(2)
actions.move_to_element(element).perform()

element2 = driver.find_element(By.XPATH, '//div[3]/div/ul/li[1]/div[1]')
# element2 = driver.find_element(By.XPATH, '//div[4]/div/ul/li[1]/div[1]')
# element2 = driver.find_element(By.XPATH, '//div[5]/div/ul/li[1]/div[1]')
# element2 = driver.find_element(By.XPATH, '//div[6]/div/ul/li[1]/div[1]')
actions2 = ActionChains(driver)
time.sleep(2)
actions2.move_to_element(element2).perform()

driver.find_element(By.LINK_TEXT, "Gudang").click()

i = 5 
while i <= len(sheetrange['A']):
	JGudang 		= sheetrange['A'+str(i)].value
	Alm 			= sheetrange['B'+str(i)].value
	Prov			= sheetrange['C'+str(i)].value
	Kotkab		  	= sheetrange['D'+str(i)].value
	luas 			= sheetrange['E'+str(i)].value
	quant 			= sheetrange['F'+str(i)].value
	ket 			= sheetrange['G'+str(i)].value
	Nama			= sheetrange['H'+str(i)].value
	jmlh			= sheetrange['I'+str(i)].value
	rangan		  	= sheetrange['J'+str(i)].value

	driver.find_element(By.ID, "createButton").click()

	try:
		gd = driver.find_element(By.ID, 'jenisGudang')
		gd.click()
		if JGudang == 'Gudang Umum Terbuka':
			driver.find_element(By.ID,'JG01').click()
		elif JGudang == 'Gudang Umum Tertutup':
			driver.find_element(By.ID,'JG02').click()
		elif JGudang == 'Gudang Berharga':
			driver.find_element(By.ID,'JG03').click()
		elif JGudang == 'Gudang Berbahaya':
			driver.find_element(By.ID,'JG04').click()
		elif JGudang == 'Gudang Hewan dan Tumbuhan':
			driver.find_element(By.ID,'JG05').click()

		driver.find_element(By.ID, 'alamat').send_keys(Alm)

		vinsi = driver.find_element(By.ID,'provinsi')
		vinsi.send_keys(Prov)
		if Prov == 'Jawa Barat':
			driver.find_element(By.ID,'12').click()
		elif Prov == 'DKI Jakarta':
			driver.find_element(By.ID,'11').click()

		paten = driver.find_element(By.ID,'kotaKabupaten')
		paten.send_keys(Kotkab)
		if Kotkab == 'Bandung':
			driver.find_element(By.ID,'122').click()
		elif Kotkab == 'Jakarta Pusat':
			driver.find_element(By.ID,'116').click()
				
		driver.find_element(By.ID, 'luas').send_keys(luas)
		driver.find_element(By.ID, 'kapasitasGudang').send_keys(quant)

		foto = driver.find_element(By.ID, 'pilihFoto').click()
		time.sleep(3)
		pyautogui.typewrite(r'C:\Users\user\Documents\TRCH\Automationpython\assets\Filefoto\Gudang.png')
		pyautogui.press('enter')

		driver.find_element(By.ID, 'keterangan').send_keys(ket)

		driver.find_element(By.ID, 'nama0').send_keys(Nama)

		driver.find_element(By.ID, 'jumlah0').send_keys(jmlh)

		driver.find_element(By.ID, 'keterangan0').send_keys(rangan)
				
		driver.find_element(By.ID, 'submitButton').click()
		WebDriverWait(driver, 90).until(EC.element_to_be_clickable((By.XPATH, 'searchButton')))

	except TimeoutException:
		pass
	i = i + 1
print ("Success Created")

# 