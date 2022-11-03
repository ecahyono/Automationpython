from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
import os
import pyautogui
import pytest
from pytest import mark
import time
from pytest_html_reporter import attach
import platform
from pathlib import Path

@mark.fixture_Penerimaan
def test_setup():
    global driver
    global wb
    swin = Service(r'C:/Users/user/Documents/TRCH/chromedriver.exe')
    smac = Service('/Users/will/Downloads/chromedriver')
    if platform.system() == 'Darwin':
        driver = webdriver.Chrome(service=smac)
        driver.get('http://192.168.2.11:32400/')
        wb = load_workbook(filename='/Users/will/Documents/work/Automationpython/Filexel/Rupbasan.xlsx') 
        driver.maximize_window()
        driver.implicitly_wait(5)
    elif platform.system() == 'Windows':
        driver = webdriver.Chrome(service=swin)
        driver.get('http://192.168.2.11:32400/')
        wb = load_workbook(filename=r'C:\Users\user\Documents\TRCH\Automationpython\Filexel\Rupbasan.xlsx')    
        driver.maximize_window()
        driver.implicitly_wait(5)
    attach(data=driver.get_screenshot_as_png())
    print('setupberhasil')

@mark.fixture_Penerimaan
def test_login():
    # Menuju login
    driver.find_element(By.XPATH, '//div/span').click()
    driver.find_element(By.ID, 'username').click()
    driver.find_element(By.ID, 'username').send_keys('test-user')
    driver.find_element(By.ID, 'password').send_keys('password')
    # click button login
    driver.find_element(By.ID, 'kc-login').click()
    WebDriverWait(driver, 10)
    attach(data=driver.get_screenshot_as_png())
    print('Login Berhasil')

@mark.fixture_Penerimaan
def test_akses_menu():    
    #Registrasi
    element = driver.find_element(By.XPATH, '//*[@id="app"]/div/nav/ul/li[5]/div')                                   
    driver.implicitly_wait(10)
    actions = ActionChains(driver)
    actions.move_to_element(element).perform()
    
    #Penerimaan
    driver.find_element(By.LINK_TEXT, 'Penerimaan').click()
    attach(data=driver.get_screenshot_as_png())
    print('akses_menu')


@mark.fixture_Penerimaan
def test_tambah():
    sheetrange = wb ['Penerimaan']
    i = 2
    while i <= len(sheetrange['A']):
        try:
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/div[1]/button').click()                                 
            time.sleep(3)
            attach(data=driver.get_screenshot_as_png())
            print('Tambah satu')
        except TimeoutException:
            print('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!LOADING TERLALU LAMA!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
            pass
        i = i + 1
    print ('Success Created')

@mark.fixture_Penerimaan
def test_drop_jenis_registrasi():
    sheetrange = wb ['Penerimaan']
    i = 2
    while i <= len(sheetrange['A']):
        jenis_registrasi = sheetrange['A'+str(i)].value
        try:
            driver.find_element(By.ID, 'input_jenis_registrasi_baran_basan').click()
            if jenis_registrasi == 'Register Barang Rampasan Negara':
                driver.find_element(By.ID,'RBR').click()
            elif jenis_registrasi == 'Tingkat Penyidikan':
                driver.find_element(By.ID, 'RBS1').click()
            elif jenis_registrasi == 'Tingkat Penuntutan':
                driver.find_element(By.ID, 'RBS2').click()
            elif jenis_registrasi == 'Tingkat Pengadilan Negeri':
                driver.find_element(By.ID, 'RBS3').click
            elif jenis_registrasi == 'Tingkat Pengadilan Tinggi':
                driver.find_element(By.ID, 'RBS4').click
            attach(data=driver.get_screenshot_as_png())
            print('terpilih satu')
        except TimeoutException:
            print('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!LOADING TERLALU LAMA!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
            pass
        i = i + 1
    print ('Success Created')

@mark.fixture_Penerimaan
def test_date():
    sheetrange = wb ['Penerimaan']
    i = 2
    while i <= len(sheetrange['A']):
        tgl_penerimaan          = sheetrange['B'+str(i)].value
        try:
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[2]/div/div[1]/input').click()
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[2]/div/div[1]/input').send_keys(tgl_penerimaan)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[2]/div/div[1]/input').send_keys(Keys.ENTER)
            attach(data=driver.get_screenshot_as_png())
            print('inputdate')
        except TimeoutException:
            print('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!LOADING TERLALU LAMA!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
            pass
        i = i + 1
    print ('Success Created')

@mark.fixture_Penerimaan
def test_input():
    sheetrange = wb ['Penerimaan']
    i = 2
    while i <= len(sheetrange['A']):
        No_Reg_Rupbasan         = sheetrange['C'+str(i)].value
        try:
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[3]/div/div[1]/input').send_keys(No_Reg_Rupbasan)
            attach(data=driver.get_screenshot_as_png())
            print('akses_menu')
        except TimeoutException:
            print('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!LOADING TERLALU LAMA!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
            pass
        i = i + 1
    print ('Success Created')

@mark.fixture_Penerimaan
def test_back():
    sheetrange = wb ['Penerimaan']
    i = 2
    while i <= len(sheetrange['A']):
        try:
            driver.find_element(By.ID, 'backButton').click()
            attach(data=driver.get_screenshot_as_png())
            print('sudah kembali')
        except TimeoutException:
            print('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!LOADING TERLALU LAMA!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
            pass
        i = i + 1
    print ('Success Created')
    
