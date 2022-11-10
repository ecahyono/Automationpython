from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from pytest_html_reporter import attach
import os
import pyautogui
import pytest
import time
import platform

@pytest.fixture()
def test_setup():
    global driver
    global wb
    swin = Service(r'C:/Users/user/Documents/TRCH/chromedriver.exe')
    smac = Service('/Users/will/Downloads/chromedriver')
    if platform.system() == 'Darwin':
        driver = webdriver.Chrome(service=smac)
        driver.get('http://192.168.2.11:32400/')
        wb = load_workbook(filename='/Users/will/Documents/work/Automationpython/Filexel/Rupbasan.xlsx') 
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
    elif platform.system() == 'Windows':
        driver = webdriver.Chrome(service=swin)
        driver.get('http://192.168.2.11:32400/')
        wb = load_workbook(filename=r'C:\Users\user\Documents\TRCH\Automationpython\Filexel\Rupbasan.xlsx')    
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
    attach(data=driver.get_screenshot_as_png())
    print('setupberhasil')

def test_akses(test_setup):
    # Menuju login
    driver.find_element(By.XPATH, '//div/span').click()
    driver.find_element(By.ID, 'username').click()
    driver.find_element(By.ID, 'username').send_keys('test-user')
    driver.find_element(By.ID, 'password').send_keys('password')
    # click button login
    driver.find_element(By.ID, 'kc-login').click()
    WebDriverWait(driver, 10)
    attach(data=driver.get_screenshot_as_png())
    print('Login Berhasil')

    #Ruobasan
    element = driver.find_element(By.XPATH, '//*[@id="app"]/div/nav/ul/li[5]/div')                                   
    ActionChains(driver).move_to_element(element).perform()
    
    #Penerimaan
    driver.find_element(By.LINK_TEXT, 'Penerimaan').click()
    
    attach(data=driver.get_screenshot_as_png())
    print('akses_menu')

    sheetrange = wb ['Penerimaan']
    i = 2
    while i <= len(sheetrange['A']):
        
        jenis_registrasi           = sheetrange['A'+str(i)].value
        tgl_penerimaan             = sheetrange['B'+str(i)].value
        No_Reg_Rupbasan            = sheetrange['C'+str(i)].value
        instansi                   = sheetrange['D'+str(i)].value
        No_Reg_instansi            = sheetrange['E'+str(i)].value
        Noizin_penyitaan           = sheetrange['F'+str(i)].value
        Tg_surat                   = sheetrange['G'+str(i)].value
        Pengadilan_Penyita         = sheetrange['H'+str(i)].value
        NoS_enyitaan               = sheetrange['I'+str(i)].value
        Tgl_S_Penyitaan            = sheetrange['J'+str(i)].value
        Pasal                      = sheetrange['K'+str(i)].value
        No_BA_sertrima             = sheetrange['L'+str(i)].value
        Keterangan                 = sheetrange['M'+str(i)].value
        Petugas_Penerima           = sheetrange['N'+str(i)].value
        Petugas_Menyerahkan        = sheetrange['O'+str(i)].value
        jmlhidentitas              = sheetrange['P'+str(i)].value
        TersangkaatauPemilik       = sheetrange['Q'+str(i)].value
        NoRegistrasi               = sheetrange['R'+str(i)].value
        Nama_Lengkap               = sheetrange['S'+str(i)].value
        No_KTP                     = sheetrange['T'+str(i)].value
        Jns_kelamin                = sheetrange['U'+str(i)].value
        Tanggal_Lahir              = sheetrange['V'+str(i)].value
        Alamat                     = sheetrange['W'+str(i)].value
        No_Tlpon                   = sheetrange['X'+str(i)].value
        Keterangan_identitas       = sheetrange['Y'+str(i)].value
        jmlhinstansimenyerahkan    = sheetrange['Z'+str(i)].value
        statuspetugas              = sheetrange['AA'+str(i)].value
        saksipenyerah              = sheetrange['AB'+str(i)].value
        jmlhSaksi_Penerimaan       = sheetrange['AC'+str(i)].value
        Status_Saksi               = sheetrange['AD'+str(i)].value
        Saksi_Penerimaan           = sheetrange['AE'+str(i)].value

        #tambah
        driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/div[1]/button').click()                                 
        
        try:
            WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/div/div/h1')))
            #Jenis Registrasi
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[1]/div/div/div/div').click()
            if jenis_registrasi == 'Register Barang Rampasan Negara':
                driver.find_element(By.ID, 'RBR').click()
            elif jenis_registrasi == 'Tingkat Penyidikan':
                driver.find_element(By.ID, 'RBS1').click()
            elif jenis_registrasi == 'Tingkat Penuntutan':
                driver.find_element(By.ID, 'RBS2').click()
            elif jenis_registrasi == 'Tingkat Pengadilan Negeri':
                driver.find_element(By.ID, 'RBS3').click()
            elif jenis_registrasi == 'Tingkat Pengadilan Tinggi':
                driver.find_element(By.ID, 'RBS4').click()
            elif jenis_registrasi == 'Tingkat Mahkamah Agung':
                driver.find_element(By.ID, 'RBS5').click()
            elif jenis_registrasi == 'Register Khusus Tingkat Penyidikan':
                driver.find_element(By.ID, 'RBSK1').click()
            elif jenis_registrasi == 'Register Khusus Tingkat Penuntutan':
                driver.find_element(By.ID, 'RBSK2').click()
            elif jenis_registrasi == 'Register Khusus Tingkat Pengadilan Negeri':
                driver.find_element(By.ID, 'RBSK3').click()
            elif jenis_registrasi == 'Register Khusus Tingkat Pengadilan Tinggi':
                driver.find_element(By.ID, 'RBSK4').click()
            elif jenis_registrasi == 'Register Khusus Tingkat Mahkamah Agung':
                driver.find_element(By.ID, 'RBSK5').click()
            
            #Tanggal Penerimaan
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[2]/div/div[1]/input').click()
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[2]/div/div[1]/input').send_keys(tgl_penerimaan)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[2]/div/div[1]/input').send_keys(Keys.ENTER)

            #Nomor Registrasi Rupbasan
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[3]/div/div[1]/input').send_keys(No_Reg_Rupbasan)

            #Instansi
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[4]/div/div/div/div/input').click()
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[4]/div/div/div/div/input').send_keys(instansi)
            if instansi == 'POLDA JABAR':
                driver.find_element(By.ID, '201507090003').click()
            elif instansi == 'POLRES BANDUNG':
                driver.find_element(By.ID, '201507090007').click()
            elif instansi == 'POLRES CIMAHI':
                driver.find_element(By.ID, '201507090014').click()
            elif instansi == 'POLRES CIREBON':
                driver.find_element(By.ID, '201507090022').click()
            elif instansi == 'POLRES KUNINGAN':
                driver.find_element(By.ID, '201507090025').click()
            
            #Nomor Registrasi Instansi
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[5]/div/div/input').send_keys(No_Reg_instansi)

            #Nomor Surat Izin Penyitaan
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[6]/div/div[1]/input').send_keys(Noizin_penyitaan)

            #Tanggal Surat Izin Penyitaan
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[7]/div/div[1]/input').send_keys(Tg_surat)
            
            #Pengadilan Penyita
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[1]/div/div/div/div/input').click()
            if Pengadilan_Penyita == 'Pengadilan Negeri Bandung':
                driver.find_element(By.ID, '10').click()
            elif Pengadilan_Penyita == 'Pengadilan Negeri Jakarta Utara':
                driver.find_element(By.ID, '4').click()
            elif Pengadilan_Penyita == 'Pengadilan Negeri Sukabumi':
                driver.find_element(By.ID, '13').click()
            elif Pengadilan_Penyita == 'Pengadilan Negeri Subang':
                driver.find_element(By.ID, '27').click()
            elif Pengadilan_Penyita == 'Pengadilan Negeri Yogyakarta':
                driver.find_element(By.ID, '156').click()

            #Nomor Surat Penyitaan
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[2]/div/div/input').send_keys(NoS_enyitaan)

            #Tanggal Surat Penyitaan
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[3]/div/div/input').click()
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[3]/div/div/input').send_keys(Tgl_S_Penyitaan)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[3]/div/div/input').send_keys(Keys.ENTER)

            #Pasal
            driver.find_element(By. XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[4]/div/div[1]/input').send_keys(Pasal)
            
            #No. BA Serah Terima
            driver.find_element(By. XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div[5]/div/div/input').send_keys(No_BA_sertrima)

            # Keterangan
            driver.find_element(By. XPATH, '//textarea').send_keys(Keterangan)

            #Petugas Penerima
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[2]/div/form/div/div/div/div/div/input').click()
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[2]/div/form/div/div/div/div/div/input').send_keys(Petugas_Penerima)
            WebDriverWait(driver, 10)
            driver.find_element(By.CSS_SELECTOR, ".el-select-dropdown__item:nth-child(1) tr:nth-child(2) > .el-descriptions__cell:nth-child(1)").click()

            #Petugas yang Menyerahkan (Jaksa, Polisi, PPNS, Pengadilan)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[3]/div[1]/form/div/div/div/div/div/input').click()
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[3]/div[1]/form/div/div/div/div/div/input').send_keys(Petugas_Menyerahkan)
            WebDriverWait(driver, 10)
            # driver.find_element(By.XPATH, '//td[contains(.,"Galih")]').click()
            driver.find_element(By.CSS_SELECTOR, ".el-select-dropdown__item:nth-child(1) tr:nth-child(2) > .el-descriptions__cell:nth-child(1)").click()
            
            #Tab Identitas
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[5]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[1]/div/div/div/div/input').click()
            WebDriverWait(driver, 10)
            if TersangkaatauPemilik == 'Tersangka':
                driver.find_element(By.XPATH, '//div[19]/div/div/div[1]/ul/li[1]').click()
            elif TersangkaatauPemilik == 'Pemilik':
                driver.find_element(By.XPATH, '//div[19]/div/div/div[1]/ul/li[2]').click()

            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[5]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[2]/div/div/input').send_keys(NoRegistrasi)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[5]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[3]/div/div[1]/input').send_keys(Nama_Lengkap)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[5]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[4]/div/div/input').send_keys(No_KTP)

            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[5]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[5]/div/div/div/div/input').click()
            if Jns_kelamin == 'Laki-laki':
                driver.find_element(By.ID, 'l').click()
            elif Jns_kelamin == 'Perempuan':
                driver.find_element(By.ID, 'p').click()    

            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[5]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[6]/div/div/input').click()
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[5]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[6]/div/div/input').send_keys(Tanggal_Lahir)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[5]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[6]/div/div/input').send_keys(Keys.ENTER)

            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[5]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[7]/div/div/textarea').send_keys(Alamat)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[5]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[8]/div/div/input').send_keys(No_Tlpon)
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[5]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div[9]/div/div/textarea').send_keys(Keterangan_identitas)

            #ganti tab ke 2
            driver.find_element(By.ID, 'tab-petugas_instansi').click()
            if statuspetugas == 'Internal':
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[6]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div/div/label[1]').click()
            elif statuspetugas == 'External':
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[6]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div/div/label[2]').click()    

            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[6]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div[2]/div[1]/div/form/div/div/div/div/div/input').send_keys(saksipenyerah)
            WebDriverWait(driver, 10)
            driver.find_element(By.CSS_SELECTOR, ".el-popper:nth-child(22) .el-select-dropdown__item:nth-child(1)").click()#untuk external
            
            #ganti tab ke 3
            driver.find_element(By.ID, 'tab-saksi_penerimaan').click()
            if Status_Saksi == 'Internal':
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[7]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div/div/label[1]').click()
            elif Status_Saksi == 'External':
                driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[7]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div/div/div/label[2]').click()

            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div[1]/div[2]/div/div/form/div[7]/div[1]/div[3]/div/div[1]/div/table/tbody/tr/td[1]/div/div[2]/div[1]/div/form/div/div/div/div/div/input').send_keys(Saksi_Penerimaan)
            WebDriverWait(driver, 10)
            driver.find_element(By.XPATH, '//*[@id="el-popper-container-284"]/div[22]/div/div/div[1]/ul/li/div').click()

            driver.find_element(By.ID, 'submitButton').click()

        except TimeoutException:
            print('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!LOADING TERLALU LAMA!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
            pass
        i = i + 1
    print ('Success Created')
