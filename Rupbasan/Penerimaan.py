from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
import os
import pyautogui
import pytest
import time
import platform
from pathlib import Path

@pytest.fixture()
def test_setup():
    global driver
    global wb
    swin = Service(r'C:/Users/user/Documents/TRCH/chromedriver.exe')
    smac = Service('/Users/will/Downloads/chromedriver')
    if platform.system() == 'Darwin':
        driver = webdriver.Chrome(service=smac)
        # url = 'http://kumbang.torche.id:32400/'
        url = 'http://192.168.2.11:32400/'
        
        driver.get(url)
        # seting windows nya jadi max
        wb = load_workbook(filename='/Users/will/Documents/work/Automationpython/Filexel/Rupbasan.xlsx')   
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
    elif platform.system() == 'Windows':
        driver = webdriver.Chrome(service=swin)
        # url = 'http://kumbang.torche.id:32400/'
        url = 'http://192.168.2.11:32400/'
        
        driver.get(url)
        # seting windows nya jadi max   
        wb = load_workbook(filename=r'C:\Users\user\Documents\TRCH\Automationpython\Filexel\Rupbasan.xlsx')
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
def test_Web(test_setup):
    # read excel
    sheetrange = wb ['Penerimaan']
    # Menuju login
    driver.find_element(By.XPATH, '//div/span').click()
    driver.find_element(By.ID, 'username').click()
    driver.find_element(By.ID, 'username').send_keys('waru')
    driver.find_element(By.ID, 'password').send_keys('waru')
    # click button login
    driver.find_element(By.ID, 'kc-login').click()
    time.sleep(3) 
    #Registrasi
    element = driver.find_element(By.XPATH, '//*[@id="app"]/div/nav/ul/li[5]/div')                                   
    time.sleep(1)
    actions = ActionChains(driver)
    time.sleep(1)
    actions.move_to_element(element).perform()
    time.sleep(2)
    #Identitas
    driver.find_element(By.LINK_TEXT, 'Penerimaan').click()
    time.sleep(1)

    i = 2
    while i <= len(sheetrange['A']):
        #deklarasi per colom pada sheet
        #------------------------------------------------------
        #Tab Biodata-------------------------------------------
        #------------------------------------------------------
        Residivis               = sheetrange['A'+str(i)].value
        Rke                     = sheetrange['B'+str(i)].value
        Nama_Lengkap            = sheetrange['C'+str(i)].value
        Nama_Alias1             = sheetrange['D'+str(i)].value
        Nama_Alias2             = sheetrange['E'+str(i)].value
        Nama_Alias3             = sheetrange['F'+str(i)].value
        Nama_Kecil1             = sheetrange['G'+str(i)].value
        Nama_Kecil2             = sheetrange['H'+str(i)].value
        Nama_Kecil3             = sheetrange['I'+str(i)].value
        chcktab1                = sheetrange['J'+str(i)].value
        chcktab2                = sheetrange['K'+str(i)].value
        Kewarganegaraan         = sheetrange['L'+str(i)].value
        nik                     = sheetrange['M'+str(i)].value
        Tempat_Asal             = sheetrange['N'+str(i)].value
        Tempat_lahir            = sheetrange['O'+str(i)].value
        Tanggal_lahir           = sheetrange['P'+str(i)].value
        Jenis_kelamin           = sheetrange['Q'+str(i)].value
        Negara                  = sheetrange['R'+str(i)].value
        Agama                   = sheetrange['S'+str(i)].value
        Agama_lain              = sheetrange['T'+str(i)].value
        suku                    = sheetrange['U'+str(i)].value
        Status_perkawinan       = sheetrange['v'+str(i)].value
        Provinsi                = sheetrange['W'+str(i)].value
        Kota                    = sheetrange['X'+str(i)].value
        Alamat_rumah            = sheetrange['Y'+str(i)].value
        Telepon                 = sheetrange['Z'+str(i)].value
        Kode_pos                = sheetrange['AA'+str(i)].value
        Alamat_lain             = sheetrange['AB'+str(i)].value
        #-------------------------------------------------------
        #Tab Pekerjaan------------------------------------------
        #-------------------------------------------------------
        Jenis_Pekerjaan         = sheetrange['AC'+str(i)].value
        namaipemerintah         = sheetrange['AD'+str(i)].value
        noindpegawai            = sheetrange['AE'+str(i)].value
        Bekerjadi               = sheetrange['AF'+str(i)].value
        Keterangan_pekerjaan    = sheetrange['AG'+str(i)].value
        Tingkat_penghasilan     = sheetrange['AH'+str(i)].value
        Tingkat_pendidikan      = sheetrange['AI'+str(i)].value
        Keahlian1               = sheetrange['AJ'+str(i)].value
        Level_keahlian1         = sheetrange['AK'+str(i)].value
        Keahlian2               = sheetrange['AL'+str(i)].value
        Levelkeahlian2          = sheetrange['AM'+str(i)].value
        Minat                   = sheetrange['AN'+str(i)].value
        latin                   = sheetrange['AO'+str(i)].value
        quran                   = sheetrange['AP'+str(i)].value
        Jenis_Pekerjaan_Lain    = sheetrange['CG'+str(i)].value
        #--------------------------------------------------------
        #Tab Keluarga--------------------------------------------
        #--------------------------------------------------------
        Nama_ayah               = sheetrange['AQ'+str(i)].value
        Alamat_ayah             = sheetrange['AR'+str(i)].value
        Nama_ibu                = sheetrange['AS'+str(i)].value
        Alamat_ibu              = sheetrange['AT'+str(i)].value
        Anak_ke                 = sheetrange['AU'+str(i)].value
        Dari                    = sheetrange['AV'+str(i)].value
        Nama_saudara1           = sheetrange['AW'+str(i)].value
        Nama_saudara2           = sheetrange['AX'+str(i)].value
        Nama_saudara3           = sheetrange['AY'+str(i)].value
        Nama_saudara4           = sheetrange['AZ'+str(i)].value
        jml_istrsuam            = sheetrange['BA'+str(i)].value
        Nm_istrsuam             = sheetrange['BB'+str(i)].value
        alm_istrsuam            = sheetrange['BC'+str(i)].value
        jumlah_anak             = sheetrange['BD'+str(i)].value
        Nama_anak1              = sheetrange['BE'+str(i)].value
        Nama_anak2              = sheetrange['BF'+str(i)].value
        Nama_anak3              = sheetrange['CH'+str(i)].value
        Telepon_keluarga        = sheetrange['BG'+str(i)].value
        #--------------------------------------------------------
        #Tab Data Fisik------------------------------------------
        # -------------------------------------------------------
        Tinggi_badan            = sheetrange['BH'+str(i)].value
        Berat_badan             = sheetrange['BI'+str(i)].value
        Bentuk_rambut           = sheetrange['BJ'+str(i)].value
        Warna_rambut            = sheetrange['BK'+str(i)].value
        Bentuk_bibir            = sheetrange['BL'+str(i)].value
        Berkacamata             = sheetrange['BM'+str(i)].value
        Bentuk_mata             = sheetrange['BN'+str(i)].value
        Warna_mata              = sheetrange['BO'+str(i)].value
        Hidung                  = sheetrange['BP'+str(i)].value
        Raut_muka               = sheetrange['BQ'+str(i)].value
        Telinga                 = sheetrange['BR'+str(i)].value
        Mulut                   = sheetrange['BS'+str(i)].value
        Lengan                  = sheetrange['BT'+str(i)].value
        Tangan                  = sheetrange['BU'+str(i)].value
        Kaki                    = sheetrange['BV'+str(i)].value
        Warna_kulit             = sheetrange['BW'+str(i)].value
        Cacat_tubuh             = sheetrange['BX'+str(i)].value
        Catatancirikhusus1      = sheetrange['BY'+str(i)].value
        Catatancirikhusus2      = sheetrange['BZ'+str(i)].value
        Catatancirikhusus3      = sheetrange['CA'+str(i)].value
        #---------------------------------------------------------
        #Tab Sidik Jari-------------------------------------------
        # --------------------------------------------------------
        Nopaspor                = sheetrange['CB'+str(i)].value
        Rumus                   = sheetrange['CC'+str(i)].value
        Nodaktolskopi           = sheetrange['CD'+str(i)].value
        Pengambilansidikjari    = sheetrange['CE'+str(i)].value
        Tanggalpengambilan      = sheetrange['CF'+str(i)].value

        time.sleep(3) 
        #button +Tambah
        driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/div[1]/button').click()                                 
        time.sleep(3) 
        try:
            WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div')))
            #======================================================================
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[1]/div/div/div/div/input').click()
            driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div[1]/div/div/div/div/input').send_keys('contohkasus.xlsx')

            break
            
        except TimeoutException:
            print('ERROR')
            pass
        time.sleep(2)   
        i = i + 1
    print ('Success Created')
